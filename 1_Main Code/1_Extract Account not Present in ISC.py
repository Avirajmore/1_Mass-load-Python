import os
import re
import pyperclip
import pandas as pd
import openpyxl
import sys


folder_path = os.path.expanduser("~/Downloads")

account_present = []

file_paths = []

if os.path.isdir(folder_path):
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path) and file.endswith(('.xlsx', '.xls')):
            file_paths.append(file_path)
else:
    print("Invalid folder path.")


print(f"\n{'='*120}\n{' ' * 30} 📝  Checking for Accounts not present in ISC 📝 {' ' * 30}\n{'='*120}\n")

print("\n📝 Files To be Processed")
for file in file_paths:
    shortened_path = file.split("/")[-1]
    print(f'\n    ✅ {shortened_path}')

print(f"\n{'='*120}")
files_processed = []

# ======================================================================
# STEP 1:- Select a file to Process
# ======================================================================

for file_path in file_paths:
    print("\n🔍 Step 1: Selecting a file")
    files_processed.append(file_path)
    # Print the selected file path
    if file_path:
        filename = str(file_path.split('/')[-1])
        print(f"\n    ✅ File selected: '{filename}'.")
    else:
        print("\n    ❌ No file selected. Exiting the program. ❌")
        sys.exit()

    # ======================================================================
    # STEP 2:- Change the headers to Lower case
    # ======================================================================

    import pandas as pd

    print("\n\n🔍 Step 2: Change the headers to Lower Case")

    # Read the Excel file with all sheets, initially treating all data as strings
    xls = pd.ExcelFile(file_path)

    # Dictionary to hold modified dataframes
    sheets_dict = {}

    # Columns that should remain numeric
    numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

    # Start processing each sheet

    for sheet_name in xls.sheet_names:

        # Read each sheet into a dataframe with all columns as strings
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        
        # Convert headers to lowercase
        df.columns = [col.lower() for col in df.columns]
        
        # Convert specific columns back to numeric types
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')  # Convert to numeric, setting errors to NaN
        
        # Save modified dataframe to dictionary
        sheets_dict[sheet_name] = df

    # Write the modified dataframes back to the Excel file

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"\n    ✅ Headers of all sheets have been converted to lowercase and data types preserved as specified. ✅")


    # ======================================================================
    # STEP 3:- Adding Blank Copy Sheets...
    # ======================================================================

    import openpyxl

    print("\n\n🔍 Step 3: Adding Blank Copy Sheets...")

    # Names of the sheets to add
    sheet_names = [
        "Opportunity_Copy",
        "Opportunity_product_Copy",
        "Opportunity_team_Copy",
        "Reporting_codes_Copy",
        "Tags_Copy"
    ]

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Add new sheets
        for name in sheet_names:
            wb.create_sheet(title=name)

        # Save the workbook
        wb.save(file_path)
        print(f"\n    ✅ Blank copy sheets added successfully. ")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e} ❌")


    # ======================================================================
    # STEP 4:-  Trimming Account id and Owner id  Values
    # ======================================================================

    print("\n\n🔍 Step 4:  Trimming Account id and Owner id  Values.")

    # Load the Excel file
    sheet_name = 'Opportunity'  
    column_to_trim = 'accountid' 

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Trim the values for whitespaces and handle internal spaces for accountid

        if column_to_trim not in df.columns:
            print(f"\n    ❌ Column '{column_to_trim}' not found in the sheet. The program will stop. ❌")
            sys.exit()  # Exits the program gracefully

        new_column_name = f'Trimmed_{column_to_trim}'

        df[new_column_name] = df[column_to_trim].str.replace(r'\s+', '', regex=True).str.strip()


        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("\n    ✅ Columns trimmed and new columns with trimmed values added successfully. ")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e} ❌")
        sys.exit()  # Exits the program gracefully

    # ======================================================================    
    # STEP 5 :- To Remove country code from the DC values
    # ======================================================================

    print("\n\n🔍 Step 5: Creating Account Number columns after removing country code from DC Accounts")

    # Define the sheet and column names
    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    accountid_column = 'Trimmed_accountid'  # Replace with the actual column name containing the account Ids
    new_column_name = 'AccountNumber'  # Name for the new column

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Define a function to process the values
        def process_value(value):
            if isinstance(value, str) and value.startswith('DC'):
                return value.split('-')[0]
            return value

        # Apply the function to the accountid column and store results in the new column
        df[new_column_name] = df[accountid_column].apply(process_value)

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"\n    ✅ New column '{new_column_name}' with formatted DC values created and added successfully. ✅")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e} ❌")
        sys.exit()  # Gracefully exit if there is an error


    # ======================================================================
    # STEP 6 :- Concatenating Valus to extract to file
    # ======================================================================


    print("\n\n🔍 Step 6: Concatenating Column Values...")

    # Define the sheet and columns to concatenate
    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    columns_to_concatenate = ['AccountNumber']  # Columns to concatenate
    new_column_names = ['Concatenatedaccountid']  # New columns with concatenated values

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the specified columns exist
        missing_columns = [col for col in columns_to_concatenate if col not in df.columns]
        if missing_columns:
            print(f"\n    ❌ The following columns are missing: {', '.join(missing_columns)} ❌")
            user_input = input(" \n   Do you want to continue? (yes/no): ").lower()
            if user_input != 'yes':
                print("\n    ❌ Operation aborted by user. ❌")
                sys.exit()  # Exiting gracefully

        # Concatenate the values of the specified columns
        column = columns_to_concatenate[0]
        if column in df.columns:
            # Convert the column to string, handle NaNs by filling with empty strings
            df[column] = df[column].astype(str).fillna('')
            df[new_column_names[0]] = "'" + df[column] + "',"

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"\n    ✅ Columns concatenated successfully and new columns added. ")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e} ❌")
        sys.exit()  # Gracefully exit on error


    # ======================================================================
    # STEP 7:- Creating Account id Extract for quering
    # ======================================================================

    print("\n\n🔍 Step 7: Creating Accountid Extract...")

    os.makedirs('Extracts', exist_ok=True)
    # Specify the sheet name and required columns
    sheet_name = "Opportunity"
    required_columns = ["Concatenatedaccountid"]

    # Check if the input file exists
    if not os.path.exists(file_path):
        print(f"\n    ❌ The input file '{file_path}' does not exist. ❌")
        # Create an empty DataFrame with the required columns
        df = pd.DataFrame(columns=required_columns)
    else:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Initialize an empty DataFrame for the output
    output_df = pd.DataFrame()

    # Process each required column
    for column in required_columns:
        if column in df.columns:
            # Remove blank and duplicate values
            cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
            # Add cleaned data to the output DataFrame
            output_df[column.replace("Concatenated", "")] = cleaned_data
        else:
            print(f"\n    ❌ Column '{column}' is missing in the input file. ❌")

    # Write the processed data to a new Excel file if there's any data to write
    if not output_df.empty:
        output_file = "Extracts/Accounts_extract.xlsx"
        output_df.to_excel(output_file, index=False)
        print(f"\n    ✅ Account Id values has been written to '{output_file}' successfully. ")
    else:
        print("\n    ❌ No columns were processed due to missing columns. ❌")

    # =========================================================================
    # Load the Excel file
    extract_file_path = "Extracts/Accounts_extract.xlsx"  # Change this to your actual file path
    df = pd.read_excel(extract_file_path)
    os.makedirs('Delete', exist_ok=True)
    # Extract the "accountid" column values
    if "accountid" in df.columns:
        account_ids = df["accountid"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/Account_ids.txt", "w") as f:
            f.write("\n".join(account_ids))

    else:
        print("Column 'accountid' not found in the sheet.")

    # ==========================================================
    def remove_last_char_from_last_line(extract_file):
        try:
            # Read all lines from the file
            with open(extract_file, 'r') as file:
                lines = file.readlines()

            # Check if the file is not empty
            if lines:
                # Remove the last character from the last line
                lines[-1] = lines[-1][:-1]

                # Write the modified content back to the file
                with open(extract_file, 'w') as file:
                    file.writelines(lines)

            # print("Last character from the last line has been removed.")
        
        except Exception as e:
            print(f"Error: {e}")
    
    # ========================================================================  
    
    # Code to remove comma from the text file
    
    remove_last_char_from_last_line('Delete/Account_ids.txt')

    # # ======================================================================
    # # STEP :- To ensure account csv file is Present
    # # ======================================================================

    # print("\n\n🔍 Step: Is the Account CSV file ready?")

    # # Ask the user for input
    # while True:
    #     user_input = input("\n    🔹 Type 'yes' to proceed: ").strip()
        
    #     if user_input.lower() == "yes":
    #         print("\n       ✅ Executing the next step...\n")
    #         break  # Exit the loop once the input is valid
    #     else:
    #         print("\n       ❌ Invalid input. Please type 'yes' to proceed. ❌")


    # ======================================================================
    # Step 8 :- Copy the data from account csv file to Copy sheet
    # ======================================================================


    import os
    import sys
    import pandas as pd
    import openpyxl

    print("\n\n🔍 Step 15: Copying extracted data to main file...")

    accounts_csv = os.path.expanduser("~/Downloads/accounts.csv")  # Specify the accounts CSV file path
    directory = os.path.expanduser("~/Downloads")
    def rename_bulkquery_file(new_name):
        """Search for a file with 'bulkQuery' in its name and rename it to the provided new name."""
        for filename in os.listdir(directory):
            if "bulkQuery" in filename and filename.endswith(".csv"):
                old_path = os.path.join(directory, filename)
                new_path = os.path.join(directory, new_name)
                os.rename(old_path, new_path)
                return True  # Indicate that renaming was successful
        return False  # No matching file found
    
    # Check if the accounts CSV file exists, and prompt to retry if not
    while not os.path.exists(accounts_csv):
        
        # Try renaming a bulkQuery file first
        if rename_bulkquery_file('accounts.csv'):
            continue  # If renaming was successful, check again if the file exists

        # Read account IDs from text file
        with open("Delete/Account_ids.txt", "r", encoding="utf-8") as file:
            cliptext = file.read()

        # Copy SQL query to clipboard
        account_query = f'Select AccountNumber,id from Account where AccountNumber in ({cliptext})'
        pyperclip.copy(account_query)
       
        print(f"\n    ❌ Error: File 'accounts.csv' does not exist. Did you query the accounts?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    # Read the accounts CSV file
    accounts_df = pd.read_csv(accounts_csv, usecols=[0, 1])  # Read first two columns

    # Define the path for the copy file (Excel)
    # file_path = "/path/to/your/excel/file.xlsx"  # Modify this with the actual file path

    # Load the Excel file or create a new one if it doesn't exist
    if os.path.exists(file_path):
        book = openpyxl.load_workbook(file_path)
        if "Opportunity_Copy" not in book.sheetnames:
            sheet = book.create_sheet(title="Opportunity_Copy")
        else:
            sheet = book["Opportunity_Copy"]
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Opportunity_Copy"

    # Write the headers to the "Opportunity_Copy" sheet
    for col_index, header in enumerate(accounts_df.columns, start=1):
        sheet.cell(row=1, column=col_index, value=header)

    # Write the data to the "Opportunity_Copy" sheet
    for row_index in range(len(accounts_df)):
        sheet.cell(row=row_index + 2, column=1, value=accounts_df.iloc[row_index, 0])
        sheet.cell(row=row_index + 2, column=2, value=accounts_df.iloc[row_index, 1])

    # Save the changes
    book.save(file_path)
    print(f"\n    ✅ 'Accounts' Data has been successfully copied to '{file_path.split('/')[-1]}'.")

    # ======================================================================
    # STEP 9:- Check if the Accounts are present in ISC or not by creating column 
    # ======================================================================


    print("\n\n🔍 Step 9: Check how many Accounts are not present in ISC ")

    opportunity_sheet_name = 'Opportunity'
    opportunity_copy_sheet_name = 'Opportunity_Copy'

    try:
        # Read data from Excel sheets
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)
        
        # Check for duplicates in AccountNumber in Opportunity_Copy
        duplicate_accounts = opportunity_copy_df[opportunity_copy_df.duplicated(subset=['AccountNumber'], keep=False)]
        
        if not duplicate_accounts.empty:
            print("\n    ❗️ Duplicate AccountNumbers found with multiple Id values:")
            
            # Group by AccountNumber and prompt user for resolution
            for account_number, group in duplicate_accounts.groupby('AccountNumber'):
                print(f"\n        🔹 AccountNumber: {account_number}")
                
                # Display the 'Id' and corresponding Excel row number
                for idx, row in group.iterrows():
                    excel_row_number = idx + 2  # Adjust for Excel row numbering
                    print(f"\n           🔸 Id: {row['Id']} (Excel Row {excel_row_number})")
                
                # Prompt user to choose the Id to keep
                valid_ids = group['Id'].tolist()
                while True:
                    chosen_id = input(f"\n        🔹 Select id for AccountNumber {account_number} from above Ids: ").strip()
                    if chosen_id in valid_ids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id from {valid_ids}. ❌")
                
                # Filter DataFrame to keep only the chosen Id for the AccountNumber
                opportunity_copy_df = opportunity_copy_df[
                    ~((opportunity_copy_df['AccountNumber'] == account_number) & 
                    (opportunity_copy_df['Id'] != chosen_id))
                ]
        
        # Merge DataFrames
        merged_df = pd.merge(opportunity_df, opportunity_copy_df[['AccountNumber', 'Id']],
                            on='AccountNumber', how='left')
        
        # Handle NaN values
        merged_df['Id'] = merged_df['Id'].fillna('Not in ISC')
        
        # Rename columns
        merged_df.rename(columns={'Id': 'In ISC or Not'}, inplace=True)
        
        # Save updated DataFrame to Excel
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            merged_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)
        
        # Count 'Not in ISC'
        sheet_name = opportunity_sheet_name
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        not_in_isc_count = (df['In ISC or Not'] == 'Not in ISC').sum()
        
        # Print count
        print(f"\n    ❗️ Count of accounts 'Not in ISC': {not_in_isc_count}")
        if not_in_isc_count == 0:
            account_present.append(os.path.basename(file_path))

    except FileNotFoundError:
        print("\n    ❌ Error: The specified file was not found. Please check the file path.")
    except ValueError as e:
        print(f"\n    ❌ Error: {e}")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")


    # ======================================================================
    # Step 10:- Creating file with the list of Accounts to be Imported
    # ======================================================================

    print("\n\nStep 10:- Creating file with the list of Accounts to be Imported")

    # Name of the specific sheet where filtering is to be done
    sheet_name = 'Opportunity'  # Update with the appropriate sheet name

    # Name of the column containing "In ISC or Not" values
    isc_column_name = 'In ISC or Not'

    # Name of the column containing "AccountNumber" values
    account_column_name = 'AccountNumber'

    # Read the input Excel file
    input_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Filter rows where the value in the "In ISC or Not" column is "Not in ISC"
    filtered_df = input_df.loc[input_df[isc_column_name] == 'Not in ISC', [account_column_name]]

    # Remove duplicate and blank values
    filtered_df.dropna(inplace=True)
    filtered_df.drop_duplicates(inplace=True)

    # Path to the output Excel file
    output_file_path = os.path.expanduser("~/Downloads/Accounts_to_be_imported.xlsx")

    # Check if the output Excel file exists
    try:
        # Read existing data from the output Excel file
        output_df = pd.read_excel(output_file_path, header=None)

        # Append the filtered data to the existing DataFrame column
        if output_df.empty:
            output_df = pd.DataFrame(columns=[0])

        # Find the index where new data should be appended
        start_index = len(output_df)
        
        for idx, val in enumerate(filtered_df[account_column_name]):
            output_df.loc[start_index + idx, 0] = val
        
        # Remove duplicate rows again (in case there were duplicates in existing data)
        output_df.drop_duplicates(inplace=True)

        # Write the updated data to the output Excel file
        output_df.to_excel(output_file_path, index=False, header=False)

        print(f"\n   ✅ Data successfully appended to '{output_file_path}'.")
    except FileNotFoundError:
        # If the file doesn't exist, create it and write the filtered data directly to it
        filtered_df.to_excel(output_file_path, index=False, header=False)
        print(f"\n    ✅ Accounts to be imported file created with the filtered data.")

    # ================================================
    # Delete CSV Files
    # ================================================

    # Hardcoded directory
    directory = os.path.expanduser("~/Downloads")
    
    files_deleted = 0
    # Get list of files in the directory
    for file_name in os.listdir(directory):
        if file_name.endswith('.csv'):
            file_path = os.path.join(directory, file_name)
            os.remove(file_path)
            files_deleted += 1
    
    if files_deleted > 0:
        print(f"\n✅ Successfully deleted {files_deleted} CSV file(s) from '{directory}'.")
    else:
        print(f"\n❌ No CSV files were found in '{directory}'.")

    print(f"\n{'='*120}\n{' ' * 30} 📝  'Accounts to be Imported' file created 📝 {' ' * 30}\n{'='*120}\n")


# ====================================================================================
# Add Header to the Column
# ====================================================================================

def insert_cell_and_shift(filepath, sheet_name):
    """
    Inserts 'Accounts' in the first row, first column and shifts all other values downward.
    
    :param filepath: Path to the Excel file
    :param sheet_name: Name of the sheet to modify
    """
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]
    
    # Insert a new row at the top
    sheet.insert_rows(1)
    
    # Write 'Accounts' in the first cell
    sheet.cell(row=1, column=1, value='Accounts')
    
    # Save the workbook
    wb.save(filepath)
    wb.close()

# Example usage
file_path = os.path.expanduser("~/Downloads/Accounts_to_be_imported.xlsx")
sheet_name = "Sheet1"
insert_cell_and_shift(file_path, sheet_name)

# ====================================================================================

# ====================================================================================

# Load the original Excel file
file_path = os.path.expanduser("~/Downloads/Accounts_to_be_imported.xlsx")
df = pd.read_excel(file_path)
# Initialize lists to store valid and invalid values
invalid_values = []
valid_values = []

if df.empty:
    pass
else:
    # Define a regular expression for country codes (e.g., "-US", "-KA", etc.)
    country_code_pattern = r'-[A-Za-z]{2,3}$'

    # Check each value in the column (assuming the values are in the first column)
    for value in df.iloc[:, 0]:
        value = str(value).strip()  # Ensure it's a string and remove leading/trailing spaces
        if not (value.lower().startswith('db') or value.lower().startswith('dc')):
            # If it doesn't start with DB or DC (case insensitive), it's invalid
            invalid_values.append(value)
        elif value.lower().startswith('db'):
            # If it starts with DB or db, it should have a country code
            if not re.search(country_code_pattern, value):
                invalid_values.append(value)
            else:
                valid_values.append(value)  # Add to valid values list
        else:
            valid_values.append(value)  # Add to valid values list for DC or other valid entries

# Count invalid values
invalid_count = len(invalid_values)

# If there are invalid values, write them to a new Excel file
if invalid_count > 0:
    invalid_df = pd.DataFrame(invalid_values, columns=['Invalid Accounts'])
    invalid_df.to_excel(os.path.expanduser("~/Downloads/Invalid_Accounts.xlsx"), index=False)


# Update the original dataframe with only valid values
valid_df = pd.DataFrame(valid_values, columns=['Accounts'])

# Save the updated dataframe back to the original file
valid_df.to_excel(file_path, index=False)


# ============================================================
# Code to write Valid accounts to txt file
# ============================================================

# Load the Excel file
extract_file_path = os.path.expanduser("~/Downloads/Accounts_to_be_imported.xlsx")  # Change this to your actual file path
df = pd.read_excel(extract_file_path, header=None)  # Load without headers

# Check if the DataFrame is empty
if df.empty or df.shape[1] == 0:
    pass
    # print("Error: The Excel file is empty or does not contain any columns.")
else:
    # Extract values from the first column
    first_column_values = df.iloc[1:, 0].dropna().astype(str)  # # Skip the first row and Drop NaN values and convert to string

    # Only proceed if there are values in the first column
    # Ensure the output directory exists
    os.makedirs('Delete', exist_ok=True)

    # Save to a text file
    with open("Delete/Accounts to be Imported.txt", "w") as f:
        f.write("\n".join(first_column_values))

    # print("First column values have been saved to Delete/1_account_ids.txt")

# ============================================================
# Code to give a Summary
# ============================================================

print('\n✅ Files Processed')
for index, file in enumerate(files_processed, start=1):
    shortened_path = file.split('/')[-1]
    print(f'\n    {index}. {shortened_path} ✅')

print("\n✅ Files where Accounts are Present")
for index, file in enumerate(account_present, start=1):
    print (f"\n    {index}. {file}")

print (f"\n❗️ Invalid Accounts:- {invalid_count} ")
with open('Delete/Accounts to be Imported.txt', 'r') as file:
    line_count = sum(1 for line in file)
print(f'\n❗️ Total Accounts to Be Imported:- {line_count}')

print("\n👋 Exiting the script. Goodbye!")
print(f"\n{'='*120}\n{' ' * 30} 📝  Script Completed 📝 {' ' * 30}\n{'='*120}\n")
sys.exit()
