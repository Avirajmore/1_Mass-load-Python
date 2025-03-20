# Importing all the necessary Libraries

import os

import re
import sys
import time
import shutil
import openpyxl
import pandas as pd
from tkinter import *
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import SheetTitleException

# =========================================================================================================================================
                                                 # FOLDER CREATION & FILE MOVEMENT
# =========================================================================================================================================


# =========================================================
# Base Path where the mass load files are stored
# =========================================================

# Base directory path (fixed part)
base_dir = "/Users/avirajmore/Documents/Office Docs/Massload Files/2025" 

# =========================================================
# Define Function To avoid invalid name of the main folder and sub folder which will be used later
# =========================================================

# Function to validate folder names
def is_valid_folder_name(name):
    invalid_chars = set('\\/:*?\"<>|')
    return name and not any(char in invalid_chars for char in name)

# =========================================================
# Folder Creation starts
# =========================================================

# Output header
print("=" * 100)
print(" " * 33 + "📂 FOLDER CREATION & FILE MOVEMENT 📂")
print("=" * 100)

# =========================================================
# Step 1: Create the main Sprint folder
# =========================================================

print("\n🔍 Step 1: Creating Sprint folder")
while True:
    Sprint_Number = input("\n    📂 Enter the Sprint number: ").strip()
    if is_valid_folder_name(Sprint_Number):
        break
    else:
        print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

main_folder_path = os.path.join(base_dir, Sprint_Number)
os.makedirs(main_folder_path, exist_ok=True)
print(f"\n        ✅ Folder '{Sprint_Number}' created successfully")

# =========================================================
# Step 2: Create subfolders for Different Week and "Copy File" and "Final iteration file" folders
# =========================================================

print("\n\n🔍 Step 2: Creating subfolders")
while True:
    subfolder_name = input("\n    📂 Enter the name of the subfolder: ").strip()
    if is_valid_folder_name(subfolder_name):
        break
    else:
        print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

subfolder_path = os.path.join(main_folder_path, subfolder_name)
os.makedirs(subfolder_path, exist_ok=True)

copy_file_path = os.path.join(subfolder_path, "Copy files")
final_iteration_file_path = os.path.join(subfolder_path, "Final iteration files")

os.makedirs(copy_file_path, exist_ok=True)
os.makedirs(final_iteration_file_path, exist_ok=True)

print(f"\n        ✅ Subfolders 'Copy files' and 'Final iteration files' created successfully")

# =========================================================
# Step 3: Move Mass load files from Downloads to Week Subfolder
# =========================================================


downloads_dir = "/Users/avirajmore/Downloads"

excel_extensions = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm')

files_moved = []

print(f"\n\n🔍 Step 3: Moving Excel files")

# Initialize excel_files as an empty list
excel_files = []

# Check if any Excel files exist in Downloads folder before moving them
if os.path.exists(downloads_dir):
    for file_name in os.listdir(downloads_dir):
        if file_name.lower().endswith(excel_extensions):
            source_path = os.path.join(downloads_dir, file_name)
            target_path = os.path.join(subfolder_path, file_name)
            shutil.move(source_path, target_path)
            files_moved.append(file_name)

# Populate the excel_files list with all Excel files in the subfolder
excel_files = [f for f in os.listdir(subfolder_path) if f.lower().endswith(excel_extensions)]

# Moved Files Display under Step 3
if not files_moved:
    print("\n    📥 Moved Files:")
    print("\n        1) ❗️ No files were moved from Downloads. ")
else:
    print("\n    📥 Moved Files:")
    for index, file_name in enumerate(files_moved, start=1):
        print(f"\n        {index}) {file_name} ✅")

# =========================================================
# Step 4: Copy files independently to the "Copy files" folder
# =========================================================

print("\n\n🔍 Step 4: Copying files")
files_copied = []

# Now, excel_files contains all the Excel files (moved or existing)
for file_name in excel_files:
    source_path = os.path.join(subfolder_path, file_name)
    copy_file_name = f"{os.path.splitext(file_name)[0]}_Copy{os.path.splitext(file_name)[1]}"
    target_path = os.path.join(copy_file_path, copy_file_name)

    if os.path.exists(target_path):
        files_copied.append((file_name, "skipped"))
    else:
        shutil.copy(source_path, target_path)
        files_copied.append((file_name, "copied"))

print("\n    📤 Copied Files:")
for index, (file_name, status) in enumerate(files_copied, start=1):
    if status == "copied":
        print(f"\n        {index}) {file_name} ✅")
    elif status == "skipped":
        print(f"\n        {index}) {file_name} - Skipped 🛑 ")


# =========================================================
# Step 5: Create Folders in "Final iteration files" to save success and error files and removed Rows data
# =========================================================


print("\n\n🔍 Step 5: Creating folders in 'Final iteration files'")
folders_created = []

# Now, excel_files contains all the Excel files (moved or existing)
for file_name in excel_files:
    folder_name = os.path.splitext(file_name)[0]
    folder_path = os.path.join(final_iteration_file_path, folder_name)

    # Ensure the main folder exists (create if necessary)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path, exist_ok=True)
        folder_status = "created"
    else:
        folder_status = "exists"

    # Create subfolders inside the main folder (always ensure they exist)
    subfolders = ["Removed Rows", "Success and error files"]
    for subfolder in subfolders:
        subfolder_path = os.path.join(folder_path, subfolder)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path, exist_ok=True)

    folders_created.append((folder_name, folder_status))

# Display folder creation results
print("\n    🗂️ Folder Created:")
for index, (folder_name, status) in enumerate(folders_created, start=1):
    if status == "created":
        print(f"\n        {index}) {folder_name} ✅ (Main folder created)")
    elif status == "exists":
        print(f"\n        {index}) {folder_name} - Main folder already exists 🛑 (Subfolders ensured)")


# =========================================================================================================================================
#                                                OPPORTUNITY SHEET EXECUTION
# =========================================================================================================================================


# ======================================================================
# Step 0:-  Decide which file to Process first, it will decide the file path
# ======================================================================

# # Hardcoded Path for debugging if the below code does not work

# file_path = "/Users/avirajmore/Downloads/Avi 3 copy.xlsx"  

# ===============================================================

# Below code Automatically selects the file based on previous code selection

# Check if 'Copy files' folder exists
# File selection loop
while True:
    # Check if 'Copy files' folder exists
    if not os.path.exists(copy_file_path):
        print("\n     ❗️ 'Copy files' folder does not exist. ")
        break

    # List available files
    files_in_copy_folder = [f for f in os.listdir(copy_file_path) if os.path.isfile(os.path.join(copy_file_path, f))]

    if not files_in_copy_folder:
        print("\n     🚫 No files found in 'Copy files' folder. ")
        break

    print("\n====================================================================================================")
    print("\n\n📂 Please select a file to process:")
    print("\n    🔸 List of Files in Copy Folder: ")

    for idx, file_name in enumerate(files_in_copy_folder, start=1):
        print(f"\n        📄 {idx}. {file_name}")

    # Get user selection
    while True:
        user_input = input("\n    👉 Enter the number of the file to process (or type 'exit' to quit): ").strip()

        if user_input.lower() == 'exit':
            print("\n        ❌ File selection has been canceled. Exiting process. ")
            sys.exit()

        try:
            selected_index = int(user_input) - 1
            if 0 <= selected_index < len(files_in_copy_folder):
                file_path = os.path.join(copy_file_path, files_in_copy_folder[selected_index])
                print(f"\n        ✅ You selected the file: {files_in_copy_folder[selected_index]} ")
                break
            else:
                print(f"\n        ❗ Invalid selection. Please select a number between 1 and {len(files_in_copy_folder)}.")
        except ValueError:
            print("\n        ❗ Invalid input. Please enter a valid number or type 'exit' to cancel.")


    # ================================================================================
    # To rename the sheets
    # ================================================================================

    print(f"\n{"="*100}")

    print("\n\n🔍 Check if all the Required Sheets are present or not")
 
    # Path to the Excel file
    # file_path = 'your_excel_file.xlsx'

    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    # List of required sheets (Tags is optional)
    required_sheets = ['Opportunity', 'Opportunity_product', 'Opportunity_Team ', 'Reporting_codes', 'Tags']

    # Get the list of sheets in the workbook
    sheets_in_file = wb.sheetnames

    # Check for missing required sheets (except Tags)
    missing_sheets = [sheet for sheet in required_sheets if sheet != 'Tags' and sheet not in sheets_in_file]


    if not missing_sheets:
        print("\n    ✅ All required sheets are already present! 🎉")
    else:
        print("\n    ❌ The following required sheets are missing: ")
        for i, sheet in enumerate(missing_sheets, 1):
            print(f"\n        {i}. {sheet}")
        # Get the list of existing sheets that are not already required sheets
        available_sheets = [s for s in sheets_in_file if s not in required_sheets]
        
        if available_sheets:
            print("\n    📋 Here are the available sheets to rename: ")
            # Display the available sheets as a numbered list
            for i, s in enumerate(available_sheets, 1):
                print(f"\n        {i}. {s}")
        
        # If any required sheet is missing, ask for renaming
        for sheet in missing_sheets:
            while True:
                choice = input(f"\n    🔸 Enter the index of the sheet to rename to '{sheet}' or type 'skip': ")

                if choice.lower() == 'skip':
                    print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
                    break  # Skip renaming this sheet
                try:
                    # Convert the choice to an integer if it's a number
                    choice = int(choice)
                    if 1 <= choice <= len(available_sheets):
                        rename_sheet = available_sheets[choice - 1]
                        # Rename the selected sheet
                        ws = wb[rename_sheet]
                        ws.title = sheet
                        print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")
                        break  # Exit the loop after successful renaming
                    else:
                        print("\n        ❗ Invalid number selected. Please choose a valid option.")
                except ValueError:
                    print("\n        ❗ Invalid input, please enter a valid number or 'skip'. 😕")
        
        # Save the modified workbook (if any renaming was done)
        wb.save(file_path)
        print("\n    💾 Workbook saved with changes!")
    
    
    # ==========================================
    # Just in case if the files have similar name but different casing
    # ==========================================

    # Mapping of old sheet names to new sheet names
    sheet_name_mapping = {
        'Opportunity1': 'Opportunity',
        'Opportunity_product1': 'Opportunity_product',
        'Opportunity_team1': 'Opportunity_team',
        'Reporting_codes1': 'Reporting_codes'
    }

    # Load the Excel file
    # file_path = 'your_excel_file.xlsx'  # Replace with the actual file path
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet in wb.sheetnames:
        if sheet in sheet_name_mapping:
            # Rename sheet if it matches the mapping
            new_name = sheet_name_mapping[sheet]
            ws = wb[sheet]
            ws.title = new_name

    # Save the workbook with the renamed sheets (the content will remain unchanged)
    wb.save(file_path)


    # ======================================================================
    # Print Opportunity Script Execution 📝                               
    # ======================================================================

    print("\n")
    print("=" * 100)
    print(" " * 33 + "📝 OPPORTUNITY SHEET EXECUTION 📝")
    print("=" * 100)


    # ======================================================================
    # Step 1: Checking File Existence
    #   • Verify if the file exists at the specified path. If not, return an error and exit.
    # ======================================================================


    print("\n\n🔍 Step 1: Checking if the file exists...")

    if os.path.exists(file_path):
        filename = str(file_path.split('/')[-1])
        print(f"\n    ✅ File '{filename}' exists at the specified path. ")
    else:
        print("\n    ❌ Error: File does not exist or the path is invalid. \n")
        sys.exit()  # Stops further execution of the program


    # ======================================================================
    # Step 2: Remove duplicate rows
    #   • Remove duplicate and blank rows.
    #   • Return an error if the "Opportunity" sheet is missing.
    # ======================================================================


    print("\n\n🔍 Step 2: Removing duplicate rows and blank rows...")

    def remove_duplicates_and_blank_rows(file_path, opportunity_sheet_name):
        try:
            # Try to read the spreadsheet with the given sheet name
            df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

            # Drop duplicate rows
            df = df.drop_duplicates()

            # Remove rows where all cells are NaN (blank rows)
            df = df.dropna(how='all')

            # Save the cleaned data back to the same file without modifying formatting
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

            print(f"\n    ✅ Removed all the duplicate rows and blank rows. ")

        except ValueError as e:
            # Handle the case where the sheet does not exist
            print(f"\n    ❌ Error: The sheet '{opportunity_sheet_name}' does not exist in the file. ")
            sys.exit()
        except Exception as e:
            # Handle any other exceptions
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
            sys.exit()

    remove_duplicates_and_blank_rows(file_path, 'Opportunity')


    # ======================================================================
    # Step 3: Check for required columns and blank values
    #   • Ensure required columns are present; prompt to continue or stop if any are missing.
    #   • Extra columns are listed but do not halt the process.
    #   • Identify blank values in critical columns, report their count, and prompt to proceed or stop.

    # ======================================================================


    print("\n\n🔍 Step 3: Checking required columns and blank values...")

    opportunity_sheet_name = 'Opportunity'

    # List of required columns (case insensitive)
    required_columns = [
        'opportunity_legacy_id_c',
        'name',
        'accountid',
        'sales_stage',
        'won reason',
        'lost category',
        'lost reason',
        'expected_close_date',
        'currency_code',
        'ownerid',
        'next_step',
        'oi_source',
        'created_by',
        'modified_by',
        'created_date',
        'modified_date',
        'pricebook',
        'recordtypeid'
    ]

    # Columns to check for blank values
    columns_to_check = [
        'name',
        'accountid',
        'sales_stage',
        'expected_close_date',
        'currency_code',
        'oi_source'
    ]

    try:
        # Read the Opportunity sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except Exception as e:
        print(f"\n    ❌ Error reading Excel file: {str(e)} ")
        sys.exit()

    # Convert column names to lowercase for case-insensitive comparison
    df.columns = df.columns.str.lower()

    # Get columns that are missing and extra
    missing_columns = [col for col in required_columns if col.lower() not in df.columns]
    extra_columns = [col for col in df.columns if col.lower() not in required_columns]

    # Display missing and extra columns (only once)
    if missing_columns:
        print("\n    ❗ The following required columns are missing in the Opportunity sheet:")
        for col in missing_columns:
            print(f"\n        🔸 {col}")

    if extra_columns:
        print("\n    ❗️ The following extra columns are present in the Opportunity sheet:")
        for col in extra_columns:
            print(f"\n        🔸 {col}")

    # Check for blank values in the specified columns
    blank_values = {}
    for col in columns_to_check:
        if col.lower() in df.columns:
            blank_rows = df[col.lower()].isnull() | (df[col.lower()] == "")
            blank_values[col] = blank_rows.sum()

    # If there are any blank values or missing columns, ask the user if they want to continue
    if any(blank_values.values()) or missing_columns:
        if any(blank_values.values()):
            print("\n    ❗️ The following columns have blank values:")
            for col, count in blank_values.items():
                if count > 0:
                    print(f"\n        🔸 {col}: {count} blank values")

        # Ask for the user's choice only once
        while True:
            choice = input("\n    👉 Do you want to continue with the operation? (yes/no): ").strip().lower()
            if choice == 'yes':
                print("\n        ✅ Continuing the operation... ")
                break
            elif choice == 'no':
                print("\n        ❌ Operation terminated as requested. \n")
                sys.exit()
            else:
                print("\n      ❗ Invalid choice. Please enter 'yes' or 'no'.")
    else:
        print("\n    ✅ All required columns are present in the Opportunity sheet. ")


    # ======================================================================
    # Step 4: Count the rows and columns in the beginning of the process
    #   • Count rows after cleaning to ensure no discrepancies.
    #   • Compare row counts before and after processing; prompt to continue or exit if mismatched
    # ======================================================================

    print("\n\n🔍 Step 4: Counting the rows and columns...")

    # Name of the sheet to target
    opportunity_sheet_name = 'Opportunity'

    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Get the number of rows and columns
    oppty_initial_num_rows = df.shape[0]     # Number of rows in the DataFrame
    oppty_num_columns = df.shape[1]          # Number of columns in the DataFrame

    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {oppty_initial_num_rows}")


    # ======================================================================

    # Step 5: Convert the headers of all the sheets to lowercase
    #   • Convert all column headers to lowercase to avoid case-sensitivity issues.
    #   • Ensure numeric columns retain their data types.

    # ======================================================================

    print("\n\n🔍 Step 5: Converting headers to lowercase...")

    # Read the Excel file with all sheets, initially treating all data as strings
    xls = pd.ExcelFile(file_path)

    # Dictionary to hold modified dataframes
    sheets_dict = {}

    # Columns that should remain numeric
    numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

    # Iterate through each sheet
    for sheet_name in xls.sheet_names:
        # Read each sheet into a dataframe with all columns as strings
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        
        # Convert headers to lowercase
        df.columns = [col.lower() for col in df.columns]
        
        # Convert specific columns back to numeric types
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')  # Convert to numeric, setting errors to NaN
        
        # Save modified dataframe to dictionary
        sheets_dict[sheet_name] = df

    # Write the modified dataframes back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\n    ✅ Headers of all sheets in the file have been converted to lowercase and data types preserved as specified.")


    # ======================================================================

    # Step 6: Convert the email ids to lowercase and fill missing values with a default value
    #   • Convert email IDs of “owner id” and “created by” column to lowercase for consistent matching (e.g., during lookups like owner ID by email).
    #   • If any of the values are blank, it should automatically fill it with Data migration id
    #   • The code will show how many blank values were filled with Data migration id for reference

    # ======================================================================


    print("\n\n🔍 Step 6: Converting email ids to lowercase and filling missing values...")

    # Name of the sheet to target
    sheet_name = 'Opportunity'

    # Columns to convert to lowercase and fill blanks
    columns_to_process = ['ownerid', 'created_by']
    fill_value = "iscdmig2@in.ibm.com"

    # Read the Excel file
    excel_data = pd.read_excel(file_path, sheet_name=None)

    # Check if the specified sheet exists
    if sheet_name in excel_data:
        # Access the specified sheet
        df = excel_data[sheet_name]
        
        # Dictionary to store count of filled blank values
        filled_counts = {col: 0 for col in columns_to_process}
        
        # Fill blank cells with specified value and count filled blanks
        for column in columns_to_process:
            if column in df.columns:
                blank_count = df[column].isnull().sum()
                filled_counts[column] = blank_count
                df[column] = df[column].fillna(fill_value)
                df[column] = df[column].apply(lambda x: x.lower() if isinstance(x, str) else x)
            else:
                print(f"\n    ❌Error: Column '{column}' not found in the '{sheet_name}' sheet. Terminating the Program.")
                sys.exit()
        
        # Replace the existing data in the sheet with the modified values
        excel_data[sheet_name] = df

        # Write the modified Excel data back to the file
        with pd.ExcelWriter(file_path) as writer:
            for sheet, data in excel_data.items():
                data.to_excel(writer, sheet_name=sheet, index=False)

        # Display the count of blank columns filled for each column
        for col, count in filled_counts.items():
            print(f"\n    ❗️ Blank Values filled with Data migration Id in {col} column: {count}")
        
    else:
        print(f"\n    ❌Error: Sheet '{sheet_name}' not found in the Excel file.")


    # ======================================================================
    # Step 7: Create Blank sheets in the excel for rough work
    #   • Add blank sheets for rough work, where queried data for vlookups can be pasted.
    # ======================================================================


    print("\n\n🔍 Step 7: Creating Blank sheets for rough work...")

    # Names of the sheets to add
    sheet_names = [
        "Opportunity_Copy",
        "Opportunity_product_Copy",
        "Opportunity_team_Copy",
        "Reporting_codes_Copy",
        "Tags_Copy"
    ]

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Add new sheets
        for name in sheet_names:
            wb.create_sheet(title=name)

        # Save the workbook
        wb.save(file_path)
        print("\n    ✅ Blank Copy Sheets added successfully.")

    except Exception as e:
        print(f"\n    📝 An error occurred: {e}")


    # ======================================================================

    # Step 8: Add Pricebook and RecordType id column in the sheet
    #   • Add two new columns, "Pricebook" and "RecordType id," with predefined values for all rows.

    # ======================================================================

    print("\n\n🔍 Step 8: Adding Pricebook and RecordType id columns...")

    sheet_name = 'Opportunity' 

    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Add the two new columns with the specified values
    df['Pricebook2Id'] = '01s3h000003KXvoAAG'
    df['RecordTypeId'] = '0123h000000kppcAAA'

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\n    ✅ 'Price_Book' and 'Record_Type' Columns added successfully.")


    # ======================================================================
    # Step 9: Change the format of the Date Column
    #   • Format the date column to YYYY-MM-DD.
    #   • If invalid dates are found, return an error and exit, as closeDate is critical
    # ======================================================================

    print("\n\n🔍 Step 9: Formatting the Date column...")

    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    date_column = 'expected_close_date'  # Replace with the actual column name containing the dates

    try:
        # Load the specific sheet into a DataFrame
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except ValueError:
            print(f"\n    ❌ Error: The sheet '{sheet_name}' is missing in the file.")
            sys.exit(1)  # Exit the script with an error code

        # Check if the specified column exists
        if date_column not in df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing in the sheet '{sheet_name}'.")
            sys.exit(1)  # Exit the script with an error code

        # Ensure the date column is in datetime format, allowing for errors to be coerced to NaT
        df[date_column] = pd.to_datetime(df[date_column], errors='coerce')

        # Check for blank values (NaT) after processing
        if df[date_column].isnull().any():
            print(f"\n    ❌ Error: The column '{date_column}' contains blank or invalid values after processing.")
            print("\n    ❗️ Please review the data, correct the issues, and try again.")
            sys.exit(1)  # Exit the script with an error code

        # Format the valid dates to YYYY-MM-DD
        df[date_column] = df[date_column].dt.strftime('%Y-%m-%d')

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("\n    ✅ Date column formatted to YYYY-MM-DD successfully.")

    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the script with an error code


    # ======================================================================
    # Step 10: Create new "legacy_opportunity_split_id_c" column if it does not exist
    #   • Skip if the column already exists.
    #   • Otherwise, create it and copy values from the "opportunity_legacy_id_c" column.

    # ======================================================================


    print("\n\n🔍 Step 10: Creating 'legacy_opportunity_split_id_c' column...")

    # Specify the sheet name where you want to add the new column and copy data
    target_sheet_name = 'Opportunity'  # Replace 'Sheet1' with the name of your target sheet

    # Read the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=target_sheet_name)

    # Check if the column "legacy_opportunity_split_id_c" already exists (case-insensitive)
    existing_columns = [col.lower() for col in df.columns]
    if "legacy_opportunity_split_id_c" in existing_columns:
        print("\n    ✅ 'legacy_opportunity_split_id_c' column already exists in the sheet.")
        
        # Check for blank (NaN) values in the 'legacy_opportunity_split_id_c' column
        if df['legacy_opportunity_split_id_c'].isnull().any():
            print("\n    ❌ Error: 'legacy_opportunity_split_id_c' column contains blank (NaN) values. Please review. Exiting process.")
            sys.exit()  # Exit the code if blank values are found
    else:
        # Check if "opportunity_legacy_id_c" column exists
        if "opportunity_legacy_id_c" not in existing_columns:
            print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found. Exiting process.")
            sys.exit()  # Exit if "opportunity_legacy_id_c" is not found
        
        # Create the new column 'legacy_opportunity_split_id_c' and populate it with 'opportunity_legacy_id_c' values
        df['legacy_opportunity_split_id_c'] = df['opportunity_legacy_id_c']

        # Write the modified DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=target_sheet_name)

        print('\n    ✅ New legacy_opportunity_split_id_c column added to sheet. Process completed successfully.')



    # ======================================================================
    # Step 11: Create new column with Trimmed Account_id and Email_id column
    #   • Remove extra spaces (including within values) from 'accountid' and 'ownerid' columns.
    #   • Create a new column with trimmed values.
    #   • Throw an error and stop if these columns are missing.
    # ======================================================================


    print('\n\n🔍 Step 11: Creating new column with Trimmed Account_id and Email_id...\n')

    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    columns_to_trim = ['accountid', 'ownerid']  # Replace with the actual column names to trim

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

    except ValueError as e:
        print(f"    ❌ Error: The sheet '{sheet_name}' was not found in the file.")
        sys.exit(1)

    # Check if specified columns exist in the DataFrame
    missing_columns = [col for col in columns_to_trim if col not in df.columns]

    if missing_columns:
        print(f"    ❌ Error: The following columns were not found in the sheet '{sheet_name}': {', '.join(missing_columns)}")
        sys.exit(1)

    # Trim the values for whitespaces and create new columns for the trimmed values
    for column in columns_to_trim:
        new_column_name = f'Trimmed_{column}'
        # If the column is 'accountid', remove internal spaces in addition to trimming
        if column == 'accountid':
            df[new_column_name] = df[column].str.replace(r'\s+', '', regex=True).str.strip()
        else:
            df[new_column_name] = df[column].str.strip()

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("    ✅ Account and Email Columns trimmed successfully, and internal spaces in 'accountid' removed.")


    # ======================================================================
    # Step 12: Remove the country code from DC Accounts
    #   • For columns with both DC and DB accounts, remove country codes from DC accounts, as they are invalid.
    #   • And keep The DB values as it is
    # ======================================================================


    print("\n\n🔍 Step 12: Processing Accounts with correct format...\n")

    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    accountid_column = 'Trimmed_accountid'  # Replace with the actual column name containing the account Ids
    new_column_name = 'AccountNumber'  # Name for the new column

    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Define a function to process the values
    def process_value(value):
        if isinstance(value, str) and value.startswith('DC'):
            return value.split('-')[0]
        return value

    # Apply the function to the accountid column and store results in the new column
    df[new_column_name] = df[accountid_column].apply(process_value)

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("    ✅ New column with formatted DC values created and added to the sheet successfully.")


    # ======================================================================
    # Step 13: Concatenate the Values
    #   • Add apostrophes and commas to account IDs and emails to format them for Salesforce query use.
    # ======================================================================

    print("\n\n🔍 Step 13: Concatenating the Values...\n")

    sheet_name = 'Opportunity'  # Replace with the actual sheet name
    columns_to_concatenate = ['AccountNumber', 'Trimmed_ownerid', 'created_by']  # Replace with the actual column names to concatenate
    new_column_names = ['Concatenatedaccountid', 'Concatenatedownerid', 'concatenatedcreatedby']  # Names for the new columns with concatenated values

    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Check if the specified columns exist and prompt the user if not found
    missing_columns = [col for col in columns_to_concatenate if col not in df.columns]
    if missing_columns:
        print(f"    ❗️ The following columns are missing: {', '.join(missing_columns)}")
        user_input = input("    📝 Do you want to continue? (yes/no): ").lower()
        if user_input != 'yes':
            print("    ❌ Operation aborted.")
            exit()

    # Concatenate the values of the desired columns with inverted commas and a comma
    for i, column in enumerate(columns_to_concatenate):
        if column in df.columns:
            # Convert the column to string, handle NaNs by filling with empty strings
            df[column] = df[column].astype(str).fillna('')
            df[new_column_names[i]] = "'" + df[column] + "',"

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("    ✅ Columns concatenated and new columns with concatenated values added successfully.")


    # ======================================================================
    # Step 14: Extract Concatenated values
    #   • Save formatted values to a separate file for easier copying and pasting into the Workbench query without opening the main file.
    # ======================================================================


    print("\n\n🔍 Step 14: Extracting Concatenated values...\n")
    os.makedirs('Extracts', exist_ok=True)
    sheet_name = "Opportunity"  # Specify the sheet name

    required_columns = ["Concatenatedaccountid", "Concatenatedownerid", "concatenatedcreatedby"]

    # Check if the input file exists
    if not os.path.exists(file_path):
        print(f"    ❌ The input file '{file_path}' does not exist.")
        df = pd.DataFrame(columns=required_columns)
    else:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Initialize an empty DataFrame for the output
    output_df = pd.DataFrame()

    # Process each required column
    for column in required_columns:
        if column in df.columns:
            # Remove blank and duplicate values
            cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
            # Add cleaned data to the output DataFrame
            output_df[column.replace("Concatenated", "")] = cleaned_data
        else:
            print(f"    ❌ Column '{column}' is missing in the input file.")

    # Write the processed data to a new Excel file if there's any data to write
    if not output_df.empty:
        output_file = "Extracts/Account_and_Ownerid_extract.xlsx"
        output_df.to_excel(output_file, index=False)
        print(f"    ✅ Processed data has been written to Account_and_Ownerid_extract.xlsx")
    else:
        print("    ❌ No columns were processed due to missing columns.")

    # ====================================================
    # To extract account id to text file 
    # ====================================================


    # Load the Excel file
    extract_file_path = "Extracts/Account_and_Ownerid_extract.xlsx"  # Change this to your actual file path
    df = pd.read_excel(extract_file_path)
    os.makedirs('Delete', exist_ok=True)
    # Extract the "accountid" column values
    if "accountid" in df.columns:
        account_ids = df["accountid"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/1_account_ids.txt", "w") as f:
            f.write("\n".join(account_ids))

    else:
        print("Column 'accountid' not found in the sheet.")
    
    # ====================================================
    # To extract user id to text file 
    # ====================================================

    # Load the Excel file
    extract_file_path = "Extracts/Account_and_Ownerid_extract.xlsx"  # Change this to your actual file path

    df = pd.read_excel(extract_file_path)

    # Extract values from 'ownerid' and 'concatenatedcreatedby' (even if their lengths differ)
    ownerid_values = df["ownerid"].dropna().astype(str).tolist() if "ownerid" in df.columns else []
    createdby_values = df["concatenatedcreatedby"].dropna().astype(str).tolist() if "concatenatedcreatedby" in df.columns else []

    # Combine both lists while maintaining all values
    all_values = ownerid_values + createdby_values  # Concatenating both lists

    # Save to a text file
    with open("Delete/2_userid.txt", "w") as f:
        f.write("\n".join(all_values))


    def remove_last_char_from_last_line(extract_file):
        try:
            # Read all lines from the file
            with open(extract_file, 'r') as file:
                lines = file.readlines()

            # Check if the file is not empty
            if lines:
                # Remove the last character from the last line
                lines[-1] = lines[-1][:-1]

                # Write the modified content back to the file
                with open(extract_file, 'w') as file:
                    file.writelines(lines)

            # print("Last character from the last line has been removed.")
        
        except Exception as e:
            print(f"Error: {e}")
    
    # ========================================================================  
    
    # Code to remove comma from the text file
    
    remove_last_char_from_last_line('Delete/1_account_ids.txt')
    
    remove_last_char_from_last_line('Delete/2_userid.txt')
    
    
    # ======================================================================
    # Step 15: To copy the extracted data to main file
    #   • Copy data from downloaded CSV files into the rough sheets of the main file for vlookups.
    #   • If the files are missing, the script repeatedly asks whether to retry or stop.
    # ======================================================================


    print("\n\n🔍 Step 15: Copying extracted data to main file...")

    accounts_csv = "/Users/avirajmore/Downloads/accounts.csv"  # Specify the accounts CSV file path
    userid_csv = "/Users/avirajmore/Downloads/userid.csv"  # Specify the userid CSV file path

    # Check if the CSV files exist, and prompt to retry if not
    while not os.path.exists(accounts_csv):
        print(f"\n    ❌ Error: File 'accounts.csv' does not exist. Did you query the accounts?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    while not os.path.exists(userid_csv):
        print(f"\n    ❌ Error: File 'userid.csv' does not exist. Did you query the Userid?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n         ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    # Read the CSV files
    accounts_df = pd.read_csv(accounts_csv, usecols=[0, 1])  # Read first two columns
    userid_df = pd.read_csv(userid_csv, usecols=[0, 1, 2, 3])  # Read first four columns

    # Load the Excel file or create a new one if it doesn't exist
    if os.path.exists(file_path):
        book = openpyxl.load_workbook(file_path)
        if "Opportunity_Copy" not in book.sheetnames:
            sheet = book.create_sheet(title="Opportunity_Copy")
        else:
            sheet = book["Opportunity_Copy"]
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Opportunity_Copy"

    # Write the headers to the "Opportunity_Copy" sheet
    for col_index, header in enumerate(accounts_df.columns, start=1):
        sheet.cell(row=1, column=col_index, value=header)
    for col_index, header in enumerate(userid_df.columns, start=4):  # Change start index to 4
        sheet.cell(row=1, column=col_index, value=header)

    # Write the data to the "Opportunity_Copy" sheet
    max_length = max(len(accounts_df), len(userid_df))

    for row_index in range(max_length):
        if row_index < len(accounts_df):
            sheet.cell(row=row_index + 2, column=1, value=accounts_df.iloc[row_index, 0])
            sheet.cell(row=row_index + 2, column=2, value=accounts_df.iloc[row_index, 1])
        if row_index < len(userid_df):
            sheet.cell(row=row_index + 2, column=4, value=userid_df.iloc[row_index, 0])  # Change column to 4
            sheet.cell(row=row_index + 2, column=5, value=userid_df.iloc[row_index, 1])  # Change column to 5
            sheet.cell(row=row_index + 2, column=6, value=userid_df.iloc[row_index, 2])  # Change column to 6
            sheet.cell(row=row_index + 2, column=7, value=userid_df.iloc[row_index, 3])  # Change column to 7

    # Save the changes
    book.save(file_path)
    print(f"\n    ✅ 'Accounts' and 'Userid' Data has been successfully copied '{file_path.split('/')[-1]}'.")


    # ======================================================================
    # Step 16: Check how many Accounts are present in ISC
    #   • Perform vlookup on the 'Accountid' column using the rough sheet to fetch Salesforce IDs.
    #   • Handle duplicate Salesforce IDs by prompting you to select one.
    #   • Populate unmatched accounts with "Not present in ISC" and display the count of such accounts.
    # ======================================================================


    print("\n\n🔍 Step 16: Checking how many Accounts are present in ISC...")

    # file_path = 'path_to_your_excel_file.xlsx'  # Replace with your actual file path
    opportunity_sheet_name = 'Opportunity'
    opportunity_copy_sheet_name = 'Opportunity_Copy'

    try:
        # Read data from Excel sheets
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)
        
        # Filter out rows where 'Id' is NaN
        opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['Id'])
        
        # Check for duplicate AccountNumbers with different Id values
        duplicate_accounts = opportunity_copy_df_no_nan[
            opportunity_copy_df_no_nan.duplicated(subset=['AccountNumber'], keep=False)
        ]
        
        if not duplicate_accounts.empty:
            print("\n    ❗️ Duplicate AccountNumbers found with multiple Id values:")
            
            # Group by AccountNumber and prompt user for resolution
            for account_number, group in duplicate_accounts.groupby('AccountNumber'):
                print(f"\n        🔹 AccountNumber: {account_number}")
                
                # Display the 'Id' and corresponding Excel row number
                for idx, row in group.iterrows():
                    excel_row_number = idx + 2  # Adjust for Excel row numbering
                    print(f"\n           🔸 Id: {row['Id']} (Excel Row {excel_row_number})")
                
                # Prompt user to choose the Id to keep
                valid_ids = group['Id'].tolist()
                while True:
                    chosen_id = input(f"\n        🔹 Select id for AccountNumber {account_number} from above Ids: ").strip()
                    if chosen_id in valid_ids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id from {valid_ids}. ")
                
                # Filter DataFrame to keep only the chosen Id for the AccountNumber
                opportunity_copy_df = opportunity_copy_df[
                    ~((opportunity_copy_df['AccountNumber'] == account_number) & 
                    (opportunity_copy_df['Id'] != chosen_id))
                ]
        
        # Merge DataFrames
        merged_df = pd.merge(opportunity_df, opportunity_copy_df[['AccountNumber', 'Id']],
                            on='AccountNumber', how='left')
        
        # Handle NaN values
        merged_df['Id'] = merged_df['Id'].fillna('Not in ISC')
        
        # Rename columns
        merged_df.rename(columns={'Id': 'In ISC or Not'}, inplace=True)
        
        # Save updated DataFrame to Excel
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            merged_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)
        
        # Count 'Not in ISC'
        sheet_name = opportunity_sheet_name
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        not_in_isc_count = (df['In ISC or Not'] == 'Not in ISC').sum()
        
        # Print count
        print(f"\n    ❗️ Count of accounts Not in ISC: {not_in_isc_count}")

    except FileNotFoundError:
        print("\n    ❌ Error: The specified file was not found. Please check the file path.")
    except ValueError as e:
        print(f"\n    ❌ Error: {e}")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")


    # ======================================================================
    # Step 17: Rename 'Id' to 'userid' in Opportunity_Copy sheet
    #   • Rename the duplicate 'Id' column (from the Userid file) to 'Userid' for clarity after Step 15 merges CSV data into rough sheets.
    # ======================================================================


    # Define constants
    DEFAULT_USERID = '0053h000000sdCVAAY'
    # file_path = 'your_file_path.xlsx'
    print("\n\nStep 17: Renaming 'Id' to 'userid' in Opportunity_Copy sheet")

    try:
        # Load the Excel workbook
        wb = load_workbook(file_path)
        sheet_name = 'Opportunity_Copy'

        # Check if sheet exists
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet['E1'] = 'userid'
            wb.save(file_path)
            print("\n    ✅ Success: Renamed 'Id' to 'userid' in Opportunity_Copy sheet.")
        else:
            print(f"\n    ❗️Warning: Sheet '{sheet_name}' not found in the workbook.")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 18: Get the IDs of the Opportunity Owner
    #   • Perform vlookup on 'Ownerid' to retrieve Salesforce IDs.
    #   • Handle duplicate IDs by prompting selection.
    #   • Populate unmatched emails with the "Datamigration" Salesforce ID and display their count.
    # ======================================================================


    print("\n\n🔍 Step 18: Fetching IDs of Opportunity Owners...")

    try:
        # Sheet names
        opportunity_sheet_name = 'Opportunity'
        opportunity_copy_sheet_name = 'Opportunity_Copy'

        # Load data from sheets
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

        # Clean 'Trimmed_ownerid' and 'Email' columns
        if 'Trimmed_ownerid' in opportunity_df.columns:
            opportunity_df['Trimmed_ownerid'] = opportunity_df['Trimmed_ownerid'].str.strip().str.lower()
            # print("\n    ✅ 'Trimmed_ownerid' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Trimmed_ownerid' not found in the Opportunity sheet.")
            sys.exit()

        if 'Email' in opportunity_copy_df.columns:
            opportunity_copy_df['Email'] = opportunity_copy_df['Email'].str.strip().str.lower()
            # print("\n    ✅ 'Email' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Email' not found in the Opportunity_Copy sheet.")
            sys.exit()

        # Handle duplicates and NaN in 'userid'
        opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['userid'])
        duplicate_emails = opportunity_copy_df_no_nan[
            opportunity_copy_df_no_nan.duplicated(subset=['Email'], keep=False)
        ]

        if not duplicate_emails.empty:
            print("\n    ❗️ Duplicate Email IDs with multiple UserIDs found:")
            for email, group in duplicate_emails.groupby('Email'):
                print(f"\n        📧 Email: {email}")
                for idx, row in group.iterrows():
                    excel_row = idx + 2  # Adjust row number for Excel
                    print(f"\n           🔸 UserID: {row['userid']} (Row {excel_row})")
                
                # Collect valid UserIDs for this Email
                valid_userids = group['userid'].tolist()
                
                # Ask user to select a valid UserID
                while True:
                    chosen_userid = input(f"\n        🔹 Select id for UserId '{email}' from above Ids: ").strip()
                    if chosen_userid in valid_userids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id . ")
                
                # Filter DataFrame to keep only the chosen UserID for the Email
                opportunity_copy_df = opportunity_copy_df[
                    ~((opportunity_copy_df['Email'] == email) & (opportunity_copy_df['userid'] != chosen_userid))
                ]
            print("\n    ✅ Duplicate emails handled successfully.")

        # Perform left join to map 'userid' to 'Trimmed_ownerid'
        result_df = pd.merge(
            opportunity_df,
            opportunity_copy_df[['Email', 'userid']],
            left_on='Trimmed_ownerid',
            right_on='Email',
            how='left'
        )
        # print("\n    ✅ Merged 'Opportunity' and 'Opportunity_Copy' sheets.")

        # Handle NaN values in 'userid'
        nan_before = result_df['userid'].isna().sum()
        result_df['userid'] = result_df['userid'].fillna(DEFAULT_USERID)
        nan_after = result_df['userid'].isna().sum()

        # Drop redundant columns and rename 'userid' to 'OwnerId'
        result_df.drop(columns=['Email'], inplace=True)
        result_df.rename(columns={'userid': 'OwnerId'}, inplace=True)

        # Save the updated data back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            result_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

        print("\n    ✅ Success: IDs for Opportunity Owners updated successfully.")
        print(f"\n    ❗️ Number of invalid 'userid' values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 19: To get IDs of the Created By
    #   • Same as Step 18, but applied to the 'Created By' column.
    # ======================================================================


    print("\n\n🔍 Step 19: Fetching IDs of 'Created By'...")

    # Name of the sheets to target
    opportunity_sheet_name = 'Opportunity'
    opportunity_copy_sheet_name = 'Opportunity_Copy'

    try:
        # Read the Excel files into DataFrames
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

        # Check if 'created_by' column exists and has non-blank values
        if 'created_by' not in opportunity_df.columns or opportunity_df['created_by'].dropna().empty:
            if 'created_by' not in opportunity_df.columns:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column does not exist in 'Opportunity' sheet.")
            elif opportunity_df['created_by'].dropna().empty:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column is empty in 'Opportunity' sheet.")
        else:
            # Filter out rows where 'Email' or 'userid' are NaN
            opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['Email', 'userid'])

            # Check for duplicate emails with multiple userids
            duplicate_emails = opportunity_copy_df_no_nan[opportunity_copy_df_no_nan.duplicated(subset=['Email'], keep=False)]

            if not duplicate_emails.empty:
                print("\n    ❗️ Duplicate Email IDs found with multiple UserIDs:")
                for email, group in duplicate_emails.groupby('Email'):
                    print(f"\n        📧 Email: {email}")
                    
                    # Display the 'userid' and corresponding Excel row number
                    for idx, row in group.iterrows():
                        excel_row_number = idx + 2  # Adjust for 0-based index and Excel rows starting from 2
                        print(f"\n           🔸 UserID: {row['userid']} (Excel Row {excel_row_number})")

                    # Prompt the user to choose the userid to keep
                    valid_userids = group['userid'].tolist()
                    while True:
                        chosen_userid = input(f"\n        🔹 Enter the UserID to keep for Email '{email}' from the above options: ").strip()
                        if chosen_userid in valid_userids:
                            break
                        else:
                            print("\n           ❌ Invalid input. Please choose a valid UserID from the options above.")

                    # Filter DataFrame to keep only the chosen UserID for the Email
                    opportunity_copy_df = opportunity_copy_df[
                        ~((opportunity_copy_df['Email'] == email) & (opportunity_copy_df['userid'] != chosen_userid))
                    ]
                print("\n    ✅ Duplicate emails handled successfully.")

            # Perform VLOOKUP operation using 'created_by' and 'Email'
            merged_df = pd.merge(
                opportunity_df,
                opportunity_copy_df[['Email', 'userid']],
                left_on='created_by',
                right_on='Email',
                how='left'
            )

            # Rename the 'userid' column to 'createdbyid'
            merged_df.rename(columns={'userid': 'createdbyid'}, inplace=True)

            # Count NaN values in 'createdbyid' column before filling
            nan_before = merged_df['createdbyid'].isna().sum()

            # Replace NaN values in 'createdbyid' column with the specified default value
            default_userid = '0053h000000sdCVAAY'
            merged_df['createdbyid'] = merged_df['createdbyid'].fillna(default_userid)

            # Count NaN values in 'createdbyid' column after filling
            nan_after = merged_df['createdbyid'].isna().sum()

            # Update the 'Opportunity' sheet with the new column
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                merged_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

            print("\n    ✅ Successfully fetched and updated 'Created By' IDs.")
            print(f"\n    ❗️ Number of invalid user values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 20: Renaming Columns
    #   • Rename all columns in the opportunity sheet to match API names for seamless mass loading.
    #   • Prompt to continue or abort if required columns are missing.

    # ======================================================================


    print("\n\n🔍 Step 20: Renaming Columns...")

    sheet_name = 'Opportunity'

    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunity_legacy_id_c': 'opportunity_legacy_id__c',
        'legacy_opportunity_split_id_c': 'Legacy_Opportunity_Split_Id__c',
        'In ISC or Not': 'AccountId',
        'sales_stage': 'StageName',
        'won reason': 'Won_Reason__c',
        'lost category': 'Lost_Category__c',
        'lost reason': 'Lost_Reason__c',
        'currency_code': 'CurrencyIsoCode',
        'next_step': 'NextStep',
        'oi_source': 'OI_Group__c',
        'expected_close_date': 'CloseDate',
        'ownerid': 'Email',
    }

    try:
        # Read the Excel file
        excel_data = pd.read_excel(file_path, sheet_name=None)

        if sheet_name in excel_data:
            df = excel_data[sheet_name]

            # Check for missing columns
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]

            if missing_columns:
                print("\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n        - {col}")
                
                while True:  # Loop until a valid input is provided
                    proceed = input("\n    ❓ Do you want to proceed with renaming the available columns? (yes/no): ").strip().lower()
                    if proceed == 'yes':
                        break  # Exit the loop and proceed with the operation
                    elif proceed == 'no':
                        print("\n        ❌ Operation aborted.")
                        sys.exit(1)  # Exit the program
                    else:
                        print("\n        ❗️ Invalid choice. Please enter 'yes' or 'no'.")  # Prompt for valid input

            # Rename columns
            df.rename(columns=column_rename_mapping, inplace=True)

            # Save the changes back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            print("\n    ✅ Columns renamed successfully.")
        else:
            print(f"\n    ❌ Sheet '{sheet_name}' not found in the Excel file.")

    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
        sys.exit(1)


    # ======================================================================
    # Step 21: Rearrange the Columns in the Opportunity Copy
    #   • Rearrange columns to prioritize important fields, grouping related ones (e.g., account number and account ID) and moving less important ones to the end.
    # ======================================================================


    print("\n\n🔍 Step 21: Rearranging Columns...")

    desired_column_order = [
        'opportunity_legacy_id__c',
        'Legacy_Opportunity_Split_Id__c',
        'name',
        'AccountNumber',
        'AccountId',
        'StageName',
        'Won_Reason__c',
        'Lost_Category__c',
        'Lost_Reason__c',
        'CloseDate',
        'CurrencyIsoCode',
        'Email',
        'OwnerId',
        'NextStep',
        'OI_Group__c',
        'created_by',
        'createdbyid',
        'Pricebook2Id',
        'RecordTypeId',
        'modified_by',
        'created_date',
        'modified_date',
        'pricebook',
        'recordtypeid',
        'Trimmed_accountid',
        'Trimmed_ownerid',
        'accountid',
        'Concatenatedaccountid',
        'Concatenatedownerid',
        'concatenatedcreatedby'
    ]

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check for missing and extra columns
        missing_columns = [col for col in desired_column_order if col not in df.columns]
        extra_columns = [col for col in df.columns if col not in desired_column_order]

        # Rearrange columns
        rearranged_columns = [col for col in desired_column_order if col in df.columns]
        rearranged_columns += extra_columns  # Add extra columns to the end

        # Save the changes back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df[rearranged_columns].to_excel(writer, sheet_name=sheet_name, index=False)

        if missing_columns:
            print("\n    ❌ The following columns are missing and were skipped:")
            for col in missing_columns:
                print(f"\n        🔸  {col}")

        if extra_columns:
            print("\n    ❗️ The following extra columns were moved to the end:")
            for col in extra_columns:
                print(f"\n        🔸  {col}")

        print("\n    ✅ Columns rearranged successfully.")
    except FileNotFoundError:
        print(f"\n    ❌ File '{file_path}' not found.")
        sys.exit(1)


    # ======================================================================
    # Step 22: Final Row and Column Count
    #   • Recount rows after processing to ensure no extra rows were added mistakenly.
    #   • If there's a mismatch, prompt to either continue or stop.
    # ======================================================================


    print("\n\n🔍 Step 22: Final Row and Column Count...")

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        oppty_final_num_rows = df.shape[0]
        oppty_final_num_columns = df.shape[1]

        print(f"\n    ✅ Final row count: {oppty_final_num_rows}")

        if oppty_initial_num_rows != oppty_final_num_rows:
            print("\n    ❗️ Row count mismatch detected!")
            print(f"\n       📊 Initial: {oppty_initial_num_rows}")
            print(f"\n       📊 Final: {oppty_final_num_rows}")

            while True:
                user_input = input("\n    ❓Do you want to continue? Type 'continue' to proceed or 'no' to abort: ").strip().lower()
                if user_input == "continue":
                    print("\n       ✅ Continuing the program...")
                    break
                elif user_input == "no":
                    print("\n       ❌ Terminating the program...")
                    sys.exit(1)
                else:
                    print("\n       ❌ Invalid input. Please type 'continue' to proceed or 'no' to stop.")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
        sys.exit(1)


    # =========================================================================================================================================
    #                                                PRODCUT SHEET EXECUTION
    # =========================================================================================================================================


    # ======================================================================
    # file_path = "/Users/avirajmore/Downloads/Avi 3 copy 2.xlsx"
    # ======================================================================
    print("\n")
    print("=" * 100)
    print(" " * 33 + "📝 PRODUCT SHEET EXECUTION 📝")
    print("=" * 100)

    # ======================================================================
    # Step 1: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 1: Counting the rows and columns...")

    # Name of the sheet to target
    product_sheet_name = 'Opportunity_product'

    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path, sheet_name = 'Opportunity_product')

    # Get the number of rows and columns
    product_initial_num_rows = df.shape[0]     # Number of rows in the DataFrame
    product_num_columns = df.shape[1]          # Number of columns in the DataFrame

    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {product_initial_num_rows}")
    # print(f"\n    ✅ Initial column count: {product_num_columns}")


    # ======================================================================
    # Step 2:- Removing duplicate rows and blank rows...
    # ======================================================================

    print("\n\n🔍 Step 2: Removing duplicate rows and blank rows...")

    def remove_blank_rows(file_path, product_sheet_name):
        try:
            # Try to read the spreadsheet with the given sheet name
            df = pd.read_excel(file_path, sheet_name='Opportunity_product')

            # Drop duplicate rows
            # df = df.drop_duplicates()

            # Remove rows where all cells are NaN (blank rows)
            df = df.dropna(how='all')

            # Save the cleaned data back to the same file without modifying formatting
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=product_sheet_name, index=False)

            print(f"\n    ✅ Removed all the blank rows from '{product_sheet_name}' sheet. ")

        except ValueError as e:
            # Handle the case where the sheet does not exist
            print(f"\n    ❌ Error: The sheet '{product_sheet_name}' does not exist in the file. ")
            sys.exit()
        except Exception as e:
            # Handle any other exceptions
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
            sys.exit()

    # Call the function for 'Opportunity_product' sheet
    remove_blank_rows(file_path, 'Opportunity_product')


    # ======================================================================
    # Step 3 :- Add Exsising column, To check if the given Opportunities are present in the Opportunity Sheet 
    # ======================================================================

    print("\n\n🔍 Step 3: Verifying opportunities in the 'Opportunity' sheet...")

    opportunity_sheet_name = 'Opportunity'
    product_sheet_name = 'Opportunity_product'

    try:
        # Load the sheets into DataFrames
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        product_df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Validate the required columns
        if 'opportunity_legacy_id__c' not in opportunity_df.columns:
            print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet. ")
            sys.exit()
        elif 'opportunityid' not in product_df.columns:
            print(f"\n    ❌ Column 'opportunityid' not found in the '{product_sheet_name}' sheet. ")
            sys.exit()

        # Perform the comparison
        product_df['Existing'] = product_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

        # Calculate the number of false values
        false_count = (~product_df['Existing']).sum()

        # Save the updated data back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            product_df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message with false count
        print(f"\n    ✅ Verification completed. 'Existing' column has been added to the '{product_sheet_name}' sheet. ")
        print(f"\n    ❗️ Number of False values in 'Existing' column: {false_count}")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File not found. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 4: Formatting the date column
    # ======================================================================

    print("\n\n🔍 Step 4: Formatting the date column in the 'Opportunity_product' sheet...")

    product_sheet_name = 'Opportunity_product'  # Replace with the actual sheet name
    date_column = 'expiration date'  # Replace with the actual column name containing the dates

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Check if the specified column exists in the DataFrame
        if date_column not in df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing from the sheet '{product_sheet_name}'. ")
            sys.exit(1)  # Exit the script if the column is missing

        # Ensure the date column is in datetime format and then format it as YYYY-MM-DD
        # print(f"\n    🔄 Formatting the '{date_column}' column...")
        df[date_column] = pd.to_datetime(df[date_column]).dt.strftime('%Y-%m-%d')

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ Date column formatted successfully in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 5: Adding Quantity Columns
    # ======================================================================

    # Code to delete quantity column first

    try:
        # Load workbook and target sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        # Check if 'quantity' column exists
        header_row = [cell.value for cell in sheet[1]]
        if 'quantity' in header_row:
            col_index = header_row.index('quantity') + 1

            # Shift columns left to delete the 'quantity' column
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=sheet.max_column):
                for cell_index, cell in enumerate(row[:-1]):
                    cell.value = row[cell_index + 1].value
                row[-1].value = None  # Clear the last column cell

        # Save changes
        wb.save(file_path)
    except Exception:
        pass

    # Main step
    print("\n\n🔍 Step 5: Adding the 'Quantity' column in the 'Opportunity_product' sheet...")

    product_sheet_name = 'Opportunity_product'  # Replace with the actual sheet name
    new_column_name = 'Quantity'  # Column name to be added
    default_value = 1  # Default value for the new column

    try:
        # Load the workbook and target the specified sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        # Add new column header at the end of existing headers
        # print(f"\n    🔄 Adding the '{new_column_name}' column header...")
        sheet.cell(row=1, column=sheet.max_column + 1, value=new_column_name)

        # Iterate over rows and set default value for the new column
        # print(f"\n    🔄 Setting the default value '{default_value}' for the new column...")
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming headers in row 1)
            sheet.cell(row=row, column=sheet.max_column, value=default_value)

        # Save the workbook
        wb.save(file_path)

        # Success message
        print(f"\n    ✅ A new column '{new_column_name}' has been added to the '{product_sheet_name}' sheet with default value '{default_value}'. ✅")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 6: To create a new Currency column
    # ======================================================================
    print("\n\n🔍 Step 6: Creating or Overwriting the 'Currency' column in the 'Opportunity_product' sheet...")

    opportunity_product_df = pd.read_excel(file_path, sheet_name="Opportunity_product")
    opportunity_df = pd.read_excel(file_path, sheet_name="Opportunity")

    try:
        # Perform VLOOKUP operation (merge data)
        # print(f"\n    🔄 Merging data to create or overwrite the 'opportunity currency' column...")
        merged_df = pd.merge(opportunity_product_df, opportunity_df,
                            left_on="opportunityid", right_on="opportunity_legacy_id__c",
                            how="left")

        # Overwrite or add the "opportunity currency" column
        opportunity_product_df["opportunity currency"] = merged_df["CurrencyIsoCode"]

        # Save the modified DataFrame back to Excel
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            opportunity_product_df.to_excel(writer, sheet_name="Opportunity_product", index=False)

        # Success message
        print(f"\n    ✅ Process completed. The 'opportunity currency' column has been successfully created in the 'Opportunity_product' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError as e:
        # Handle missing column error
        print(f"\n    ❌ Error: The required column '{e.args[0]}' is missing. ")
        sys.exit()
    except Exception as e:
        # Handle any other unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 7: To delete unwanted columns from the sheet
    # ======================================================================

    print("\n\n🔍 Step 7: Deleting unwanted columns from the 'Opportunity_product' sheet...")

    # Sheet name and columns to delete
    product_sheet_name = "Opportunity_product"
    columns_to_delete = [
        "created_by",
        "current quarter revenue",
        "modified_by",
        "created_date",
        "modified_date",
        "product_code_family",
        "pricebookentryid"
    ]

    try:
        # Load the workbook and target the specified sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        columns_deleted = []

        # Check and delete each specified column
        for col_name in columns_to_delete:
            found = False
            for col in sheet.iter_cols():
                if col[0].value == col_name:
                    sheet.delete_cols(col[0].column)
                    columns_deleted.append(col_name)
                    found = True
                    # print(f"\n        🔸 '{col_name}' deleted. ")
                    break
            if not found:
                print(f"\n        🔸 '{col_name}' not found. ")

        # Save the workbook
        wb.save(file_path)

        # Success or no deletion message
        if columns_deleted:
            print(f"\n    ✅ Successfully deleted mentioned columns from the '{product_sheet_name}' sheet:")
        else:
            print(f"\n    ❌ No columns from the specified list were found in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 8: To get the "Product_Code_Family" column
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'Product_Code_Family' column in the 'Opportunity_product' sheet...")

    product_sheet_name = "Opportunity_product"

    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Check if required columns exist
        if "product" not in df.columns:
            print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "product_type" not in df.columns:
            print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Concatenate values from 'product' and 'product_type' columns with a hyphen
        # print(f"    🔄 Creating 'Product_Code_Family' column by concatenating 'product' and 'product_type'...")
        df["Product_Code_Family"] = df["product"] + "-" + df["product_type"]

        # Save the updated DataFrame back to the same sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ The 'Product_Code_Family' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 9: To get the "Practise_Multiple country" column
    # ======================================================================

    print("\n\n🔍 Step 9: Creating 'Practise_Multiple country' column in the 'Opportunity_product' sheet...")

    product_sheet_name = "Opportunity_product"

    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Check if required columns exist
        if "product" not in df.columns:
            print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "product_type" not in df.columns:
            print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "opportunity currency" not in df.columns:
            print(f"\n    ❌ Error: Column 'opportunity currency' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Concatenate values from 'product', 'product_type', and 'opportunity currency' columns with a hyphen
        # print(f"    🔄 Creating 'Practise_Multiple country' column by concatenating 'product', 'product_type', and 'opportunity currency'...")
        df["Practise_Multiple country"] = df["product"] + "-" + df["product_type"] + "-" + df["opportunity currency"]

        # Save the updated DataFrame back to the same sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ The 'Practise_Multiple country' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 10: To Concatenate the Currency and Product Family
    # ======================================================================

    print("\n\n🔍 Step 10: Concatenating 'Currency' and 'Product Family' columns...")

    product_sheet_name = "Opportunity_product"

    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Check if required columns exist
        if "Product_Code_Family" not in df.columns:
            print(f"\n    ❌ Error: Column 'Product_Code_Family' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "opportunity currency" not in df.columns:
            print(f"\n    ❌ Error: Column 'opportunity currency' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Task 1: Concatenate values from 'Product_Code_Family' column with inverted commas and commas
        # print(f"    🔄 Concatenating 'Product_Code_Family' column with inverted commas and commas...")
        df["Concatenated Product Family"] = "'" + df["Product_Code_Family"] + "',"

        # Task 2: Concatenate values from 'opportunity currency' column with inverted commas and commas
        # print(f"    🔄 Concatenating 'opportunity currency' column with inverted commas and commas...")
        df["Concatenated Currency"] = "'" + df["opportunity currency"] + "',"

        # Save the updated DataFrame back to the same sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ 'Concatenated Product Family' and 'Concatenated Currency' columns have been added successfully. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 11: To keep the decimal values as 2
    # ======================================================================

    print("\n\n🔍 Step 11: Formatting decimal values to two decimal places...")

    product_sheet_name = 'Opportunity_product'
    headers_to_format = ['unitprice', 'expiring amount']

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        # Function to format numbers to two decimal places and convert to float
        def format_to_float(value):
            if isinstance(value, (int, float)):
                return float(f'{value:.2f}')
            return value

        # Find column indices based on headers
        column_indices = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header in headers_to_format:
                column_indices[header] = col

        # Iterate through each header and format numbers
        for header in headers_to_format:
            col = column_indices.get(header)
            if col:
                # print(f"    🔄 Formatting '{header}' column...")
                for row in range(2, sheet.max_row + 1):  # Start from row 2 assuming headers in row 1
                    cell = sheet.cell(row=row, column=col)
                    formatted_value = format_to_float(cell.value)
                    sheet.cell(row=row, column=col, value=formatted_value)

        # Save the workbook
        wb.save(file_path)

        # Success message
        print(f"\n    ✅ Numbers in columns {', '.join(headers_to_format)} have been formatted to two decimal places. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 12: To extract the concatenated values
    # ======================================================================

    print("\n\n🔍 Step 12: Extracting concatenated values...")

    def process_excel_file(file_path, sheet_name, required_columns, output_file):
        # Check if the input file exists
        if not os.path.exists(file_path):
            print(f"\n    ❌ Error: The input file '{file_path}' does not exist. ")
            return

        try:
            # Read the Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            print(f"\n    ❌ Error: Failed to read the Excel file. Details: {e} ")
            return

        # Initialize a dictionary to hold cleaned data for each column
        cleaned_data_dict = {}

        # Process each required column
        for column in required_columns:
            if column in df.columns:
                # Remove blank and duplicate values
                cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
                cleaned_data_dict[column.replace("Concatenated", "").strip()] = cleaned_data
            else:
                print(f"\n    ❌ Error: Column '{column}' is missing. ")

        # Create an empty DataFrame for the output
        output_df = pd.DataFrame()

        # Add each cleaned column as a separate DataFrame and concatenate them
        for key, cleaned_data in cleaned_data_dict.items():
            output_df = pd.concat([output_df, pd.DataFrame({key: cleaned_data})], axis=1, ignore_index=False)

        # Write the processed data to a new Excel file if there's any data to write
        if not output_df.empty:
            try:
                output_df.to_excel(output_file, index=False)
                print(f"\n    ✅ Data written to '{output_file}'. ")
            except Exception as e:
                print(f"\n    ❌ Error: Failed to write the Excel file. Details: {e} ")
        else:
            print("\n    ❌ Error: No data to process. ")

    # Specify the input file path, sheet name, required columns, and output file path
    sheet_name = "Opportunity_product"  # Specify the sheet name
    required_columns = ["Concatenated Product Family", "Concatenated Currency"]
    output_file = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Specify the output file path

    # Process the Excel file
    process_excel_file(file_path, sheet_name, required_columns, output_file)


    # ======================================================================
    # To extact product code to text file
    # ======================================================================

    import pandas as pd
    # Load the Excel file
    extract_file_path = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Change this to your actual file path
    df = pd.read_excel(extract_file_path)

    # Extract the "accountid" column values
    if "Product Family" in df.columns:
        Product_Family = df["Product Family"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/3_product_code.txt", "w") as f:
            f.write("\n".join(Product_Family))

    else:
        print(f"Column not found in the sheet.")

    # ======================================================================
    # To extact currency code to text file
    # ======================================================================
    # Load the Excel file
    extract_file_path = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Change this to your actual file path
    df = pd.read_excel(extract_file_path)

    # Extract the "accountid" column values
    if "Currency" in df.columns:
        Product_Family = df["Currency"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/4_currency.txt", "w") as f:
            f.write("\n".join(Product_Family))

    else:
        print(f"Column not found in the sheet.")
    
    # ========================================================================  
    
    # Code to remove comma from the text file
    
    remove_last_char_from_last_line('Delete/3_product_code.txt')
    
    remove_last_char_from_last_line('Delete/4_currency.txt')
    
    # ======================================================================
    # Step 13:- To copy the data from CSV file
    # ======================================================================

    print("\n\n🔍 Step 13: Copying data from CSV file to Excel...")

    # Define the CSV file path
    csv_file_path = "/Users/avirajmore/Downloads/productfamily.csv"

    # Check if the CSV file exists, and prompt to retry if not
    while not os.path.exists(csv_file_path):
        print(f"\n    ❌ Error: The CSV file at path '{csv_file_path}' does not exist.")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        if try_again == 'no':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    # Read data from the CSV file
    df = pd.read_csv(csv_file_path)

    # Specify the Excel file path and sheet name
    sheet_name = "Opportunity_product_Copy"

    # Write data to the specified sheet in the Excel file
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"\n    ✅ Data from the CSV file has been successfully copied to the '{sheet_name}' sheet.")


    # ======================================================================
    # Step 14:- Create 'Practise_Multiple country' column in Product Copy sheet
    # ======================================================================

    print("\n\n🔍 Step 14: Create 'Practise_Multiple country' column in Product Copy sheet" )

    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name="Opportunity_product_Copy")

    # Concatenate values from "product" and "product_type" columns with a hyphen
    df["Practise_Multiple country"] = df["Product2.Product_Code_Family__c"] + "-" + df["CurrencyIsoCode"]

    # Save the updated DataFrame to the same sheet
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="Opportunity_product_Copy", index=False)

    print(f"\n    ✅ The values from have been successfully concatenated and saved in the 'Practise_Multiple country' column.")


    # ======================================================================
    # Step 15:- Getting the PricebookEntry id
    # ======================================================================

    print("\n\n🔍 Step 15: Getting the PricebookEntry id ...")

    # New code to check if the pricebook entry is active or not
    product_sheet_name = 'Opportunity_product'
    opportunity_copy_sheet_name = 'Opportunity_product_Copy'

    # Function to standardize column names
    def standardize_columns(df):
        df.columns = df.columns.str.strip().str.lower()
        return df

    # Function to standardize column values
    def standardize_column_values(df, column_name):
        df[column_name] = df[column_name].str.strip().str.lower()
        return df

    # Read the data from both sheets
    opportunity_df = pd.read_excel(file_path, sheet_name=product_sheet_name)
    opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

    # Standardize column names
    opportunity_df = standardize_columns(opportunity_df)
    opportunity_copy_df = standardize_columns(opportunity_copy_df)

    # Standardize column values for the merge key
    opportunity_df = standardize_column_values(opportunity_df, 'practise_multiple country')
    opportunity_copy_df = standardize_column_values(opportunity_copy_df, 'practise_multiple country')

    # Check if 'pricebookentryid' already exists in opportunity_df
    if 'pricebookentryid' in opportunity_df.columns:
        raise KeyError("❌ Error: Column 'pricebookentryid' already exists in 'Opportunity_product'. Please check your data processing steps.")

    # Filter the opportunity_copy_df based on 'IsActive' column
    opportunity_copy_df['pricebookentryid'] = opportunity_copy_df.apply(
        lambda row: row['id'] if row['isactive'] else 'Not Active', axis=1
    )

    # Perform a left join to get PriceBookEntryid for each Practise_Multiple country
    merged_df = pd.merge(opportunity_df, 
                        opportunity_copy_df[['practise_multiple country', 'pricebookentryid']], 
                        left_on='practise_multiple country', 
                        right_on='practise_multiple country',
                        how='left')

    # Fill missing values with 'No Pricebookid found'
    merged_df['pricebookentryid'] = merged_df['pricebookentryid'].fillna('No Pricebookid found')

    # Count occurrences of 'No Pricebookid found' and 'Not Active'
    count_no_pricebookid_found = (merged_df['pricebookentryid'] == 'No Pricebookid found').sum()
    count_not_active = (merged_df['pricebookentryid'] == 'Not Active').sum()

    # Save the updated DataFrame back to the same Excel file
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        merged_df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    print(f"\n    ✅ The 'Opportunity_product' sheet has been successfully updated with the 'PriceBookEntryid' column.")
    print(f"\n        ❗️ Count of 'No Pricebookid found': {count_no_pricebookid_found}")
    print(f"\n        ❗️ Count of 'Not Active': {count_not_active}")

    # ======================================================================
    # Step 16: Rearranging the Columns in Sequence
    # ======================================================================

    print("\n\n🔍 Step 16: Rearranging Columns in the 'Opportunity_product' Sheet...")

    # Name of the sheet to target
    sheet_name = 'Opportunity_product'

    # Specify the desired order of columns
    desired_column_order = [
        'opportunityid',
        'existing',
        'quantity',
        'product',
        'product_type',
        'product_code_family',
        'opportunity currency',
        'practise_multiple country',
        'pricebookentryid'
    ]

    try:
        # Read the Excel file
        excel_data = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Check if the sheet exists
        if isinstance(excel_data, pd.DataFrame):
            # Check if all specified columns exist
            missing_columns = [col for col in desired_column_order if col not in excel_data.columns]
            extra_columns = [col for col in excel_data.columns if col not in desired_column_order]

            # Rearrange columns
            rearranged_columns = [col for col in desired_column_order if col in excel_data.columns]

            # Add extra columns to the end
            rearranged_columns += extra_columns

            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                excel_data[rearranged_columns].to_excel(writer, sheet_name=sheet_name, index=False)

            # Notify the user about the changes
            if missing_columns:
                print("\n    ❌ The following columns were missing and were skipped:")
                for col in missing_columns:
                    print(f"\n      🔸 {col}")
            
            # if extra_columns:
            #     print("\n    🔄 The following extra columns were moved to the end of the sheet:")
            #     for col in extra_columns:
            #         print(f"\n      🔸 {col}")

            print(f"\n    ✅ Columns successfully rearranged in the '{sheet_name}' sheet of the file: {file_path.split('/')[-1]}")
        else:
            print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found.")


    # ======================================================================
    # Step 17: Rename the Columns
    # ======================================================================

    print("\n\n🔍 Step 17: Renaming Columns in the 'Opportunity_product' Sheet...")

    # Name of the sheet to target
    sheet_name = 'Opportunity_product'

    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunityid': 'Legacy_Opportunity_Split_Id__c',
        'quantity': 'Quantity',
        'product_code_family': 'Product_Family__c',
        'pricebookentryid': 'PricebookEntryId',
        'unitprice': 'UnitPrice',
        'term': 'Term__c',
        'classification type': 'Classification__c',
        'type': 'Type__c',
        'renewal type': 'Renewal_Type__c',
        'renewal status': 'Renewal_Status__c',
        'expiration date': 'Expiration_Date__c',
        'expiring term': 'Expiring_Term__c',
        'expiring amount': 'Expiring_Amount__c',
        'external id': 'External_IDs__c',
    }

    # Read the Excel file
    try:
        excel_data = pd.read_excel(file_path, sheet_name=None)
        
        # Check if the specified sheet exists
        if sheet_name in excel_data:
            # Access the specified sheet
            df = excel_data[sheet_name]
            
            # Check if all specified columns exist
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]
            
            # If any specified column is missing, notify the user and ask if they want to proceed
            if missing_columns:
                print("\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n      🔸 {col}")
                
                # Loop until a valid response is entered
                while True:
                    proceed = input("\n    🔹 Do you want to proceed with the execution? (yes/no): ").lower()
                    if proceed == 'yes':
                        break
                    elif proceed == 'no':
                        print("\n      🚫 Operation aborted by the user.")
                        exit()
                    else:
                        print("\n      ❗️ Invalid input. Please enter 'yes' or 'no'.")
            
            # Rename specified columns
            df.rename(columns=column_rename_mapping, inplace=True)
            
            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"\n    ✅ Columns renamed successfully in the '{sheet_name}' sheet of the file: {file_path.split('/')[-1]}")
        
        else:
            print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path.split('/')[-1]}' not found.")

    # ======================================================================
    # Step 18: Final Row and Column Count
    # ======================================================================

    print("\n\n🔍 Step 18: Final Row and Column Count...")

    # Name of the sheet to target
    product_sheet_name = 'Opportunity_product'

    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path, sheet_name=product_sheet_name)

    # Get the number of rows and columns in the DataFrame
    product_final_num_rows = df.shape[0]
    product_final_num_columns = df.shape[1]

    # Display the final row count
    print(f"\n    ✅ Final row count: {product_final_num_rows}")
    # print(f"\n    ✅ Final column count: {product_final_num_columns}")

    # Check if the number of rows has changed
    if product_initial_num_rows != product_final_num_rows:
        print(f"\n    ❗️ Row count mismatch detected!")
        print(f"\n       📊 Initial: {product_initial_num_rows}")
        print(f"\n       📊 Final: {product_final_num_rows}")

        while True:
            # Ask the user whether to continue or stop
            user_input = input(
                f"\n    🔹 Do you want to continue? Type 'continue' to proceed or 'no' to stop: "
            ).strip().lower()

            if user_input == "continue":
                print("\n      🔄 Continuing the program...")
                break  # Exit the loop and continue execution
            elif user_input == "no":
                print("\n      🚫 Terminating the program...")
                exit()  # Terminate the program
            else:
                print("\n      ❗️ Invalid input. Please type 'continue' to proceed or 'no' to stop.")


    # ======================================================================
    print("\n")
    print("=" * 100)
    print(" " * 33 + "📝 PRODUCT SHEET COMPLETED 📝")
    print("=" * 100)
    # ======================================================================


    # =========================================================================================================================================
    #                                                TEAM MEMBER SHEET EXECUTION
    # =========================================================================================================================================
    
    
    # Display the header once
    print("\n\n📄 Execute Next Sheet:")

    sheet_name = 'Opportunity_Team '
    
    def is_sheet_empty(file_path, sheet_name):
        """
        Checks if a given sheet in an Excel file contains any data beyond headers.
        Returns:
            - True, None if the sheet is empty or has only headers.
            - False, DataFrame (first 4 rows) if the sheet contains data.
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Check if the sheet is empty or only contains headers
            if df.empty or df.dropna(how='all').shape[0] == 0:
                return True, None  # Sheet is empty or has only headers
            
            return False, df.head(4)  # Sheet contains data, return first 4 rows
        except Exception as e:
            print(f"\n⚠️ Error reading sheet '{sheet_name}': {e}\n")
            return None, None
    is_empty, preview = is_sheet_empty(file_path, sheet_name)

    if is_empty:
        print(f"\n📂 The sheet '{sheet_name}' is empty or contains only headers.\n")
    elif is_empty is None:
        print("\n⚠️ Could not process the sheet due to an error.\n")
    else:
        print(f"\n✅ The sheet '{sheet_name}' contains data. Here are the first 4 rows:\n")
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    while True:
    
        choice = input("\n    🔹 Do you want to execute the Team member Sheet ? (yes/no): ").strip().lower()
        
        if choice == "yes":
            print(f"\n        ⏳ Executing the Sheet: Teammember sheet ")
            
            # ======================================================================
            print("\n")
            print("=" * 100)
            print(" " * 33 + "📝 TEAM MEMBER SHEET EXECUTION 📝")
            print("=" * 100)

            # ======================================================================

            # ======================================================================
            # Step 1:-  To rename the sheet to Opportunity Team
            # ======================================================================

            print("\n\n🔍 Step 1: Renaming Team Sheet...")

            # Target sheet name to find and rename
            target_name = 'Opportunity_team'  # Modified to lowercase 'team'

            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)

            # Function to normalize and clean sheet names for comparison
            def normalize_sheet_name(name):
                return name.strip().replace(' ', '').replace('-', '').lower().replace('_', '')

            # Flag to track if the sheet is found
            sheet_found = False
            sheet_to_rename = None

            # Iterate through sheet names and rename if found
            for sheet_name in workbook.sheetnames:
                if normalize_sheet_name(sheet_name) == normalize_sheet_name(target_name):
                    workbook[sheet_name].title = 'Opportunity_team_2'  # Renaming to lowercase 'team'
                    sheet_found = True
                    sheet_to_rename = sheet_name
                    print(f"\n    ✅ Sheet '{sheet_name}' has been successfully renamed to 'Opportunity_team_2'.")
                    break

            # If sheet not found, raise an error with the name of the sheet
            if not sheet_found:
                sheet_names = ", ".join(workbook.sheetnames)
                print(f"\n    ❌ ERROR: Sheet '{target_name}' not found.")
                sys.exit()

            # Save the workbook
            workbook.save(file_path)

            # ======================================================================
            # Step 2:- Check count of rows and columns
            # ======================================================================


            print("\n\n🔍 Step 2: Checking the Number of Rows and Columns...")

            # Name of the sheet to target
            opportunity_sheet_name = 'Opportunity_team_2'

            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

                # Get the number of rows and columns
                team__initial_num_rows = df.shape[0]
                team__initial_num_columns = df.shape[1]

                # Print the number of rows and columns
                print(f"\n    ✅ Initial rows count '{opportunity_sheet_name}' sheet: {team__initial_num_rows}\n")


            except FileNotFoundError:
                print("\n    ❌ ERROR: File not found. Please check the file path.\n")
            except ValueError:
                print(f"\n    ❌ ERROR: Sheet '{opportunity_sheet_name}' not found in the workbook.\n")
            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}\n")


            # ======================================================================
            # Step 3: Creating New Records for Multiple Emails in a Cell
            # ======================================================================


            print("\n\n🔍 Step 3: Creating New Records for Multiple Emails in a Cell...")

            # Load the existing workbook
            wb = load_workbook(file_path)

            # Load the DataFrame from the 'Opportunity_team_2' sheet
            df_opportunity_team_2 = pd.read_excel(file_path, sheet_name='Opportunity_team_2')

            # Initialize an empty list to store rows for the new DataFrame
            new_rows = []
            duplicate_count = 0  # Counter for duplicate rows
            skipped_blank_count = 0  # Counter for rows skipped due to blank values

            # Set to track duplicate rows
            seen_rows = set()

            # Iterate through each row in the original DataFrame
            for index, row in df_opportunity_team_2.iterrows():
                # Check if the row is already seen (duplicate)
                row_tuple = tuple(row.items())
                if row_tuple in seen_rows:
                    duplicate_count += 1
                    continue  # Skip duplicate rows
                seen_rows.add(row_tuple)

                # Handle blank or NaN values in 'opportunityid' and 'email'
                opportunity_id = row.get('opportunityid', None)
                emails = row.get('email', None)

                if pd.isna(opportunity_id) or pd.isna(emails):
                    skipped_blank_count += 1  # Increment skipped rows count
                    continue  # Skip rows with missing 'opportunityid' or 'email'

                # Split emails if multiple are present
                emails = str(emails).split(',')

                if len(emails) > 1:
                    for email in emails:
                        email = email.strip()  # Remove any whitespace
                        if email:  # Skip blank emails
                            new_row = row.copy()
                            new_row['email'] = email  # Assign a single email
                            new_rows.append(new_row)
                else:
                    # Handle rows with a single email
                    if emails[0].strip():  # Skip rows with blank single email
                        new_rows.append(row)

            # Create a new DataFrame with the processed rows
            df_Opportunity_team = pd.DataFrame(new_rows, columns=df_opportunity_team_2.columns)

            # Total row counts before and after removing duplicates
            total_rows_before = len(df_opportunity_team_2)
            total_rows_after = len(df_Opportunity_team)

            # Drop duplicate rows (if necessary)
            df_Opportunity_team = df_Opportunity_team.drop_duplicates()

            # Create a new sheet in the workbook
            sheet_name = 'Opportunity_team'
            ws = wb.create_sheet(title=sheet_name)

            # Convert DataFrame to rows and append to the new sheet
            for r_idx, row in enumerate(dataframe_to_rows(df_Opportunity_team, index=False, header=True), 1):
                ws.append(row)

            # Save the workbook
            wb.save(file_path)

            # Print results
            print(f"\n    ✅ New sheet '{sheet_name}' has been successfully created in the Excel file.\n")
            print(f"\n        🔸 Total rows before processing: {total_rows_before}")
            print(f"\n        🔸 Total rows after processing: {total_rows_after}")
            print(f"\n        🔸 Duplicate rows removed: {total_rows_before - total_rows_after}")
            print(f"\n        🔸 Rows skipped due to blank values in 'opportunityid' or 'email': {skipped_blank_count}")


            # ======================================================================
            # Step 4: Concatenating Email Values
            # ======================================================================

            print("\n\n🔍 Step 4: Concatenating Email Values...")

            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Select the sheet
            sheet = wb['Opportunity_team']

            # Find the column index of 'email'
            email_column_index = None
            for col in sheet.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    if cell.value == 'email':
                        email_column_index = cell.column
                        break
                if email_column_index is not None:
                    break

            if email_column_index is None:
                print("\n    ❌ ERROR: Column 'email' not found in the 'Opportunity_team' sheet.")
                raise ValueError("Column 'email' not found.")

            # Define the column header for the new column
            Concat_T_M_column_header = 'Concat_T_M'

            # Calculate the max row in the email column
            max_row = sheet.max_row

            # Process each row starting from the second row (assuming the first row is the header)
            rows_processed = 0  # Counter for processed rows
            for row in range(2, max_row + 1):
                # Get the value from the email column
                email_value = sheet.cell(row=row, column=email_column_index).value

                # Check if the email_value is not None
                if email_value is not None:
                    # Concatenate with inverted commas and comma
                    concatenated_value = f"'{email_value}',"

                    # Write the concatenated value to the new column
                    Concat_T_M_cell = sheet.cell(row=row, column=email_column_index + 1)
                    Concat_T_M_cell.value = concatenated_value

                    rows_processed += 1

            # Add the header for the new column
            sheet.cell(row=1, column=email_column_index + 1, value=Concat_T_M_column_header)

            # Save the workbook
            wb.save(file_path)

            # Print completion message
            print(f"\n    ✅ Concatenated email values.")


            # ======================================================================
            # Step 5 :- Checking if Opportunities Exist in the 'Opportunity' Sheet
            # ======================================================================

            print("\n\n🔍 Step 5: Checking if Opportunities Exist in the 'Opportunity' Sheet...")

            # File path and sheet names
            # file_path = 'your_file_path.xlsx'
            opportunity_sheet_name = 'Opportunity'
            Opportunity_team_sheet_name = 'Opportunity_team'

            try:
                # Load the sheets into DataFrames
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                Opportunity_team_df = pd.read_excel(file_path, sheet_name=Opportunity_team_sheet_name)

                # Check for required columns
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ ERROR: Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet.")
                elif 'opportunityid' not in Opportunity_team_df.columns:
                    print(f"\n    ❌ ERROR: Column 'opportunityid' not found in the '{Opportunity_team_sheet_name}' sheet.")
                else:
                    # Add a new 'Existing' column
                    Opportunity_team_df['Existing'] = Opportunity_team_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])
                    
                    # Count rows where 'Existing' is False
                    false_count = len(Opportunity_team_df[~Opportunity_team_df['Existing']])

                    # Save the updated DataFrame back to the sheet
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        Opportunity_team_df.to_excel(writer, sheet_name=Opportunity_team_sheet_name, index=False)
                    
                    # Display results
                    print(f"\n    ✅ The 'Existing' column has been added to the '{Opportunity_team_sheet_name}' sheet.")
                    print(f"\n    ❗️ Number of Opportnities Missing in Team sheet: {false_count}")

            except FileNotFoundError:
                print(f"\n    ❌ ERROR: File '{file_path}' not found.")
            except KeyError as e:
                print(f"\n    ❌ ERROR: {str(e)}")
            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {str(e)}")

            # ======================================================================
            # Step 6 :- Extracting Concatenated Values
            # ======================================================================

            print("\n\n🔍 Step 6: Extracting Concatenated Values...")

            # Define the input Excel file path and sheet name
            # file_path = "file_path.xlsx"  # Replace with your input file path
            sheet_name = "Opportunity_team"

            # Check if the input file exists
            if not os.path.exists(file_path):
                print(f"\n    ❌ ERROR: The input file '{file_path}' does not exist. Please check the file path and try again.")
                exit()

            try:
                # Read the Excel file
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Specify the column to extract concatenated values from
                column_name = "Concat_T_M"

                # Check if the column exists in the dataframe
                if column_name not in df.columns:
                    print(f"\n    ❌ ERROR: Column '{column_name}' is missing in the sheet '{sheet_name}' of the input file.")
                    exit()

                # Remove blank values and drop duplicates
                cleaned_data = df[column_name].dropna().drop_duplicates().reset_index(drop=True)

                # Create a new DataFrame for the output
                output_df = pd.DataFrame({column_name: cleaned_data})

                # Define the output file path and name
                output_file_path = "Extracts/Team_Member_extract.xlsx"

                # Write the processed data to a new Excel file
                output_df.to_excel(output_file_path, index=False)

                # Success message
                print(f"\n    ✅ Created 'Team_Member_extract' file and saved in Downloads")


            except Exception as e:
                print(f"\n    ❌ ERROR: An unexpected error occurred: {str(e)}")


            # ==================================================================================================================
            # Load the Excel file
            extract_file_path = "Extracts/Team_Member_extract.xlsx"  # Change this to your actual file path
            df = pd.read_excel(extract_file_path)

            # Extract the "accountid" column values
            if "Concat_T_M" in df.columns:
                account_ids = df["Concat_T_M"].dropna().astype(str)  # Drop NaN values and convert to string

                # Save to a text file
                with open("Delete/5_teammember.txt", "w") as f:
                    f.write("\n".join(account_ids))

            else:
                print("Column 'accountid' not found in the sheet.")
            # ==================================================================================================================
            remove_last_char_from_last_line('Delete/5_teammember.txt')
            # ======================================================================
            # 🔍 Step 7: Copying Data from CSV File
            # ======================================================================

            print("\n\n🔍 Step 7: Copying Data from CSV File...")


            # Define the file path for the CSV file
            csv_file_path = "/Users/avirajmore/Downloads/teammember.csv"

            # Loop until the file is found or the user decides to exit

            while not os.path.exists(csv_file_path):
                print(f"\n    ❌ The file '{csv_file_path}' is not present. Did you Query the Team member?")

                try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
                
                while try_again not in ['yes', 'no']:
                    print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
                if try_again != 'yes':
                    print("\n          🚫 Exiting the program.")
                    sys.exit()

            # Process the file if it exists
            if os.path.exists(csv_file_path):
                try:
                    # Read data from the CSV file
                    df = pd.read_csv(csv_file_path)

                    # Specify the Excel file path and sheet name
                    sheet_name = "Opportunity_team_Copy"

                    # Write data to the specified sheet in the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Success message
                    print(f"\n    ✅ Data from the CSV file has been successfully copied to the '{sheet_name}' sheet in the Excel file:")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: Excel file '{file_path}' not found.")
                except Exception as e:
                    print(f"\n    ❌ Error: An unexpected error occurred: {e}")



            # ======================================================================
            # 🔍 Step 8: Fetching User IDs of Team Members...
            # ======================================================================

            print ('\n\n🔍 Step 8: Fetching User IDs of Team Members...')

            # Define the sheet names
            opportunity_team_sheet_name = "Opportunity_team"
            opportunity_team_copy_sheet_name = "Opportunity_team_Copy"

            try:
                # Load data from the specified sheets
                opportunity_team_df = pd.read_excel(file_path, sheet_name=opportunity_team_sheet_name)
                opportunity_team_copy_df = pd.read_excel(file_path, sheet_name=opportunity_team_copy_sheet_name)

                # Clean and normalize the 'email' columns for consistency
                opportunity_team_df["email"] = opportunity_team_df["email"].str.strip().str.lower()
                opportunity_team_copy_df["Email"] = opportunity_team_copy_df["Email"].str.strip().str.lower()

                # Perform a left join to match emails and retrieve IDs
                result_df = pd.merge(
                    opportunity_team_df,
                    opportunity_team_copy_df[["Email", "Id"]],
                    left_on="email",
                    right_on="Email",
                    how="left"
                )

                # Count NaN values in the 'Id' column before filling
                nan_before = result_df["Id"].isna().sum()

                # Replace NaN values with "Inactive"
                result_df["Id"] = result_df["Id"].fillna("Inactive")

                # Count NaN values after filling (should be 0)
                nan_after = result_df["Id"].isna().sum()

                # Calculate the number of NaN values replaced
                nan_replaced = nan_before - nan_after

                # Drop the redundant 'Email' column from the result DataFrame
                result_df.drop(columns=["Email"], inplace=True)

                # Rename the 'Id' column to 'OwnerId'
                result_df.rename(columns={"Id": "OwnerId"}, inplace=True)

                # Save the updated data back to the 'Opportunity_team' sheet
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    result_df.to_excel(writer, sheet_name=opportunity_team_sheet_name, index=False)

                # Output success message
                print(f"\n    ❗️ Number of 'Inactive' values : {nan_replaced}")

            except FileNotFoundError:
                print(f"\n    ❌ File not found at path: {file_path}.")

            except KeyError as e:
                print(f"\n    ❌ KeyError: Column '{e}' not found. Please check the column names in your Excel sheets.")

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")


            # ======================================================================
            # 🔍 Step 9: Rearranging Columns in Sequence...
            # ======================================================================


            print("\n\n🔍 Step 9: Rearranging Columns in Sequence...")

            # Define the sheet name to target
            sheet_name = "Opportunity_team"  # Replace with the name of your sheet

            # Specify the desired order of columns
            desired_column_order = [
                "opportunityid",
                "Existing",
                "opportunityaccesslevel",
                "teammemberrole",
                "email",
                "OwnerId",
                "Concat_T_M"
            ]

            try:
                # Load the data from the specified sheet
                excel_data = pd.read_excel(file_path, sheet_name=sheet_name)

                # Verify that the sheet exists
                if isinstance(excel_data, pd.DataFrame):
                    # Identify missing and extra columns
                    missing_columns = [col for col in desired_column_order if col not in excel_data.columns]
                    extra_columns = [col for col in excel_data.columns if col not in desired_column_order]

                    # Determine the rearranged column order
                    rearranged_columns = [col for col in desired_column_order if col in excel_data.columns]

                    # Append extra columns (if any) to the end
                    rearranged_columns += extra_columns

                    # Rearrange the columns and save the updated DataFrame back to the Excel file
                    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                        excel_data[rearranged_columns].to_excel(writer, sheet_name=sheet_name, index=False)

                    # Display success message with details
                    print(f"\n    ✅ Columns rearranged successfully in the team sheet")

                    # Notify about missing columns
                    if missing_columns:
                        print("\n    ❗️ The following columns were missing and skipped:")
                        for col in missing_columns:
                            print(f"        🔸 {col}")

                    # Notify about extra columns
                    if extra_columns:
                        print("\n    🔹 The following extra columns were moved to the end:")
                        for col in extra_columns:
                            print(f"        🔸 {col}")
                else:
                    print(f"\n    ❌ Sheet '{sheet_name}' not found in the Excel file.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")


            # ======================================================================
            # 🔍 Step 10: Renaming Columns...
            # ======================================================================

            print("\n\n🔍 Step 10: Renaming Columns...")

            # Name of the sheet to target
            sheet_name = 'Opportunity_team'

            # Dictionary mapping old column names to new column names
            column_rename_mapping = {
                'opportunityid': 'OpportunityId',
                'teammemberrole': 'TeamMemberRole',
                'opportunityaccesslevel': 'OpportunityAccessLevel',
                'OwnerId': 'UserId',
            }

            # file_path = "your_excel_file.xlsx"  # Specify the path to your Excel file

            try:
                # Read the Excel file
                excel_data = pd.read_excel(file_path, sheet_name=None)

                # Check if the specified sheet exists
                if sheet_name in excel_data:
                    
                    # Access the specified sheet
                    df = excel_data[sheet_name]

                    # Check if all specified columns exist
                    missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]

                    # If any specified column is missing, notify the user and ask if they want to proceed
                    if missing_columns:
                        
                        print("\n    ❗️The following columns are missing and cannot be renamed:")
                        
                        for col in missing_columns:
                            print(f"\n      🔸 {col}")

                        while True:  # Loop until a valid response is given
                            proceed = input("\n    🔹 Do you want to proceed with the execution? (yes/no): ").strip().lower()

                            if proceed == 'yes':
                                break  # Exit the loop and proceed
                            elif proceed == 'no':
                                print("\n      ⛔️ Operation aborted. Exiting...")
                                exit()  # Exit the program
                            else:
                                print("\n      ❗️ Invalid input. Please enter 'yes' or 'no'.")

                    # Rename specified columns
                    df.rename(columns=column_rename_mapping, inplace=True)

                    # Write the modified DataFrame back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Success message
                    print(f"\n    ✅ Columns renamed successfully in the Team sheet.")

                else:
                    print(f"\n    ❌ Sheet '{sheet_name}' not found in the Excel file.")

            except FileNotFoundError:
                
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

            except Exception as e:
                
                print(f"\n    ❌ An unexpected error occurred: {e}")


            # ======================================================================
            # Step 11: Count the Number of Rows and Columns
            # ======================================================================

            print("\n\n🔍 Step 11: Counting the Number of Rows and Columns...")

            # Name of the sheet to target
            opportunity_sheet_name = 'Opportunity_team'

            # Try to read the Excel file into a DataFrame
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

                # Get the number of rows and columns
                team_final_num_rows = df.shape[0]
                team_final_num_columns = df.shape[1]

                # Print the number of rows and columns
                print(f"\n    ✅ Final rows count: {team_final_num_rows}")

            except FileNotFoundError:
                print(f"\n❌ Error: The file '{file_path}' was not found. Please check the file path.")

            except ValueError:
                print(f"\n❌ Error: Sheet '{opportunity_sheet_name}' not found in the Excel file.")

            except Exception as e:
                print(f"\n❌ An unexpected error occurred: {e}")


            # ======================================================================
            # Team Member Sheet Completed  
            # ======================================================================
            print("\n")
            print("=" * 100)
            print(" " * 33 + "📝 TEAM MEMBER SHEET COMPLETED 📝")
            print("=" * 100)

            break

        elif choice == "no":
            print("\n        🚫 Team Member sheet execution skipped!")
            print("\n")
            print("=" * 100)
            print(" " * 33 + "📝 TEAM MEMBER SHEET SKIPPED 📝")
            print("=" * 100)

            break  # Just breaking without running Block 2

        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")


    # =========================================================================================================================================
    #                                                STRATEGY SHEET EXECUTION
    # =========================================================================================================================================

    # Display the header once
    print("\n\n📄 Execute Next Sheet:")
    sheet_name = 'Reporting_codes'
    is_empty, preview = is_sheet_empty(file_path, sheet_name)

    if is_empty:
        print(f"\n📂 The sheet '{sheet_name}' is empty or contains only headers.\n")
    elif is_empty is None:
        print("\n⚠️ Could not process the sheet due to an error.\n")
    else:
        print(f"\n✅ The sheet '{sheet_name}' contains data. Here are the first 4 rows:\n")
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    while True:

        choice = input("\n    🔹 Do you want to execute the Strategy Sheet ? (yes/no): ").strip().lower()
        
        if choice == "yes":
            print(f"\n        ⏳ Executing the Sheet: Strategy sheet ")

            print("\n")
            print("=" * 100)
            print(" " * 33 + "📝 STRATEGY SHEET EXECUTION 📝")
            print("=" * 100)

            # =========================================
            # Step 1: Rename Reporting Codes Sheet
            # =========================================

            print("\n\n🔍 Step 1: Renaming Reporting Codes Sheet...")

            def rename_reporting_codes(wb):
                for sheetname in wb.sheetnames:
                    # Normalize sheet name by removing spaces and underscores, and converting to lowercase
                    normalized_name = sheetname.strip().replace('_', '').replace(' ', '').lower()
                    if normalized_name == 'reportingcodes':
                        wb[sheetname].title = 'Reporting_codes_2'
                        return True
                return False

            try:
                # Try to load the workbook
                wb = openpyxl.load_workbook(file_path)

                # Attempt to rename the sheet
                if rename_reporting_codes(wb):
                    # Save the changes to the file
                    wb.save(file_path)
                    print("\n    ✅ Sheet 'Reporting_codes' renamed to 'Reporting_codes_2' successfully.")
                else:
                    print("\n    ❌ Error: Sheet 'Reporting_codes' not found in the workbook.")
                    sys.exit()

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found. Please check the file path and try again.")
            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")


            # ======================================================================
            # Step 2: Rename the Tag_2 Sheet
            # ======================================================================

            print("\n\n🔍 Step 2: Renaming 'Tags' Sheet to 'Tags_2'...")

            # file_path = 'example.xlsx'  # Replace with your Excel file path

            try:
                # Try loading the workbook
                wb = openpyxl.load_workbook(file_path)

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found.")
                sys.exit()  # Exit the program if file is not found
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"\n    ❌ Error: The file '{file_path}' is not a valid Excel file.")
                sys.exit()  # Exit the program if the file is not valid

            # Check if 'Tags' sheet exists and rename it
            if 'Tags' in wb.sheetnames:
                original_sheet = wb['Tags']
                original_sheet.title = 'Tags_2'
                wb.save(file_path)
                print("\n    ✅ Sheet 'Tags' renamed to 'Tags_2' successfully.")
            else:
                print("\n    ❗️ Sheet 'Tags' not found. No action taken.")

            # Close the workbook
            wb.close()

            # ========================================================================
            # Step 3: Renaming Columns in "Tags_2" Sheet
            # ========================================================================

            print("\n\n🔍 Step 3: Renaming Columns in 'Tags_2' Sheet...")


            try:
                # Load the workbook
                wb = load_workbook(file_path)

                # Define the target sheet name
                target_sheet_name = 'Tags_2'

                # Check if the target sheet exists in the workbook
                if target_sheet_name in wb.sheetnames:
                    # Load the target sheet
                    ws = wb[target_sheet_name]

                    # Define mappings of different variations of column names
                    column_name_mappings = {
                        'tags': 'tag',
                        'Opportunity Id': 'opportunityid',
                        'opportunity_id': 'opportunityid',
                        'Opportunityid': 'opportunityid',
                        'opportunityid': 'opportunityid'
                    }

                    # Iterate over the first row to find and rename columns
                    renamed_columns = []
                    for col_idx, cell in enumerate(ws[1], start=1):  # start=1 for 1-based index
                        if isinstance(cell.value, str):
                            normalized_name = cell.value.strip().lower()  # Normalize to lowercase and strip whitespace
                            if normalized_name in column_name_mappings:
                                new_column_name = column_name_mappings[normalized_name]
                                ws.cell(row=1, column=col_idx, value=new_column_name)  # Rename the column header
                                renamed_columns.append(f"'{cell.value}' ➔ '{new_column_name}'")

                    # Save the modified workbook
                    wb.save(file_path)

                    if renamed_columns:
                        print("\n    ✅ Columns renamed successfully in the 'Tags_2' sheet.")

                    else:
                        print("\n    ❗️ No columns were renamed. All columns were already in the desired format.")

                else:
                    print(f"\n    ❗️ '{target_sheet_name}' sheet not found in the Excel file. No action taken.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found.")
                sys.exit()

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")
                sys.exit()


            # ========================================================================
            # Step 4: Rename the columns in "Reporting_codes_2" sheet
            # ========================================================================

            print("\n\n🔍 Step 4: Renaming Columns in 'Reporting_codes_2' Sheet...\n")


            # Load the workbook
            try:
                wb = load_workbook(file_path)

                # Define the target sheet name
                target_sheet_name = 'Reporting_codes_2'

                # Check if the target sheet exists in the workbook
                if target_sheet_name in wb.sheetnames:
                    # Load the target sheet
                    ws = wb[target_sheet_name]

                    # Define mappings of different variations of column names
                    column_name_mappings = {
                        'tags': 'tag',
                        'Opportunity Id': 'opportunityid',
                        'opportunity_id': 'opportunityid',
                        'Opportunityid': 'opportunityid',
                        'opportunityid': 'opportunityid'
                    }

                    # List to track renamed columns
                    renamed_columns = []

                    # Iterate over the first row to find and rename columns
                    for col_idx, cell in enumerate(ws[1], start=1):  # start=1 for 1-based index
                        if isinstance(cell.value, str):
                            normalized_name = cell.value.strip().lower()  # normalize to lowercase and strip whitespace
                            if normalized_name in column_name_mappings:
                                new_column_name = column_name_mappings[normalized_name]
                                ws.cell(row=1, column=col_idx, value=new_column_name)  # Rename the column header
                                renamed_columns.append(f"'{cell.value}' ➔ '{new_column_name}'")

                    # Save the modified workbook
                    wb.save(file_path)

                    # Provide output
                    if renamed_columns:
                        print(f"\n    ✅ Columns renamed successfully in '{target_sheet_name}' sheet:")

                    else:
                        print(f"\n    ❗️ No columns were renamed. All columns were already in the desired format.")
                    
                else:
                    print(f"\n    ❗️ '{target_sheet_name}' sheet not found in the Excel file. No action taken.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found. Please check the file path and try again.")
                exit()

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")
                exit()

            # ========================================================================
            # Step 5: Create separate sheet for "Tags"
            # ========================================================================

            print("\n\n🔍 Step 5: Creating 'Tags' Sheet...")

            # Load the Excel file
            try:
                wb = load_workbook(file_path)
            except FileNotFoundError:
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")
                exit()

            # Load the specific sheet into a DataFrame
            sheet_name = 'Reporting_codes_2'
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            except KeyError:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in '{file_path}'.")
                wb.close()
                exit()

            # Normalize column names to lowercase for consistency
            df.columns = df.columns.str.lower()

            # Check if 'opportunity_id' and 'tag' columns are present
            expected_columns = ['opportunityid', 'tag']
            missing_columns = [col for col in expected_columns if col not in df.columns]
            if missing_columns:
                print(f"\n    ❌ Error: Missing columns: {', '.join(missing_columns)}. ")
                wb.close()
                exit()

            # Filter rows where 'tag' column contains a value
            filtered_df = df[df['tag'].notnull()]

            # Check if 'Tags_2' sheet already exists
            if 'Tags_2' in wb.sheetnames:
                overwrite_input = input("\n    ❗️ Sheet 'Tags_2' already exists. Do you want to overwrite it? (yes/no): ").lower().strip()
                if overwrite_input == 'yes':
                    # Remove the existing 'Tags_2' sheet
                    existing_sheet = wb['Tags_2']
                    wb.remove(existing_sheet)
                    print(f"\n       ✅ 'Tags_2' sheet cleared for overwrite.")
                else:
                    print("\n       ❌ Not overwriting 'Tags_2' sheet. Aborting further processing.")
                    wb.close()
                    exit()

            # Create a new sheet named 'Tags_2' and write filtered data
            ws = wb.create_sheet('Tags_2')

            # Append header row
            ws.append(expected_columns)

            # Append filtered data rows
            copied_rows = 0
            for r_idx, row in enumerate(dataframe_to_rows(filtered_df[expected_columns], index=False, header=False)):
                ws.append(row)
                copied_rows += 1

            # Save the updated workbook
            try:
                wb.save(file_path)
                print(f"\n    ✅ Filtered data ({copied_rows} rows) has been successfully written to 'Tags_2' sheet.")
            except Exception as e:
                print(f"\n    ❌ Error occurred while saving: {e}")
            finally:
                wb.close()

            # ========================================================================
            # Step 6: To Delete the "Tag" Column from Reporting Codes Sheet
            # ========================================================================

            print("\n\n🔍 Step 6: Deleting 'Tag' Column from Reporting Codes Sheet...")

            # Load the Excel file
            try:
                wb = load_workbook(file_path)

            except FileNotFoundError:
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")
                exit()

            # Load a specific sheet
            sheet_name = 'Reporting_codes_2'
            if sheet_name not in wb.sheetnames:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in '{file_path}'. Aborting further processing.")
                wb.close()
                exit()

            # Get the sheet
            ws = wb[sheet_name]

            # Find and delete 'tag' column if it exists (case insensitive)
            tag_column_found = False
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                if col[0].value and col[0].value.strip().lower() == 'tag':
                    ws.delete_cols(col_idx)
                    tag_column_found = True
                    print(f"\n    ✅ 'Tag' column found and deleted from '{sheet_name}' sheet.")
                    break

            if not tag_column_found:
                print(f"\n    ❗️ Column 'tag' not found in '{sheet_name}' sheet. No action taken.")

            # Save and close the workbook
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"\n    ❌ Error occurred while saving: {e}")
            finally:
                wb.close()

            # ========================================================================
            # Step 7: To Remove Comma Separated Values from 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 7: Removing Comma Separated Values from 'Reporting_codes' Sheet...\n")


            def process_reporting_codes(file_path, sheet_name):
                try:
                    # Load the existing workbook
                    wb = load_workbook(file_path)
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to load workbook. {e}")
                    return
                
                try:
                    # Read the specific sheet into a DataFrame
                    df_reporting_codes_2 = pd.read_excel(file_path, sheet_name=sheet_name)
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to read sheet '{sheet_name}'. {e}")
                    return
                
                # Initialize an empty list to store rows for the new DataFrame
                new_rows_reporting = []
                
                # Iterate through each row in the original DataFrame
                for index, row in df_reporting_codes_2.iterrows():
                    reporting_codes = []
                    if isinstance(row['reporting_codes'], str) and row['reporting_codes'].strip():
                        reporting_codes = row['reporting_codes'].split(',')
                    
                    # If there are multiple reporting codes in the cell, create new rows
                    if reporting_codes:
                        for code in reporting_codes:
                            new_row_reporting = row.copy()
                            new_row_reporting['reporting_codes'] = code.strip()
                            new_rows_reporting.append(new_row_reporting)
                    else:
                        new_rows_reporting.append(row)
                
                # Create a new DataFrame from the list of rows
                df_Reporting_codes = pd.DataFrame(new_rows_reporting, columns=df_reporting_codes_2.columns)
                
                # Drop duplicate rows

                df_Reporting_codes.drop_duplicates(inplace=True)
                
                # Create a new sheet in the existing workbook
                sheet_name_reporting = 'Reporting_codes'
                if sheet_name_reporting in wb.sheetnames:
                    print(f"    ❗️ Sheet '{sheet_name_reporting}' already exists. It will be replaced.")
                    ws_reporting = wb[sheet_name_reporting]
                    wb.remove(ws_reporting)
                    ws_reporting = wb.create_sheet(title=sheet_name_reporting)
                else:
                    ws_reporting = wb.create_sheet(title=sheet_name_reporting)
                
                # Convert DataFrame to rows and append to the sheet

                for r_idx, row in enumerate(dataframe_to_rows(df_Reporting_codes, index=False, header=True), 1):
                    ws_reporting.append(row)
                
                # Save the workbook with the new sheet
                try:
                    wb.save(file_path)
                    print(f"\n    ✅ New sheet '{sheet_name_reporting}' has been successfully created in the Excel file.")
                    print(f"\n        🔸 Total rows before removing duplicates: {len(df_reporting_codes_2)}")
                    print(f"\n        🔸 Total rows after removing duplicates: {len(df_Reporting_codes)}")
                    print(f"\n        🔸 Duplicated Rows removed: {len(df_reporting_codes_2) - len(df_Reporting_codes)}.")
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to save workbook. {e}")

            # Define the file path to your Excel file
            # file_path = '/Users/avirajmore/Downloads/demo1 copy 2.xlsx'
            process_reporting_codes(file_path, 'Reporting_codes_2')

            # ========================================================================
            # Step 8: To Remove Comma Separated Values from 'Tags_2' Sheet
            # ========================================================================

            print("\n\n🔍 Step 8: Removing Comma Separated Values from 'Tags_2' Sheet...")


            # Define the file path to your Excel file
            # file_path = 'path/to/your/excel_file.xlsx'

            # Load the existing workbook
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"\n    ❌ Error: Failed to load workbook. {e}")
                exit()

            # Initialize variables for sheet and DataFrame
            sheet_name = 'Tags_2'
            df_tag_2 = None

            # Attempt to load the sheet using openpyxl, handle if it doesn't exist
            try:
                sheet = wb[sheet_name]
                df_tag_2 = pd.DataFrame(sheet.values)
                df_tag_2.columns = df_tag_2.iloc[0]  # Assuming the first row contains column headers
                df_tag_2 = df_tag_2[1:]  # Skip the first row (headers)
            except KeyError:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the workbook.")
                exit()

            # If sheet exists and DataFrame is loaded, process the data
            if df_tag_2 is not None:
                new_rows_tag = []
                duplicate_count_tag = 0  # Initialize count for duplicate rows
                seen_rows_tag = set()

                # Iterate through each row in the original dataframe
                for index, row in df_tag_2.iterrows():
                    row_tuple_tag = tuple(row.items())
                    if row_tuple_tag in seen_rows_tag:
                        duplicate_count_tag += 1  # Increment duplicate count
                    else:
                        seen_rows_tag.add(row_tuple_tag)
                        # Check if 'tag' column exists and is a non-empty string
                        if isinstance(row['tag'], str) and row['tag'].strip():
                            tags = row['tag'].split(',')
                        else:
                            tags = []
                        
                        # If there are multiple tags in the cell, create new rows
                        if len(tags) > 0:
                            for tag in tags:
                                new_row_tag = row.copy()  # Create a copy of the original row
                                new_row_tag['tag'] = tag.strip()  # Assign a single tag
                                new_rows_tag.append(new_row_tag)
                        else:
                            new_rows_tag.append(row)

                # Create a new DataFrame from the list of rows
                df_Tag = pd.DataFrame(new_rows_tag, columns=df_tag_2.columns)

                # Count total rows before and after removing duplicates
                total_rows_before_tag = len(df_tag_2)
                total_rows_after_tag = len(df_Tag)

                # Drop duplicate rows based on all columns (if necessary)
                df_Tag = df_Tag.drop_duplicates()

                # Count how many duplicates were removed
                removed_duplicates_count_tag = total_rows_before_tag - total_rows_after_tag

                # Create a new sheet in the existing workbook
                sheet_name_tag = 'Tags'
                if sheet_name_tag in wb.sheetnames:
                    print(f"\n    ❗️ Sheet '{sheet_name_tag}' already exists. It will be replaced.")
                    ws_tag = wb[sheet_name_tag]
                    wb.remove(ws_tag)
                ws_tag = wb.create_sheet(title=sheet_name_tag)

                # Convert DataFrame to rows and append to the sheet
                for r_idx, row in enumerate(dataframe_to_rows(df_Tag, index=False, header=True), 1):
                    ws_tag.append(row)

                # Save the workbook with the new sheet
                try:
                    wb.save(file_path)
                    print(f"\n    ✅ New sheet '{sheet_name_tag}' has been successfully created in the Excel file.")
                    print(f"\n        🔸 Total rows before removing duplicates: {total_rows_before_tag}")
                    print(f"\n        🔸 Total rows after removing duplicates: {total_rows_after_tag}")
                    print(f"\n        🔸 {removed_duplicates_count_tag} duplicate rows were removed.")
                
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to save workbook. {e}")

            # ========================================================================
            # Step 9: Add Existing Column to 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 9: Adding 'Existing' Column to 'Reporting_codes' Sheet...")

            # Specify the file path of the Excel file
            # file_path = "/Users/avirajmore/Downloads/your_excel_file.xlsx"

            # Specify the sheet names
            opportunity_sheet_name = 'Opportunity'
            reporting_codes_sheet_name = 'Reporting_codes'

            try:
                # Read data from the specified sheets
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                reporting_codes_df = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)

                # Check if required columns exist
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in 'Reporting codes' sheet.")
                    sys.exit()
                elif 'opportunityid' not in reporting_codes_df.columns:
                    print(f"\n    ❌ Column 'opportunityid' not found in 'Reporting codes' sheet.")
                    sys.exit()
                else:
                    # Create a new column 'Existing' in reporting_codes_df
                    reporting_codes_df['Existing'] = reporting_codes_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

                    # Write back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        reporting_codes_df.to_excel(writer, sheet_name=reporting_codes_sheet_name, index=False)

                    # Notify the successful operation
                    print(f"\n    ✅ 'Existing' column has been successfully added to the 'Reporting codes' sheet.")

            except FileNotFoundError:
                print(f"    ❌ File not found.")
                sys.exit()
            except Exception as e:
                print(f"    ❌ Error: {e}")
                sys.exit()


            # ========================================================================
            # Step 10: Add Existing Column to 'Tags' Sheet
            # ========================================================================

            print("\n\n🔍 Step 10: Adding 'Existing' Column to 'Tags' Sheet...")


            # Specify the file path of the Excel file
            # file_path = "/Users/avirajmore/Downloads/your_excel_file.xlsx"

            # Specify the sheet names
            opportunity_sheet_name = 'Opportunity'
            tags_sheet_name = 'Tags'

            try:
                # Read data from the specified sheets
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                tags_df = pd.read_excel(file_path, sheet_name=tags_sheet_name)

                # Check if required columns exist
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in '{opportunity_sheet_name}' sheet.")
                    sys.exit()
                elif 'opportunityid' not in tags_df.columns:
                    print(f"\n    ❌ Column 'opportunityid' not found in '{tags_sheet_name}' sheet.")
                    sys.exit()
                else:
                    # Create a new column 'Existing' in tags_df
                    tags_df['Existing'] = tags_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

                    # Write back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        tags_df.to_excel(writer, sheet_name=tags_sheet_name, index=False)

                    # Notify the successful operation
                    print(f"\n    ✅ 'Existing' column has been successfully added to the '{tags_sheet_name}' sheet.")

            except FileNotFoundError:
                print(f"    ❌ File '{file_path}' not found.")
                sys.exit()
            except Exception as e:
                print(f"    ❌ Error: {e}")
                sys.exit()

            # ========================================================================
            # Step 11: To Concatenate Values in 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 11: Concatenating Values in 'Reporting_codes' Sheet...")


            # Replace 'your_file_path.xlsx' with the path to your Excel file
            # file_path = 'your_file_path.xlsx'
            sheet_name = 'Reporting_codes'

            try:
                # Load the workbook and specify append mode to modify the sheet
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    
                    # Read the specified sheet into a DataFrame
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Check if 'reporting_codes' column exists (case-insensitive)
                    lowercase_columns = [col.lower() for col in df.columns]
                    
                    if 'reporting_codes' in lowercase_columns:
                        # Find the actual column name in the original case
                        actual_reporting_codes_column = df.columns[lowercase_columns.index('reporting_codes')]
                        
                        # Create a new column 'Concatcodes' with concatenated values
                        df['Concatcodes'] = df[actual_reporting_codes_column].apply(lambda x: f"'{x}'," if pd.notnull(x) else x)
                        
                        # Write the modified DataFrame back to the specified sheet
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Notify successful operation
                        print(f"\n    ✅ The 'Concatcodes' column has been successfully added to '{sheet_name}' sheet.")
                    else:
                        print(f"\n    ❗️ The '{sheet_name}' sheet does not have a column named 'reporting_codes'.")
            except FileNotFoundError:
                print(f"\n    ❌ File '{file_path}' not found.")
                sys.exit()
            except Exception as e:
                print(f"\n    ❌ Error: {e}")
                sys.exit()

            # ========================================================================
            # Step 12:- To concatenate values in "Tags" sheet
            # ========================================================================

            print("\n\n🔍 Step 12: Adding 'Concattags' Column to 'Tags' Sheet...")

            # file_path = 'your_file_path.xlsx'
            sheet_name = 'Tags'

            # Check if the sheet exists in the workbook
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                if sheet_name in sheet_names:
                    # Load the entire workbook
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # Read the specified sheet into a DataFrame
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # Check if 'tag' column exists (case-insensitive)
                        lowercase_columns = [col.lower() for col in df.columns]
                        if 'tag' in lowercase_columns:
                            # Find the actual column name in the original case
                            actual_tag_column = df.columns[lowercase_columns.index('tag')]

                            # Create a new column 'Concattags' with concatenated values
                            df['Concattags'] = df[actual_tag_column].apply(lambda x: f"'{x}'," if pd.notnull(x) else x)

                            # Write the modified DataFrame back to the specified sheet
                            df.to_excel(writer, sheet_name=sheet_name, index=False)

                            # Print a meaningful message
                            print("\n    ✅ 'Concattags' column has been successfully added to the 'Tags' sheet.")
                        else:
                            print(f"\n    ❗️ The '{sheet_name}' sheet does not contain a column named 'tag'. Please verify the column name.")
                else:
                    print(f"\n    ❗️ The '{sheet_name}' sheet does not exist in the workbook. Please check the sheet name.")
                    sys.exit()
            except FileNotFoundError:
                print(f"\n    ❌ The file '{file_path}' was not found. Please verify the file path.")
                sys.exit()
            except Exception as e:
                print(f"\n    ❌ An error occurred: {str(e)}")
                sys.exit()

            # ========================================================================
            # Step 13:- To extract concatenated values
            # ========================================================================


            # Step 13:- Extracting Data from 'Reporting_codes' and 'Tags' Sheets

            print("\n\n🔍 Step 13: Extracting Data from 'Reporting_codes' and 'Tags' Sheets...")

            # Define the input file path and sheet names
            # file_path = "path/to/your/file_path.xlsx"
            reporting_codes_sheet_name = "Reporting_codes"
            tags_sheet_name = "Tags"

            # Define the columns to extract from each sheet
            reporting_codes_column = "Concatcodes"
            tags_column = "Concattags"

            # Check if the input file exists
            if not os.path.exists(file_path):
                print(f"\n    ❌ The input file '{file_path}' does not exist. Please verify the file path.")
                exit()

            # Initialize flags to check if sheets are found
            reporting_codes_found = False
            tags_found = False

            # Read data from the Reporting_codes sheet if it exists
            if reporting_codes_sheet_name in pd.ExcelFile(file_path).sheet_names:
                df_reporting_codes = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)
                reporting_codes_found = True

            else:
                print(f"\n    ❗️ Sheet '{reporting_codes_sheet_name}' not found.")
                sys.exit()
            # Read data from the Tags sheet if it exists
            if tags_sheet_name in pd.ExcelFile(file_path).sheet_names:
                df_tags = pd.read_excel(file_path, sheet_name=tags_sheet_name)
                tags_found = True
            else:
                print(f"    ❗️ Sheet '{tags_sheet_name}' not found.")
                sys.exit()
            # Extract the required columns if sheets are found
            concatcodes_values = []
            concattags_values = []

            if reporting_codes_found:
                concatcodes_values = df_reporting_codes[reporting_codes_column].dropna().unique()
            else:
                print(f"\n    ❗️ Column '{reporting_codes_column}' not found in '{reporting_codes_sheet_name}' sheet.")

            if tags_found:
                concattags_values = df_tags[tags_column].dropna().unique()
            else:
                print(f"\n    ❗️ Column '{tags_column}' not found in '{tags_sheet_name}' sheet.")

            # Determine the maximum length to align columns
            max_length = max(len(concatcodes_values), len(concattags_values))

            # Extend shorter list to match max_length with a placeholder
            if len(concatcodes_values) < max_length:
                concatcodes_values = list(concatcodes_values) + [None] * (max_length - len(concatcodes_values))
            if len(concattags_values) < max_length:
                concattags_values = list(concattags_values) + [None] * (max_length - len(concattags_values))

            # Create a new DataFrame for the extracted values
            output_df = pd.DataFrame({
                reporting_codes_column: concatcodes_values,
                tags_column: concattags_values
            })

            # Define the output file path and name
            output_file_path = "Extracts/Reporting_codes_and_Tags_extract.xlsx"

            # Write the extracted data to a new Excel file
            output_df.to_excel(output_file_path, index=False)

            # Final success message
            print(f"\n    ✅ Extracted data has been successfully written to '{output_file_path}'.")

            # ==================================================================================================================
            # To extract tags and code to text file
            # ==================================================================================================================
            # Load the Excel file
            extract_file_path = "Extracts/Reporting_codes_and_Tags_extract.xlsx"  # Change this to your actual file path

            df = pd.read_excel(extract_file_path)

            # Extract values from 'ownerid' and 'concatenatedcreatedby' (even if their lengths differ)
            Concatcodes_values = df["Concatcodes"].dropna().astype(str).tolist() if "Concatcodes" in df.columns else []
            Concattags_values = df["Concattags"].dropna().astype(str).tolist() if "Concattags" in df.columns else []

            # Combine both lists while maintaining all values
            all_values = Concatcodes_values + Concattags_values  # Concatenating both lists

            # Save to a text file
            with open("Delete/6_strategy.txt", "w") as f:
                f.write("\n".join(all_values))
            
            # ========================================================================  
            # Code to remove comma from the text file

            remove_last_char_from_last_line('Delete/6_strategy.txt')

            # ========================================================================
            # Step 14:- Processing CSV File and Adding Filtered Data to Excel
            # ========================================================================

            print("\n\n🔍 Step 14: Processing CSV File and Adding Filtered Data to Excel...")

            # Define the file path for the CSV file
            csv_file_path = "/Users/avirajmore/Downloads/tags.csv"

            # Loop until the file is found or the user decides to exit
            while not os.path.exists(csv_file_path):
                print(f"\n    ❌ Error: The file '{csv_file_path}' does not exist.")
                try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                while try_again not in ['yes', 'no']:
                    print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                if try_again == 'no':
                    print("\n          🚫 Exiting the program.")
                    sys.exit()  # Exit the program if the user chooses 'no'

            # If the file exists, process the file
            try:
                # Read CSV file into a DataFrame
                df = pd.read_csv(csv_file_path)

                # Specify the column name to filter
                filter_column = "Record_Type_Name__c"
                filter_value = "Reporting codes"

                # Check if the filter column exists in the CSV file
                if filter_column not in df.columns:
                    print(f"\n    ❌ Error: Column '{filter_column}' not found in the CSV file.")
                    sys.exit()  # Exit if the required column is not found

                # Filter rows where the specified column equals the specified value
                df_filtered = df[df[filter_column] == filter_value]

                # Specify the sheet name in the Excel file
                sheet_name = "Reporting_codes_Copy"

                # Write filtered data to the specified sheet in the Excel file
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

                print(f"\n    ✅ Filtered data has been successfully copied to the '{sheet_name}' sheet in the Excel file.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: The Excel file '{file_path}' was not found.")
            except KeyError:
                print(f"\n    ❌ Error: The column '{filter_column}' is not found in the CSV file.")
            except Exception as e:
                print(f"\n    ❌ Error: {e}")

            # ========================================================================
            # Step 15:- To copy data from Tag csv file
            # ========================================================================


            # Step 15:- Processing CSV File and Adding Filtered Data to Excel
            print("\n\n🔍 Step 15: Processing CSV File and Adding Filtered Data to Excel...")

            # Define the file path for the CSV file
            csv_file_path = "/Users/avirajmore/Downloads/tags.csv"

            # Loop until the file is found or the user decides to exit
            while not os.path.exists(csv_file_path):
                print(f"\n    ❌ Error: The file '{csv_file_path}' is not found.")
                try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                while try_again not in ['yes', 'no']:
                    print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                if try_again == 'no':
                    print("\n          🚫 Exiting the program.")
                    sys.exit()  # Exit the program if the user chooses 'no'

            # If the file exists, process the file
            try:
                # Read CSV file into a DataFrame
                df = pd.read_csv(csv_file_path)

                # Specify the column name to filter
                filter_column = "Record_Type_Name__c"
                filter_value = "Tags"

                # Check if the filter column exists in the CSV file
                if filter_column not in df.columns:
                    print(f"\n    ❌ Error: Column '{filter_column}' not found in the CSV file.")
                    sys.exit()  # Exit if the required column is not found

                # Filter rows where the specified column equals the specified value
                df_filtered = df[df[filter_column] == filter_value]

                # Specify the sheet name in the Excel file
                sheet_name = "Tags_Copy"

                # Write filtered data to the specified sheet in the Excel file
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

                print(f"\n    ✅ Filtered data has been successfully copied to the '{sheet_name}' sheet in the Excel file.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: The Excel file '{file_path}' was not found.")
            except KeyError:
                print(f"\n    ❌ Error: The column '{filter_column}' is not found in the CSV file.")
            except Exception as e:
                print(f"\n    ❌ Error: {e}")

            # =======================================================================
            # Step 16: To get strategy Ids for Codes
            # ========================================================================

            print("\n\n🔍 Step 16: To get strategy Ids for Codes...")

            reporting_codes_sheet_name = 'Reporting_codes'
            reporting_codes_copy_sheet_name = 'Reporting_codes_Copy'

            def vlookup_operation(file_path, reporting_codes_sheet_name, reporting_codes_copy_sheet_name):
                # Check if the file exists
                if not os.path.exists(file_path):
                    print(f"\n    ❌ Error: File '{file_path}' not found.")
                    return

                try:
                    # Read the data from both sheets
                    reporting_codes_df = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)
                    reporting_codes_copy_df = pd.read_excel(file_path, sheet_name=reporting_codes_copy_sheet_name)

                    # Ensure column names are standardized and lowercase
                    reporting_codes_df.columns = reporting_codes_df.columns.str.strip().str.lower()
                    reporting_codes_copy_df.columns = reporting_codes_copy_df.columns.str.strip().str.lower()

                    # Convert relevant columns to lowercase for case-insensitive merge
                    reporting_codes_df['reporting_codes'] = reporting_codes_df['reporting_codes'].str.lower()
                    reporting_codes_copy_df['name'] = reporting_codes_copy_df['name'].str.lower()

                    # Check if the necessary columns exist (case-insensitive)
                    if 'reporting_codes' not in reporting_codes_df.columns:
                        print(f"\n    ❌ Column 'reporting_codes' not found in '{reporting_codes_sheet_name}' sheet.")
                        sys.exit()
                    if 'name' not in reporting_codes_copy_df.columns:
                        print(f"\n    ❌ Column 'name' not found in '{reporting_codes_copy_sheet_name}' sheet.")
                        sys.exit()
                    if 'id' not in reporting_codes_copy_df.columns:
                        print(f"\n    ❌ Column 'id' not found in '{reporting_codes_copy_sheet_name}' sheet.")
                        sys.exit()

                    # Perform merge using standardized column names
                    merged_df = pd.merge(reporting_codes_df,
                                        reporting_codes_copy_df[['name', 'id']],
                                        left_on='reporting_codes', right_on='name',
                                        how='left')

                    # Count NaN values before replacing them
                    na_count = merged_df['id'].isna().sum()

                    # Create a new column 'StrategyId' and fill missing values with 'Not found'
                    merged_df['StrategyId'] = merged_df['id'].fillna('Not found')

                    # Drop unnecessary columns after merging
                    merged_df.drop(['name', 'id'], axis=1, inplace=True)

                    # Save the updated DataFrame back to the same Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        merged_df.to_excel(writer, sheet_name=reporting_codes_sheet_name, index=False)
                    print(f"\n    ❗️ Count of Codes  'Not found': {na_count}")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: File '{file_path}' not found.")
                    sys.exit()
                except KeyError as e:
                    print(f"\n    ❌ {str(e)}")
                    sys.exit()
                except Exception as e:
                    print(f"\n    ❌ An error occurred: {str(e)}")
                    sys.exit()

            # Example usage:
            # file_path = 'your_file_path.xlsx'
            vlookup_operation(file_path, reporting_codes_sheet_name, reporting_codes_copy_sheet_name)

            # ========================================================================
            # Step 17:- To get strategy Ids for tags
            # ========================================================================

            print("\n\n🔍 Step 17: To get strategy Ids for Tags...")

            tags_sheet_name = 'Tags'
            tags_copy_sheet_name = 'Tags_Copy'

            def vlookup_tags(file_path, tags_sheet_name, tags_copy_sheet_name):
                # Check if the file exists
                if not os.path.exists(file_path):
                    print(f"\n    ❌ Error: File '{file_path}' not found.")
                    return

                try:
                    # Read the data from both sheets
                    tags_df = pd.read_excel(file_path, sheet_name=tags_sheet_name)
                    tags_copy_df = pd.read_excel(file_path, sheet_name=tags_copy_sheet_name)

                    # Ensure column names are standardized and lowercase
                    tags_df.columns = tags_df.columns.str.strip().str.lower()
                    tags_copy_df.columns = tags_copy_df.columns.str.strip().str.lower()

                    # Convert relevant columns to lowercase for case-insensitive merge
                    tags_df['tag'] = tags_df['tag'].str.lower()
                    tags_copy_df['name'] = tags_copy_df['name'].str.lower()

                    # Check if the necessary columns exist (case-insensitive)
                    if 'tag' not in tags_df.columns:
                        print(f"\n    ❌ Column 'tag' not found in '{tags_sheet_name}' sheet.")
                    if 'name' not in tags_copy_df.columns:
                        print(f"\n    ❌ Column 'name' not found in '{tags_copy_sheet_name}' sheet.")
                    if 'id' not in tags_copy_df.columns:
                        print(f"\n    ❌ Column 'id' not found in '{tags_copy_sheet_name}' sheet.")

                    # Perform merge using standardized column names

                    merged_df = pd.merge(tags_df,
                                        tags_copy_df[['name', 'id']],
                                        left_on='tag', right_on='name',
                                        how='left')

                    # Count NaN values before replacing them
                    na_count = merged_df['id'].isna().sum()

                    # Create a new column 'Strategy_tag_Id' and fill missing values with 'Not found'
                    merged_df['Strategy_tag_Id'] = merged_df['id'].fillna('Not found')

                    # Drop unnecessary columns after merging
                    merged_df.drop(['name', 'id'], axis=1, inplace=True)

                    # Save the updated DataFrame back to the same Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        merged_df.to_excel(writer, sheet_name=tags_sheet_name, index=False)

                    print(f"\n    ❗️ Count of 'NaN' values replaced with 'Not found': {na_count}")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: File '{file_path}' not found.")
                    sys.exit()
                except KeyError as e:
                    print(f"\n    ❌ {str(e)}")
                    sys.exit()
                except Exception as e:
                    print(f"\n    ❌ An error occurred: {str(e)}")
                    sys.exit()

            # Example usage:
            # file_path = 'your_file_path.xlsx'
            vlookup_tags(file_path, tags_sheet_name, tags_copy_sheet_name)



            # ========================================================================
            print("\n")
            print("=" * 100)
            print(" " * 33 + "📝 STRATEGY SHEET COMPLETED 📝")
            print("=" * 100)

            # ========================================================================


            # ========================================================================
            # Rearrainging the sheets!
            # ========================================================================


            # Step 18: Rearranging Sheets in Workbook
            print("\n\n📄 Rearranging Sheets in Workbook...")

            # Define the file paths
            # file_path = "/Users/avirajmore/Downloads/avi.xlsx"
            new_file_path = "/Users/avirajmore/Downloads/Rearranged_file.xlsx"

            # Check if the input file exists
            if not os.path.exists(file_path):
                print(f"\n    ❌ Error: The file '{file_path}' does not exist.")
            else:
                try:
                    # Load the workbook
                    wb = openpyxl.load_workbook(file_path)

                    # Desired order of sheets
                    desired_order = [
                        "opportunity",
                        "opportunity_Copy",
                        "opportunity_product",
                        "opportunity_product_Copy",
                        "Opportunity_team",
                        "Opportunity_team_Copy",
                        "Reporting_codes",
                        "Reporting_codes_Copy",
                        "Tags",
                        "Tags_Copy"
                    ]

                    # Normalize sheet names to lower case for comparison
                    sheet_names = {sheet.title.lower(): sheet for sheet in wb.worksheets}

                    # List to store ordered sheets
                    ordered_sheets = []

                    # Track missing sheets
                    missing_sheets = []

                    for sheet_name in desired_order:
                        # Check if the normalized name exists in the workbook
                        normalized_name = sheet_name.lower()
                        if normalized_name in sheet_names:
                            ordered_sheets.append(sheet_names[normalized_name])
                        else:
                            missing_sheets.append(sheet_name)

                    # If any sheets are missing, show the missing sheets and exit
                    if missing_sheets:
                        print(f"\n    ❌ The following sheets were missing and will be skipped:\n")
                        for missing_sheet in missing_sheets:
                            print(f"\n        🔸 {missing_sheet}")
                    else:
                        # Append any remaining sheets that were not in the desired order
                        remaining_sheets = [sheet for sheet in wb.worksheets if sheet not in ordered_sheets]
                        ordered_sheets.extend(remaining_sheets)

                        # Create a new workbook to hold the sheets in the desired order
                        new_wb = openpyxl.Workbook()
                        new_wb.remove(new_wb.active)  # Remove the default sheet created by openpyxl

                        for sheet in ordered_sheets:
                            new_sheet = new_wb.create_sheet(title=sheet.title)
                            for row in sheet.iter_rows(values_only=True):
                                new_sheet.append(row)

                        # Save the new workbook
                        new_wb.save(new_file_path)
                        print(f"\n    ✅ Rearranged workbook saved as '{new_file_path}'.")

                except Exception as e:
                    print(f"\n    ❌ An error occurred: {str(e)}")


            break
        
        elif choice == "no":
            print("\n        🚫 Strategy Sheet execution skipped!")

            break  # Just breaking without running Block 2
        
        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")

    # =========================================================================================================================================
    #                                                FINAL FILE EXECUTION
    # =========================================================================================================================================


    # ========================================================================
    print("\n")
    print("=" * 100)
    print(" " * 33 + "📝 FINAL SHEET EXECUTION 📝")
    print("=" * 100)
    # ========================================================================



    # ========================================================================
    # Step 1:-Initialize file name
    # ========================================================================


    print("\n\n🔍 Initializing file name and determining output path...")

    # Assume the file selected in Code 2 is stored in 'file_path'
    # Example:
    # file_path = "/Users/avirajmore/Downloads/ProductFamily_and_Currency_extract.xlsx"

    # Extract the selected file name from the file path
    selected_file_name = os.path.basename((file_path).split("/")[-1])

    # Replace '_Copy' (case-insensitive) and remove the extension
    folder_name = os.path.splitext(re.sub(r'_Copy', '', selected_file_name))[0]

    # Define the base path where the 'Final iteration files' folder exists (created by Code 1)
    base_dir = "/Users/avirajmore/Documents/Office Docs/Massload Files/2025"
    final_iteration_file_path = os.path.join(base_dir, Sprint_Number, subfolder_name, "Final iteration files")

    # Construct the path to the corresponding folder
    output = os.path.join(final_iteration_file_path, folder_name)


    # Validate the existence of the folder
    if os.path.exists(output) and os.path.isdir(output):
        print(f"\n    ✅ Output folder selected automatically:")
        # To remove unecessary path from the file path
        print(f"\n        📂 {output.split(base_dir, 1)[-1]}")

    else:
        print("\n    ❌ Error: The corresponding output folder does not exist.")
        print(f"\n       ❗️ Please check the path: {output}")


    # Generate CSV file names based on the folder name
    file_name = output.split('/')[-1]

    opportunity = "1_" + file_name + '_Opportunity load.csv'

    opportunity_product = "2_" + file_name + '_Opportunity Product load.csv'

    opportunity_team = "3_" + file_name + '_Opportunity Team member Load.csv'

    reporting_codes = "4_" + file_name + '_Opportunity Reporting_ codes.csv'

    tags = "5_" + file_name + '_Opportunity Reporting_codes_Tags.csv'

    print("\n    📄 CSV File Names Generated:")
    print(f"\n        1️⃣ {opportunity}")
    print(f"\n        2️⃣ {opportunity_product}")
    print(f"\n        3️⃣ {opportunity_team}")
    print(f"\n        4️⃣ {reporting_codes}")
    print(f"\n        5️⃣ {tags}")


    # ======================================================================
    # Step 1: Creating Opportunity Sheet
    # ======================================================================

    print("\n\n🔍 CREATING OPPORTUNITY FILE")

    output_file = output + "/" + opportunity  # Path for the processed CSV
    removed_rows_file = output+'/Removed Rows/Removed_Rows - Oppty.csv'  # Path for the removed rows CSV

    # Predefined columns to delete without asking
    predefined_columns_oppty = [
        'AccountNumber', 'Email', 'created_by', 'modified_by', 'created_date',
        'modified_date', 'Trimmed_accountid', 'Trimmed_ownerid', 'Type Of Opportunity',
        'Concatenatedaccountid', 'Concatenatedownerid', 'concatenatedcreatedby'
    ]

    # Columns to exclude from user deletion prompt
    excluded_columns = [
        'opportunity_legacy_id__c', 'Legacy_Opportunity_Split_Id__c', 'name', 'StageName',
        'Won_Reason__c', 'Lost_Category__c', 'Lost_Reason__c', 'CloseDate', 'CurrencyIsoCode',
        'OwnerId', 'NextStep', 'OI_Group__c'
    ]

    # Read the "Opportunity" sheet into a DataFrame
    try:
        opportunity_df = pd.read_excel(file_path, sheet_name='Opportunity')
    except Exception as e:
        print(f"\n    ❌ Error reading the file: {e}")
        exit()

    # Step 0: Remove blank columns, rows, and duplicates
    opportunity_df.dropna(axis=1, how='all', inplace=True)  # Remove blank columns
    opportunity_df.dropna(axis=0, how='all', inplace=True)  # Remove blank rows
    opportunity_df.drop_duplicates(inplace=True)  # Remove duplicate rows

    # Initialize a DataFrame to store removed rows with a "Removal Reason" column
    removed_rows_df = pd.DataFrame(columns=opportunity_df.columns.tolist() + ['Reason'])

    # Track all columns that will be dropped
    all_dropped_columns = []

    # Step 1: Remove predefined columns if they exist
    columns_to_delete = [col for col in predefined_columns_oppty if col in opportunity_df.columns]
    if columns_to_delete:
        opportunity_df.drop(columns=columns_to_delete, inplace=True)
        all_dropped_columns.extend(columns_to_delete)

    else:
        print("\n    ❗️ No predefined columns found for deletion.")

    # Step 2: User interface for column deletion with checkboxes
    root = Tk()
    root.title("Select Columns to Delete")
    root.geometry("500x600")
    root.resizable(False, False)

    canvas = Canvas(root)
    scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    frame = Frame(canvas)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    checkboxes = {}

    for column in opportunity_df.columns:
        if column not in excluded_columns:
            var = IntVar()
            checkboxes[column] = var
            checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
            checkbutton.pack(anchor="w", pady=5)

    button_frame = Frame(root)
    submit_button = Button(button_frame, text="Submit", command=root.quit, 
                        font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
    submit_button.pack(side="right")
    button_frame.pack(anchor="ne", padx=20, pady=10)
    frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

    root.mainloop()
    root.destroy()

    columns_to_delete_from_user = [col for col, var in checkboxes.items() if var.get() == 1]
    if columns_to_delete_from_user:
        opportunity_df.drop(columns=columns_to_delete_from_user, inplace=True)
        all_dropped_columns.extend(columns_to_delete_from_user)
        print("\n    ✅ Additional columns deleted:")
        for col in columns_to_delete_from_user:
            print(f"\n        🔸 {col}")
    else:
        print("\n    ✅ No additional columns selected for deletion.")

    # Step 3: Remove rows where AccountId is "Not in ISC"
    if 'AccountId' in opportunity_df.columns:
        rows_dropped_not_in_isc = opportunity_df[opportunity_df['AccountId'] == "Not in ISC"].copy()
        count_not_in_isc = len(rows_dropped_not_in_isc)  # Count the rows to be removed
        if not rows_dropped_not_in_isc.empty:
            rows_dropped_not_in_isc['Reason'] = "AccountId is 'Not in ISC'"
            opportunity_df = opportunity_df[opportunity_df['AccountId'] != "Not in ISC"]
            removed_rows_df = pd.concat([removed_rows_df, rows_dropped_not_in_isc], ignore_index=True)
            print(f"\n    ❗️ Rows with Account 'Not in ISC' : {count_not_in_isc}.")
        else:
            print("\n    ❗️ Rows with Account 'Not in ISC' : 0 .")

    # Step 4: Handle invalid PricebookEntryId rows
    try:
        opportunity_product_df = pd.read_excel(file_path, sheet_name='Opportunity_product')
        invalid_pricebook_ids = opportunity_product_df[
            opportunity_product_df['PricebookEntryId'].isin(['Not Active', 'No Pricebookid found'])
        ]['Legacy_Opportunity_Split_Id__c'].unique()

        rows_to_remove_invalid_pricebook = opportunity_df[
            opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
        ].copy()
        count_invalid_pricebook = len(rows_to_remove_invalid_pricebook)  # Count the rows to be removed
        if not rows_to_remove_invalid_pricebook.empty:
            rows_to_remove_invalid_pricebook['Reason'] = "Invalid PricebookEntryId"
            opportunity_df = opportunity_df[
                ~opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
            ]
            removed_rows_df = pd.concat([removed_rows_df, rows_to_remove_invalid_pricebook], ignore_index=True)
            print(f"\n    ❗️ Opportunities removed due to invalid PricebookEntryId: {count_invalid_pricebook}.")
        else:
            print("\n    ❗️ No opportunities found with invalid PricebookEntryId.")
    except Exception as e:
        print(f"\n    ❌ Error processing invalid PricebookEntryId rows: {e}")

    # Step 5: Drop columns from removed_rows_df, except "Removal Reason"
    columns_to_drop_from_removed = [col for col in all_dropped_columns if col in removed_rows_df.columns]
    if columns_to_drop_from_removed:
        removed_rows_df.drop(columns=columns_to_drop_from_removed, inplace=True)

    # Step 6: Save the processed DataFrame
    try:
        opportunity_df.to_csv(output_file, index=False)
        print("\n    ✅ Processed data saved to:")
        print(f"\n        📂 {"/".join(output_file.split("/")[-5:])}")
    except Exception as e:
        print(f"\n    ❌ Error saving the processed file: {e}")

    # Step 7: Save the removed rows DataFrame
    if not removed_rows_df.empty:
        try:
            removed_rows_df.to_csv(removed_rows_file, index=False)
            print(f"\n    ✅ Removed rows saved to:")
            print(f"\n        📂 {"/".join(removed_rows_file.split("/")[-5:])}")
        except Exception as e:
            print(f"\n    ❌ Error saving the removed rows file: {e}")

    # =======================================================
    # Step 2:- Creating product file
    # =======================================================


    print("\n\n🔍 CREATING PRODUCT FILE")
    # Step 1: Define the input Excel file path (commented out as requested)
    # file_path = 'your_file_path_here.xlsx'  # Replace with your actual file path
    sheet_name = 'Opportunity_product'

    # Define the predefined columns to delete
    predefined_columns_product = [
        'existing', 'product', 'product_type', 'Product_Family__c', 
        'opportunity currency', 'practise_multiple country', 
        'quantity.1', 'concatenated product family', 'concatenated currency'
    ]

    # Initialize lists to track deleted columns and dropped rows
    deleted_columns = []  # To store names of columns deleted
    rows_dropped = 0  # To count total rows dropped
    rows_dropped_existing_false = 0  # Track rows where 'existing' == False

    # Initialize a DataFrame to store removed rows where 'existing' == False
    removed_rows_df = pd.DataFrame()

    # Try to read data from the "Opportunity_product" sheet into a DataFrame
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Step 2: Remove all rows where 'existing' == False and store them in removed_rows_df
        initial_row_count = len(df)
        removed_rows_df = df[df['existing'] == False].copy()
        rows_dropped_existing_false = len(removed_rows_df)  # Track the number of rows removed
        df = df[df['existing'] == True]

        # Print the count of rows removed where 'existing' == False
        print(f"\n    ❗️ Rows removed where 'existing' == False: {rows_dropped_existing_false}")

        # Add a "Reason" column to the removed rows to specify why they were removed
        removed_rows_df['Reason'] = "Opportunity Missing From Main sheet"

        # Step 3: Remove predefined columns from both the main DataFrame and removed rows DataFrame
        columns_to_delete_predefined = [col for col in predefined_columns_product if col in df.columns]
        if columns_to_delete_predefined:
            df.drop(columns=columns_to_delete_predefined, inplace=True)
            removed_rows_df.drop(columns=columns_to_delete_predefined, inplace=True, errors='ignore')
            deleted_columns.extend(columns_to_delete_predefined)

        # Step 4: Set up a graphical interface (GUI) to select columns to delete
        root = Tk()
        root.title("Select Columns to Delete")

        # Set window size and make it fixed
        root.geometry("500x600")
        root.resizable(False, False)

        # Scrollbar setup
        canvas = Canvas(root)
        scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        frame = Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor="nw")

        # Dictionary to hold IntVar for each checkbox (for column selection)
        checkboxes = {}

        # Step 5: Add checkboxes for each column in the DataFrame
        for column in df.columns:
            var = IntVar()
            checkboxes[column] = var
            checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
            checkbutton.pack(anchor="w", pady=5)

        # Create a frame for the submit button and place it at the top right
        button_frame = Frame(root)
        submit_button = Button(button_frame, text="Submit", command=root.quit, 
                            font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
        submit_button.pack(side="right")

        # Place the button frame in the grid to ensure it stays at the top right
        button_frame.pack(anchor="ne", padx=20, pady=10)  # 'ne' positions it top-right

        # Update the scroll region to fit all elements
        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

        # Run the Tkinter main loop
        root.mainloop()
        root.destroy()

        # Step 6: Process the selected columns to delete after user submits
        columns_to_delete_from_user = [col for col, var in checkboxes.items() if var.get() == 1]

        if columns_to_delete_from_user:
            # Remove the selected columns from the main DataFrame (df)
            df.drop(columns=columns_to_delete_from_user, inplace=True)
            # Also remove the same columns from the removed rows DataFrame (removed_rows_df)
            removed_rows_df.drop(columns=columns_to_delete_from_user, inplace=True, errors='ignore')
            deleted_columns.extend(columns_to_delete_from_user)
            print("\n    ✅ Additional columns deleted:")
            for col in columns_to_delete_from_user:
                print(f"\n        🔸 {col}")
        else:
            print("\n    ✅ No additional columns selected for deletion.")


        # Step 7: Remove any rows that contain only blank values in the main DataFrame
        df.dropna(axis=0, how='all', inplace=True)

        # Step 8: Remove any columns that contain only blank values in the main DataFrame
        df.dropna(axis=1, how='all', inplace=True)

        # Step 9: Remove any duplicate rows based on all columns in the main DataFrame
        # df.drop_duplicates(inplace=True)

        # Step 10: Define the output CSV file path (updated with new values)
        output_file = output + "/" + opportunity_product  # Path for the processed CSV

        # Step 11: Check if the directory exists before saving
        output_dir = os.path.dirname(output_file)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist

        # Step 12: Save the processed DataFrame to the specified CSV file
        df.to_csv(output_file, index=False)
        print("\n    ✅ Processed data saved to")
        print(f"\n        📂 {"/".join(output_file.split("/")[-5:])}")

        # Step 13: Save the removed rows (where 'existing' == False) to a separate CSV file
        removed_rows_file2 = output+'/Removed Rows/Removed_Rows - Product.csv'  # Path for the removed rows CSV
        if not removed_rows_df.empty:
            removed_rows_df = removed_rows_df[removed_rows_df.columns.intersection(df.columns.tolist() + ['Reason'])]
            removed_rows_df.to_csv(removed_rows_file2, index=False)
            print("\n    ✅ Removed rows saved to:")
            print(f"\n        📂 {"/".join(removed_rows_file2.split("/")[-5:])}")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e}")



    # =======================================================
    # Step 3: Processing Opportunity Team Data
    # =======================================================

    # Ask user if they want to run the Team member sheet

    while True:
        print("\n================================================================================")
        user_input = input("\n📄 Do you want to run the Opportunity Team Member Sheet? (yes/no): ").strip().lower()
        if user_input == "yes":
            print("\n    ⏳ Running Opportunity Team Member Sheet...")
            print("\n================================================================================")

            # =======================================================
            # MAIN CODE
            # =======================================================

            print("\n\n🔍 CREATING TEAM MEMBER FILE")

            # Predefined columns to delete
            predefined_columns_Team = ['Existing', 'email', 'Concat_T_M']

            # File paths
            # file_path = '/path/to/your/input_file.xlsx'  # Uncomment and set your file path
            sheet_name = 'Opportunity_team'
            output_file = output + '/' + opportunity_team  # Path for the processed CSV
            removed_rows_file = output+'/Removed Rows/Removed_Rows - Team.csv'  # Path for removed rows CSV

            # Initialize tracking variables
            deleted_columns = []
            rows_dropped = 0
            removed_rows_saved = False  # Flag to track if removed rows were saved

            try:
                # Step 1: Read the Excel file into a DataFrame
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Step 2: Remove rows where "Existing" == False
                if 'Existing' in df.columns:
                    removed_rows = df[df['Existing'] == False].copy()
                    df = df[df['Existing'] == True]
                    rows_dropped = len(removed_rows)

                    if rows_dropped > 0:
                        # Add a "Reason" column to removed rows
                        removed_rows['Reason'] = "Opportunity Missing From Main sheet"

                        # Remove predefined columns from removed rows
                        removed_rows.drop(columns=[col for col in predefined_columns_Team if col in removed_rows.columns], inplace=True)

                        # Save removed rows to CSV
                        removed_rows.to_csv(removed_rows_file, index=False)
                        removed_rows_saved = True  # Mark that removed rows were saved
                    else:
                        pass
                else:
                    print("\n    ❌ 'Existing' column not found in the DataFrame. Ensure it exists before processing.")
                    sys.exit()

                # Step 3: Remove predefined columns
                predefined_to_delete = [col for col in predefined_columns_Team if col in df.columns]
                if predefined_to_delete:
                    df.drop(columns=predefined_to_delete, inplace=True)
                    deleted_columns.extend(predefined_to_delete)

                # Step 4: GUI for selecting additional columns to delete
                root = Tk()
                root.title("Select Columns to Delete")
                root.geometry("500x600")
                root.resizable(False, False)

                # Scrollable UI
                canvas = Canvas(root)
                scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
                canvas.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side="right", fill="y")
                canvas.pack(side="left", fill="both", expand=True)

                frame = Frame(canvas)
                canvas.create_window((0, 0), window=frame, anchor="nw")

                # Checkboxes for columns
                checkboxes = {}
                for column in df.columns:
                    var = IntVar()
                    checkboxes[column] = var
                    checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
                    checkbutton.pack(anchor="w", pady=5)

                # Submit button
                submit_button = Button(frame, text="Submit", command=root.quit, 
                                    font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
                submit_button.pack(pady=20)

                # Run the GUI
                frame.update_idletasks()
                canvas.config(scrollregion=canvas.bbox("all"))
                root.mainloop()
                root.destroy()

                # Step 5: Process user-selected columns for deletion
                user_selected_columns = [col for col, var in checkboxes.items() if var.get() == 1]
                if user_selected_columns:
                    df.drop(columns=user_selected_columns, inplace=True)
                    deleted_columns.extend(user_selected_columns)
                    print("\n    ✅ Additional columns deleted:")
                    for col in user_selected_columns:
                        print(f"\n        🔸 {col}")
                else:
                    print("\n    ✅ No additional columns selected for deletion.")

                # Step 6: Cleaning up the DataFrame
                df.dropna(axis=0, how='all', inplace=True)  # Remove rows with all blank values
                df.dropna(axis=1, how='all', inplace=True)  # Remove columns with all blank values
                df.drop_duplicates(inplace=True)  # Remove duplicate rows

                # Step 7: Save the processed DataFrame
                output_dir = os.path.dirname(output_file)
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                df.to_csv(output_file, index=False)

                # Final Print Statements
                print(f"\n    ❗️ Total rows dropped where 'Existing' == False: {rows_dropped}")
                if removed_rows_saved:
                    print("\n    ✅ Removed rows saved to:")
                    print(f"\n        📂 {"/".join(removed_rows_file.split("/")[-5:])}")

                print(f"\n    ✅ Processed data saved to: ")
                print(f"\n        📂 {"/".join(output_file.split("/")[-5:])}")

            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break
        
        elif user_input == "no":
            print("\n    🛑 Skipping team member Sheet...")
            print("\n================================================================================")
            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 4: Processing Reporting code
    # =======================================================

    while True:
        print("\n================================================================================")
        user_input = input("\n📄 Do you want to run the Reporting codes Sheet? (yes/no): ").strip().lower()
        if user_input == "yes":
            print("\n    ⏳ Running Reporting codes Sheet...")
            print("\n================================================================================")

            print("\n\n🔍 CREATING REPORTING CODES FILE")
            # Predefined columns
            predefined_columns_Reportingcode = ['reporting_codes', 'existing', 'concatcodes']

            # File paths
            # file_path = 'your_file_path_path.xlsx'  # Uncomment and set your file path
            sheet_name = 'Reporting_codes'
            output_file = output + "/" + reporting_codes  # Path for the processed CSV
            removed_rows_file = output+'/Removed Rows/Removed_Rows - ReportingCodes.csv'  # Path for removed rows CSV

            # Initialize variables
            deleted_columns = []
            rows_dropped = 0

            try:
                # Step 1: Read the Excel file into a DataFrame
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Step 2: Convert all column names to lowercase
                df.columns = df.columns.str.lower()

                # Step 3: Check if 'existing' column exists
                if 'existing' not in df.columns:
                    raise ValueError(f"\n    ❌ Column 'existing' not found in the DataFrame from sheet '{sheet_name}'. Please check your input data.")

                # Step 4: Remove rows where "existing" == False and save them to another file
                removed_rows = df[df['existing'].astype(str).str.lower() != 'true'].copy()
                df = df[df['existing'].astype(str).str.lower() == 'true']
                rows_dropped = len(removed_rows)

                if rows_dropped > 0:
                    removed_rows['Reason'] = "Opportunity Missing From Main sheet"
                    removed_rows.drop(columns=[col for col in predefined_columns_Reportingcode if col in removed_rows.columns], inplace=True)
                    removed_rows.to_csv(removed_rows_file, index=False)


                # Step 5: Remove predefined columns from the main data
                predefined_to_delete = [col for col in predefined_columns_Reportingcode if col in df.columns]
                if predefined_to_delete:
                    df.drop(columns=predefined_to_delete, inplace=True)
                    deleted_columns.extend(predefined_to_delete)

                # Step 6: Create UI for selecting additional columns to delete
                root = Tk()
                root.title("Select Columns to Delete")

                # Set window size
                root.geometry("500x600")
                root.resizable(False, False)

                # Scrollable UI
                canvas = Canvas(root)
                scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
                canvas.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side="right", fill="y")
                canvas.pack(side="left", fill="both", expand=True)

                frame = Frame(canvas)
                canvas.create_window((0, 0), window=frame, anchor="nw")

                # Checkbox setup
                checkboxes = {}
                for column in df.columns:
                    var = IntVar()
                    checkboxes[column] = var
                    checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
                    checkbutton.pack(anchor="w", pady=5)

                # Submit button
                submit_button = Button(frame, text="Submit", command=root.quit, font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
                submit_button.pack(pady=20)

                # Run the UI
                frame.update_idletasks()
                canvas.config(scrollregion=canvas.bbox("all"))
                root.mainloop()
                root.destroy()

                # Step 7: Process user-selected columns for deletion
                user_selected_columns = [col for col, var in checkboxes.items() if var.get() == 1]
                if user_selected_columns:
                    df.drop(columns=user_selected_columns, inplace=True)
                    deleted_columns.extend(user_selected_columns)
                    print("\n    ✅ Additional columns deleted:")
                    for col in user_selected_columns:
                        print(f"\n        🔸 {col}")
                else:
                    print("\n    ✅ No additional columns selected for deletion.")

                # Step 8: Clean up the DataFrame
                df.dropna(axis=0, how='all', inplace=True)  # Remove rows with all blank values
                df.dropna(axis=1, how='all', inplace=True)  # Remove columns with all blank values
                df.drop_duplicates(inplace=True)  # Remove duplicate rows

                # Step 9: Save the processed DataFrame
                output_dir = os.path.dirname(output_file)
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                df.to_csv(output_file, index=False)

                # Summary of deletions
                print(f"\n    ❗️ Total rows removed where 'existing' == False: {rows_dropped}")

                # Final summary messages
                print("\n    ✅ Processed data saved to:")
                print(f"\n        📂 {"/".join(output_file.split("/")[-5:])}")
                if rows_dropped > 0:
                    print("\n    ✅ Removed rows saved to:")
                    print(f"\n        📂 {"/".join(removed_rows_file.split("/")[-5:])}")

            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break
        elif user_input == "no":
            print("\n    🛑 Skipping Reporting Code sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 5: Processing Tags sheet
    # =======================================================

    while True:
        print("\n================================================================================")
        user_input = input("\n📄 Do you want to run the Tags Sheet? (yes/no): ").strip().lower()
        if user_input == "yes":
            print("\n    ⏳ Running Tags Sheet...")
            print("\n================================================================================")

            # Predefined columns
            predefined_columns_tags = ['tag', 'existing', 'concattags']

            # File paths
            # file_path = 'your_file_path_path.xlsx'  # Uncomment and set your file path
            sheet_name = 'Tags'
            output_file = output + "/" + tags  # Path for the processed CSV
            removed_rows_file = output+'/Removed Rows/Removed_Rows - Tags.csv'  # Path for removed rows CSV

            # Initialize variables
            deleted_columns = []
            rows_dropped = 0

            try:
                print("\n\n🔍 CREATING TAGS FILE")

                # Step 1: Read the Excel file into a DataFrame
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Step 2: Convert all column names to lowercase
                df.columns = df.columns.str.lower()

                # Step 3: Check if 'existing' column exists
                if 'existing' not in df.columns:
                    raise ValueError(f"\n    ❌ Column 'existing' not found in the DataFrame from sheet '{sheet_name}'. Please check your input data.")

                # Step 4: Remove rows where "existing" == False and save them to another file
                removed_rows = df[df['existing'].astype(str).str.lower() != 'true'].copy()
                df = df[df['existing'].astype(str).str.lower() == 'true']
                rows_dropped = len(removed_rows)

                if rows_dropped > 0:
                    # Add a "Reason" column to removed rows
                    removed_rows['Reason'] = "Opportunity Missing From Main sheet"

                    # Save removed rows without predefined columns
                    removed_rows.drop(columns=[col for col in predefined_columns_tags if col in removed_rows.columns], inplace=True)
                    removed_rows.to_csv(removed_rows_file, index=False)

                # Step 5: Remove predefined columns from the main data
                predefined_to_delete = [col for col in predefined_columns_tags if col in df.columns]
                if predefined_to_delete:
                    df.drop(columns=predefined_to_delete, inplace=True)
                    deleted_columns.extend(predefined_to_delete)

                # Step 6: Create UI for selecting additional columns to delete
                root = Tk()
                root.title("Select Columns to Delete")

                # Set window size
                root.geometry("500x600")
                root.resizable(False, False)

                # Scrollable UI
                canvas = Canvas(root)
                scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
                canvas.configure(yscrollcommand=scrollbar.set)
                scrollbar.pack(side="right", fill="y")
                canvas.pack(side="left", fill="both", expand=True)

                frame = Frame(canvas)
                canvas.create_window((0, 0), window=frame, anchor="nw")

                # Checkbox setup
                checkboxes = {}
                for column in df.columns:
                    var = IntVar()
                    checkboxes[column] = var
                    checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
                    checkbutton.pack(anchor="w", pady=5)

                # Submit button
                submit_button = Button(frame, text="Submit", command=root.quit, font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
                submit_button.pack(pady=20)

                # Run the UI
                frame.update_idletasks()
                canvas.config(scrollregion=canvas.bbox("all"))
                root.mainloop()
                root.destroy()

                # Step 7: Process user-selected columns for deletion
                user_selected_columns = [col for col, var in checkboxes.items() if var.get() == 1]
                if user_selected_columns:
                    df.drop(columns=user_selected_columns, inplace=True)
                    deleted_columns.extend(user_selected_columns)
                    print("\n    ✅ Additional columns deleted:")
                    for col in user_selected_columns:
                        print(f"\n        🔸 {col}")

                # Step 8: Clean up the DataFrame
                df.dropna(axis=0, how='all', inplace=True)  # Remove rows with all blank values
                df.dropna(axis=1, how='all', inplace=True)  # Remove columns with all blank values
                df.drop_duplicates(inplace=True)  # Remove duplicate rows

                # Step 9: Save the processed DataFrame
                output_dir = os.path.dirname(output_file)
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                df.to_csv(output_file, index=False)

                # Final summary of deletions
                print("\n    🔸 Total rows removed where 'existing' == False: {rows_dropped}")

                # Final summary messages
                print("\n    ✅ Processed data saved to:")
                print(f"\n        📂 {"/".join(output_file.split("/")[-5:])}")
                if rows_dropped > 0:
                    print("\n    ✅ Removed rows saved to:")
                    print(f"\n        📂 {"/".join(removed_rows_file.split("/")[-5:])}")

            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break

        elif user_input == "no":
            print("\n    🛑 Skipping Tag Sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")


    # ========================================================================
    # Last Step: Copy the Summary File to the Folder
    # ========================================================================


    print("\n\n🔍 Copying the Summary File to the Selected Folder...")

    # Path to the reference file
    reference_file_path = "/Users/avirajmore/Documents/Office Docs/Massload Files/Reference File/Reference_Summary_file.xlsx"

    # Check if the reference file exists
    if not os.path.exists(reference_file_path):
        print("\n    ❌ Error: Reference file does not exist at the specified path.")
        print(f"\n       📂 Path: {reference_file_path}\n")
    else:
        try:
            # Get the folder name selected in Code 3
            selected_folder_name = os.path.basename(output)

            # Define the destination file path with the new name
            destination_file_path = os.path.join(output, f"{selected_folder_name}_summary file.xlsx")

            # Copy and rename the reference file to the destination
            shutil.copy(reference_file_path, destination_file_path)
            print(f"\n    ✅ Reference file copied successfully to the folder: {selected_folder_name}")

            # Load the copied summary file using openpyxl
            wb = load_workbook(destination_file_path)

            # Look for the sheet named "Summary"
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]  # Access the "Summary" sheet

                # Write the folder name (extracted from output path) to cell D4
                ws['D4'] = selected_folder_name

                # Save the updated file (ensuring no other changes are made)
                wb.save(destination_file_path)
                print(f"\n    ✅ Folder name '{selected_folder_name}' written to cell D4 in the 'Summary' sheet.")
            else:
                print("\n    ❌ Error: 'Summary' sheet not found in the copied file.")
                print(f"\n       ❗️ Please check the file structure.")

        except Exception as e:
            print("\n    ❌ An error occurred while copying or modifying the file:")
            print(f"\n    ❗️ Details: {e}")


    # ========================================================================
    print("\n")
    print("=" * 100)
    print(" " * 33 + "📝 FINAL SHEET COMPLETED 📝")
    print("=" * 100)
    # ========================================================================

    # =====================================================
    # Delete CSV Files
    # =====================================================
    # Hardcoded directory
    directory = "/Users/avirajmore/Downloads"

    print("\n\n🔍 Delete the extract files")

    response = input(f"\n    👉 Do you want to delete all the extract files? (yes/no): ").strip().lower()
    if response == 'yes':
        # Get list of files in the directory
        def delete_folder(folder_path):
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
                print(f"\n        🗑️Folder '{folder_path.split('/')[-1]}' and its contents have been deleted.")
            else:
                print(f"\n        ❗️ Folder '{folder_path}' does not exist.")


        delete_folder('/Users/avirajmore/Documents/Office Docs/Python file/1_Mass load Python/Extracts')
        delete_folder('/Users/avirajmore/Documents/Office Docs/Python file/1_Mass load Python/Delete')
    else:
        print("\n        🛑 No files were deleted.")

    # Ask the user for confirmation
    print("\n\n🔍 Delete CSV Files in Downloads folder")

    response = input(f"\n    👉 Do you want to delete all CSV files in the Downloads folder? (yes/no): ").strip().lower()

    if response == 'yes':

        try:
            files_deleted = 0
            for file_name in os.listdir(directory):
                if file_name.endswith('.csv'):
                    file_path = os.path.join(directory, file_name)
                    os.remove(file_path)
                    files_deleted += 1

            if files_deleted > 0:
                print("\n        🗑️ All CSV files have been successfully deleted.")
            else:
                print("\n        ❗️ No CSV files found in the directory.")
        except Exception as e:
            print(f"\n        ❌ An error occurred while deleting files:{e}")

    else:
        print("\n        🛑 No files were deleted.")

    print("\n")
    print("=" * 100)
    print(f"\n ✅ File Prepared: {filename} ✅\n")
    print("=" * 100)
    


    while True:  # Inner loop
        continue_processing = input("\n 👉 Do you want to process another file? (yes/no): ").strip().lower()
        
        if continue_processing == 'yes':
            # Add logic for processing another file here
            print("\n    ⏳ Processing the file...")
            break  # Exit the inner loop and continue to the next iteration of the outer loop
        
        elif continue_processing == 'no':
            print(f"\n     🔚 End of Script\n")
            print("=" * 100)
            print("\n")
            break  # Exit the inner loop
        else:
            print("\n    ❗️ Invalid input. Please select 'yes' or 'no'.")
    
    if continue_processing == 'no':
        break  # Exit the outer loop if the user selects 'no'
