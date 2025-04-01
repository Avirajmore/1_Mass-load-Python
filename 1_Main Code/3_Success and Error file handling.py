import os
import shutil
from tkinter import Tk, filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tabulate import tabulate

# =======================================================================================================
print("=" * 100)
print(" " * 33 + "📝 Success And Error Files EXECUTION 📝")
print("=" * 100)
# =======================================================================================================

# Copying all the success and error files as backup
print("\n\n🔍 Step 1: Select CSV files and a destination folder to copy.")

root = Tk()
root.withdraw()  # Hide the root window
root.attributes('-topmost', True)  # Bring dialog to front

# Select CSV files
csv_files = root.tk.splitlist(filedialog.askopenfilenames(
    title="📂 Select CSV Files to Copy",
    filetypes=[("CSV Files", "*.csv")],
    initialdir=os.path.expanduser("~/Downloads")
))

# Handle no files selected
if not csv_files:
    print("\n    ❗️ No files selected. Skipping to the next step.")
else:
    # Select destination folder
    destination_folder = filedialog.askdirectory(
        title="📁 Select Destination Folder to Paste Success/Error Files"
    )

    # Handle no destination folder selected
    if not destination_folder:
        print("\n    ❗️ No destination folder selected. Skipping file copy.")
    else:
        success_files = []
        failed_files = []

        # Copy each CSV file to the selected folder
        for file in csv_files:
            try:
                shutil.copy(file, destination_folder)
                success_files.append(os.path.basename(file))
            except Exception:
                failed_files.append(os.path.basename(file))

        # Print results for copied files
        if success_files:
            print("\n    ✅ Successfully Copied Files:")
            for file in success_files:
                print(f"\n        📄 {file}")

        if failed_files:
            print("\n    ❌ Failed to Copy Files:")
            for file in failed_files:
                print(f"\n        ❗️ {file}")


# =======================================================================================================
# Selecting a summary file and adding CSV data
print("\n\n🔍 Step 2: Select a summary file where you want to copy the CSV data.")

# Select summary file
summary_file_path = filedialog.askopenfilename(
    title="📂 Select a Summary File"
)

# Handle no file selected
if not summary_file_path:
    print("\n\n    ❌ No file selected. Operation canceled.")
    exit()

print(f"\n    ✅ Selected file:")
shortened_path = "/".join(summary_file_path.split("/")[-5:])
print(f"\n         📂 {shortened_path}")


# Taking the file name automatically from Summary file name
file_name = summary_file_path.split("/")[-1].split("_summary file.xlsx")[0]



# Define the CSV files and their corresponding sheet names
files_and_sheets = {
    "opptysuccess.csv"  : "Opportunity Success",
    "opptyerror.csv"    : "Opportunity Failures",
    "productsuccess.csv": "Opportunity Product Success",
    "producterror.csv"  : "Opportunity Product Failures",
    "teamsuccess.csv"   : "Team Member Success",
    "teamerror.csv"     : "Team Member Failure",
    "codessuccess.csv"  : "Reporting Code Success",
    "codeserror.csv"    : "Reporting Code Failure",
    "tagssuccess.csv"   : "Tag Success",
    "tagserror.csv"     : "Tag Failure",
    "feedsuccess.csv"   : "Feed Item Success",
    "feederror.csv"     : "Feed Item Failure",
    "contactsuccess.csv": "Contact Success",
    "contacterror.csv"  : "Contact Failure"
}

# Define Success and error folder path

half_path = '/'.join(summary_file_path.split('/')[:-1]) #Automatically take the path from Summary file selected above 
# for example 
# if summary file path is '/Users/avirajmore/Documents/Office Docs/Massload Files/2025/Demo/demo/Final iteration files/TSS Massload - Batch 133-10012025/Mass_load_template_CHEVA - IAPP 2024_summary file.xlsx'
# Half path will be:- '/Users/avirajmore/Documents/Office Docs/Massload Files/2025/Demo/demo/Final iteration files/TSS Massload - Batch 133-10012025

Success_error_path = half_path + '/Success and error files'
# For example:- 
# '/Users/avirajmore/Documents/Office Docs/Massload Files/2025/Demo/demo/Final iteration files/TSS Massload - Batch 133-10012025' + '/Success and error files'

# Check if the Excel file exists
if os.path.exists(summary_file_path):
    workbook = load_workbook(summary_file_path)
else:
    print(f"\n    ❌ Error: Excel file '{summary_file_path}' not found. Exiting...")
    exit()

# Check if any sheet already has data
data_exists = any(
    sheet.max_row > 1 for sheet in workbook.worksheets if sheet.title in files_and_sheets.values()
)

# Ask user once if they want to continue, keep asking until valid input
if data_exists:
    while True:
        user_response = input(
            "\n        ❗️ Some sheets already have data. Do you want to continue and append data to all sheets? (yes/no): "
        ).strip().lower()

        if user_response == 'yes':
            break
        elif user_response == 'no':
            print("\n    ❌ Operation canceled by the user.")
            exit()
        else:
            print("\n           ❌ Invalid choice. Please enter 'yes' or 'no'.")

# Start processing each CSV file
# Process each CSV file
processed_files = []
skipped_files = []

for csv_filename, sheet_name in files_and_sheets.items():
    csv_file_path = os.path.join(Success_error_path, csv_filename)

    # Check if the CSV file exists
    if not os.path.exists(csv_file_path):
        skipped_files.append((csv_filename, "File not found"))
        continue

    # Load the CSV data
    csv_data = pd.read_csv(csv_file_path)

    # Create the sheet if it doesn't exist
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    # Overwrite the CSV data in the sheet
    sheet = workbook[sheet_name]

    # Clear existing sheet content
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.value = None
    # Clear all rows in the sheet
    sheet.delete_rows(1, sheet.max_row)

    # Write the new CSV data
    for row in dataframe_to_rows(csv_data, index=False, header=True):
        sheet.append(row)

    processed_files.append(csv_filename)

# Save the workbook after all modifications
workbook.save(summary_file_path)


print(f"\n    ✅ {len(processed_files)} CSV file(s) successfully written to the Excel file.")

print(f"\n    ⭕️ {len(skipped_files)} file(s) were not processed.")
for i, (file, reason) in enumerate(skipped_files, 1):
    print(f"\n        ❗️ {file}")



# ==============================================================
# Step 3 :- Copy the data from Removed rows to summary file
# ==============================================================

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
print("\n\n🔍 Step 3: Copy the data from Removed rows to summary file")

# Define the CSV-to-Sheet mapping
files_and_sheets = {

    "Removed_Rows - Oppty.csv"                                  : "Opportunity Failures",

    "Removed_Rows - Product.csv"                                : "Opportunity Product Failures",
    
    "Removed_Rows - Product (Duplicate Opportunity).csv"        : "Opportunity Product Failures",
    
    "Removed_Rows - Team.csv"                                   : "Team Member Failure",
    
    "Removed_Rows - Team member (Duplicate Opportunity).csv"    : "Team Member Failure",
    
    "Removed_Rows - ReportingCodes.csv"                         : "Reporting Code Failure",
    
    "Removed_Rows - Codes (Duplicate Opportunity).csv"          : "Reporting Code Failure",
    
    "Removed_Rows - Tags.csv"                                   : "Tag Failure",
    
    "Removed_Rows - Tags (Duplicate Opportunity).csv"           : "Tag Failure",
    
    "Removed_Rows - Feed Item.csv"                              : "Feed Item Failure",
    
    "Removed_Rows - Contact.csv"                                : "Contact Failure",
}

# Define column mapping (CSV column name -> Summary sheet column name)
column_mapping = {
    "Reason": "ERRORS",  # Example: Map 'Reason' in CSV to 'Error' in the summary sheet
    # Add more mappings if needed
}

# Define the folder path where the CSV files are stored
Remove_rows_path = half_path + '/Removed Rows'

# Ask user for the summary file path
# summary_file_path = input("Enter the full path of the summary Excel file: ").strip()

# Check if the Excel file exists
if not os.path.exists(summary_file_path):
    print(f"\n    ❌ Error: Excel file '{summary_file_path}' not found. Exiting...")
    exit()

# Load the workbook
workbook = load_workbook(summary_file_path)

# Process each CSV file
processed_files = []
skipped_files = []

for csv_filename, sheet_name in files_and_sheets.items():
    csv_file_path = os.path.join(Remove_rows_path, csv_filename)

    # Check if the CSV file exists
    if not os.path.exists(csv_file_path):
        skipped_files.append((csv_filename, "File not found"))
        continue

    # Load the CSV data
    csv_data = pd.read_csv(csv_file_path)

    # Apply column mapping
    csv_data.rename(columns=column_mapping, inplace=True)

    # Check if the sheet exists in the workbook
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    # Get the sheet
    sheet = workbook[sheet_name]

    # Get the header of the sheet (summary file)
    sheet_header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

    # If the sheet is empty, populate it with CSV headers
    if not sheet_header or all(h is None for h in sheet_header):
        sheet_header = list(csv_data.columns)  # Convert pandas Index to a list
        for col_idx, header in enumerate(sheet_header, start=1):
            sheet.cell(row=1, column=col_idx, value=header)  # Write headers to the first row

    # Reorder CSV columns to match the sheet's columns
    csv_data = csv_data.reindex(columns=sheet_header, fill_value=None)

    # Append data row by row (excluding headers from CSV)
    for row in csv_data.itertuples(index=False):
        sheet.append(row)

    processed_files.append(csv_filename)

# Save the workbook after all modifications
workbook.save(summary_file_path)

# Print Summary
print(f"\n    ✅ {len(processed_files)} CSV file(s) successfully appended to the Excel file.")
if skipped_files:
    print(f"\n    ⭕ {len(skipped_files)} file(s) were not processed:")
    for i, (file, reason) in enumerate(skipped_files, 1):
        print(f"\n        {i}. {file}")


# ==============================================================
# Step 4 :- Mention the count of Success and Error files
# ==============================================================

import os
from openpyxl import load_workbook

print("\n\n🔍 Step 4: Mention the count of Success and Error files")

# Define the mapping of sheets to cells in the "Summary" sheet
sheet_to_cell_mapping = {
    "Opportunity Success": "E5",
    "Opportunity Failures": "F5",
    "Opportunity Product Success": "E6",
    "Opportunity Product Failures": "F6",
    "Team Member Success": "E7",
    "Team Member Failure": "F7",
    "Reporting Code Success": "E8",
    "Reporting Code Failure": "F8",
    "Tag Success": "E9",
    "Tag Failure": "F9",
}

# Define the cell ranges for summation in the "Summary" sheet
sum_mappings = {
    ("E5", "F5"): "G5",
    ("E6", "F6"): "G6",
    ("E7", "F7"): "G7",
    ("E8", "F8"): "G8",
    ("E9", "F9"): "G9",
}

# Ask user for the summary file path
# summary_file_path = input("Enter the full path of the summary Excel file: ").strip()

# Check if the file exists
if not os.path.exists(summary_file_path):
    print(f"\n    ❌ Error: Summary file '{summary_file_path}' not found. Exiting...")
    exit()

# Load the workbook
workbook = load_workbook(summary_file_path)

# Check if the "Summary" sheet exists
if "Summary" not in workbook.sheetnames:
    print(f"\n    ❌ Error: 'Summary' sheet not found in the workbook. Exiting...")
    exit()

# Get the "Summary" sheet
summary_sheet = workbook["Summary"]

# List to collect names of missing sheets
missing_sheets = []

# Process each sheet and count rows
for sheet_name, cell_address in sheet_to_cell_mapping.items():
    if sheet_name not in workbook.sheetnames:
        missing_sheets.append(sheet_name)  # Collect missing sheet names
        continue

    # Get the sheet
    sheet = workbook[sheet_name]

    # Count the number of rows (excluding the header)
    row_count = sum(1 for _ in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True) if any(_))

    # Write the row count to the specified cell in the "Summary" sheet
    summary_sheet[cell_address] = row_count

# Perform the summation in the "Summary" sheet
for (cell_1, cell_2), result_cell in sum_mappings.items():
    value_1 = summary_sheet[cell_1].value or 0
    value_2 = summary_sheet[cell_2].value or 0
    summary_sheet[result_cell] = value_1 + value_2

# Save the updated workbook
workbook.save(summary_file_path)

# Output the result
if missing_sheets:
    print(f"\n    ❗️ Skipped as the sheet was not found")
    for i, sheet in enumerate(missing_sheets, 1):
        print(f"\n        {i}. {sheet}")

print(f"\n    ✅ Row counts and summations updated successfully in the 'Summary' sheet of :")
shortened_path = "/".join(summary_file_path.split('/')[-5:])
print(f"\n         📂 {shortened_path}")


# ============================================================
# Show the summary data as table
# ============================================================
from collections import defaultdict
from tabulate import tabulate

# Group data by high-level categories
categories = defaultdict(lambda: {"Success": 0, "Failures": 0, "Total": 0})

# Assuming `sheet_to_cell_mapping` is already defined as provided
sheet_to_cell_mapping = {
    "Opportunity Success": "E5",
    "Opportunity Failures": "F5",
    "Opportunity Product Success": "E6",
    "Opportunity Product Failures": "F6",
    "Team Member Success": "E7",
    "Team Member Failure": "F7",
    "Reporting Code Success": "E8",
    "Reporting Code Failure": "F8",
    "Tag Success": "E9",
    "Tag Failure": "F9"
}

# Example summary_sheet (assuming it's already loaded)
# summary_sheet = your_loaded_summary_sheet_object

for sheet_name, cell_address in sheet_to_cell_mapping.items():
    # Extract high-level category from the sheet name (e.g., remove "Success"/"Failures")
    category = " ".join(sheet_name.split()[:-1])
    
    # Identify Success or Failures based on the sheet name
    if "Success" in sheet_name:
        value = summary_sheet[cell_address].value or 0
        categories[category]["Success"] += value
    elif "Failure" in sheet_name:
        value = summary_sheet[cell_address].value or 0
        categories[category]["Failures"] += value

# Calculate the Total for each category
for category, counts in categories.items():
    counts["Total"] = counts["Success"] + counts["Failures"]

# Prepare the table data without the "Comments" column
table_data = [
    [category, counts["Success"], counts["Failures"], counts["Total"]]
    for category, counts in categories.items()
]

# Define table headers with bold formatting (headers are bold)
headers = ["Summary", "Success", "Failures", "Total"]

# Print the table with enhanced formatting
print("\n\n🔍 Aggregated Counts from the 'Summary' Sheet:\n")
table = tabulate(table_data, headers=headers, tablefmt="fancy_grid", numalign="right", stralign="center")
# Bold headers
table = table.replace(headers[0], f"\033[1m{headers[0]}\033[0m")
table = table.replace(headers[1], f"\033[1m{headers[1]}\033[0m")
table = table.replace(headers[2], f"\033[1m{headers[2]}\033[0m")
table = table.replace(headers[3], f"\033[1m{headers[3]}\033[0m")

print(table)

print("\n    ✅ Simulation Complete. 'Summary' Sheet Updated Successfully.\n")

# ============================================================
# ============================================================

print("\n")
print("=" * 100)
print(" " * 33 + "📝 Success And Error Files Completed 📝")
print("=" * 100)
print("\n")