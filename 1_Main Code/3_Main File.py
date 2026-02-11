# Importing all the necessary Libraries
import time
import os
import re
import sys
import shutil
import openpyxl
import pyperclip
import numpy as np
import pandas as pd
from tkinter import *
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import SheetTitleException

# =========================================================================================================================================
                                                 # FOLDER CREATION & FILE MOVEMENT
# =========================================================================================================================================
start_time = time.time()   # Record start time
# =========================================================
# Define the base paths for storing mass load files
# =========================================================

# Path of the folder Where you want to save the Mass Load Files. 
# ❗️ Change this path if you want to store files in a different location
BASE_DIR = os.path.expanduser("~/Documents/Office Docs/Massload Files/2026") 

# for Example :-  os.path.expanduser("~/Downloads") == "/Users/avirajmore/Downloads" 

# Path of the folder Where you have saved a the template of the Summary file
# ❗️ Change this path if your summary file is stored elsewhere
REF_SUMMARY_FILE_PATH = os.path.expanduser("~/Documents/Office Docs/Massload Files/Reference File/Reference_Summary_file.xlsx") #Change it to where you want to store the summary file

# Path for the Downloads Folder
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")

# =========================================================
# Folder Creation starts
# =========================================================

# Function to display a title with a decorative line
def show_title(title):

    line_width = 100
    line = "=" * line_width
    print(f"\n{line}")
    print(title.center(line_width))
    print(f"{line}\n")

# Display the title for the folder creation and file movement process
title = "📂 FOLDER CREATION & FILE MOVEMENT 📂"
show_title(title)

# =========================================================
# Function to validate folder names 
# =========================================================

# Function to validate folder names
def is_valid_folder_name(name):
    invalid_chars = set('\\/:*?\"<>|')
    return name and not any(char in invalid_chars for char in name) # return True if name is not empty and Does not have Invalid Characters

# =========================================================
print("\n\n🔍 Step 1: Creating Main Folder")

# Prompt the user for a valid folder name, Keep prompting the user until a valid folder name is provided 
while True:

    Current_Iteration_Folder_Name  = input("\n    📂 Enter the name of the Folder: ").strip()
    
    # Check if the folder name is valid
    if is_valid_folder_name(Current_Iteration_Folder_Name ): 
        break # If valid, exit the loop and proceed
    
    else:
        # If invalid, display an error message and prompt again
        print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

# Create the main folder and subfolders

# Construct a Full path for the user-specified Folder inside BASE_DIR
Current_iteration_Folder = os.path.join(BASE_DIR, Current_Iteration_Folder_Name )

# Create the main subfolder (if it doesn't already exist)
os.makedirs(Current_iteration_Folder, exist_ok=True)

# Define paths for additional subfolders "Copy Files and Final Iteration File"
copy_file_path = os.path.join(Current_iteration_Folder, "Copy files")
final_iteration_file_path = os.path.join(Current_iteration_Folder, "Final iteration files")

# Create the additional subfolders "Copy Files and Final Iteration File"
os.makedirs(copy_file_path, exist_ok=True)
os.makedirs(final_iteration_file_path, exist_ok=True)

print(f"\n        ✅ Subfolders 'Copy files' and 'Final iteration files' created successfully")

# =========================================================
# Step 2: Move Mass load files from Downloads to Current Iteration Folder
# =========================================================

# Define supported Excel file extensions
excel_extensions = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm')

# List to keep track of successfully moved files
files_moved = []

print(f"\n\n🔍 Step 2: Moving Excel files")

# Check if any Excel files exist in Downloads folder before moving them
if os.path.exists(DOWNLOADS_DIR):
    
    # Loop through all files in the Downloads folder
    for file_name in os.listdir(DOWNLOADS_DIR):
    
        # Check if the file has an Excel file extension (case-insensitive)    
        if file_name.lower().endswith(excel_extensions):
            
            # Construct full source and target file paths
            source_path = os.path.join(DOWNLOADS_DIR, file_name)
            target_path = os.path.join(Current_iteration_Folder, file_name)
            
            # Move the file from Downloads to the target folder
            shutil.move(source_path, target_path)

            # Add the file name to the list of moved files
            files_moved.append(file_name)


# Populate the excel_files list with all Excel files in the Current Iteration Folder

excel_files_Curr_Iter = [] # Initialize an empty List to store All the File

for f in os.listdir(Current_iteration_Folder):
    if f.lower().endswith(excel_extensions): # Check if the file has a valid Excel extension
        excel_files_Curr_Iter.append(f)

# Display the files that were moved from the Downloads folder (if any)

if not files_moved:
    print("\n    📥 Moved Files:")
    print("\n        1) ❗️ No files were moved from Downloads. ")
else:
    print("\n    📥 Moved Files:")
    for index, file_name in enumerate(files_moved, start=1):
        print(f"\n        {index}) {file_name} ✅") # Show each moved file with an index

# =========================================================
# Step 4: Make a Copy of Original Files and save them to "Copy Folder" 
# =========================================================

print("\n\n🔍 Step 3: Copying files")

files_copied = [] # Initialize an empty List

# Iterate through all Excel files in the current iteration folder

for file_name in excel_files_Curr_Iter: 
    
    # Rename Oriiginal File By adding "_Copy"
    copy_file_name = f"{os.path.splitext(file_name)[0]}_Copy{os.path.splitext(file_name)[1]}" 
    
    # Construct full source and target file paths
    source_path = os.path.join(Current_iteration_Folder, file_name)
    target_path = os.path.join(copy_file_path, copy_file_name)
    

    if os.path.exists(target_path):
        files_copied.append((file_name, "skipped"))
    else:
        shutil.copy(source_path, target_path)
        files_copied.append((file_name, "copied"))

# Display the results of the file copying operation

print("\n    📤 Copied Files:")
for index, (file_name, status) in enumerate(files_copied, start=1):
    if status == "copied":
        print(f"\n        {index}) {file_name} ✅") # File successfully copied
    elif status == "skipped":
        print(f"\n        {index}) {file_name} - Skipped 🛑 ") # File was skipped because it already existed


# =========================================================
# Step 5: Create Folders in "Final iteration files" to store Final processed Files and other Data
# =========================================================

print("\n\n🔍 Step 4: Creating folders in 'Final iteration files'")
Final_File_Folders_created = []

# Iterate through all Excel files in Currenct Iteration to create "Final Files" folders with that File's Name
for file_name in excel_files_Curr_Iter:

    # Extract file name without extension (.xlxs) by spliting the file name base on "." (dot) to Rename the Folders in Final Files Foldrt
    Final_File_Folder_Name = os.path.splitext(file_name)[0] 

    # Define the Folder path of Each File
    Final_File_Folder_Path = os.path.join(final_iteration_file_path, Final_File_Folder_Name) 

    # Ensure that the final folder, named after the Excel file, exists (create it if necessary).
    if not os.path.exists(Final_File_Folder_Path):
        os.makedirs(Final_File_Folder_Path, exist_ok=True)
        folder_status = "created"
    else:
        folder_status = "exists"

    # Create Given subfolders inside the main folder (always ensure they exist)
    subfolders = ["Removed Rows", "Success and error files" , "CSV Files"]
    
    for subfolder in subfolders:
        SubFolder_Path = os.path.join(Final_File_Folder_Path, subfolder)
        if not os.path.exists(SubFolder_Path):
            os.makedirs(SubFolder_Path, exist_ok=True)

    Final_File_Folders_created.append((Final_File_Folder_Name, folder_status))

# Display folder creation results
print("\n    🗂️ Folder Created:")
for index, (Final_File_Folder_Name, status) in enumerate(Final_File_Folders_created, start=1):
    if status == "created":
        print(f"\n        {index}) {Final_File_Folder_Name} ✅ (Main folder created)")
    elif status == "exists":
        print(f"\n        {index}) {Final_File_Folder_Name} - Main folder already exists 🛑 (Subfolders ensured)")

# Check if the "Copy files" folder exists before proceeding
if not os.path.exists(copy_file_path):
    print("\n     ❗️ 'Copy files' folder does not exist. ")
    sys.exit()

# Retrieve the list of files present in the "Copy files" folder

files_in_copy_folder = [] # Initialize an empty List to Save File in Copy Folder

# Iterate over all items in the directory
for f in os.listdir(copy_file_path):
    file_path = os.path.join(copy_file_path, f)
    
    # Check if the item is a file
    if os.path.isfile(file_path):
        files_in_copy_folder.append(f)

if not files_in_copy_folder:
    print("\n     🚫 No files found in 'Copy files' folder. ")
    sys.exit()

count_of_files_copy_folder = len(files_in_copy_folder)

# =========================================================================================================================================
#                                                OPPORTUNITY SHEET EXECUTION
# =========================================================================================================================================


# ======================================================================
# Determine Which File to Process and Set the File Path Accordingly
# ======================================================================

while True:
    
    # Display available files for selection

    print("\n====================================================================================================")
    print("\n📂 Please select a file to process:")
    print("\n    🔸 List of Files in Copy Folder: ")

    for idx, file_name in enumerate(files_in_copy_folder, start=1):
        print(f"\n        📄 {idx}. {file_name}")

    # Prompt the user to select a file from the list
    while True:
        
        user_input = input("\n    👉 Enter the number of the file to process (or type 'exit' to quit): ").strip()
        
        # Allow the user to exit the selection process
        if user_input.lower() == 'exit':
            print("\n        ❌ File selection has been canceled. Exiting process. \n")
            sys.exit()

        try:
            # Convert user input to an index and validate selection
            selected_index = int(user_input) - 1
            
            if 0 <= selected_index < len(files_in_copy_folder):
                file_path = os.path.join(copy_file_path, files_in_copy_folder[selected_index])
                print(f"\n        ✅ You selected the file: {files_in_copy_folder[selected_index]} \n")
                break # Exit the loop if a valid file is selected

            else:
                print(f"\n        ❗ Invalid selection. Please select a number between 1 and {len(files_in_copy_folder)}.")
        
        except ValueError:
            print("\n        ❗ Invalid input. Please enter a valid number or type 'exit' to cancel.")

    # ================================================================================
    # Code to Construct Path to Folder of current Running file in "Final Iteration file" Folder
    # ================================================================================
    file_start_time = time.time()   # Record start time
    # Extract the selected file name from the full file path
    selected_file_name = os.path.basename(file_path.split("/")[-1])

    # Remove '_Copy' from the file name (if present) and remove the file extension
    folder_file_name = os.path.splitext(re.sub(r'_Copy', '', selected_file_name))[0]

    # Define the output folder path where processed data will be stored
    output = os.path.join(final_iteration_file_path, folder_file_name)

    # Define paths for subdirectories where different types of processed data will be saved
    csv_file_dir = os.path.join(output, "CSV Files") # Folder for storing CSV Files
    removed_rows_dir = os.path.join(output, "Removed Rows") # Folder for storing Removed Rows Files

    symbol = "="
    print(symbol*100)
    # # ================================================================================
    # # Check for missing required sheets and rename if necessary
    # # ================================================================================

    # print("\n\n🔍 Check if all the Required Sheets are present or not")
 
    # # Load the Excel workbook 
    # '''
    #     Info:
    #     We use openpyxl here instead of pandas because:
    #     - openpyxl gives access to sheet names, cell formatting, and workbook structure.
    #     - pandas is mainly for working with data tables (DataFrames), not the workbook structure.
    # '''
    # wb = openpyxl.load_workbook(file_path)
    # # 📌 [NEW] Auto-Rename Known Variants to Correct Names

    # # Define the list of required sheet names
    # # 'Tags' is considered optional and will not be treated as missing if absent
    # required_sheets = ['Opportunity', 'Opportunity_product','Opportunity_Team ', 'Reporting_codes', 'Tags']
    
    # variant_mapping = {
    #     'Opportunity_products': 'Opportunity_product',
    #     'Opportunity_Team': 'Opportunity_Team '  # note the trailing space
    # }

    # for sheet_name in wb.sheetnames:
    #     if sheet_name in variant_mapping:
    #         ws = wb[sheet_name]
    #         correct_name = variant_mapping[sheet_name]
    #         ws.title = correct_name
    #         print(f"\n    ✅  Renamed '{sheet_name}' to '{correct_name}' automatically.")

    # # Get the list of sheet names present in the current workbook
    # sheets_in_file = wb.sheetnames

    # # Identify missing required sheets (excluding 'Tags' which is optional)
    # missing_sheets = [] # Initialize Missing Sheet List

    # for sheet in required_sheets:
    #     # Check if the sheet is not 'Tags' and is not in the list of sheets in the file
    #     if sheet != 'Tags' and sheet not in sheets_in_file:
    #         missing_sheets.append(sheet)

    # # Check if all required sheets are present
    # if not missing_sheets:
    #     # All required sheets are present — no further action needed
    #     print("\n    ✅ All required sheets are already present! 🎉")

    # else:
    #     print("\n    ❌ The following required sheets are missing: ")

    #     # Print the missing sheets
    #     for i, sheet in enumerate(missing_sheets, 1):
    #         print(f"\n        {i}. {sheet}")

    #     # Identify extra sheets in the workbook that are NOT in the required list
    #     available_sheets = [] # Initialize Available Sheet List

    #     for s in sheets_in_file:
    #         # Check if the sheet is not in the list of required sheets
    #         if s not in required_sheets:
    #             available_sheets.append(s)

    #     # If there are available sheets, prompt user to rename them
    #     if available_sheets:
    #         print("\n    📋 Here are the available sheets to rename: ")
    #         # Display the available sheets as a numbered list
    #         for i, s in enumerate(available_sheets, 1):
    #             print(f"\n        {i}. {s}")
        
    #     used_indices = []  # keep track of already used sheet indices

    #     # Loop through each missing sheet and ask the user if they want to rename one of the available sheets
    #     for sheet in missing_sheets:
    #         if len(used_indices) == len(available_sheets):
    #             print(f"\n    ⏭️  No sheets available to rename. Automatically skipping '{sheet}'!")
    #             continue

    #         while True:
    #             choice = input(f"\n    🔸 Enter the index of the sheet to rename to '{sheet}' or type 'skip': ")

    #             if choice.lower() == 'skip':
    #                 print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
    #                 break

    #             try:
    #                 choice = int(choice)
    #                 if 1 <= choice <= len(available_sheets):
    #                     if choice in used_indices:
    #                         print("\n        ❗ That sheet has already been used. Choose a different one.")
    #                         continue

    #                     rename_sheet = available_sheets[choice - 1]
    #                     ws = wb[rename_sheet]
    #                     ws.title = sheet

    #                     print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")

    #                     used_indices.append(choice)
    #                     break

    #                 else:
    #                     print("\n        ❗ Invalid number selected. Please choose a valid option.")
    #             except ValueError:
    #                 print("\n        ❗ Invalid input, please enter a valid number or 'skip'.")
    #     # Save the modified workbook (if any renaming was done)
    #     wb.save(file_path)
    #     print("\n    💾 Workbook saved with changes!")
    
    
    # # ==========================================
    # # To Handle Sheets with Similar Names but Different Casing
    # # ==========================================

    # # Create a mapping of sheet names that may have different cases to their standard names
    # sheet_name_mapping = {
    #     'Opportunity1': 'Opportunity',
    #     'Opportunity_product1': 'Opportunity_product',
    #     'Opportunity_team1': 'Opportunity_team',
    #     'Reporting_codes1': 'Reporting_codes'
    # }

    # # Load the Excel file
    # wb = openpyxl.load_workbook(file_path)

    # # Iterate through all sheets in the workbook
    # for sheet in wb.sheetnames:

    #     # Check if the sheet's name is present in the mapping dictionary
    #     if sheet in sheet_name_mapping:

    #         # If a match is found, get the corresponding new name
    #         new_name = sheet_name_mapping[sheet]
            
    #         # Access the worksheet with the current name
    #         ws = wb[sheet]

    #         # Rename the sheet to the mapped new name
    #         ws.title = new_name

    # # Save the workbook with the renamed sheets (the content will remain unchanged)
    # wb.save(file_path)
    
    # ==========================================
    # Convert all column headers to lowercase
    # ==========================================
    
    print("\n\n🔍 Converting all columns headers to lowercase...")

    # Read the Excel file with all sheets, initially treating all data as strings
    xls = pd.ExcelFile(file_path)

    # Create a dictionary to store updated DataFrames for each sheet
    sheets_dict = {}

    # List of column names that should be kept as numeric (e.g., for calculations)
    numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

    # Iterate through each sheet
    for sheet_name in xls.sheet_names:
        # Read each sheet into a dataframe with all columns as strings
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        
        # Convert all column headers to lowercase for consistency
        df.columns = [col.lower() for col in df.columns]
        
        # Convert specified columns back to numeric (e.g., prices, amounts, terms)
        for col in numeric_columns:
            if col in df.columns:
                # Store original values before conversion
                original_values = df[col].copy()
                
                # Convert the column to numeric, replacing errors with NaN
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Identify and display values that failed to convert
                invalid_mask = df[col].isna() & original_values.notna() & (original_values != "")
                if invalid_mask.any():
                    print(f"\n    ❗️ Warning: The following values in column '{col}' could not be converted to numeric and were set to NaN:")
                    for i, val in original_values[invalid_mask].items():
                        print(f"\n       🔸 Row {i + 2}: '{val}'")  # +2 to adjust for header and 0-based index
                
        # Save modified dataframe to dictionary
        sheets_dict[sheet_name] = df

    # Write the modified dataframes back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\n    ✅ Headers of all sheets in the file have been converted to lowercase and data types preserved as specified.")

    # ======================================================================
    # Print Opportunity Script Execution 📝                               
    # ======================================================================
    
    # Display the title for the Opportunity Sheet Execution
    title = "📝 OPPORTUNITY SHEET EXECUTION 📝"
    show_title(title)
    
    # ======================================================================
    # Step 1: File Existence Check
    # ======================================================================

    print("\n\n🔍 Step 1: Checking if the file exists...")
    
    def check_file_exists(file_path):
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            print(f"\n    ✅ File '{filename}' exists at the specified path.")
            return (filename)
        else:
            print("\n    ❌ Error: File does not exist or the path is invalid.\n")
            sys.exit()  # Exit the program if file is not found

    # Example usage:
    filename = check_file_exists(file_path)
    
    # ======================================================================
    #  Extract the Mpp_Number__c column and save it to a new Excel file
    # ======================================================================

    opportunity_df = pd.read_excel(file_path, sheet_name="Opportunity", dtype={"Mpp_Number__c": str})

    if 'Mpp_Number__c' in opportunity_df.columns:
        # Extract the two required columns
        columns_to_extract = ["opportunity_legacy_id_c", "Mpp_Number__c"]
        mpp_df = opportunity_df[columns_to_extract]

        # Save to a new Excel file, ensuring Mpp_Number__c remains text
        mpp_output_file = os.path.join(csv_file_dir,"MPP_Column.xlsx")
        with pd.ExcelWriter(mpp_output_file, engine='openpyxl') as writer:
            mpp_df.to_excel(writer, index=False)

    # ======================================================================
    # Step 2: Removing Duplicates and Blank Rows
    # ======================================================================

    print("\n\n🔍 Step 2: Removing duplicate rows and blank rows...")
    
    # Define Opportunity Sheet Name
    opportunity_sheet_name = "Opportunity"

    def clean_sheet(df, sheet_name, remove_duplicates=False):
        try:
            # Remove duplicate rows if requested
            if remove_duplicates:
                df = df.drop_duplicates()

            # Remove rows where all cells are NaN (blank rows)
            df = df.dropna(how='all')

            # Specific check for Opportunity sheet
            if sheet_name == "Opportunity":
                df = df.dropna(subset=['opportunity_legacy_id_c'], how='all')

            # Trim whitespace from all string columns
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

            if remove_duplicates:
                print(f"\n    ✅ Removed duplicate and blank rows from '{sheet_name}' sheet. ")
            else:
                print(f"\n    ✅ Removed blank rows from '{sheet_name}' sheet. ")

        except Exception as e:
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
            sys.exit()
        return df

    opportunity_df = clean_sheet(opportunity_df, opportunity_sheet_name, remove_duplicates=True)


    # ======================================================================
    # Function to check for duplicates in the 'opportunity_legacy_id_c' column of the 'Opportunity' sheet
    # ======================================================================

    # Check if 'opportunity_legacy_id_c' column exists
    if 'opportunity_legacy_id_c' not in opportunity_df.columns:
        print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found in the sheet.")
        sys.exit(1)

    # Check for duplicate values
    if opportunity_df['opportunity_legacy_id_c'].duplicated().any():
        print("\n    ❌ Error: Duplicate values still Present in 'opportunity_legacy_id_c' column.")
        sys.exit(1)

    # ======================================================================
    # Step 3: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 3: Counting Initial rows in Opportunity Sheet...")

    def count_row(df):
        # Get the number of rows and columns
        num_rows = df.shape[0]     # Number of rows in the DataFrame

        return num_rows

    oppty_initial_num_rows = count_row(opportunity_df)
    
    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {oppty_initial_num_rows}")

    # =============================
    # Step 4:- Check for Already Existing Opportunities in ISC
    # =============================
    
    print("\n\n🔍 Step 4: Looking for Already existing Oppties")

    # Set paths
    legacy_csv = os.path.expanduser("~/Downloads/legacyid.csv") 

    # while not os.path.exists(legacy_csv):

    #     print(f"\n    ❌ File 'legacyid.csv' does not exist. Did you query the Legacy Id?")

    #     legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

    #     while legacy_choice not in ['yes', 'exit']:
    #         print("\n          ❗️ Invalid input. Please enter 'yes' or 'exit'.")
    #         legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

    #     if legacy_choice != 'yes':
    #         print ("\n           🚫 Skipping this Step")
    #         break
    
    if not os.path.exists(legacy_csv):
        
        print(f"\n    ❌ File 'legacyid.csv' does not exist. Did you query the Legacy Id?")
    
    if os.path.exists(legacy_csv):
        # Read CSV file
        csv_df = pd.read_csv(legacy_csv)
        csv_ids = set(csv_df['Opportunity_Legacy_Id__c'].dropna().astype(str))

        try:
            
            if 'opportunity_legacy_id_c' in opportunity_df.columns:
                # Ensure the column is string for safe comparison
                opportunity_df['opportunity_legacy_id_c'] = opportunity_df['opportunity_legacy_id_c'].astype(str)
                
                # Add Found/Not Found column
                opportunity_df['Already Exist'] = opportunity_df['opportunity_legacy_id_c'].apply(
                    lambda x: 'Already Exist in ISC' if x in csv_ids else 'Does not Exist in ISC'
                )

                count_not_exist = (opportunity_df['Already Exist'] == 'Already Exist in ISC').sum()

                print(f"\n    ✅ Added 'Already Exist' column ")
                if count_not_exist > 0:
                    print(f"\n    ❗️ Count of Already Exsisting Opportunities in ISC' : {count_not_exist}")
                else:
                    print(f"\n    ✅ All Opptys are new")
                
            else:
                print(f"\n    ❌ 'opportunity_legacy_id_c' not found ")
        except Exception as e:
            print(f"\n    ❌ Error processing : {e}")
    
    
    # ======================================================================
    # Step 5: Convert the email ids to lowercase and fill missing values with a default value
    # ======================================================================

    print("\n\n🔍 Step 5: Converting email ids to lowercase and filling missing values...")

    # Columns to convert to lowercase and fill blanks
    columns_to_process = ['ownerid', 'created_by']
    
    # Default value to fill in case of missing email IDs
    data_migration_email = "iscdmig2@in.ibm.com"

    
    # Initialize a dictionary to track how many blanks are filled per column
    filled_counts = {}
    for col in columns_to_process:
        filled_counts[col] = 0
    
    # Fill blank cells with specified value and count filled blanks
    for column in columns_to_process:
        if column in opportunity_df.columns:
            # Count missing values before filling
            blank_count = opportunity_df[column].isnull().sum()
            filled_counts[column] = blank_count
            
            # Fill missing values with default and convert all to lowercase
            opportunity_df[column] = opportunity_df[column].fillna(data_migration_email)
            opportunity_df[column] = opportunity_df[column].apply(lambda x: x.lower() if isinstance(x, str) else x)

        else:
            # If the column doesn't exist, print an error and exit
            print(f"\n    ❌ Error: Column '{column}' not found in the '{opportunity_sheet_name}' sheet. Terminating the Program.")
            sys.exit()
    

    # Display the count of blank columns filled for each column
    for col, count in filled_counts.items():
        if count > 0:
            print(f"\n    ❗️ Blank Values filled with Data migration Id in {col} column: {count}")
        else:
            print(f"\n    ✅ All Valid Email ids in {col} column")

    # ======================================================================
    # Step 6: Add Pricebook and RecordType id column in the sheet
    # ======================================================================

    print("\n\n🔍 Step 6: Adding Pricebook and RecordType id columns...")

    # Add the two new columns with the specified values
    opportunity_df['Pricebook2Id'] = '01s3h000003KXvoAAG'
    opportunity_df['RecordTypeId'] = '0123h000000kppcAAA'

    print("\n    ✅ 'Price_Book' and 'Record_Type' Columns added successfully.")

    # ======================================================================
    # Step 7: Change the format of the Date Column
    # ======================================================================

    print("\n\n🔍 Step 7: Formatting the Date column...")

    date_column = 'expected_close_date'

    try:
        # Check if the specified column exists
        if date_column not in opportunity_df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing in the sheet '{opportunity_sheet_name}'.")
            sys.exit(1)  # Exit the script with an error code

        # Ensure the date column is in datetime format, allowing for errors to be coerced to NaT
        opportunity_df[date_column] = pd.to_datetime(opportunity_df[date_column], errors='coerce')

        # Check for blank values (NaT) after processing
        if opportunity_df[date_column].isnull().any():
            print(f"\n    ❌ Error: The column '{date_column}' contains blank or invalid values after processing.")
            print("\n    ❗️ Please review the data, correct the issues, and try again.")
            sys.exit(1)  # Exit the script with an error code

        # Format the valid dates to YYYY-MM-DD
        opportunity_df[date_column] = opportunity_df[date_column].dt.strftime('%Y-%m-%d')

        print("\n    ✅ Date column formatted to YYYY-MM-DD successfully.")
    
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")


    # ======================================================================
    # Step 8: Create new "legacy_opportunity_split_id_c" column if it does not exist
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'legacy_opportunity_split_id_c' column...")

    # Check if the column "legacy_opportunity_split_id_c" already exists (case-insensitive)
    existing_columns = [col.lower() for col in opportunity_df.columns]
    if "legacy_opportunity_split_id_c" in existing_columns:
        print("\n    ✅ 'legacy_opportunity_split_id_c' column already exists in the sheet.")
        
        # Check for blank (NaN) values in the 'legacy_opportunity_split_id_c' column
        if opportunity_df['legacy_opportunity_split_id_c'].isnull().any():
            print("\n    ❌ Error: 'legacy_opportunity_split_id_c' column contains blank (NaN) values. Please review. Exiting process.")
            sys.exit()  # Exit the code if blank values are found
    else:
        # Check if "opportunity_legacy_id_c" column exists
        if "opportunity_legacy_id_c" not in existing_columns:
            print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found. Exiting process.")
            sys.exit()  # Exit if "opportunity_legacy_id_c" is not found
        
        # Create the new column 'legacy_opportunity_split_id_c' and populate it with 'opportunity_legacy_id_c' values
        opportunity_df['legacy_opportunity_split_id_c'] = opportunity_df['opportunity_legacy_id_c']

        print('\n    ✅ New legacy_opportunity_split_id_c column added to sheet. Process completed successfully.')

    # ======================================================================
    # Step 9: Create new column with Trimmed Account_id and Email_id column
    # ======================================================================

    print('\n\n🔍 Step 9: Creating new column with Trimmed Account_id and Email_id...\n')

    columns_to_trim = ['accountid', 'ownerid']  

    # Check if specified columns exist in the DataFrame
    missing_columns = [col for col in columns_to_trim if col not in opportunity_df.columns]

    if missing_columns:
        print(f"    ❌ Error: The following columns were not found in the sheet '{opportunity_sheet_name}': {', '.join(missing_columns)}")
        sys.exit(1)

    # Trim the values for whitespaces and create new columns for the trimmed values
    for column in columns_to_trim:
        # If the column is 'accountid', remove internal spaces in addition to trimming
        if column == 'accountid':
            opportunity_df[column] = opportunity_df[column].str.replace(r'\s+', '', regex=True).str.strip()
        else:
            opportunity_df[column] = opportunity_df[column].str.strip()

    print("    ✅ Account and Email Columns trimmed successfully, and internal spaces in 'accountid' removed.")


    # ======================================================================
    # Step 10: Remove the country code from DC Accounts
    #   • For columns with both DC and DB accounts, remove country codes from DC accounts, as they are invalid.
    #   • And keep The DB values as it is
    # ======================================================================

    print("\n\n🔍 Step 10: Processing Accounts with correct format...\n")

    accountid_column = 'accountid'
    new_column_name = 'AccountNumber'  

    # Define a function to process the values
    def process_value(value):
        if isinstance(value, str) and value.startswith('DC'):
            return value.split('-')[0]
        return value

    # Apply the function to the accountid column and store results in the new column
    opportunity_df[new_column_name] = opportunity_df[accountid_column].apply(process_value)

    print("    ✅ New column with formatted DC values created and added to the sheet successfully.")


    # ==========================================================================================
    # Step 11: Copy Extracted Data to the Main Excel File
    # ==========================================================================================


    print("\n\n🔍 Step 11: Checking for Extract Files...")

    # Define expected CSV file paths
    accounts_csv = DOWNLOADS_DIR+"/accounts.csv"  
    userid_csv = DOWNLOADS_DIR+"/userid.csv" 
    
    # Check if the CSV files exist, and prompt to retry if not

    while not os.path.exists(accounts_csv):

        print(f"\n    ❌ File 'accounts.csv' does not exist. Did you query the accounts?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            break
    if os.path.exists(accounts_csv):
        print("\n    ✅ Account CSV files is present.")

    # Check if the CSV files exist, and prompt to retry if not
    while not os.path.exists(userid_csv):
    
        print(f"\n    ❌ File 'userid.csv' does not exist. Did you query the Userid?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n         ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()
    if os.path.exists(userid_csv):
        print("\n    ✅ Userid CSV files is present.")

    # ======================================================================
    # Step 12: Check how many Accounts are present in ISC
    # ======================================================================

    print("\n\n🔍 Step 12: Checking how many Accounts are present in ISC...")


    try:
        
        # Load data from the account Extract sheets
        accounts_df = pd.read_csv(accounts_csv, usecols=[0, 1])  # Read first two columns

        # Remove rows from accounts_df where 'Id' is missing
        account_df_no_nan = accounts_df.dropna(subset=['Id'])
        
        # Identify AccountNumbers that appear more than once with different Id values
        duplicate_accounts = account_df_no_nan[
            account_df_no_nan.duplicated(subset=['AccountNumber'], keep=False)
        ]
        
        if not duplicate_accounts.empty:
            print("\n    ❗️ Duplicate AccountNumbers found with multiple Id values:")
            
            # Loop through each group of duplicates by AccountNumber
            for account_number, group in duplicate_accounts.groupby('AccountNumber'):
                print(f"\n        🔹 AccountNumber: {account_number}")
                
                # Display Ids and their corresponding row numbers in Excel
                for idx, row in group.iterrows():
                    excel_row_number = idx + 2  # Adjust for Excel row numbering
                    print(f"\n           🔸 Id: {row['Id']} (Excel Row {excel_row_number})")
                
                # Ask user to select the correct Id to retain for this AccountNumber
                valid_ids = group['Id'].tolist()
                while True:
                    chosen_id = input(f"\n        🔹 Select id for AccountNumber {account_number} from above Ids: ").strip()
                    if chosen_id in valid_ids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id from {valid_ids}. ")
                
                # Keep only the row with the chosen Id for the current AccountNumber
                accounts_df = accounts_df[
                    ~((accounts_df['AccountNumber'] == account_number) & 
                    (accounts_df['Id'] != chosen_id))
                ]
        
        # Merge the original opportunity_df with the filtered Ids from accounts_df
        merged_df = pd.merge(opportunity_df, accounts_df[['AccountNumber', 'Id']],
                            on='AccountNumber', how='left')
        
        # Count how many AccountNumbers are missing (i.e., not found in ISC)
        not_in_isc_count = merged_df["Id"].isna().sum()

        # Replace missing Ids with a placeholder text
        merged_df['Id'] = merged_df['Id'].fillna('Not in ISC')
        
        # Optionally, you could replace with AccountNumber instead of 'Not in ISC' using combine_first
        # merged_df['Id'] = merged_df['Id'].combine_first(opportunity_df['AccountNumber'])
        
        # Rename the 'Id' column to indicate ISC status
        merged_df.rename(columns={'Id': 'In ISC or Not'}, inplace=True)
        opportunity_df = merged_df
        
        # Display summary of missing accounts
        if not_in_isc_count > 0:
            print(f"\n    ❗️ Count of accounts Not in ISC: {not_in_isc_count}")
        else:
            print(f"\n    ✅ All Accounts are Present")
    
    except FileNotFoundError:
        print("\n    ❌ Error: The specified file was not found. Please check the file path.")
    except ValueError as e:
        print(f"\n    ❌ Error: {e}")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
    
    # ======================================================================
    # Step 13: Get the IDs of the Opportunity Owner
    # ======================================================================

    print("\n\n🔍 Step 13: Fetching IDs of Opportunity Owners...")
    DEFAULT_USERID = '0053h000000sdCVAAY'

    try:

        userid_df = pd.read_csv(userid_csv)  # Read first two columns
        
        # Clean and normalize the 'ownerid' column in opportunity_df
        if 'ownerid' in opportunity_df.columns:
            opportunity_df['ownerid'] = opportunity_df['ownerid'].str.strip().str.lower()
            # print("\n    ✅ 'ownerid' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'ownerid' not found in the Opportunity sheet.")
            sys.exit()

        # Clean and normalize the 'Email' column in userid_df
        if 'Email' in userid_df.columns:
            userid_df['Email'] = userid_df['Email'].str.strip().str.lower()
            # print("\n    ✅ 'Email' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Email' not found in the Opportunity_Copy sheet.")
            sys.exit()

        # Remove rows with missing 'userid' and detect duplicate Email entries with different 'userid' values
        ownerid_copy_df = userid_df.dropna(subset=['Id'])
        duplicate_emails = ownerid_copy_df[
            ownerid_copy_df.duplicated(subset=['Email'], keep=False)
        ]

        # If duplicate emails are found, prompt the user to resolve them
        if not duplicate_emails.empty:
            print("\n    ❗️ Duplicate Email IDs with multiple UserIDs found:")
            for email, group in duplicate_emails.groupby(['Email']):
                print(f"\n        📧 Email: {email}")

                # Show all UserIDs associated with the duplicated email
                for idx, row in group.iterrows():
                    excel_row = idx + 2  # Adjust for Excel row indexing (1-based + header)
                    print(f"\n           🔸 UserID: {row['Id']} (Row {excel_row})")
                
                # Collect valid UserIDs for this Email
                valid_userids = group['Id'].tolist()
                
                # Ask the user to select the correct UserID for the current email
                while True:
                    chosen_userid = input(f"\n        🔹 Select Id for UserId '{email}' from above Ids: ").strip()
                    if chosen_userid in valid_userids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id . ")
                
                # Keep only the row with the selected UserID for the current email
                userid_df = userid_df[
                    ~((userid_df['Email'] == email) & (userid_df['Id'] != chosen_userid))
                ]
            print("\n    ✅ Duplicate emails handled successfully.")

        # Perform a left join to map the 'ownerid' from opportunity_df to 'userid' in userid_df
        result_df = pd.merge(
            opportunity_df,
            userid_df[['Email', 'Id']],
            left_on='ownerid',
            right_on='Email',
            how='left'
        )
        
        # Count how many 'userid' entries are missing (NaN) before filling them
        nan_before = result_df['Id'].isna().sum()
        
        # Fill missing userids with a default fallback value
        result_df['Id'] = result_df['Id'].fillna(DEFAULT_USERID)

        # Remove redundant 'Email' column and rename 'userid' to 'OwnerId'
        result_df.drop(columns=['Email'], inplace=True)
        result_df.rename(columns={'Id': 'OwnerId'}, inplace=True)
        opportunity_df = result_df

        print("\n    ✅ Success: IDs for Opportunity Owners updated successfully.")
        
        # Notify user if any invalid userids were replaced
        if nan_before > 0:
            print(f"\n    ❗️ Number of invalid 'ownerid' values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")

    # ======================================================================
    # Step 14: To get IDs of the Created By
    # ======================================================================

    print("\n\n🔍 Step 14: Fetching IDs of 'Created By'...")

    try:

        # Check if 'created_by' column exists and contains any non-empty values
        if 'created_by' not in opportunity_df.columns or opportunity_df['created_by'].dropna().empty:
            if 'created_by' not in opportunity_df.columns:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column does not exist in 'Opportunity' sheet.")
            elif opportunity_df['created_by'].dropna().empty:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column is empty in 'Opportunity' sheet.")
        else:

            # Perform a left join (VLOOKUP-like) to map 'created_by' email to 'Id'
            merged_df = pd.merge(
                opportunity_df,
                userid_df[['Email', 'Id']],
                left_on='created_by',
                right_on='Email',
                how='left'
            )

            # Rename 'Id' to 'createdbyid' for clarity
            merged_df.rename(columns={'Id': 'createdbyid'}, inplace=True)

            # Count how many 'createdbyid' values are missing before filling them
            nan_before = merged_df['createdbyid'].isna().sum()

            # Replace missing IDs with the default fallback user ID
            merged_df['createdbyid'] = merged_df['createdbyid'].fillna(DEFAULT_USERID)

            opportunity_df = merged_df
            print("\n    ✅ Success: IDs for  'Created By' updated successfully.")
            
            # Save the updated dataframe back to the original Opportunity sheet
            if nan_before > 0:
                print(f"\n    ❗️ Number of invalid 'createdbyid' values replaced with Data Migration Id: {nan_before}")
        

    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")

    # ======================================================================
    # Step 15: Renaming Columns
    # ======================================================================


    print("\n\n🔍 Step 15: Renaming Columns...")



    def safe_rename_columns(df, column_rename_mapping):
        try:
            # Find missing columns
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]

            if missing_columns:
                print(f"\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n        🔸 {col}")
                
                while True:
                    proceed = input("\n    🔹 Do you want to proceed with renaming the available columns? (yes/no): ").strip().lower()
                    if proceed == 'yes':
                        break
                    elif proceed == 'no':
                        print("\n        ❌ Operation aborted.")
                        sys.exit(1)
                    else:
                        print("\n        ❗️ Invalid choice. Please enter 'yes' or 'no'.")

            # Rename columns
            df.rename(columns=column_rename_mapping, inplace=True)
            print(f"\n    ✅ Columns renamed successfully.")

        except Exception as e:
            print(f"\n    ❌ An unexpected error occurred while renaming columns in {e}")
            sys.exit(1)
        return df

    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunity_legacy_id_c': 'opportunity_legacy_id__c',
        'legacy_opportunity_split_id_c': 'Legacy_Opportunity_Split_Id__c',
        'In ISC or Not': 'AccountId',
        'sales_stage': 'StageName',
        'won reason': 'Won_Reason__c',
        'lost category': 'Lost_Category__c',
        'lost reason': 'Lost_Reason__c',
        'currency_code': 'CurrencyIsoCode',
        'next_step': 'NextStep',
        'oi_source': 'OI_Group__c',
        'expected_close_date': 'CloseDate',
        'ownerid': 'Email',
    }

    # Call Function to Rename the columns
    opportunity_df = safe_rename_columns(opportunity_df, column_rename_mapping)

    # ======================================================================
    # Step 16: Rearrange the Columns in the Opportunity Copy
    #   • Rearrange columns to prioritize important fields, grouping related ones (e.g., account number and account ID) and moving less important ones to the end.
    # ======================================================================


    print("\n\n🔍 Step 16: Rearranging Columns...")

    def rearrange_and_save_columns(df, desired_column_order, sheet_name):

        try:
            # Check for missing and extra columns
            missing_columns = [col for col in desired_column_order if col not in df.columns]
            extra_columns = [col for col in df.columns if col not in desired_column_order]

            # Rearrange columns
            rearranged_columns = [col for col in desired_column_order if col in df.columns]
            rearranged_columns += extra_columns  # Add extra columns to the end

            if missing_columns:
                print("\n    ❌ The following columns are missing and were skipped:")
                for col in missing_columns:
                    print(f"\n        🔸  {col}")

            if sheet_name.lower() == 'opportunity':
                if extra_columns:
                    print("\n    🔷 The following extra columns were moved to the end:")
                    for col in extra_columns:
                        print(f"\n        🔸  {col}")

            df = df[rearranged_columns]

            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"\n    ✅ Columns successfully rearranged in the '{sheet_name}' sheet of the file: {file_path.split('/')[-1]}")

        except FileNotFoundError:
            print(f"\n    ❌ File '{file_path}' not found.")
            sys.exit(1)

        return df  # Return in case you want to use it afterward


    desired_column_order = [
        'opportunity_legacy_id__c','Legacy_Opportunity_Split_Id__c','name','AccountNumber','AccountId',
    'StageName','Won_Reason__c','Lost_Category__c','Lost_Reason__c','CloseDate','CurrencyIsoCode','Email','OwnerId',
    'NextStep','OI_Group__c','created_by','createdbyid','Pricebook2Id','RecordTypeId','modified_by','created_date',
    'modified_date','pricebook','recordtypeid','accountid'
    ]
    
    # Call Function to Rearrange Columns in a sheet
    opportunity_df =rearrange_and_save_columns(opportunity_df, desired_column_order, opportunity_sheet_name)

    # ======================================================================
    # Step 17: Final Row and Column Count
    # ======================================================================


    print("\n\n🔍 Step 17: Counting Final Row in Opportunity Sheet...")

    oppty_final_num_rows = count_row(opportunity_df)

    print(f"\n    ✅ Final row count: {oppty_final_num_rows}")

    # Code to Check if the row count has changed
    def check_row_count(initial_num_rows, final_num_rows,label = ""):
        """
        Check if the row count of a DataFrame has changed.
        """
        if initial_num_rows != final_num_rows:
            print(f"\n    ❗️ Row count mismatch detected{f' in {label}' if label else ''}!")
            print(f"\n       📊 Initial: {initial_num_rows}")
            print(f"\n       📊 Final: {final_num_rows}")

            while True:
                user_input = input("\n    🔹 Do you want to continue? Type 'continue' to proceed or 'no' to abort: ").strip().lower()
                if user_input == "continue":
                    print("\n       ✅ Continuing the program...")
                    break
                elif user_input == "no":
                    print("\n       ❌ Terminating the program...")
                    sys.exit(1)
                else:
                    print("\n       ❌ Invalid input. Please type 'continue' to proceed or 'no' to stop.")

    check_row_count(oppty_initial_num_rows, oppty_final_num_rows, label = "Opportunity")
    # =========================================================================================================================================
    #                                                PRODUCT SHEET EXECUTION
    # =========================================================================================================================================


    title = "📝 PRODUCT SHEET EXECUTION 📝"
    show_title(title)

    product_sheet_name = 'Opportunity_product'

    product_df = pd.read_excel(file_path, sheet_name=product_sheet_name)
    # ======================================================================
    # Step 1: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 1: Counting Initial rows in Product Sheet...")

    # Function to count rows in a specific sheet
    product_initial_num_rows = count_row(product_df)

    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {product_initial_num_rows}")
    
    # ======================================================================
    # Step 2:- Removing duplicate rows and blank rows...
    # ======================================================================

    print("\n\n🔍 Step 2: Removing blank rows...")

    product_df = clean_sheet(product_df, product_sheet_name, remove_duplicates=False)


    # ======================================================================
    # Step 3 :- Add Exsising column, To check if the given Opportunities are present in the Opportunity Sheet 
    # ======================================================================

    print("\n\n🔍 Step 3: Verifying opportunities in the 'Opportunity' sheet...")
    
    def verify_opportunity(df,sheet_name):
        try:
            # opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

            # Validate the required columns
            if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet. ")
                sys.exit()
            elif 'opportunityid' not in df.columns:
                print(f"\n    ❌ Column 'opportunityid' not found in the '{sheet_name}' sheet. ")
                sys.exit()

            # Perform the comparison
            df['existing'] = df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

            # Calculate the number of false values
            false_count = (~df['existing']).sum()
        
            if false_count > 0:
                print(f"\n    ❗️ Number of False values in 'existing' column: {false_count}")
            else:
                print(f"\n    ✅ All Opportunities Exist In Opportunity Sheet")
        
        except Exception as e:
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e}")
            sys.exit()
        return df

    # Call the function to verify opportunities
    product_df = verify_opportunity(product_df,product_sheet_name) 

    

    # ======================================================================
    # Step 4: Formatting the date column
    # ======================================================================

    print("\n\n🔍 Step 4: Formatting the date column in the 'Opportunity_product' sheet...")

    date_column = 'expiration date'  # The column containing the dates to be formatted

    try:
        # Check if the specified column exists in the DataFrame
        if date_column not in product_df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing from the sheet '{product_sheet_name}'. ")
            sys.exit(1)   # Exit the script if the required column is not found

        # Convert the values in the date column to datetime format and standardize to 'YYYY-MM-DD'
        product_df[date_column] = pd.to_datetime(product_df[date_column]).dt.strftime('%Y-%m-%d')

        # Success message
        print(f"\n    ✅ Date column formatted successfully in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()

    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 5: Adding Quantity Columns
    # ======================================================================

    print("\n\n🔍 Step 5: Adding the 'Quantity' column in the 'Opportunity_product' sheet...")
    
    # Attempt to delete existing 'quantity' column if it already exists
    if 'quantity' in product_df.columns:
        product_df = product_df.drop(columns=['quantity'])

    # Now add a fresh 'Quantity' column to the sheet
    new_column_name = 'Quantity'  # Column name to be added
    default_value = 1  # Default value for the new column

    # Add the new column with the default value
    product_df[new_column_name] = default_value
    
    # Success message
    print(f"\n    ✅ A new column '{new_column_name}' has been added to the '{product_sheet_name}' sheet with default value '{default_value}'.")



    # ======================================================================
    # Step 6: To create a new Currency column
    # ======================================================================
    
    print("\n\n🔍 Step 6: Creating or Overwriting the 'opportunity currency' column in the 'Opportunity_product' sheet...")

    try:
        # Read both relevant sheets from the Excel file into pandas DataFrames
        # opportunity_df = pd.read_excel(file_path, sheet_name = opportunity_sheet_name)

        # Create a mapping from 'opportunity_legacy_id__c' to 'CurrencyIsoCode'
        currency_mapping = opportunity_df.set_index("opportunity_legacy_id__c")["CurrencyIsoCode"]

        # Use this mapping to populate the 'opportunity currency' column in the product sheet
        product_df["opportunity currency"] = product_df["opportunityid"].map(currency_mapping).fillna("Not Found")

        # Success message
        print("\n    ✅ Process completed. The 'opportunity currency' column has been successfully created in the 'Opportunity_product' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found.")
        sys.exit()

    except KeyError as e:
        # Handle missing column error
        print(f"\n    ❌ Error: The required column '{e.args[0]}' is missing.")
        sys.exit()

    except Exception as e:
        # Handle any other unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e}")
        sys.exit()


    # ======================================================================
    # Step 7: To delete unwanted columns from the sheet
    # ======================================================================

    print("\n\n🔍 Step 7: Deleting unwanted columns from the 'Opportunity_product' sheet...")

    # Define columns to be removed
    columns_to_delete = [
        "created_by","current quarter revenue","modified_by",
        "created_date","modified_date","product_code_family","pricebookentryid"
    ]

    try:
        unnamed_columns = []
        for col in product_df.columns:
            if pd.isna(col) or str(col).strip() == "" or str(col).startswith("Unnamed"):
                unnamed_columns.append(col)

        # Combine all columns to remove
        all_columns_to_delete = unnamed_columns + columns_to_delete
        
        # Track deleted columns
        deleted_columns = []

        # Drop only the columns that exist
        for col in all_columns_to_delete:
            if col in product_df.columns:
                product_df = product_df.drop(columns=[col])
                deleted_columns.append(col)
                # print(f"\n    🔸 '{col}'")

        if deleted_columns:
            print(f"\n    ✅ Successfully deleted {len(deleted_columns)} columns from the '{product_sheet_name}' sheet.")
        else:
            print(f"\n    ✅ No columns were deleted from the '{product_sheet_name}' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()



    # ======================================================================
    # Step 8: To get the "Product_Code_Family" column
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'Product_Code_Family' column in the 'Opportunity_product' sheet...")

    try:
        # Validate required columns
        if "product" not in product_df.columns:
            print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "product_type" not in product_df.columns:
            print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Create new column by concatenating product and product_type
        product_df["Product_Code_Family"] = product_df["product"] + "-" + product_df["product_type"]
        
        # Success message
        print(f"\n    ✅ The 'Product_Code_Family' column has been created and saved. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()



    # ======================================================================
    # Step 9: To get the "Practise_Multiple country" column
    # ======================================================================

    print("\n\n🔍 Step 9: Creating 'Practise_Multiple country' column in the 'Opportunity_product' sheet...")


    try:

        # Validate required columns
        required_columns = ["product", "product_type", "opportunity currency"]
        
        for col in required_columns:
            if col not in product_df.columns:
                print(f"\n    ❌ Error: Column '{col}' not found in '{product_sheet_name}' sheet.")
                sys.exit()

        # Create new column by concatenating product, product_type, and opportunity currency
        product_df["Practise_Multiple country"] = product_df["product"] + "-" + product_df["product_type"] + "-" + product_df["opportunity currency"]

        # Success message
        print(f"\n    ✅ The 'Practise_Multiple country' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 11: To keep the decimal values as 2
    # ======================================================================

    print("\n\n🔍 Step 11: Formatting decimal values to two decimal places...")

    headers_to_format = ['unitprice', 'expiring amount']

    try:
        # Format each specified column to 2 decimal places
        for col in headers_to_format:
            if col in product_df.columns:
                product_df[col] = pd.to_numeric(product_df[col], errors='coerce').round(2)

        print(f"\n    ✅ Numbers in columns {', '.join(headers_to_format)} have been formatted to two decimal places. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 12:- To copy the data from CSV file
    # ======================================================================

    print("\n\n🔍 Step 12: Check If the Product Extract File is present...")

    # Define the path to the 'productfamily.csv' file
    product_family_csv = DOWNLOADS_DIR+"/productfamily.csv"

    # Continuously check if the file exists
    # If not, prompt the user to retry or exit
    while not os.path.exists(product_family_csv):

        print(f"\n    ❌ File 'productFamily.csv' does not exist. Did you query the ProductFamily?")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        if try_again == 'no':
            print("\n          🚫 Exiting the program.")
            sys.exit()
    
    if os.path.exists(accounts_csv):
        print(f"\n    ✅ Product CSV file is present.")




    # ======================================================================
    # Step 13:- Create 'Practise_Multiple country' column in Product Copy sheet
    # ======================================================================

    print("\n\n🔍 Step 13: Create 'Practise_Multiple country' column in Product Copy sheet" )


    product_copy_df = pd.read_csv(product_family_csv)  # Read first two columns
    # Create a new column by combining Product Code Family and Currency with a hyphen
    product_copy_df["Practise_Multiple country"] = product_copy_df["Product2.Product_Code_Family__c"] + "-" + product_copy_df["CurrencyIsoCode"]

    print(f"\n    ✅ The values from have been successfully concatenated and saved in the 'Practise_Multiple country' column.")


    # ======================================================================
    # Step 14:- Getting the PricebookEntry id
    # ======================================================================

    print("\n\n🔍 Step 14: Getting the PricebookEntry id ...")

    # Function to standardize column names
    def standardize_columns(df):
        df.columns = df.columns.str.strip().str.lower()
        return df

    # Function to standardize column values
    def standardize_column_values(df, column_name):
        df[column_name] = df[column_name].str.strip().str.lower()
        return df

    # Standardize column names and merge key values in both DataFrames
    product_df = standardize_columns(product_df)
    product_copy_df = standardize_columns(product_copy_df)

    # Standardize column values for the merge key
    product_df = standardize_column_values(product_df, 'practise_multiple country')
    product_copy_df = standardize_column_values(product_copy_df, 'practise_multiple country')

    # Prevent accidental overwrite if 'pricebookentryid' already exists
    if 'pricebookentryid' in product_df.columns:
        raise KeyError("❌ Error: Column 'pricebookentryid' already exists in 'Opportunity_product'. Please check your data processing steps.")

    # 1️⃣ Keep only active entries and get their ids
    active_product_copy_df = product_copy_df[product_copy_df['isactive'] == True].copy()
    active_product_copy_df['pricebookentryid'] = active_product_copy_df['id']
    active_product_copy_df = active_product_copy_df[['practise_multiple country', 'pricebookentryid']].drop_duplicates(subset='practise_multiple country', keep='first')

    # 2️⃣ Merge active pricebook ids into product_df
    merged_df = pd.merge(product_df, 
                        active_product_copy_df, 
                        on='practise_multiple country', 
                        how='left')

    # 3️⃣ Find rows with no active pricebook id found (null values)
    missing_pricebook_mask = merged_df['pricebookentryid'].isna()

    # 4️⃣ Now check if any inactive entries exist for those countries
    inactive_product_copy_df = product_copy_df[product_copy_df['isactive'] == False].copy()
    inactive_countries = inactive_product_copy_df['practise_multiple country'].unique()

    # 5️⃣ Assign 'Not Active' to countries with inactive records
    merged_df.loc[merged_df['practise_multiple country'].isin(inactive_countries) & missing_pricebook_mask, 'pricebookentryid'] = 'Not Active'

    # 6️⃣ Fill remaining missing values with 'No Pricebookid found'
    merged_df['pricebookentryid'] = merged_df['pricebookentryid'].fillna('No Pricebookid found')

    # 7️⃣ Count stats
    count_no_pricebookid_found = (merged_df['pricebookentryid'] == 'No Pricebookid found').sum()
    count_not_active = (merged_df['pricebookentryid'] == 'Not Active').sum()

    product_df = merged_df

    # 9️⃣ Print completion message and stats
    print(f"\n    ✅ The 'Opportunity_product' sheet has been successfully updated with the 'PriceBookEntryid' column.")

    if count_no_pricebookid_found > 0 or count_not_active > 0:
        print(f"\n        ❗️ Count of 'No Pricebookid found': {count_no_pricebookid_found}")
        print(f"\n        ❗️ Count of 'Not Active': {count_not_active}")
    else:
        print(f"\n    ✅ All Products are Valid")





    # ======================================================================
    # Step 15: Rename the Columns
    # ======================================================================

    print("\n\n🔍 Step 15: Renaming Columns in the 'Opportunity_product' Sheet...")


    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunityid': 'Legacy_Opportunity_Split_Id__c',
        'quantity': 'Quantity',
        'product_code_family': 'Product_Family__c',
        'pricebookentryid': 'PricebookEntryId',
        'unitprice': 'UnitPrice',
        'term': 'Term__c',
        'classification type': 'Classification__c',
        'type': 'Type__c',
        'renewal type': 'Renewal_Type__c',
        'renewal status': 'Renewal_Status__c',
        'expiration date': 'Expiration_Date__c',
        'expiring term': 'Expiring_Term__c',
        'expiring amount': 'Expiring_Amount__c',
        'external id': 'External_IDs__c',
    }

    product_df = safe_rename_columns(product_df, column_rename_mapping)
    # ======================================================================
    # Step 16: Rearranging the Columns in Sequence
    # ======================================================================

    print("\n\n🔍 Step 16: Rearranging Columns in the 'Opportunity_product' Sheet...")

    # Specify the desired order of columns
    desired_column_order = [
        'Legacy_Opportunity_Split_Id__c','existing','Quantity','product','product_type','PricebookEntryId',
        'Product_Family__c','opportunity currency','practise_multiple country'
    ]

    # Call Function to Rearrange Columns in a sheet
    product_df =rearrange_and_save_columns(product_df, desired_column_order, product_sheet_name)


    # ======================================================================
    # Step 18: Final Row and Column Count
    # ======================================================================

    print("\n\n🔍 Step 18: Counting Final rows in Product Sheet...")

    product_final_num_rows = count_row(product_df)

    # Display the final row count
    print(f"\n    ✅ Final row count: {product_final_num_rows}")

    # Check if the number of rows has changed
    check_row_count(product_initial_num_rows, product_initial_num_rows, label = "Opportunity_product")

    # ======================================================================

    title = "📝 PRODUCT SHEET COMPLETED 📝"
    show_title(title)
    # ======================================================================



    # =========================================================================================================================================
    #                                                TEAM MEMBER SHEET EXECUTION
    # =========================================================================================================================================
    
    
    # Display the header once
    print("\n\n📄 Execute Next Sheet:")

    original_team_sheet = 'Opportunity_Team '
    
    team_df = pd.read_excel(file_path, sheet_name=original_team_sheet)
    
    def is_sheet_empty(df, original_team_sheet):
        """
        Check if a given Excel sheet is empty or contains only headers.

            Returns:
                - (True, None): If the sheet is empty or has only headers.
                - (False, DataFrame): If the sheet contains data (returns first 4 rows).
                - (None, None): If an error occurs while reading the sheet.
        """
        try:
            # Check if the sheet is empty or only contains headers
            if df.empty or df.dropna(how='all').shape[0] == 0:
                return True, None  # Sheet is empty or has only headers
            
            return False, df.head(4)  # Sheet contains data, return first 4 rows
        except Exception as e:
            print(f"\n❗️ Error reading sheet '{original_team_sheet}': {e}\n")
            return None, None
        
    # Check if Opportunity_Team sheet has data
    is_empty, preview = is_sheet_empty(team_df, original_team_sheet)

    # Display results based on sheet contents
    if is_empty:
        print(f"\n📂 The sheet '{original_team_sheet}' is empty or contains only headers.")
        choice = 'no'
    elif is_empty is None:
        print("\n❗️ Could not process the sheet due to an error.\n")
        choice = 'no'
    else:
        choice = 'yes'
        print(f"\n✅ The sheet '{original_team_sheet}' contains data. Here are the first 4 rows:\n")
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    # ==============================================
    # Function to skip values in one cell
    # ==============================================
    
    def process_dataframe(df, column_name, split_column):
        """
        Process a DataFrame by handling duplicates, blank values, and splitting column values into multiple rows.
        
        Parameters:
        - df: The DataFrame to be processed.
        - column_name: The name of the column to check for missing values (e.g., 'opportunityid').
        - split_column: The column to split if it contains multiple values (e.g., 'email', 'reporting_codes').
        - sheet_name: The name of the sheet being processed (used in print statements).
        
        Returns:
        - A new DataFrame with the processed rows.
        """
        # Initialize an empty list to store rows for the new DataFrame
        new_rows = []
        duplicate_count = 0  # Counter for duplicate rows
        skipped_blank_count = 0  # Counter for rows skipped due to blank values

        # Set to track duplicate rows
        seen_rows = set()

        # Iterate through each row in the original DataFrame
        for index, row in df.iterrows():
            # Check if the row is already seen (duplicate)
            row_tuple = tuple(row.items())
            if row_tuple in seen_rows:
                duplicate_count += 1
                continue  # Skip duplicate rows
            seen_rows.add(row_tuple)

            # Handle blank or NaN values in specified column
            column_value = row.get(column_name, None)
            split_values = row.get(split_column, None)

            if pd.isna(column_value) or pd.isna(split_values):
                skipped_blank_count += 1  # Increment skipped rows count
                continue  # Skip rows with missing column values

            # Split the values if multiple are present
            split_values = str(split_values).split(',')

            if len(split_values) > 1:
                for value in split_values:
                    value = value.strip()  # Remove any whitespace
                    if value:  # Skip blank values
                        new_row = row.copy()
                        new_row[split_column] = value  # Assign a single value
                        new_rows.append(new_row)
            else:
                # Handle rows with a single value
                if split_values[0].strip():  # Skip rows with blank single value
                    new_rows.append(row)

        total_rows_before = len(df)
        # Create a new DataFrame with the processed rows
        df_processed = pd.DataFrame(new_rows, columns=df.columns)

        # Total row counts before and after removing duplicates
        total_rows_after = len(df_processed)

        # Drop duplicate rows (if necessary)
        df_processed = df_processed.drop_duplicates()

        # Print results
        print(f"\n    ✅ Values Separated.")
        print(f"\n        🔸 Total rows before processing: {total_rows_before}")
        print(f"\n        🔸 Total rows after processing: {total_rows_after}")
        rows_difference =  total_rows_before - total_rows_after
        if rows_difference > 0 :
            print(f"\n        🔸 Rows removed: {rows_difference}")
        else:
            print(f"\n        🔸 Rows Added: {abs(rows_difference)}")
        
        return df_processed
    # ==============================================
    # Function Kept outside Loop
    # ==============================================

    # Ask user to proceed with execution of Team Member sheet
    while True:
        print(f"\n    🔹 Do you want to execute the Team member Sheet ? (yes/no): {choice}")
        if choice == "yes":
            team_member_choice = 'yes'
            print(f"\n        ⏳ Executing the Sheet: Teammember sheet ")
            
            # ======================================================================

            title = "📝 TEAM MEMBER SHEET EXECUTION 📝"
            show_title(title)

            # ======================================================================
            
            Opportunity_team_sheet_name = 'Opportunity_team'

            # ======================================================================
            # Step 1:-  Check count of rows and columns
            # ======================================================================
            
            print("\n\n🔍 Step 1: Counting Initial rows in Team Sheet ...")

            # Function to count rows in a specific sheet
            team__initial_num_rows = count_row(team_df)
            
            # Print the number of rows and columns
            print(f"\n    ✅ Initial rows count in Team sheet: {team__initial_num_rows}\n")

            # ======================================================================
            # Step 2: Creating New Records for Multiple Emails in a Cell
            # ======================================================================

            print("\n\n🔍 Step 2: Creating New Records for Multiple Emails in a Cell...")

            team_df = process_dataframe(team_df, 'opportunityid', 'email')

            # ======================================================================
            # Step 5 :- Checking if Opportunities Exist in the 'Opportunity' Sheet
            # ======================================================================

            print("\n\n🔍 Step 5: Checking if Opportunities Exist in the 'Opportunity' Sheet...")

            # Call Function to verify opportunities
            team_df = verify_opportunity(team_df,Opportunity_team_sheet_name)

            # ======================================================================
            # 🔍 Step 3: Fetching User IDs of Team Members...
            # ======================================================================

            print ('\n\n🔍 Step 3: Fetching User IDs of Team Members...')

            try:

                team_df["email"] = team_df["email"].astype(str).str.strip().str.lower()
                userid_df["Email"] = userid_df["Email"].astype(str).str.strip().str.lower()
                userid_df["Email"] = userid_df["Email"].str.strip().str.lower()

                # Perform a left join:
                result_df = pd.merge(
                    team_df,
                    userid_df[["Email", "Id"]],
                    left_on="email",
                    right_on="Email",
                    how="left"
                )

                # Count how many user IDs are missing (NaN) before replacement
                nan_before = result_df["Id"].isna().sum()

                # Fill missing IDs with "Inactive" to flag unmatched users
                result_df["Id"] = result_df["Id"].fillna("Inactive")

                # Count missing IDs after replacement (should be zero now)
                nan_after = result_df["Id"].isna().sum()

                # Calculate how many IDs were marked as "Inactive"
                nan_replaced = nan_before - nan_after

                # Remove the now redundant 'Email' column (came from the right dataframe)
                result_df.drop(columns=["Email"], inplace=True)

                # Rename 'Id' column to 'UserId' for clarity or standardization
                result_df.rename(columns={"Id": "UserId"}, inplace=True)
                team_df = result_df
                # Output success message
                if nan_replaced > 0:
                    print(f"\n    ❗️ Number of 'Inactive' values : {nan_replaced}")
                else:
                    print(f"\n    ✅ All Team Member are valid")

            except FileNotFoundError:
                print(f"\n    ❌ File not found at path: {file_path}.")

            except KeyError as e:
                print(f"\n    ❌ KeyError: Column '{e}' not found. Please check the column names in your Excel sheets.")

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")

            # ======================================================================
            # 🔍 Step 4: Renaming Columns...
            # ======================================================================

            print("\n\n🔍 Step 4: Renaming Columns...")

            # Dictionary mapping old column names to new column names
            column_rename_mapping = {
                'opportunityid': 'OpportunityId',
                'teammemberrole': 'TeamMemberRole',
                'opportunityaccesslevel': 'OpportunityAccessLevel',
            }
            team_df = safe_rename_columns(team_df, column_rename_mapping)
            
            # ======================================================================
            # 🔍 Step 9: Rearranging Columns in Sequence...
            # ======================================================================

            print("\n\n🔍 Step 9: Rearranging Columns in Sequence...")

            # Specify the preferred order for key columns in the sheet
            desired_column_order = [
                "OpportunityId",
                "existing",
                "OpportunityAccessLevel",
                "TeamMemberRole",
                "email",
                "UserId"
            ]

            team_df =rearrange_and_save_columns(team_df, desired_column_order,Opportunity_team_sheet_name )    

            # ======================================================================
            # Step 11: Count the Number of Rows and Columns
            # ======================================================================

            print("\n\n🔍 Step 11: Counting the Number of Rows and Columns...")

            # check count of rows
            team_final_num_rows = count_row(team_df)
            # Print the number of rows and columns
            print(f"\n    ✅ Final rows count: {team_final_num_rows}")


            # ======================================================================
            # Team Member Sheet Completed  
            # ======================================================================

            title = "📝 TEAM MEMBER SHEET COMPLETED 📝"
            show_title(title)

            break
        elif choice == "no":
            team_member_choice = 'no'
            print("\n        🚫 Team Member sheet execution skipped!")
            print("\n")

            title = "📝 TEAM MEMBER SHEET SKIPPED 📝"
            show_title(title)

            break  # Just breaking without running Block 2

        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")


    # =========================================================================================================================================
    #                                                STRATEGY SHEET EXECUTION
    # =========================================================================================================================================

    # Display the header once
    print("\n\n📄 Execute Next Sheet:")
    
    original_codes_sheet_name = 'Reporting_codes'
    tag_sheet_name = 'Tags'

    try:

        codes_df = pd.read_excel(file_path, sheet_name = original_codes_sheet_name)
        
        # Check if the 'Reporting_codes' sheet is empty or contains only headers
        is_empty, preview = is_sheet_empty(codes_df, original_codes_sheet_name)

        if is_empty:
            # If the sheet is empty or contains only headers, skip execution
            print(f"\n📂 The sheet '{original_codes_sheet_name}' is empty or contains only headers.\n")
            choice = 'no'

        elif is_empty is None:
            # If an error occurred while processing the sheet
            print("\n❗️ Could not process the sheet due to an error.\n")
            choice = 'no'

        else:
            # Sheet contains data, proceed with processing
            choice = 'yes'
            
            # ----------------------------- Code to create Tags Column if missing --------------------------------- 
            
            tag_column = 'tags'
            
            # Ensure 'tags' column exists; if not, create it with empty values
            if tag_column not in codes_df.columns:
                codes_df['tags'] = None  # Create an empty 'tags' column

            # Check if 'tags' column exists (case-insensitive)
            tags_column = [col for col in codes_df.columns if col.strip().lower() == 'tags']

            # Check if 'Reporting Codes' column exists (case-insensitive)
            reporting_codes_column = [col for col in codes_df.columns if col.strip().lower() == 'reporting_codes']

            # Check if 'Opportunity_id' column exists (case-insensitive)
            opportunity_id_column = [col for col in codes_df.columns if col.strip().lower() == 'opportunity_id']

            # If 'tags' column doesn't exist or is empty, add or update the 'tags' column
            if reporting_codes_column and opportunity_id_column:
                reporting_codes_column = reporting_codes_column[0]
                opportunity_id_column = opportunity_id_column[0]

                # If the 'tags' column doesn't exist or is completely empty
                if not tags_column or codes_df[tags_column[0]].isna().all():
                    # Convert 'tags' column to string type to avoid issues when assigning text
                    codes_df['tags'] = codes_df['tags'].astype(str)  # Convert to string type to avoid dtype mismatch
                    
                    # Populate 'tags' column by copying 'Reporting Codes' values for each unique opportunity_id
                    for opportunity_id in codes_df[opportunity_id_column].unique():
                        
                        # Get all rows for the current opportunity_id
                        opportunity_rows = codes_df[codes_df[opportunity_id_column] == opportunity_id]
                        
                        # Assign corresponding 'Reporting Codes' values to the 'tags' column
                        codes_df.loc[codes_df[opportunity_id_column] == opportunity_id, 'tags'] = opportunity_rows[reporting_codes_column].values
            
            print(f"\n✅ The sheet '{original_codes_sheet_name}' contains data. Here are the first 4 rows:\n")
            preview = codes_df.head(4)
            print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))
    except ValueError:
        print(f"\n⚠️ The sheet '{original_codes_sheet_name}' is not present in the Excel file.\n")
        choice = 'no'
        
    while True:

        print(f"\n    🔹 Do you want to execute the Strategy Sheet ? (yes/no): {choice}")
        
        if choice == "yes": # Automatically taken

            strategy_choice = 'yes'
            print(f"\n        ⏳ Executing the Sheet: Strategy sheet ")

            print("\n")
            title = "📝 STRATEGY SHEET EXECUTION 📝"
            show_title(title)

            # ========================================================================
            # Step 1: Rename the columns in "Reporting_codes_2" sheet
            # ========================================================================

            print("\n\n🔍 Step 1: Renaming Columns in 'Reporting_codes' Sheet...\n")

            # Load the workbook

            # Define normalization mapping for column headers
            column_name_mappings = {
                'tags': 'tag',
                'Opportunity Id': 'opportunityid',
                'opportunity_id': 'opportunityid',
                'Opportunityid': 'opportunityid',
                'opportunityid': 'opportunityid'
            }


            # Normalize column names: strip and lowercase for matching
            renamed_columns = {}
            new_columns = []

            for col in codes_df.columns:
                normalized_col = col.strip().lower()
                if normalized_col in column_name_mappings:
                    new_col = column_name_mappings[normalized_col]
                    renamed_columns[col] = new_col
                    new_columns.append(f"'{col}' ➔ '{new_col}'")

            # Rename columns
            codes_df.rename(columns=renamed_columns, inplace=True)
            
            if new_columns:
                print(f"\n    ✅ Columns renamed successfully:")
            else:
                print("\n    ❗️ No columns were renamed. All columns were already in the desired format.")

            # ========================================================================
            # Step 2: Create separate sheet for "Tags"
            # ========================================================================

            print("\n\n🔍 Step 2: Creating 'Tags' Sheet...")

            # Normalize column names to lowercase for consistency
            codes_df.columns = codes_df.columns.str.lower()

            # Check if 'opportunity_id' and 'tag' columns are present
            expected_columns = ['opportunityid', 'tag']
            missing_columns = [col for col in expected_columns if col not in codes_df.columns]
            if missing_columns:
                print(f"\n    ❌ Error: Missing columns: {', '.join(missing_columns)}. ")
                exit()

            # Filter rows where 'tag' column contains a value
            tags_df = codes_df[codes_df['tag'].notnull()][expected_columns]
            codes_df.drop(columns='tag', inplace=True)

            if not tags_df.empty:
                print('\n    ✅ Tags Created Separately')
            else:
                print('\n    ❌ Something went wrong')

            # ========================================================================
            # Step 3: To Remove Comma Separated Values from 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 3: Removing Comma Separated Values and Duplicates from 'Reporting_codes' Sheet...\n")
            
            # Splitting Values for Reporting Codes
            codes_df = process_dataframe(codes_df, 'opportunityid', 'reporting_codes')
            
            # Splitting Values for Tags
            tags_df = process_dataframe(tags_df, 'opportunityid', 'tag')

            # ========================================================================
            # Step 4: Add existing Column to 'Tags' Sheet
            # ========================================================================

            print("\n\n🔍 Step 4: Adding 'existing' Column to Strategy...")

            # Add Existing Column to Reporting Codes
            codes_df = verify_opportunity(codes_df,original_codes_sheet_name) 
            
            # Add Existing Column to tags
            tags_df = verify_opportunity(tags_df, tag_sheet_name)
            
            # ========================================================================
            # Step 5:- Processing CSV File and Adding Filtered Data to Excel
            # ========================================================================

            print("\n\n🔍 Step 5: Getting Id for codes and Tags...")

            # Define the file path for the Exported csv file
            tags_csv = DOWNLOADS_DIR+ "/tags.csv"

            # Flag to determine whether to proceed with processing
            run_code_strategy = False
            
            # Check if the file already exists
            if os.path.exists(tags_csv):
                run_code_strategy = True  # Proceed if the file is already available
            else:         
                
                # Keep trying to find or rename the file until it's available or user chooses to skip
                while not os.path.exists(tags_csv):
                    
                    # Inform that the file is still missing and prompt for next action
                    print(f"\n    ❌ File 'tags.csv' does not exist. Did you query the tags?")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                    # Keep asking until valid input is received
                    while try_again not in ['yes', 'no']:
                        print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                    if try_again == 'yes':
                        if os.path.exists(tags_csv):
                            run_code_strategy = True
                            break
                        else:
                            continue

                    if try_again == 'no':
                        run_code_strategy = False
                        strategy_choice = 'no'
                        print("\n          🚫 Skipping Strategy Sheet.")
                        break

            # If the csv file exists or becomes available, proceed with processing it
            if run_code_strategy:
                
                # ----------------------------------------
                # To copy the extract data from tags csv into dataframe 
                # ----------------------------------------

                def extract_data_copy_df(df,filter_value):
                    # Specify the column name to filter
                    filter_column = "Record_Type_Name__c"

                    # Check if the filter column exists in the CSV file
                    if filter_column not in df.columns:
                        print(f"\n    ❌ Error: Column '{filter_column}' not found in the CSV file.")
                        sys.exit()  # Exit if the required column is not found

                    # Filter rows where the specified column equals the specified value
                    df = df[df[filter_column] == filter_value]
                    return df
            
                # Read CSV file into a DataFrame
                code_copy_df = pd.read_csv(tags_csv)
                code_copy_df = extract_data_copy_df(code_copy_df,"Reporting codes")

                # Read CSV file into a DataFrame
                tag_copy_df = pd.read_csv(tags_csv)
                tag_copy_df = extract_data_copy_df(tag_copy_df,"Tags")

                # ----------------------------------------
                # To do Vlookup and get id for strategy 
                # ----------------------------------------

                def merge_strategy_ids(base_df, lookup_df, base_key, lookup_key='Name', id_column='Id', new_column='StrategyId', label='Codes'):
                    """
                    Merges a base DataFrame with a lookup DataFrame on lowercase-matched keys,
                    adds a new column with strategy IDs, and logs missing matches.

                    Parameters:
                        base_df (pd.DataFrame): DataFrame containing the key to match.
                        lookup_df (pd.DataFrame): Lookup DataFrame with 'Name' and 'Id' columns.
                        base_key (str): Column name in base_df to match on.
                        lookup_key (str): Column name in lookup_df to match on (default 'Name').
                        id_column (str): ID column to extract from lookup_df (default 'Id').
                        new_column (str): Name of the new column to store results (default 'StrategyId').
                        label (str): Text label for error/log messages (e.g. 'Codes', 'Tags').

                    Returns:
                        pd.DataFrame: Updated base_df with new_column added.
                    """
                    try:
                        # Normalize both merge keys
                        base_df[base_key] = base_df[base_key].astype(str).str.strip().str.lower()
                        lookup_df[lookup_key] = lookup_df[lookup_key].str.lower()

                        # Merge on lowercase-matched columns
                        result_df = pd.merge(
                            base_df,
                            lookup_df[[lookup_key, id_column]],
                            left_on=base_key,
                            right_on=lookup_key,
                            how='left'
                        )

                        # Count and report unmatched items
                        na_count = result_df[id_column].isna().sum()

                        # Create new ID column, replacing missing with placeholder
                        result_df[new_column] = result_df[id_column].fillna('Not found')

                        # Drop lookup columns
                        result_df.drop([lookup_key, id_column], axis=1, inplace=True)

                        # Log results
                        if na_count > 0:
                            print(f"\n    ❗️ Count of {label} 'Not found': {na_count}")
                        else:
                            print(f"\n    ✅ All {label} are Present")

                        return result_df

                    except FileNotFoundError:
                        print(f"\n    ❌ Error: Required file not found.")
                        sys.exit()
                    except KeyError as e:
                        print(f"\n    ❌ KeyError: {str(e)}")
                        sys.exit()
                    except Exception as e:
                        print(f"\n    ❌ An unexpected error occurred: {str(e)}")
                        sys.exit()
                
                # Get Ids for Reporting Codes
                codes_df = merge_strategy_ids(codes_df, code_copy_df, base_key='reporting_codes', label='Codes')
                
                # Get Ids for tags
                tags_df = merge_strategy_ids(tags_df, tag_copy_df, base_key='tag', label='Tags')

                # ========================================================================
                # Step :- If any Tags are not Found, Create Tags to Be inserted Csv file
                # ========================================================================
                
                output_for_tags = csv_file_dir + "/Tags_to_be_inserted.csv" # Change to your desired output file path

                def process_excel(df,output_for_tags):
                    
                    # Check if required columns exist
                    required_columns = {'opportunityid', 'tag', 'existing', 'StrategyId'}
                    if required_columns.issubset(df.columns):
                        
                        # Filter rows where StrategyId has 'Not found'
                        filtered_df = df[df['StrategyId'] == 'Not found']

                        # Drop duplicate tags
                        filtered_df = filtered_df.drop_duplicates(subset='tag')
                        
                        # Create the output DataFrame
                        output_df = pd.DataFrame({
                            'Name': filtered_df['tag'],
                            'Strategy_Full_Name__c': '',
                            'RecordTypeId': '0123h000000kqchAAA',
                            'Record_Type_Name__c': 'Tags',
                            'IsDeleted': False,
                            'Active__c': True
                        })
                        
                        # Save to new Excel file
                        if not output_df.empty:
                            output_df.to_csv(output_for_tags, index=False)
                        return
                    
                process_excel(tags_df, output_for_tags)

                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                    codes_df.to_excel(writer, sheet_name=original_codes_sheet_name, index=False)
                
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                    tags_df.to_excel(writer, sheet_name=tag_sheet_name, index=False)


                # ========================================================================
                print("\n")
                title = "📝 STRATEGY SHEET COMPLETED 📝"
                show_title(title)
                # ========================================================================

            break
        elif choice == "no":
            strategy_choice = 'no'
            print("\n        🚫 Strategy Sheet execution skipped!")

            print("\n")
            title = "📝 STRATEGY SHEET SKIPPED 📝"
            show_title(title)

            break  # Just breaking without running Block 2
        
        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")

    # ==========================================================================
                        # Contact Role Sheet Execution
    # ==========================================================================

    # Display the header once
    print("\n\n📄 Execute Next Sheet:")
    required_columns = {'opportunityid', 'contactid'}
    original_contact_sheet = 'Contact Roles'

    wb = load_workbook(file_path)
    if original_contact_sheet in wb.sheetnames:
        contact_df = pd.read_excel(file_path, sheet_name=original_contact_sheet)
        # Check if 'role' column exists
        if 'role' in contact_df.columns:
            # Replace blank or NaN values with 'Other'
            contact_df['role'] = contact_df['role'].replace(r'^\s*$', np.nan, regex=True)  # Treat empty strings as NaN
            contact_df['role'] = contact_df['role'].fillna('Other')

        is_empty, preview = is_sheet_empty(contact_df, original_contact_sheet)

        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    if original_contact_sheet not in wb.sheetnames:
        print(f"\n📂 The sheet '{original_contact_sheet}' is missing.")
        print("\n    🚫 Contact Role sheet execution skipped!")
        contact_choice = 'no'

    elif is_empty:
        print(f"\n📂 The sheet '{original_contact_sheet}' is empty or contains only headers.")
        print("\n    🚫 Contact Role sheet execution skipped!")
        contact_choice = 'no'

    elif not required_columns.issubset(preview.columns.str.lower()):
        print(f"\n    🚫 Required columns {required_columns} not found in the sheet. Skipping execution!")
        contact_choice = 'no'

    elif preview[['opportunityid', 'contactid']].isnull().all().any():
        # Check if either column is completely NaN or blank
        print(f"\n    🚫 One of the required columns has no data. Skipping execution!")
        contact_choice = 'no'

    elif is_empty is None:
        print("\n    ❗️ Could not process the sheet due to an error.")
        contact_choice = 'no'

    else:
        while True:
            user_choice = input(f"\n🔹 Do you want to execute the Contact Roles Sheet ? (yes/no): ").strip().lower()
            
            if user_choice == 'yes':
                
                contact_choice = 'yes'
                
                print(f"\n        ⏳ Executing the Sheet: Contact Roles")
                
                # ======================================================================
                print("\n")
                title = "📝 CONTACT ROLES SHEET EXECUTION 📝"
                show_title(title)

                # ======================================================================

                print("\n🔍 Step 1: Processing Contact sheet...")

                contact_sheet_name = 'Contact Roles'
                contact_df = verify_opportunity(contact_df,contact_sheet_name)

                # Adjust 'contactid' to match the correct column name from the printout
                if 'contactid' in contact_df.columns:
                    contact_df['contactid'] = contact_df['contactid'].apply(lambda x: str(int(x)))
                else:
                    print("Column 'contactid' not found!")

                # Use ExcelWriter to write the changes back to the same file, replacing the existing "Contact Roles" sheet
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    contact_df.to_excel(writer, sheet_name='Contact Roles', index=False)

                title = "📝 CONTACT ROLES SHEET COMPLETED 📝"
                show_title(title)

                break

            elif user_choice == 'no':
                contact_choice = 'no'
                print("\n        🚫 Contact Role sheet execution skipped!\n")
                
                print("\n")
                title = "📝 CONTACT ROLE SHEET SKIPPED 📝"
                show_title(title)
                break
            else:
                print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.\n")

    # =========================================================================================================================================
    #                                                FINAL FILE EXECUTION
    # =========================================================================================================================================


    # ========================================================================
    print("\n")
    title = "📝 FINAL SHEET EXECUTION 📝"
    show_title(title)
    # ========================================================================

    # ========================================================================
    # Step 1:-Initialize file name
    # ========================================================================


    print("\n\n🔍 Initializing file name and determining output path...")

    # Check if the output directory exists and is valid
    if os.path.exists(output) and os.path.isdir(output):
        print(f"\n    ✅ Output folder selected automatically:")
        
        # Show a relative path by trimming the base directory for cleaner display
        print(f"\n        📂 {output.split(BASE_DIR, 1)[-1]}")

    else:
        # If the path doesn't exist or is invalid, show an error and print full path
        print("\n    ❌ Error: The corresponding output folder does not exist.")
        print(f"\n       ❗️ Please check the path: {output}")


    # Extract the folder name from the output path to use in naming CSV files
    file_name = output.split('/')[-1]
    
    # Dynamically generate standardized CSV file names for different sheets
    opportunity = "1" + '_Opportunity load.csv'

    opportunity_product = "2" + '_Opportunity Product load.csv'

    opportunity_team = "3" + '_Opportunity Team member Load.csv'

    reporting_codes = "4" + '_Reporting_codes.csv'

    tags = "5" + '_Tags.csv'

    Contact_role = "6" + '_Contact_Roles.csv'

    # Print the generated file names for confirmation
    print("\n    📄 CSV File Names Generated:")
    print(f"\n        1️⃣ {opportunity}")
    print(f"\n        2️⃣ {opportunity_product}")
    print(f"\n        3️⃣ {opportunity_team}")
    print(f"\n        4️⃣ {reporting_codes}")
    print(f"\n        5️⃣ {tags}")
    print(f"\n        6️⃣ {Contact_role}")


    # ======================================================================
    # Step 1: Creating Opportunity Sheet
    # ======================================================================
    print()
    print("=" * 100)
    print("\n\n🔍 CREATING OPPORTUNITY FILE")
    

    # --------------------------- Define Path for Processed data and Removed Rows Data  ----------------------

    def get_file_paths(sheet_name, output_dir, removed_rows_dir, output_filename, removed_row_filename):
        # Construct the full paths based on provided filenames
        output_file = os.path.join(output_dir, output_filename)  # Processed CSV output path
        removed_rows_file = os.path.join(removed_rows_dir, removed_row_filename)  # Removed rows CSV
        
        return output_file, removed_rows_file

    sheet_name = 'Opportunity'
    opppty_processed_file, removed_rows_oppty = get_file_paths(sheet_name, output, removed_rows_dir,opportunity,'Removed_Rows - Oppty.csv')


    # -------------------------------------------- Define Columns --------------------------------------------

    predefined_columns_oppty = [
        'Already Exist','AccountNumber', 'Email', 'created_by', 'modified_by', 'created_date',
        'modified_date', 'Trimmed_accountid', 'Trimmed_ownerid', 'Type Of Opportunity',"mpp_number__c",
        'Concatenatedaccountid', 'Concatenatedownerid', 'concatenatedcreatedby','accountid','type of opportunity','Email.1'
    ]

    # List of columns to exclude from user selection in the GUI
    excluded_columns_oppty = [
        'opportunity_legacy_id__c', 'Legacy_Opportunity_Split_Id__c', 'name', 'StageName','description',
        'Won_Reason__c', 'Lost_Category__c', 'Lost_Reason__c', 'CloseDate', 'CurrencyIsoCode',
        'OwnerId', 'NextStep', 'OI_Group__c','AccountId','createdbyid','Pricebook2Id','RecordTypeId'
    ]

    # ---------------------- Read Opportunity Sheet ----------------------
    
    opportunity_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # ---------------------- Clean DataFrame ----------------------

    def clean_dataframe(df: pd.DataFrame, drop_duplicates: bool = True) -> pd.DataFrame:
        """
        Cleans the given DataFrame by:
        - Dropping columns that are completely empty
        - Dropping rows that are completely empty
        - Optionally dropping duplicate rows

        Parameters:
            df (pd.DataFrame): The DataFrame to clean.
            drop_duplicates (bool): Whether to drop duplicate rows. Defaults to True.

        Returns:
            pd.DataFrame: A cleaned DataFrame.
        """
        df.dropna(axis=1, how='all', inplace=True)  # Drop blank columns
        df.dropna(axis=0, how='all', inplace=True)  # Drop blank rows
        if drop_duplicates:
            df.drop_duplicates(inplace=True)        # Drop duplicate rows
        return df


    # call the function to clean the DataFrame
    opportunity_df = clean_dataframe(opportunity_df)

    # Initialize DataFrame to store removed rows and reasons
    removed_rows_oppty_df = pd.DataFrame(columns=opportunity_df.columns.tolist() + ['Reason'])

    # Track all columns that are dropped (predefined + user-selected)
    dropped_columns_oppty = []

    # ---------------------- Remove Rows: Already Exist is 'Already Exist in ISC' ----------------------

    def remove_rows_based_on_condition(df, removed_rows_df , column, value,reason):
    
        # Initialize the count to 0 for tracking removed rows
        row_removed_count = 0

        if column in df.columns:
            # Identify rows to remove based on the condition
            rows_to_remove = df[df[column] == value].copy()
            
            # Count the rows that are being removed
            row_removed_count = len(rows_to_remove)  # Count the rows to be removed
            
            if not rows_to_remove.empty:
                # Add a reason for removal
                rows_to_remove['Reason'] = reason
                # Add the removed rows to the removed_rows_df
                removed_rows_df = pd.concat([removed_rows_df, rows_to_remove], ignore_index=True)
            
            # Update the original DataFrame to keep only rows that don't match the condition
            df = df[df[column] != value]    # Keep only unmatched rows

        return df, removed_rows_df, row_removed_count

    opportunity_df, removed_rows_oppty_df, count_duplicate_oppty = remove_rows_based_on_condition(opportunity_df, removed_rows_oppty_df , 'Already Exist' , "Already Exist in ISC" ,  "Duplicate Opportunity")

    # ---------------------- Delete Predefined Columns ----------------------

    def delete_predefined_columns(df, predefined_columns,dropped_columns):
            

        columns_to_delete_predefined = []
        for col in predefined_columns:
            if col in df.columns:
                columns_to_delete_predefined.append(col)

        if columns_to_delete_predefined:
            df.drop(columns=columns_to_delete_predefined, inplace=True)
            dropped_columns.extend(columns_to_delete_predefined)
        
        return df, dropped_columns
    
    # Call the function to delete predefined columns
    opportunity_df,dropped_columns_oppty = delete_predefined_columns(opportunity_df, predefined_columns_oppty,dropped_columns_oppty)

    # ---------------------- GUI for User-Selectable Column Deletion ----------------------

    def show_column_deletion_gui(df, excluded_columns, dropped_columns):
        """
        Displays a GUI for selecting columns to delete from a DataFrame.

        Parameters:
        - opportunity_df (pd.DataFrame): The DataFrame from which columns may be deleted.
        - excluded_columns (list): List of columns that should not be shown in the GUI.
        - dropped_columns (list): A list to extend with any columns deleted by the user.

        Returns:
        - pd.DataFrame: The modified DataFrame with selected columns removed.
        """
        # Filter columns to be shown in the GUI (excluding important ones)

        columns_for_ui = []
        for col in df.columns:
            if col not in excluded_columns:
                columns_for_ui.append(col)
        
        checkboxes = {}
        if columns_for_ui:
            root = Tk()
            root.title("Select Columns to Delete")
            root.geometry("500x600")
            root.resizable(False, False)

            # Scrollable canvas setup
            canvas = Canvas(root)
            scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas to hold checkboxes
            frame = Frame(canvas)
            canvas.create_window((0, 0), window=frame, anchor="nw")

            for column in columns_for_ui:
                var = IntVar()
                checkboxes[column] = var
                checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
                checkbutton.pack(anchor="w", pady=5)

            # Submit button to close GUI
            button_frame = Frame(root)
            submit_button = Button(button_frame, text="Submit", command=root.quit,
                                font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
            submit_button.pack(side="right")
            button_frame.pack(anchor="ne", padx=20, pady=10)
            frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))

            root.mainloop()
            root.destroy()
        else:
            print("\n    ✅ No user-selectable columns available for deletion. Skipping GUI.")

        # Delete selected columns
        columns_to_delete_from_user = []
        for col, var in checkboxes.items():
            if var.get() == 1:
                columns_to_delete_from_user.append(col)

        if columns_to_delete_from_user:
            df.drop(columns=columns_to_delete_from_user, inplace=True)
            dropped_columns.extend(columns_to_delete_from_user)
            print("\n    ✅ Additional columns deleted:")
            for col in columns_to_delete_from_user:
                print(f"\n        🔸 {col}")
        else:
            print("\n    ✅ No additional columns selected for deletion.")

        return df
    # Call the GUI function to show the column deletion options
    
    opportunity_df = show_column_deletion_gui(opportunity_df, excluded_columns_oppty, dropped_columns_oppty)

# ---------------------- Remove Rows: AccountId is 'Not in ISC' ----------------------
    
    opportunity_df, removed_rows_oppty_df, count_not_in_isc = remove_rows_based_on_condition(opportunity_df, removed_rows_oppty_df , 'AccountId' , "Not in ISC" ,  "AccountId is 'Not in ISC'")

# ---------------------- Remove Rows with Invalid PricebookEntryId ----------------------

    count_invalid_pricebook = 0

    try:
        opportunity_product_df = pd.read_excel(file_path, sheet_name='Opportunity_product')
        
        # Identify invalid rows based on PricebookEntryId
        invalid_pricebook_ids = opportunity_product_df[
            opportunity_product_df['PricebookEntryId'].isin(['Not Active', 'No Pricebookid found'])
        ]['Legacy_Opportunity_Split_Id__c'].unique()

        rows_to_remove_invalid_pricebook = opportunity_df[
            opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
        ].copy()
        count_invalid_pricebook = len(rows_to_remove_invalid_pricebook)  # Count the rows to be removed
        
        if count_invalid_pricebook > 0:
            rows_to_remove_invalid_pricebook['Reason'] = "Invalid PricebookEntryId"
            opportunity_df = opportunity_df[
                ~opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
            ]
            removed_rows_oppty_df = pd.concat([removed_rows_oppty_df, rows_to_remove_invalid_pricebook], ignore_index=True)

    except Exception as e:
        print(f"\n    ❌ Error processing invalid PricebookEntryId rows: {e}")

# ---------------------- Row Removal Summary ----------------------

    def print_removal_summary(reason_counts: dict, label: str = "rows"):
        total_removed = sum(reason_counts.values())

        if total_removed > 0:
            print(f"\n    ❗️ Total {label} removed: {total_removed}")
            for reason, count in reason_counts.items():
                print(f"\n        🔸 {reason}: {count}")
        else:
            print(f"\n    ✅ No {label} removed")
    
    print_removal_summary({
        "Removed due to invalid PricebookEntryId": count_invalid_pricebook,
        "Remove due Account Not in ISC": count_not_in_isc,
        "Removed due to Duplicate Opportunity": count_duplicate_oppty
    })
        
# ---------------------- Clean Removed Rows DF ----------------------
   
    def drop_dropped_columns_from_removed_rows(dropped_columns,removed_rows_df):
        columns_to_drop_from_removed = []
        for col in dropped_columns:
            if col in removed_rows_df.columns:
                columns_to_drop_from_removed.append(col)

        if columns_to_drop_from_removed:
            removed_rows_df.drop(columns=columns_to_drop_from_removed, inplace=True)

        return removed_rows_df

    removed_rows_oppty_df = drop_dropped_columns_from_removed_rows(dropped_columns_oppty,removed_rows_oppty_df)

    # ---------------------- Save Cleaned Opportunity Data ----------------------

    def save_dataframe(df, file_path, label):
        if not df.empty:
            try:
                df.to_csv(file_path, index=False)
                shortened = "/".join(file_path.split("/")[-4:])
                print(f"\n    ✅ {label} data saved to:\n\n        📂 {shortened}")
            except Exception as e:
                print(f"\n    ❌ Error saving the {label.lower()} file: {e}")
        else:
            print(f"\n    ❌ {label} file is empty. Skipped The file")
    
    save_dataframe(opportunity_df, opppty_processed_file, "Opportunity")


# ---------------------- Save Removed Rows Data ----------------------

    def save_removed_rows(df_removed, removed_path, label):
        if not df_removed.empty:
            try:
                df_removed.to_csv(removed_path, index=False)
                shortened = "/".join(removed_path.split("/")[-5:])
                print(f"\n    ✅ Removed rows saved to:\n\n        📂 {shortened}")
            except Exception as e:
                print(f"\n    ❌ Error saving the removed rows file for {label.lower()}: {e}")

    save_removed_rows(removed_rows_oppty_df, removed_rows_oppty, "Removed Rows_Opportunity")

    mpp_output_file = os.path.join(csv_file_dir,"MPP_Column.xlsx")
    
    if os.path.exists(mpp_output_file):
        csv_df = pd.read_csv(opppty_processed_file, dtype={'opportunity_legacy_id__c': str})
        excel_df = pd.read_excel(mpp_output_file, dtype={'opportunity_legacy_id_c': str, 'Mpp_Number__c': str})
        # Read the files, preserve leading zeros
        csv_df = pd.read_csv(opppty_processed_file, dtype={'opportunity_legacy_id__c': str})
        excel_df = pd.read_excel(mpp_output_file, dtype={'opportunity_legacy_id_c': str, 'Mpp_Number__c': str})

        # Convert matching columns to lowercase for case-insensitive comparison
        csv_df['opportunity_legacy_id__c_lower'] = csv_df['opportunity_legacy_id__c'].str.lower()
        excel_df['opportunity_legacy_id_c_lower'] = excel_df['opportunity_legacy_id_c'].str.lower()

        # Merge the Excel Mpp_Number__c into the CSV based on lowercased IDs
        csv_df = csv_df.merge(
            excel_df[['opportunity_legacy_id_c_lower', 'Mpp_Number__c']],
            how='left',
            left_on='opportunity_legacy_id__c_lower',
            right_on='opportunity_legacy_id_c_lower'
        )

        # Drop the helper lowercase columns
        csv_df.drop(columns=['opportunity_legacy_id__c_lower', 'opportunity_legacy_id_c_lower'], inplace=True)

        # Save back to the **same CSV file** (overwrite)
        csv_df.to_csv(opppty_processed_file, index=False)

        print(f"CSV file '{opppty_processed_file}' updated successfully with 'Mpp_Number__c' values.")

    # =======================================================
    # Step 2:- Creating the Opportunity Product File
    # =======================================================
    print()
    print("=" * 100)
    print("\n\n🔍 CREATING PRODUCT FILE")

    # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------

    sheet_name = 'Opportunity_product'
    product_processed_file, removed_rows_product = get_file_paths(sheet_name, output, removed_rows_dir,opportunity_product,'Removed_Rows - Product.csv')
    
    # -------------------------------------------- Define Columns --------------------------------------------

    predefined_columns_product = [
        'existing', 'product', 'product_type', 'Product_Family__c', 
        'opportunity currency', 'practise_multiple country', 
        'quantity.1', 'concatenated product family', 'concatenated currency'
    ]

    # Define columns that should never be shown to the user for deletion
    excluded_columns_product = [
    'Type__c','Renewal_Type__c','Renewal_Status__c','Expiration_Date__c','Expiring_Term__c','Expiring_Amount__c',
    'External_IDs__c','month 1 revenue','month 2 revenue','month 3 revenue','next quarter revenue','first 12 months revenue',
    'pre-contract planned revenue','pre-contract start date','pre-contract end date','loss reason/attition reason',
    'Legacy_Opportunity_Split_Id__c','PricebookEntryId','UnitPrice','Term__c','Classification__c','Quantity'
    ]

    try:
        # ---------------------- Load Product Sheet ----------------------

        product_df = pd.read_excel(file_path, sheet_name=sheet_name)

        # ---------------------- Clean DataFrame ----------------------
        
        product_df = clean_dataframe(product_df, drop_duplicates=False)  # Clean the opportunity_df without dropping duplicates
    
        # Initialize DataFrame to track removed rows with reasons
        removed_rows_product_df = pd.DataFrame(columns=product_df.columns.tolist() + ['Reason'])
        
        # Track all columns that are dropped (predefined + user-selected)
        dropped_columns_product = []
        
        # ---------------------- Remove Rows Based on existing ----------------------
        
        # Call the function to remove non-existing rows
        product_df, removed_rows_product_df, rows_dropped_existing_count = remove_rows_based_on_condition(product_df, removed_rows_product_df,'existing',False,"Opportunity Missing From Main sheet")

        # ---------------------- Remove Rows Based on Removed Oppties ----------------------

        def load_removed_rows_oppty(removed_rows_oppty,df,removed_rows_df,column_match):
            
            rows_dropped_legacy_count = 0
            if os.path.exists(removed_rows_oppty):
                
                oppty_removed_rows_df = pd.read_csv(removed_rows_oppty, usecols=["opportunity_legacy_id__c"])
                opportunity_ids_set = set(oppty_removed_rows_df["opportunity_legacy_id__c"].dropna().astype(str))

                # Filter rows where "Legacy_Opportunity_Split_Id__c" exists in opportunity_ids_set
                removed_rows_legacy = df[df[column_match].astype(str).isin(opportunity_ids_set)].copy()
                rows_dropped_legacy_count = len(removed_rows_legacy)

                if not removed_rows_legacy.empty:
                    removed_rows_legacy["Reason"] = "Opportunity not loaded"
                    removed_rows_df = pd.concat([removed_rows_df, removed_rows_legacy], ignore_index=True)
                    df = df[~df[column_match].astype(str).isin(opportunity_ids_set)]  # Keep only unmatched rows
            return df, removed_rows_df, rows_dropped_legacy_count

        product_df,removed_rows_product_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, product_df, removed_rows_product_df, "Legacy_Opportunity_Split_Id__c")
        
        # ---------------------- Remove Predefined Columns ----------------------

        product_df, dropped_columns_product = delete_predefined_columns(product_df, predefined_columns_product, dropped_columns_product)

        # ---------------------- User-Guided Column Deletion (GUI) ----------------------
        
        product_df = show_column_deletion_gui(product_df, excluded_columns_product, dropped_columns_product)
        
        # ---------------------- Row Removal Summary ----------------------

        print_removal_summary({
            "Due to 'existing' == False": rows_dropped_existing_count,
            "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
        })
        
        # ---------------------- Clean Removed Rows DF ----------------------
        
        # Drop all dropped columns from removed_rows_df (retain only 'Reason' and important columns)
        
        removed_rows_product_df = drop_dropped_columns_from_removed_rows(dropped_columns_product,removed_rows_product_df)
        
        # ---------------------- Save Processed Product File ----------------------
        
        save_dataframe(product_df, product_processed_file, "Product")

        # ---------------------- Save Removed Rows Data ----------------------
        
        save_removed_rows(removed_rows_product_df, removed_rows_product, "Removed Rows_Product")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e}")

    # =======================================================
    # Step 3: Processing Opportunity Team Data
    # =======================================================

    # Prompt: Check if user wants to run the Opportunity Team Member Sheet processing

    while True:
        print("\n================================================================================")
        print(f'\n📄 Do you want to run the team member Sheet? (yes/no): {team_member_choice}')
        
        # Use value from previous logic (assumed auto-input)
        user_input = team_member_choice # Automatically taking the user input from the above question from file processing
        
        if user_input == "yes":
            print("\n    ⏳ Running Opportunity Team Member Sheet...")
            print("\n================================================================================")

            # =======================================================
            # MAIN CODE
            # =======================================================

            print("\n\n🔍 CREATING TEAM MEMBER FILE")

            # ------------------------- Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Opportunity_team'
            team_procesed_file, removed_rows_team = get_file_paths(sheet_name, output, removed_rows_dir,opportunity_team,'Removed_Rows - Team.csv')
            
            # ---------------------- Define Columns ----------------------

            predefined_columns_Team = ['existing', 'email', 'Concat_T_M']
            
            # Columns that should not appear in the deletion GUI
            excluded_columns_Team=['OpportunityId','OpportunityAccessLevel','TeamMemberRole','UserId']

            try:
                
                # ---------------------- Load Sheet ----------------------

                team_df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # ---------------------- Clean DataFrame ----------------------

                team_df = clean_dataframe(team_df)

                #---------------------- Initialize Removed Rows df ----------------------
                
                removed_rows_team_df = pd.DataFrame(columns=team_df.columns.tolist() + ['Reason'])
                
                # ---------------------- columns that are dropped ----------------------

                dropped_columns_team = [] 
                
                # ---------------------- Remove rows where 'existing' is False ----------------------

                team_df,removed_rows_team_df,rows_dropped_existing_count =  remove_rows_based_on_condition(team_df, removed_rows_team_df,'existing',False,"Opportunity Missing From Main sheet")
                        
                # ---------------------- Remove Rows Based on Removed Oppties ----------------------
                
                team_df,removed_rows_team_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, team_df, removed_rows_team_df, "OpportunityId")

                # ---------------------- Remove Predefined Columns ----------------------
                
                team_df, dropped_columns_team = delete_predefined_columns(team_df, predefined_columns_Team, dropped_columns_team)
                
                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------

                # Call the GUI function to show the column deletion options
                team_df = show_column_deletion_gui(team_df, excluded_columns_Team, dropped_columns_team)
                
                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })                

                # ---------------------- Clean Removed Rows DF ----------------------
                
                removed_rows_team_df = drop_dropped_columns_from_removed_rows(dropped_columns_team,removed_rows_team_df)

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(team_df, team_procesed_file, "Team Member")
                
                # ---------------------- Save Removed Rows Data ----------------------

                save_removed_rows(removed_rows_team_df, removed_rows_team, "Removed Rows_Team")
                

            except Exception as e:
            
                print(f"\n    ❌ An error occurred: {e}")
            
            break # Exit loop if "yes" block executed successfully
        
        elif user_input == "no":
            print("\n    🛑 Skipping team member Sheet...")
            print("\n================================================================================")
            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 4: Processing Reporting code
    # =======================================================

    while True:
        print("\n================================================================================")
        print(f"\n📄 Do you want to run the Reporting codes Sheet? (yes/no)?: {strategy_choice}")
        user_input = strategy_choice
        if user_input == "yes":
            print("\n    ⏳ Running Reporting codes Sheet...")
            print("\n================================================================================")

            print("\n\n🔍 CREATING REPORTING CODES FILE")
            
            # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Reporting_codes'
            code_processed_file, removed_rows_codes = get_file_paths(sheet_name, output, removed_rows_dir,reporting_codes,'Removed_Rows - ReportingCodes.csv')
            
            # -------------------------------- Define Columns --------------------------------------
                      
            predefined_columns_Reportingcode = [ 'existing', 'concatcodes',"reporting_codes"]

            # Columns excluded from user deletion selection
            excluded_columns_codes = ['opportunityid','StrategyId']


            try:
                # ---------------------- Load Sheet ----------------------
                codes_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------

                codes_df = clean_dataframe(codes_df)
                
                #---------------------- Initialize Removed Rows df ----------------------
                
                removed_rows_code_df = pd.DataFrame(columns=codes_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------

                dropped_columns_codes = [] 
                
                # ---------------------- Remove all Rows If all Codes are not found ----------------------

                # Check if all StrategyId values are 'Not Found'
                def remove_if_strategy_not_found(df, removed_rows_df, removed_rows_path,reason,predefined_columns):                    
                    if 'StrategyId' in df.columns:
                        strategy_values = df['StrategyId'].astype(str).str.lower().dropna()

                        if not strategy_values.empty and strategy_values.nunique() == 1 and strategy_values.iloc[0] == 'not found':
                            df['Reason'] = reason
                            removed_rows_df = pd.concat([removed_rows_df, df], ignore_index=True)
                            removed_rows_df.drop(columns=[col for col in predefined_columns if col in removed_rows_df.columns], inplace=True)
                            # removed_rows_df.to_csv(removed_rows_path, index=False)

                            print(f"\n    ❗️ All StrategyId values are 'Not Found'. Sheet skipped and data moved to removed rows.")
                            return df, removed_rows_df, True  # Signal to break or skip
                    return df, removed_rows_df, False
                                
                codes_df,removed_rows_code_df,should_break =  remove_if_strategy_not_found(codes_df, removed_rows_code_df, removed_rows_codes,'Code not found',predefined_columns_Reportingcode)

                if should_break:
                    break
                
                # ---------------------- Remove rows where 'existing' is False ----------------------
                
                codes_df, removed_rows_code_df, rows_dropped_existing_count = remove_rows_based_on_condition(codes_df, removed_rows_code_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------


                codes_df,removed_rows_code_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, codes_df, removed_rows_code_df, "opportunityid")
                
                # ---------------------- Remove Predefined Columns ----------------------
                
                codes_df, dropped_columns_codes = delete_predefined_columns(codes_df, predefined_columns_Reportingcode, dropped_columns_codes)
       
                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------

                codes_df = show_column_deletion_gui(codes_df, excluded_columns_codes, dropped_columns_codes)

                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })      
            
                # ---------------------- Clean Removed Rows DF ----------------------                

                removed_rows_code_df = drop_dropped_columns_from_removed_rows(dropped_columns_codes,removed_rows_code_df)

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(codes_df, code_processed_file, "Reporting Codes")
                
                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_code_df, removed_rows_codes, "Removed Rows_Codes")

            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break
        elif user_input == "no":
            print("\n    🛑 Skipping Reporting Code sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 5: Processing Tags sheet
    # =======================================================

    while True:
        print("\n================================================================================")
        
        print(f"\n📄 Do you want to run the Tags Sheet? (yes/no)?: {strategy_choice}")
        
        user_input = strategy_choice # Automatically take user's earlier choice
        
        if user_input == "yes":
            print("\n    ⏳ Running Tags Sheet...")
            print("\n================================================================================")

            # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------
            
            sheet_name = 'Tags'
            tag_processed_file, removed_rows_tags = get_file_paths(sheet_name, output, removed_rows_dir,tags,'Removed_Rows - Tags.csv')
            
            # -------------------------------- Define Columns --------------------------------------
            
            predefined_columns_tags = ['existing', 'concattags','tag']
            
            excluded_columns_tags = ['opportunityid','StrategyId']


            try:
                print("\n\n🔍 CREATING TAGS FILE")

                # ---------------------- Load Sheet ----------------------

                tags_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------
                
                tags_df = clean_dataframe(tags_df)
                
                #---------------------- Initialize Removed Rows df ----------------------
                removed_rows_tag_df = pd.DataFrame(columns=tags_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------

                dropped_columns_tags = [] 

                # ---------------------- Remove all Rows If all Tags are not found ----------------------
                
                tags_df,removed_rows_tag_df,should_break =  remove_if_strategy_not_found(tags_df, removed_rows_tag_df, removed_rows_tags,'Tag not found',predefined_columns_tags)

                if should_break:
                    break
                
                # ---------------------- Remove rows where 'existing' is False ----------------------
                
                tags_df, removed_rows_tag_df, rows_dropped_existing_count = remove_rows_based_on_condition(tags_df, removed_rows_tag_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------
                
                tags_df,removed_rows_tag_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, tags_df, removed_rows_tag_df, "opportunityid")
                
                # ---------------------- Remove Predefined Columns ----------------------
                
                tags_df, dropped_columns_tags = delete_predefined_columns(tags_df, predefined_columns_tags, dropped_columns_tags)

                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------
                
                # Call the GUI function to show the column deletion options
                tags_df = show_column_deletion_gui(tags_df, excluded_columns_tags, dropped_columns_tags)

                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })      
                
                # ---------------------- Clean Removed Rows DF ----------------------
                
                removed_rows_tag_df = drop_dropped_columns_from_removed_rows(dropped_columns_tags,removed_rows_tag_df)    

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(tags_df, tag_processed_file, "Tags")
                
                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_tag_df, removed_rows_tags, "Removed Rows_Tags")


            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break

        elif user_input == "no":
            print("\n    🛑 Skipping Tag Sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")


    # =======================================================
    # Step 6: Processing Contact Role
    # =======================================================
    while True:
        print("\n================================================================================")
        print(f'\n📄 Do you want to run the Contact Role Sheet? (yes/no): {team_member_choice}')
        
        # Automatically using the user input captured earlier from file processing
        user_input = contact_choice 
        
        if user_input == "yes":
            print("\n    ⏳ Running Contact Role Member Sheet...")
            print("\n================================================================================")

            print("\n🔍 CREATING CONTACTS FILE")
            
            # ------------------------ Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Contact Roles'
            contact_processed_file, removed_rows_contact = get_file_paths(sheet_name, output, removed_rows_dir,Contact_role,'Removed_Rows - contact.csv')

            # Define hardcoded columns to delete from contact sheet

            predefined_columns_contact = ['existing']
            
            excluded_columns_contact = ["opportunityid","contactid","role"]

            
            try:
                # ---------------------- Load Product Sheet ----------------------

                contact_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------
                
                contact_df = clean_dataframe(contact_df)

                #---------------------- Initialize Removed Rows df ----------------------
                removed_rows_contact_df = pd.DataFrame(columns=contact_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------
                
                dropped_columns_contacts = []
                
                # ---------------------- Remove rows where 'existing' is False ----------------------

                contact_df, removed_rows_contact_df, rows_dropped_existing_count = remove_rows_based_on_condition(contact_df, removed_rows_contact_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------

                contact_df,removed_rows_contact_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, contact_df, removed_rows_contact_df, "opportunityid")

                # ---------------------- Remove Predefined Columns ----------------------
                
                contact_df, dropped_columns_contacts = delete_predefined_columns(contact_df, predefined_columns_contact, dropped_columns_contacts)

                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------
                
                contact_df = show_column_deletion_gui(contact_df, excluded_columns_contact, dropped_columns_contacts)
                
                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })   
                
                # ---------------------- Clean Removed Rows DF ----------------------   

                removed_rows_contact_df = drop_dropped_columns_from_removed_rows(dropped_columns_contacts,removed_rows_contact_df)    
            
                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(contact_df, contact_processed_file, "Contact")

                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_contact_df, removed_rows_contact, "Removed Rows_Contact")

        
            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
        
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
        
            break  # Exit loop if "yes" block executed successfully
        
        elif user_input == "no":
            print("\n    🛑 Skipping Contact Role Sheet...")
            print("\n================================================================================")
            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # ========================================================================
    # Last Step: Copy the Summary File to the Folder
    # ========================================================================

    print("\n\n🔍 Copying the Summary File to the Selected Folder...")

    # Check if the reference summary file exists at the specified location
    if not os.path.exists(REF_SUMMARY_FILE_PATH):
        print("\n    ❌ Error: Reference file does not exist at the specified path.")
        print(f"\n       📂 Path: {REF_SUMMARY_FILE_PATH}\n")
    else:
        try:
            # Extract the name of the selected folder (from the output path)
            selected_folder_name = os.path.basename(output)

            # Create the destination file path and rename the copied file accordingly
            destination_file_path = os.path.join(output, f"{selected_folder_name}_summary file.xlsx")

            # Copy the reference file to the destination location
            shutil.copy(REF_SUMMARY_FILE_PATH, destination_file_path)
            print(f"\n    ✅ Reference file copied successfully to the folder: {selected_folder_name}")

            # Load the copied file using openpyxl
            wb = load_workbook(destination_file_path)

            # Check if the 'Summary' sheet exists in the copied workbook
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]  # Access the "Summary" sheet

                # Write the selected folder name into cell D4
                ws['D4'] = selected_folder_name

                # Save the updated file (ensuring no other changes are made)
                wb.save(destination_file_path)
                print(f"\n    ✅ Folder name '{selected_folder_name}' written to cell D4 in the 'Summary' sheet.")
            else:
                print("\n    ❌ Error: 'Summary' sheet not found in the copied file.")
                print(f"\n       ❗️ Please check the file structure.")

        except Exception as e:
            print("\n    ❌ An error occurred while copying or modifying the file:")
            print(f"\n    ❗️ Details: {e}")


    # ========================================================================
    print("\n")
    title = "📝 FINAL SHEET COMPLETED 📝"
    show_title(title)
    # ========================================================================

    # =====================================================
    # Delete CSV Files
    # =====================================================
    # Hardcoded directory


    print("\n\n🔍 Delete the extract files")

    response = 'yes'
    # response = 'no'
    if response == 'yes':
        # Get list of files in the DOWNLOADS_DIR
        def delete_folder(folder_path):
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
                print(f"\n        🗑️ Folder '{folder_path.split('/')[-1]}' and its contents have been deleted.")
            else:
                print(f"\n        ❗️ Folder '{folder_path}' does not exist.")


        delete_folder('Extracts')
        delete_folder('Delete')
    else:
        print("\n        🛑 No files were deleted.")

    print("\n")
    title = f"✅ File Prepared: {filename} ✅"
    show_title(title)
    
    file_end_time = time.time()   # Record start time
    elapsed_time = file_end_time - file_start_time
    print(f"\n    ✅ Total time taken: {elapsed_time:.2f} seconds")

    files_in_copy_folder.remove(files_in_copy_folder[selected_index])

    if files_in_copy_folder:
        while True:  # Inner loop
            continue_processing = input("\n 👉 Do you want to process another file? (yes/no): ").strip().lower()
            
            if continue_processing == 'yes':
                # Add logic for processing another file here
                print("\n    ⏳ Processing the file...")
                break  # Exit the inner loop and continue to the next iteration of the outer loop
            
            elif continue_processing == 'no':
                print(f"\n     🔚 End of Script\n")
                print("=" * 100)
                print("\n")
                break  # Exit the inner loop
            else:
                print("\n    ❗️ Invalid input. Please select 'yes' or 'no'.")
        
        if continue_processing == 'no':
            break  # Exit the outer loop if the user selects 'no'
    else:
        print(f"\n     🔚 End of Script\n")
        print("=" * 100)
        print("\n")
        end_time = time.time()   # Record start time
        elapsed_time = end_time - start_time
        print(f"\n    ✅ Total time taken for {count_of_files_copy_folder} files: {elapsed_time:.2f} seconds\n")

        break  # Exit the inner loop
