import os
import re
import time
import shutil
import openpyxl
import pyperclip
import itertools
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook

# ==================================================
# Function to display title
def show_title(title):

    line_width = 100
    line = "=" * line_width
    print(f"\n{line}")
    print(title.center(line_width))
    print(f"{line}\n")

title = "📝  Extract Data and create Queries 📝"

show_title(title)
# ==================================================

# Directory where extracted files will be saved
EXTRACT_FOLDER = "Extracted_Files"

if os.path.exists(EXTRACT_FOLDER):
    shutil.rmtree(EXTRACT_FOLDER)
# Create the folder if it doesn't exist
os.makedirs(EXTRACT_FOLDER)

# Path where source Excel files are located
DOWNLOAD_DIR = os.path.expanduser("~/Downloads")
# Path for the consolidated extracted data
EXTRACT_OUTPUT_FILE = "Extracted_Files/Extracted_data.xlsx"


print("\n🔍 Files to Process")
for file in os.listdir(DOWNLOAD_DIR):
    if file.endswith(".xlsx"):
        file_path = os.path.join(DOWNLOAD_DIR, file)
        xls = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
        if 'Opportunity' in xls or 'Opportunity_product' in xls:
            print(f"\n       📁 {file}")
        else:
            print(f"\n       ❌ {file}")

print("\n🔍 Step 1: Extract Data from Files:")

# Loop through each file in the source directory
for file in os.listdir(DOWNLOAD_DIR):
    if file.endswith(".xlsx"):
        file_path = os.path.join(DOWNLOAD_DIR, file)

        wb = openpyxl.load_workbook(file_path)
        variant_mapping = {
            'Opportunity_products': 'Opportunity_product',
            'Opportunity_Team': 'Opportunity_Team '  # trailing space is intended here
        }

        # Iterate through the workbook sheets and rename if there's a variant
        for sheet_name in wb.sheetnames:
            if sheet_name in variant_mapping:
                ws = wb[sheet_name]
                correct_name = variant_mapping[sheet_name]
                ws.title = correct_name
                print(f"\n    🔄 Renamed '{sheet_name}' to '{correct_name}' automatically.")
        # Save the workbook with the new sheet names
        wb.save(file_path)

        # Load all sheets from the current Excel file
        xls = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')

        collected_data = {}

        # Process 'Opportunity_product' sheet if it exists
        if 'Opportunity_product' in xls:
            df = xls['Opportunity_product']
            df['product_family'] = df['Product'].astype(str) + '-' + df['product_type'].astype(str)

        # Define how columns should be extracted and grouped from various sheets
        sheet_config = {
            "Opportunity": {
                "columns": ["accountid", "ownerid", "created_by", "currency_code", "opportunity_legacy_id_c"],
                "output": {
                    "Accounts": ["accountid"],
                    "Email_id": ["ownerid", "created_by"],
                    "Currency": ["currency_code"],
                    "Legacy_Ids": ["opportunity_legacy_id_c"]
                }
            },
            "Opportunity_product": {
                "columns": ["product_family"],
                "output": {
                    "Product": ["product_family"]
                }
            },
            "Opportunity_Team ": {
                "columns": ["Email"],
                "output": {
                    "Email_id": ["Email"]
                }
            }
        }

        # Process all other sheets as per config
        for sheet_key, config in sheet_config.items():
            if sheet_key not in xls:
                # print(f"\n   ❗️'{file}' does not have correct format.Skipping the file.")
                continue

            df = xls[sheet_key]
            df.columns = df.columns.str.strip()  # Clean column names

            for out_sheet, needed_cols in config["output"].items():
                valid_cols = [col for col in needed_cols if col in df.columns]
                if not valid_cols:
                    continue

                if out_sheet == "Email_id":
                    stacked = pd.concat([df[col].dropna().astype(str) for col in valid_cols], ignore_index=True).to_frame(name="Email_id")
                    collected_data.setdefault(out_sheet, []).append(stacked)

                else:
                    subset = df[valid_cols].dropna(how='all')
                    if not subset.empty:
                        collected_data.setdefault(out_sheet, []).append(subset)

        # Special handling for 'Reporting_codes' to extract 'Tags'
        if 'Reporting_codes' in xls:
            df = xls['Reporting_codes']
            df.columns = df.columns.str.strip()  # Clean column names

            tags_columns = [col for col in df.columns if col.lower() == 'tags']
            reporting_codes_columns = [col for col in df.columns if col.lower() == 'reporting_codes']

            if reporting_codes_columns:
                reporting_codes_column = reporting_codes_columns[0]

                # If 'Tags' doesn't exist, create it using 'reporting_codes'
                if not tags_columns:
                    df['Tags'] = df[reporting_codes_column]
                    tags_column = 'Tags'
                else:
                    tags_column = tags_columns[0]
                    # If 'Tags' exists but is all NaN or empty strings
                    if df[tags_column].isna().all() or (df[tags_column].astype(str).str.strip() == '').all():
                        df[tags_column] = df[reporting_codes_column]

                # Now extract 'Tags' values (drop NaNs and blanks)
                strategy_data = df[tags_column].dropna().astype(str).str.strip()
                strategy_data = strategy_data[strategy_data != '']  # remove empty strings

                if not strategy_data.empty:
                    strategy_df = strategy_data.to_frame(name='Strategy')
                    collected_data.setdefault('Strategy', []).append(strategy_df)

        # Append or create a new extracted data Excel file
        file_exists = os.path.exists(EXTRACT_OUTPUT_FILE)

        with pd.ExcelWriter(EXTRACT_OUTPUT_FILE, engine="openpyxl", mode="a" if file_exists else "w", if_sheet_exists="overlay" if file_exists else None) as writer:
            for sheet_name, dfs in collected_data.items():
                combined = pd.concat(dfs, ignore_index=True)

                if file_exists:
                    try:
                        existing = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=sheet_name)
                        combined = pd.concat([existing, combined], ignore_index=True)
                    except Exception:
                        pass  # Sheet doesn't exist, just write new

                combined.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"\n   ✅ Data Extracted and stored in {EXTRACT_OUTPUT_FILE}:")

# ============================
# Trim whitespace in all cells and column names
# ============================

print("\n🔍 Step 2: Cleaning data")

def trim_columns_all_sheets(EXTRACT_OUTPUT_FILE):
    # Read all sheets into a dictionary
    sheets = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=None, dtype=str)  # Read all as strings to ensure trimming works
    
    trimmed_sheets = {}

    for sheet_name, df in sheets.items():
        new_df = pd.DataFrame()
        
        for column in df.columns:
            new_column_name = column.strip()  # Also trim column names
            if column.lower().strip() == 'accountid':
                new_df[new_column_name] = df[column].astype(str).str.replace(r'\s+', '', regex=True).str.strip()
            else:
                new_df[new_column_name] = df[column].astype(str).str.strip()
        
        trimmed_sheets[sheet_name] = new_df
    
    # Save the trimmed data back to a new Excel file
    with pd.ExcelWriter(EXTRACT_OUTPUT_FILE, engine='openpyxl') as writer:
        for sheet_name, df in trimmed_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


trim_columns_all_sheets(EXTRACT_OUTPUT_FILE)

print("\n   ✅ Trimmed Data")

# ============================
# Expand comma-separated values into separate rows via cartesian product
# ============================

# Load the Excel file
sheets_to_process = ['Email_id', 'Strategy']

# Read all sheets
all_sheets = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=None)

# Process the Email_id and Strategy sheets
for sheet_name in sheets_to_process:
    if sheet_name in all_sheets:
        df = all_sheets[sheet_name]
        new_rows = []

        for _, row in df.iterrows():
            # Split values in each cell by comma, strip whitespace
            split_row = [str(cell).split(',') if isinstance(cell, str) and ',' in cell else [cell] for cell in row]
            split_row = [[val.strip() for val in values] for values in split_row]

            # Create cartesian product (every combination from split lists)
            for combination in itertools.product(*split_row):
                new_rows.append(combination)

        # Create a new DataFrame from the expanded rows
        all_sheets[sheet_name] = pd.DataFrame(new_rows, columns=df.columns)
    else:
        print(f"Sheet '{sheet_name}' not found in the Excel file.")

# Write all sheets back to the same Excel file
with pd.ExcelWriter(EXTRACT_OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
    for sheet_name, df in all_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n   ✅ Separated Values based on Comma")

# ============================
# Process Account IDs - strip extra parts from values starting with 'DC'
# ============================

accountid_column = 'accountid'
new_column_name = 'accountid'

# Load the specific sheet into a DataFrame
df = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name="Accounts")

# Define a function to process the values
def process_value(value):
    if isinstance(value, str) and value.startswith('DC'):
        return value.split('-')[0]
    return value

# Apply the function to the accountid column and store results in the new column
df[new_column_name] = df[accountid_column].apply(process_value)

# Save the updated DataFrame back to the Excel file
with pd.ExcelWriter(EXTRACT_OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='Accounts', index=False)

print("\n   ✅ Formated Account values to the correct Format")

# ============================
# Remove duplicates from each sheet based on the first column
# ============================

# Load all sheets as a dict of DataFrames
xls = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=None)

# Remove duplicates in each sheet
for sheet_name, df in xls.items():
    # Check if there's at least one column
    if not df.empty:
        first_col = df.columns[0]
        df = df.drop_duplicates(subset=first_col)
        xls[sheet_name] = df

# Save back to the same file
with pd.ExcelWriter(EXTRACT_OUTPUT_FILE, engine='openpyxl', mode='w') as writer:
    for sheet_name, df in xls.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n   ✅ Removed Duplicates")

# ============================
# Concatenate each cell value with quotes and a comma for SQL formatting
# ============================

print("\n🔍 Step 3: Create file with concatenated Data")
xls = pd.ExcelFile(EXTRACT_OUTPUT_FILE)

# Define the output file path to avoid overwriting the original
CONCATENATE_OUTPUT_FILE = 'Extracted_Files/Concatenated_excel_file.xlsx'

# Create an Excel writer to save the modified data
with pd.ExcelWriter(CONCATENATE_OUTPUT_FILE, engine='openpyxl') as writer:
    # Iterate through all sheets
    for sheet_name in xls.sheet_names:
        # Load the sheet into a DataFrame
        df = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Iterate through each column in the DataFrame
        for column in df.columns:
            # Concatenate each value with single quotes and a comma
            df[column] = df[column].apply(lambda x: f"'{x}',")
        
        # Save the modified DataFrame to the new Excel file
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"\n   ✅ Concated values saved in {CONCATENATE_OUTPUT_FILE}")

# ============================
# Function to generate a query file from a single sheet column
# ============================

print("\n🔍 Step 4: Generating Queries")
def generate_query_from_sheet(EXTRACT_OUTPUT_FILE, sheet_name, column_name, query_template, output_txt_base):
    # Read specific sheet
    df = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=sheet_name, dtype=str)

    # Drop empty values and clean
    values = df[column_name].dropna().astype(str).str.strip().tolist()

    # Query parts length
    base_query_length = len(query_template.replace("{values}", ""))
    max_query_length = 80000

    # Estimate max value chunk size per query
    value_strings = [f"'{val}'" for val in values]

    chunks = []
    current_chunk = []
    current_length = base_query_length

    for val in value_strings:
        val_length = len(val) + 1  # for comma and newline
        if current_length + val_length > max_query_length:
            # Save current chunk
            chunks.append(current_chunk)
            # Start new chunk
            current_chunk = [val]
            current_length = base_query_length + val_length
        else:
            current_chunk.append(val)
            current_length += val_length

    if current_chunk:
        chunks.append(current_chunk)

    # Ensure output folder exists
    os.makedirs(os.path.dirname(output_txt_base), exist_ok=True)

    # Write each query to separate file
    for idx, chunk in enumerate(chunks, start=1):
        formatted_values = ",\n".join(chunk)
        query = query_template.replace("{values}", formatted_values)
        if len(chunks) == 1:
                output_txt = output_txt_base
        else:
            output_txt = output_txt_base.replace(".txt", f"_part{idx}.txt")
        with open(output_txt, "w") as file:
            file.write(query)
    
    query_file = output_txt_base.split("/")[-1]
    print(f"\n       📁 Generated {len(chunks)} query file(s) for {query_file}.")

# Queries generation (various sheets & templates)

account_output_txt = 'Extracted_Files/Queries/1_Account_query.txt'
account_query = "SELECT AccountNumber, id FROM Account WHERE AccountNumber IN ({values})"
generate_query_from_sheet(EXTRACT_OUTPUT_FILE, 'Accounts', 'accountid', account_query,account_output_txt)

Email_output_txt = 'Extracted_Files/Queries/2_Userid_query.txt'
userid_query = "select email,id,Profile.Name,isactive from user where email in ({values}) and Profile.Name != 'IBM Partner Community Login User' and IsActive = true"
generate_query_from_sheet(EXTRACT_OUTPUT_FILE, 'Email_id', 'Email_id', userid_query,Email_output_txt)

strategy_output_txt = 'Extracted_Files/Queries/4_Strategy_query.txt'
strategy_query = "Select id,name,Record_Type_Name__c from Strategy__c where name in ({values})"
generate_query_from_sheet(EXTRACT_OUTPUT_FILE, 'Strategy', 'Strategy', strategy_query,strategy_output_txt)

legacy_output_txt = 'Extracted_Files/Queries/5_Legacy_query.txt'
legacy_query = "SELECT Opportunity_Legacy_Id__c, Id,Name,Owned_By_Name__c,OwnerId FROM Opportunity WHERE Opportunity_Legacy_Id__c IN ({values})"
generate_query_from_sheet(EXTRACT_OUTPUT_FILE, 'Legacy_Ids', 'opportunity_legacy_id_c', legacy_query,legacy_output_txt)

# ============================
# Function to generate query using two sheets and columns
# ============================

def generate_query_from_two_sheets(EXTRACT_OUTPUT_FILE, sheet1, column1, sheet2, column2, query_template, output_txt):
    # Read both sheets
    df1 = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=sheet1, dtype=str)
    df2 = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name=sheet2, dtype=str)

    # Clean and dropna
    values1 = df1[column1].dropna().astype(str).str.strip().tolist()
    values2 = df2[column2].dropna().astype(str).str.strip().tolist()

    # Format values for SQL IN clause
    formatted_values1 = ",\n".join([f"'{val}'" for val in values1])
    formatted_values2 = ",\n".join([f"'{val}'" for val in values2])

    # Replace placeholders in query template
    query = query_template.replace("{product}", formatted_values1).replace("{currency}", formatted_values2)

    # Ensure output folder exists
    os.makedirs(os.path.dirname(output_txt), exist_ok=True)

    # Write query to file
    with open(output_txt, "w") as file:
        file.write(query)


# Product & currency query generation
sheet1 = 'Product'
column1 = 'product_family'
sheet2 = 'Currency'
column2 = 'currency_code'
product_query = "SELECT Product2.Product_Code_Family__c, CurrencyIsoCode, id, isactive FROM PricebookEntry WHERE Product2.Product_Code_Family__c IN ({product}) AND CurrencyIsoCode IN ({currency})"
product_output_txt = 'Extracted_Files/Queries/3_PricebookEntry_query.txt'

generate_query_from_two_sheets(EXTRACT_OUTPUT_FILE, sheet1, column1, sheet2, column2, product_query, product_output_txt)

print("\n   ✅ Queries Generated")

# ============================
# ============================

file_mapping = {
    "1_Account_query.txt": "accounts.csv",
    "2_Userid_query.txt": "userid.csv",
    "3_PricebookEntry_query.txt": "productfamily.csv",
    "4_Strategy_query.txt": "tags.csv",
    "5_Legacy_query.txt":"legacyid.csv"
}
# ============================
# Function to find & rename 'bulkQuery' CSV file
# ============================
def wait_and_rename_bulkquery_file(new_name, dir_path, part_number=None, timeout=60):
    # print(f"\n⏳ Waiting for 'bulkQuery' file to appear in {dir_path}... (timeout: {timeout} seconds)")
    
    start_time = time.time()
    while time.time() - start_time < timeout:
        for filename in os.listdir(dir_path):
            if "bulkQuery" in filename and filename.endswith(".csv"):
                old_path = os.path.join(dir_path, filename)
                
                # If part number is present, adjust new_name
                final_name = new_name.replace(".csv", f"{part_number}.csv") if part_number else new_name
                new_path = os.path.join(dir_path, final_name)

                if os.path.exists(new_path):
                    backup_path = new_path.replace(".csv", f"_backup_{int(time.time())}.csv")
                    os.rename(new_path, backup_path)
                    print(f"\n   ⚠️ Existing file '{final_name}' backed up as '{os.path.basename(backup_path)}'")

                shutil.move(old_path, new_path)
                print(f"\n   ✅ File renamed to '{final_name}'")

                return True
        time.sleep(2)
    print("\n   ❌ Timeout: No 'bulkQuery' file found. Please Rename it manually")
    return False


# ==================================================
title = "📝 Files Processed 📝"
show_title(title)
# ==================================================

break_outer_loop = False

while not break_outer_loop:
    Copy_query = input("\n📝 Do you want to proceed With Copying the Queries? (yes/no): ").strip().lower()

    if Copy_query == 'yes':
        query_dir = os.path.join(EXTRACT_FOLDER, "Queries")
        query_files = sorted([f for f in os.listdir(query_dir) if f.endswith(".txt")])  # sorted alphabetically

        if not query_files:
            print("\n❌ No query files found.")
            break

        print("\n📑 Available Query Files:")
        for idx, file_name in enumerate(query_files, start=1):
            print(f"\n   {idx}) {file_name}")

        while True:
            choose = input('\n📝 Enter the number of the query you want to copy or type "s" to Skip: ').strip()

            if choose.lower() == 's':
                print("\n    🚫 Skipping Copying Queries")
                break_outer_loop = True
                break

            if choose.isdigit() and 1 <= int(choose) <= len(query_files):
                selected_file = query_files[int(choose) - 1]
                with open(os.path.join(query_dir, selected_file), "r") as f:
                    query = f.read()

                print(f"\n   ✅ {selected_file} copied to clipboard")
                pyperclip.copy(query)
                
                # Remove _partX if present
                base_filename = selected_file.split("_part")[0] + ".txt" if "_part" in selected_file else selected_file

                new_csv_name = file_mapping.get(base_filename)

                # Extract part number if any
                part_number = None
                if "_part" in selected_file:
                    part_number = selected_file.split("_part")[-1].split(".txt")[0]
                if new_csv_name:
                    print(f"\n   📥 Please download extract — it will automatically Renamed to '{new_csv_name}' when it appears.")
                    wait_and_rename_bulkquery_file(new_csv_name, DOWNLOAD_DIR, part_number)
                else:
                    print("\n   ⚠️ No mapping found for this query file. Skipping rename.")

            else:
                print("\n   ❗️ Invalid Selection")
                continue

    elif Copy_query == 'no':
        print("\n   🚫 Skipping Copying Queries")
        break

    else:
        print("\n   ❗️ Invalid Choice")


title = "📝  Merge CSV Files 📝"
show_title(title)
while True:

    merge = input("\n📝 Do you want to merge CSV?(yes/no): ").strip().lower()

    if merge == 'yes':

        # Function to select multiple CSV files
        def select_files(title):
            root = tk.Tk()
            root.withdraw()
            file_paths = filedialog.askopenfilenames(title=title, filetypes=[("CSV files", "*.csv")])

            if file_paths:
                return file_paths
            else:
                print("\n   ❌ No files selected.")
                return None

        # Function to merge multiple CSV files
        def merge_csv_files(file_paths, output_file):
            merged_df = pd.DataFrame()  # Empty DataFrame to start with

            for file_path in file_paths:
                df = pd.read_csv(file_path)
                merged_df = pd.concat([merged_df, df], ignore_index=True)

            merged_df.to_csv(output_file, index=False)
            print(f"\n   ✅ Files merged successfully into: {output_file}")

        print("\n🔍 Select the csv files to merge")
        # Main flow
        file_paths = select_files("Select CSV files to merge")

        print("\n🔍 Name of new Csv File")
        if file_paths:
            output_name = input("\n📄 Enter the name for the merged CSV file (without .csv extension): ").strip()
            if not output_name:
                output_name = "Merged_file"  # Default name if left blank

            output_file = os.path.expanduser(f"~/Downloads/{output_name}.csv")

            merge_csv_files(file_paths, output_file)
        else:
            print("\n   ⚠️ No files selected. Exiting.")
        
        break

    elif merge == 'no':
        print("\n    🚫 Skipping Merging CSV Files")
        break
    else:   
        print("\n    ❗️ Invalid Choice")    

# ======================

title = "📝  Missing Accounts and Tags 📝"
show_title(title)

while True:

    vlookup = input("\n📝 Do you want to check For Accounts and Tags Missing?(yes/no): ").strip().lower()

    if vlookup == 'yes':

        # Load the Excel sheet with account IDs
        accounts_df = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name='Accounts')

        # Load the CSV file with AccountNumber and Id
        csv_file_path = os.path.expanduser("~/Downloads/accounts.csv")  # replace with your CSV file path
        lookup_df = pd.read_csv(csv_file_path)

        # Convert both to lowercase for case-insensitive matching
        accounts_df['accountid_lower'] = accounts_df['accountid'].astype(str).str.lower()
        lookup_df['AccountNumber_lower'] = lookup_df['AccountNumber'].astype(str).str.lower()

        # Perform the left join like VLOOKUP
        merged_df = pd.merge(
            accounts_df,
            lookup_df[['AccountNumber_lower', 'Id']],
            left_on='accountid_lower',
            right_on='AccountNumber_lower',
            how='left'
        )

        # Fill NaN values with 'Not found in ISC'
        merged_df['Id'] = merged_df['Id'].fillna('Not found in ISC')

        # Optional: drop helper columns if not needed
        merged_df.drop(columns=['accountid_lower', 'AccountNumber_lower'], inplace=True)

        # Save to a new Excel file
        ACCOUNT_VLOOKUP_FILE = 'Extracted_Files/Accounts_vlookup.xlsx'
        merged_df.to_excel(ACCOUNT_VLOOKUP_FILE, index=False)

        # print(f"\n    ✅ VLOOKUP completed for Accounts.")

        try:

            # Load the Excel file's "Strategy" sheet
            strategy_df = pd.read_excel(EXTRACT_OUTPUT_FILE, sheet_name='Strategy')

            # Load the tags CSV file
            tags_csv_path = os.path.expanduser("~/Downloads/tags.csv")  # replace with your tags CSV file path
            tags_df = pd.read_csv(tags_csv_path)

            # Convert both columns to lowercase for case-insensitive matching
            strategy_df['Strategy_lower'] = strategy_df['Strategy'].astype(str).str.lower()
            tags_df['Name_lower'] = tags_df['Name'].astype(str).str.lower()

            # Perform left join (like VLOOKUP)
            merged_strategy_df = pd.merge(
                strategy_df,
                tags_df[['Name_lower', 'Id']],
                left_on='Strategy_lower',
                right_on='Name_lower',
                how='left'
            )

            # Fill NaN values with 'Not found in ISC'
            merged_strategy_df['Id'] = merged_strategy_df['Id'].fillna('Not found in ISC')

            # Drop helper columns if you don't want them in final output
            merged_strategy_df.drop(columns=['Strategy_lower', 'Name_lower'], inplace=True)

            # Save to a new Excel file
            TAG_VLOOKUP_FILE = 'Extracted_Files/tags_vlookup.xlsx'
            merged_strategy_df.to_excel(TAG_VLOOKUP_FILE, index=False)

            # print(f"\n    ✅ VLOOKUP for Strategy completed.")

        except FileNotFoundError:
            print(f"\n    ⚠️ tags.csv not found — skipping tags VLOOKUP.")
        except Exception as e:
            print(f"\n    ⚠️ Error during tags VLOOKUP: {e}")

        # Load the Excel file
        excel_file_path = 'Extracted_Files/Accounts_vlookup.xlsx'  # replace with your file path
        df = pd.read_excel(excel_file_path)

        # Filter rows where Id is 'Not found in ISC'
        not_found_df = df[df['Id'] == 'Not found in ISC']

        # Select only the 'accountid' column
        accountids_not_found = not_found_df[['accountid']]

        # Save to a new Excel file
        ACCOUNT_TO_IMPORT = os.path.expanduser("~/Downloads")+'/Accounts_to_be_imported.xlsx'
        accountids_not_found.to_excel(ACCOUNT_TO_IMPORT, index=False)

        # print(f"\n    ✅ Missing accountids saved to {ACCOUNT_TO_IMPORT}")
        try:
            # Load the Excel file
            excel_file_path = 'Extracted_Files/tags_vlookup.xlsx'  # replace with your file path
            df = pd.read_excel(excel_file_path)

            # Filter rows where Id is 'Not found in ISC'
            not_found_df = df[df['Id'] == 'Not found in ISC']

            # Select only the 'Strategy' column
            Strategy_not_found = not_found_df[['Strategy']]

            # Create a DataFrame with the desired structure, similar to Code 1
            output_df = pd.DataFrame({
                'Name': Strategy_not_found['Strategy'],
                'Strategy_Full_Name__c': '',
                'RecordTypeId': '0123h000000kqchAAA',
                'Record_Type_Name__c': 'Tags',
                'IsDeleted': False,
                'Active__c': True
            })

            # Save to a new Excel file
            if not output_df.empty:
                csv_output_path = os.path.expanduser("~/Downloads/Tags_to_be_inserted.csv")
                output_df.to_csv(csv_output_path, index=False)
                # print(f"\n    ❗️ Missing tags saved to {csv_output_path}")

        except FileNotFoundError:
            print(f"\n    ⚠️ tags_vlookup.xlsx not found — skipping missing tags export.")
        except Exception as e:
            print(f"\n    ⚠️ Error while handling tags_vlookup.xlsx: {e}")


        # Load the original Excel file
        ACCOUNT_TO_IMPORT = os.path.expanduser("~/Downloads/Accounts_to_be_imported.xlsx")
        df = pd.read_excel(ACCOUNT_TO_IMPORT)
        # Initialize lists to store valid and invalid values
        invalid_values = []
        valid_values = []

        if not df.empty:
            # Define a regular expression for country codes (e.g., "-US", "-KA", etc.)
            country_code_pattern = r'-[A-Za-z]{2,3}$'

            # Check each value in the column (assuming the values are in the first column)
            for value in df.iloc[:, 0]:
                value = str(value).strip()  # Ensure it's a string and remove leading/trailing spaces
                if not (value.lower().startswith('db') or value.lower().startswith('dc')):
                    # If it doesn't start with DB or DC (case insensitive), it's invalid
                    invalid_values.append(value)
                elif value.lower().startswith('db'):
                    # If it starts with DB or db, it should have a country code
                    if not re.search(country_code_pattern, value):
                        invalid_values.append(value)
                    else:
                        valid_values.append(value)  # Add to valid values list
                else:
                    valid_values.append(value)  # Add to valid values list for DC or other valid entries

        # Count invalid values
        invalid_count = len(invalid_values)
        valid_count = len(valid_values)

        # If there are invalid values, write them to a new Excel file
        if invalid_count > 0:
            invalid_df = pd.DataFrame(invalid_values, columns=['Invalid Accounts'])
            invalid_df.to_excel(os.path.expanduser("~/Downloads/Invalid_Accounts.xlsx"), index=False)


        # Update the original dataframe with only valid values
        valid_df = pd.DataFrame(valid_values, columns=['Accounts'])

        # Save the updated dataframe back to the original file
        valid_df.to_excel(ACCOUNT_TO_IMPORT, index=False)

        # ===========================================
        # ===========================================
        
        account_list_df = pd.read_csv(csv_file_path)
        account_numbers_set = set(account_list_df['AccountNumber'].astype(str).str.strip())
        
        # Summary stats
        total_valid = 0
        total_invalid = 0
        files_with_no_matches = []

        # Process each Excel file in the folder
        for file_name in os.listdir(DOWNLOAD_DIR):
            if file_name.endswith('.xlsx'):
                file_path = os.path.join(DOWNLOAD_DIR, file_name)
                try:
                    df = pd.read_excel(file_path, sheet_name='Opportunity')
                except Exception as e:
                    continue

                if 'accountid' not in df.columns:
                    continue
                # Clean and process accountid
                df['accountid'] = df['accountid'].astype(str).str.replace(r'\s+', '', regex=True).str.strip()
                df['accountid'] = df['accountid'].apply(process_value)

                # Find values not in CSV
                not_in_csv = df[~df['accountid'].isin(account_numbers_set)]['accountid'].dropna().unique()

                file_valid = []
                file_invalid = []

                for value in not_in_csv:
                    value = str(value).strip()
                    if not (value.lower().startswith('db') or value.lower().startswith('dc')):
                        file_invalid.append(value)
                    elif value.lower().startswith('db'):
                        if not re.search(country_code_pattern, value):
                            file_invalid.append(value)
                        else:
                            file_valid.append(value)
                    else:
                        file_valid.append(value)

                # Track if file had no matches
                if not file_valid and not file_invalid:
                    files_with_no_matches.append(file_name)
                        # Append to cumulative DataFrames
                if file_valid:
                    valid_df = pd.concat([valid_df, pd.DataFrame(file_valid, columns=['Accounts'])], ignore_index=True)
                    total_valid += len(file_valid)
                if file_invalid:
                    invalid_df = pd.concat([invalid_df, pd.DataFrame(file_invalid, columns=['Invalid Accounts'])], ignore_index=True)
                    total_invalid += len(file_invalid)
                
        # Final summary
        print("\n   ✅ Processing Complete!")
        print(f"\n       ❗️ Total tags to be inserted: {len(Strategy_not_found)}")
        print(f"\n       ❗️ Total Accounts to Be Imported: {valid_count}")
        print(f"\n       ❗️ Invalid Accounts: {invalid_count}")
        if files_with_no_matches:
            print("\n    ✅ Files with No Missing Accounts:")
            for fname in files_with_no_matches:
                print(f"\n       📄 {fname}")
        else:
            print("\n    ✅ All files had some valid or invalid accounts.\n")
 
        break

    elif vlookup == 'no' :
        print('\n   🚫 Skipping ') 
        break

    else:
        print('\n   ❗️ Invalid Choice')

import shutil

# Source and destination file paths
source_file = '/Users/avirajmore/Downloads/userid.csv'
destination_file = '/Users/avirajmore/Downloads/teammember.csv'

# Check if source file exists
if os.path.exists(source_file):
    shutil.copy(source_file, destination_file)

print("\n👋 Exiting the script. Goodbye!")
title = "📝  Script Completed 📝"
show_title(title)