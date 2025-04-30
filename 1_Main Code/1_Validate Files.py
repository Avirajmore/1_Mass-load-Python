import os
import pandas as pd
from tabulate import tabulate
import openpyxl

# ------------------------------------ Function to Display Section Titles Nicely ------------------------------------

def show_title(title):

    line_width = 100
    line = "=" * line_width
    print(f"\n{line}")
    print(title.center(line_width))
    print(f"{line}\n")

# ------------------------------------ Start Program ------------------------------------

# Display the main title
title = "📝  File Data Validation 📝"
show_title(title)

# Define the folder path to look for Excel files
folder_path = os.path.expanduser("~/Downloads")

# Create a list to store file validation summaries
summary_list = []

# ------------------------------------ Show Available Excel Files ------------------------------------

print(f"\n🔍 Files Available: ")
for file in os.listdir(folder_path):
    if file.endswith(".xlsx"):
        file_path = os.path.join(folder_path, file)
        print(f"\n    ✅ {file}")

# ------------------------------------ Start Validation Process for Each Excel File ------------------------------------


for file in os.listdir(folder_path):
    if file.endswith(".xlsx"):
        
        file_path = os.path.join(folder_path,file)
        
        # Show current file being processed 
        title = f"✅ {file} ✅"
        show_title(title)

        # Initialize status for the file
        file_status = "✅ All Good"
        
        # ------------------------------------ Setup Validation Parameters ------------------------------------

        # --- List of Valid API Names ---
        # Used to validate values in specific columns (e.g., sales stages, types, etc.)
        valid_api_names = [
            'Engage', 'Qualify', 'Design', 'Propose', 'Negotiate', 'Closing', 'Prospecting', 'Developing', 'Negotiation', 'Won', 'Lost',
            'RNL_MAINT_ADMIN', 'RNL_MAINT_NOPAY', 'RNL_MAINT_CATCH', 'RNL_MAINT_ENDSUPP', 'RNL_MAINT_OUTPROD',
            'RNL_MAINT_REPLACE', 'RNL_MAINT_SELFSERV', 'RNL_MAINT_THIRD', 'RNL_COMP_PRICE', 'RNL_CNTR_EARLY',
            'RNL_CNTR_MOVE', 'RNL_CUST_NORESP', 'RNL_CUST_NOBDGT', 'RNL_CUST_OUTOFBUS', 'RNL_CUST_PROJEND',
            'RNL_CUST_TEMP', 'RNL_IBM_PRODCHNG', 'RNL_IBM_NOPROD', 'RNL_PROD_NOSUPP', 'RNL_PROD_NOTRELB',
            'RNL_PROD_NOSOLN', 'RNL_REVN_MOVETERM', 'RNL_REVN_MOVESAAS', 'RNL_REVN_MOVEPERP', 'RNL_REVN_MOVESUBS',
            'RNL_REVN_MOVEPERS', 'RNL_REVN_MOVETYPE', 'RNL_REVN_MOVESALE', 'Business Partner', 'Digital Sales',
            'Field Sales', 'Digital Sales Development', 'TYPEESA', 'TYPEOEM', 'TYPEPCR', 'TYPERFS', 'TYPESRV',
            'TYPEASP', 'TYPEEST', 'TYPESSP', 'TYPEMNL', 'TYPERBK', 'TYPEZSW', 'TYPEESP', 'TYPEASE', 'TYPEPLN',
            'TYPEINS', 'TYPEBLD', 'TYPEMIG', 'TYPEPER', 'TYPEESM', 'TYPEEXC', 'TYPELSU', 'TYPELIN', 'TYPECUS',
            'CLASSNEW', 'CLASSEXP', 'CLASSREN', 'CLASSUPG', 'CLASSDEP', 'CLASSREI', 'CLASSWEX', 'RENEW_QUALIFY',
            'RENEW_DESIGN', 'RENEW_PROPOSE', 'RENEW_NEGOTIATE', 'RENEW_CLOSING', 'RENEW_LOST', 'RENEW_WON',
            'AUTOREN_ORIG', 'AUTOREN_12M', 'AUTOREN_24M', 'AUTOREN_36M', 'TERMINATE', 'Bill CONTINUOUS', 'NOTAPPL',
            'AUTOREN_SUBNEW', 'AUTOREN_SUBUPG', 'TERMINATE_SUBNEW', 'TERMINATE_SUBUPG','SW','SWSUBSCR','SWSVC','SAAS','MAINT',
            'WONSKILL',	'WONBRAND',	'WONPRICE',	'WONINC',	'WONTERMS',	'WONRFP',	'WONISV',	'WONEXEC',	'WONNEEDS',	'WONFIN',
            'LCDNP',	'LIBMDNP',	'LLTC',	'LOSTTLS',	'LOSTFIN',	'LOSTREN',	'FINPART','IBMRESC',	'FINBP', 'IBMREXP',	'CUSPRIOR',	
            'FINCASH',	'CUSSTATU',	'COMPEXP',	'COMPINC',	'COMPTRMS',	'COMPSOLN',	'FINREJECT', 'FINCUST',	'COMPSAT',	'FIND',	'IBMDUP',	
            'IBMERR',	'CUSSPONS',	'FINGOE',	'FINIBM',	'COMPREL',	'COMPPART',	'IBMBUDG',	'FINK',	'IBMSUPP',	'FINCOMP',	'CUSTBUDG',	
            'IBMLOW',	'CUSTRISK',	'COMPMOVE',	'CUSTACT',	'FINNH',	'FINNONF',	'COMPPRCE',	'RNL_COMP_PRICE',	'RNL_CNTR_EARLY',	'RNL_CNTR_MOVE',	
            'RNL_CUST_NOBDGT',	'RNL_CUST_NORESP',	'RNL_CUST_OUTOFBUS',	'RNL_CUST_PROJEND',	'RNL_CUST_TEMP',	'RNL_IBM_PRODCHNG',	'RNL_IBM_NOPROD',	
            'RNL_PROD_NOSUPP',	'RNL_PROD_NOTRELB',	'RNL_PROD_NOSOLN',	'RNL_REVN_MOVESAAS',	'RNL_REVN_MOVETERM',	'RNL_REVN_MOVEPERP',	'RNL_REVN_MOVESUBS',	
            'RNL_REVN_MOVEPERS',	'RNL_REVN_MOVETYPE',	'RNL_REVN_MOVESALE',	'IBMUNDES',	'IBMUNRSP',	'TLSTRIBM',	'TLSTRNON',	'TLSCLOUD',	'TLSOTHER',	'TLSSELFP',	
            'TLSSELFE',	'TLSSELFO',	'TLSTPMP',	'TLSTPME',	'TLSTPMSD',	'TLSMVS',	'TLSIBMNA',	'TLSNORES',	'TLSBUDG',	'TLSDUPE',	'TLSNONE',	'BPDREXP',	'BPSUPP',	
            'BPESA',	'BPQUA',	'BPOTHER',	'TLSRENCO'
        ]

        # --- Required Columns for Validation ---
        # Define the mandatory columns for each shee
        required_columns = {
            "Opportunity": [
                "opportunity_legacy_id_c", "name", "accountid", "sales_stage",
                "expected_close_date", "currency_code", "ownerid", "OI_Source",'created_by'
            ],
            "Opportunity_product": [
                "OpportunityId", "Product", "product_type", "unitprice", "Term",
                "Classification Type", "Type"
            ]
        }

        # --- Columns That Must Have Values from the Valid API List ---
        columns_to_validate = [
            "sales_stage", "OI_Source", "Classification Type", "Type", "Renewal type", "Renewal Status","product_type",
            "Won Reason","Lost Category","Lost Reason"
        ]

        # --- Columns to Check for Missing (Blank) Values ---
        blank_values_columns = {
            "Opportunity": [
                "opportunity_legacy_id_c", "name", "accountid", "sales_stage", "expected_close_date", "currency_code", "ownerid", "OI_Source"
            ],
            "Opportunity_product": [
                "Product", "product_type", "unitprice", "Classification Type"
            ]
        }

        # ---------------------- Step 1:- Check if all required sheets are present ----------------------


        print(f"\n🔍 Step 1:Check if all the Required Sheets are present or not")

        import openpyxl

        def check_and_rename_sheets(file_path):
            # Load the Excel workbook
            wb = openpyxl.load_workbook(file_path)

            # Define the required sheets and their variants
            required_sheets = ["Opportunity", "Opportunity_product", "Opportunity_Team ", "Reporting_codes"]
            
            # Mapping of variants to the correct names
            variant_mapping = {
                'Opportunity_products': 'Opportunity_product',
                'Opportunity_Team': 'Opportunity_Team '  # trailing space is intended here
            }

            # Iterate through the workbook sheets and rename if there's a variant
            for sheet_name in wb.sheetnames:
                if sheet_name in variant_mapping:
                    ws = wb[sheet_name]
                    correct_name = variant_mapping[sheet_name]
                    ws.title = correct_name
                    print(f"\n    🔄 Renamed '{sheet_name}' to '{correct_name}' automatically.")

            # Get the list of sheet names in the current workbook after renaming
            sheets_in_file = wb.sheetnames

            # Identify missing required sheets (excluding 'Tags' which is optional)
            missing_sheets = [sheet for sheet in required_sheets if sheet not in sheets_in_file]

            # Print results based on the status of sheets
            if not missing_sheets:
                print("\n    ✅ All required sheets are already present! 🎉")
                # Identify extra sheets in the workbook that are NOT in the required list
                extra_sheets = [s for s in sheets_in_file if s not in required_sheets]

                if extra_sheets:
                    print(f"\n    📋 There are {len(extra_sheets)} extra sheet(s) in the file:")
                    for i, s in enumerate(extra_sheets, 1):
                        print(f"\n        {i}. {s}")
                else:
                    print("\n    📄 No extra sheets found. Everything looks clean!")
            else:
                print("\n    ❌ The following required sheets are missing: ")
                for i, sheet in enumerate(missing_sheets, 1):
                    print(f"\n        {i}. {sheet}")

                # Identify extra sheets in the workbook that are NOT in the required list
                available_sheets = [s for s in sheets_in_file if s not in required_sheets]

                # If there are available sheets, show them
                if available_sheets:
                    print("\n    📋 Here are the available extra sheets to rename: ")
                    for i, s in enumerate(available_sheets, 1):
                        print(f"\n        {i}. {s}")
                
                used_indices = []  # To keep track of already used sheet indices

                # Loop through each missing sheet and ask the user if they want to rename an available sheet
                for sheet in missing_sheets:
                    if len(used_indices) == len(available_sheets):
                        print(f"\n    ⏭️  No sheets available to rename. Skipping '{sheet}'!")
                        continue

                    while True:
                        choice = input(f"\n    🔸 Enter the index of the sheet to rename to '{sheet}' or type 'skip': ")

                        if choice.lower() == 'skip':
                            print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
                            break

                        try:
                            choice = int(choice)
                            if 1 <= choice <= len(available_sheets) and choice not in used_indices:
                                rename_sheet = available_sheets[choice - 1]
                                ws = wb[rename_sheet]
                                ws.title = sheet
                                print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")
                                used_indices.append(choice)
                                break
                            else:
                                print("\n        ❗ Invalid choice or sheet already used. Please choose a different one.")
                        except ValueError:
                            print("\n        ❗ Invalid input, please enter a valid number or 'skip'.")
                
                print("\n    💾 Workbook saved with changes!")
            # Save the modified workbook only once after all operations
            wb.save(file_path)

        # Call the function to check and rename sheets
        check_and_rename_sheets(file_path)

        # ------------------------------ Check for Mandatory Columns -------------------------

        xls = pd.ExcelFile(file_path)
        
        print(f"\n🔍 Step 2: Checking Required Columns: ")
        for sheet, columns in required_columns.items():
            if sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                df_columns_lower = [col.strip().lower() for col in df.columns]
                missing = [col for col in columns if col.lower() not in df_columns_lower]
                if missing:
                    file_status = "❌ Issues Found"
                    print(f"\n    ❌ Missing columns in '{sheet}':")
                    for col in missing:
                        print(f"\n        🔸 {col}")
                else:
                    print(f"\n    ✅ All required columns present in '{sheet}'")
            else:
                file_status = "❌ Issues Found"
                print(f"\n    ❌ Sheet '{sheet}' not found.")

        # ----------------------  Check for Columns with Invalid API names ---------------------

        print(f"\n🔍 Step 3: Checking for Invalid API names:")
        file_path = os.path.join(folder_path, file)
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        if 'Opportunity' in sheet_names and 'Opportunity_product' in sheet_names:
            # Convert valid API names to lowercase for case-insensitive check
            valid_api_names_lower = [val.strip().lower() for val in valid_api_names]
            # Track invalid values per sheet and column
            invalid_report = {}

            for sheet_name in ['Opportunity', 'Opportunity_product']:
                if sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)

                    for col in columns_to_validate:
                        if col in df.columns:
                            non_blank_values = df[col].dropna().astype(str)
                            non_blank_values = non_blank_values[non_blank_values.str.strip() != ""]

                            invalid_values = non_blank_values[~non_blank_values.str.strip().str.lower().isin(valid_api_names_lower)]

                            if not invalid_values.empty:
                                if sheet_name not in invalid_report:
                                    invalid_report[sheet_name] = {}
                                invalid_report[sheet_name][col] = invalid_values.unique()

            # Display results
            if invalid_report:

                for sheet, columns in invalid_report.items():
                    file_status = "❌ Issues Found"
                    print(f"\n    ❗️ Sheet: {sheet}")
                    for col, values in columns.items():
                        print(f"\n        🔹 Column: {col}")
                        for val in values:
                            print(f"\n            🔸 {val}")

                print("\n    ✅ All other values are valid.\n")

            else:
                print("\n    🎉 All API values are valid in both sheets.")
        else:
            file_status = "❌ Issues Found"
            print(f"\n    ❌ One or both sheets 'Opportunity' and 'Opportunity_product' not found.")

        # ----------------------  Check for missing Lost Catergory, Loss reason and won reason ---------------------

        print(f"\n🔍 Step 4: Check if Won Reason, Lost category and Lost reason are present")
        if "Opportunity" in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name='Opportunity')
                df['Excel Row'] = df.index + 2  # Excel row numbers

                identifier_cols = ['Opportunity Name'] if 'Opportunity Name' in df.columns else []
                
                df['sales_stage'] = df['sales_stage'].str.lower()

                # Check for missing reasons based on stage
                missing_won_reason = df[(df['sales_stage'] == 'won') & (df['Won Reason'].isnull() | (df['Won Reason'].astype(str).str.strip() == ''))]
                missing_lost_info = df[(df['sales_stage'] == 'lost') & (
                    (df['Lost Category'].isnull() | (df['Lost Category'].astype(str).str.strip() == '')) |
                    (df['Lost Reason'].isnull() | (df['Lost Reason'].astype(str).str.strip() == ''))
                )]

                # 🔴 Check for invalid values in 'Lost' fields when stage is 'Won'
                invalid_lost_fields_in_won = df[(df['sales_stage'] == 'Won') & (
                    (df['Lost Category'].notnull() & (df['Lost Category'].astype(str).str.strip() != '')) |
                    (df['Lost Reason'].notnull() & (df['Lost Reason'].astype(str).str.strip() != ''))
                )]

                # 🔴 Check for invalid 'Won Reason' when stage is 'Lost'
                invalid_won_reason_in_lost = df[(df['sales_stage'] == 'Lost') & (
                    (df['Won Reason'].notnull() & (df['Won Reason'].astype(str).str.strip() != ''))
                )]

                # Reporting
                if not missing_won_reason.empty:
                    file_status = "❌ Issues Found"
                    print(f"\n    ❗️ Missing 'Won Reason' where sales_stage = 'Won'")
                    # table_str = tabulate(
                    #     missing_won_reason[['Excel Row'] + identifier_cols + ['sales_stage', 'Won Reason']],
                    #     headers='keys', tablefmt='pretty', showindex=False
                    # )
                    # indented_table = "\n".join("        " + line for line in table_str.splitlines())
                    # print(indented_table)
                else:
                    print("\n    ✅ No missing 'Won Reason' for any rows with sales_stage = 'Won'.")

                if not missing_lost_info.empty:
                    file_status = "❌ Issues Found"
                    print(f"\n    ❗️ Missing 'Lost Category' or 'Lost Reason' for these rows where sales_stage = 'Lost'")
                    # table_str = tabulate(
                    #     missing_lost_info[['Excel Row'] + identifier_cols + ['sales_stage', 'Lost Category', 'Lost Reason']],
                    #     headers='keys', tablefmt='pretty', showindex=False
                    # )
                    # indented_table = "\n".join("        " + line for line in table_str.splitlines())
                    # print(indented_table)
                else:
                    print("\n    ✅ No missing 'Lost Category' or 'Lost Reason' for any rows with sales_stage = 'Lost'.")

                # 🔴 Invalid values in Lost fields for 'Won' stage
                if not invalid_lost_fields_in_won.empty:
                    file_status = "❌ Issues Found"
                    print(f"\n    ❗️ Invalid 'Lost Category' or 'Lost Reason' present for these rows where sales_stage = 'Won'")
                    table_str = tabulate(
                        invalid_lost_fields_in_won[['Excel Row'] + identifier_cols + ['sales_stage', 'Lost Category', 'Lost Reason']],
                        headers='keys', tablefmt='pretty', showindex=False
                    )
                    indented_table = "\n".join("        " + line for line in table_str.splitlines())
                    print(indented_table)
                else:
                    print("\n    ✅ No invalid 'Lost Category' or 'Lost Reason' for any rows with sales_stage = 'Won'.")

                # 🔴 Invalid 'Won Reason' in 'Lost' stage
                if not invalid_won_reason_in_lost.empty:
                    file_status = "❌ Issues Found"
                    print(f"\n    ❗️ Invalid 'Won Reason' present for these rows (sales_stage = 'Lost'):\n")
                    table_str = tabulate(
                        invalid_won_reason_in_lost[['Excel Row'] + identifier_cols + ['sales_stage', 'Won Reason']],
                        headers='keys', tablefmt='pretty', showindex=False
                    )
                    indented_table = "\n".join("        " + line for line in table_str.splitlines())
                    print(indented_table)
                else:
                    print("\n    ✅ No invalid 'Won Reason' for any rows with sales_stage = 'Lost'.")
            except Exception as e:
                file_status = "❌ Issues Found"
                print(f"\n    ❗️ Error: {e}")

        else:
            file_status = "❌ Issues Found"
            print(f"\n    ❌ Sheet 'Opportunity' not found.")

        # ----------------------  Check for Blank values in Important Columns ---------------------

        print(f"\n🔍 Step 5: Check if there are any blank values in Important columns")
        def check_blank_values(df, column_names):
            blank_columns = []
            if df.empty:
                # if no rows, consider all columns as blank
                return column_names
            for col in column_names:
                if col in df.columns:
                    blanks_found = df[col].isnull().any() or (df[col].apply(lambda x: isinstance(x, str) and x.strip() == '')).any()
                    if blanks_found:
                        blank_columns.append(col)
            return blank_columns

        if 'Opportunity' in xls.sheet_names:
            try:
                df_opportunity = pd.read_excel(xls, sheet_name='Opportunity')
                df_opportunity = df_opportunity.dropna(axis=0,how='all')
                df_opportunity = df_opportunity.dropna(axis=1,how='all')
                df_opportunity = df_opportunity.dropna(subset=['opportunity_legacy_id_c'], how='all')
                blank_cols_opp = check_blank_values(df_opportunity, blank_values_columns["Opportunity"])
                if blank_cols_opp:
                    file_status = "❌ Issues Found"
                    print("\n    ❗️ Blank values found in columns of 'Opportunity':")
                    for col in blank_cols_opp:
                        print(f'\n        🔸 {col}')
                else:
                    print("\n    ✅ No blank values in required columns of 'Opportunity'.")
            except Exception as e:
                file_status = "❌ Issues Found"
                print(f"\n    ❗️ Error: {e}")
        else:
            file_status = "❌ Issues Found"
            print(f"\n    ❌ Sheet 'Opportunity' not found.")

        if 'Opportunity_product' in xls.sheet_names:
            try:
    
                df_opportunity_product = pd.read_excel(xls, sheet_name='Opportunity_product')
                blank_cols_prod = check_blank_values(df_opportunity_product, blank_values_columns["Opportunity_product"])
                if blank_cols_prod:
                    file_status = "❌ Issues Found"
                    print("\n    ❗️ Blank values found in columns of 'Opportunity_product':")
                    for col in blank_cols_prod:
                        print(f'\n        🔸 {col}')
                else:
                    print("\n    ✅ No blank values in required columns of 'Opportunity_product'.")
            except Exception as e:
                file_status = "❌ Issues Found"
                print(f"\n    ❗️ Error: {e}")        
        else:  
            file_status = "❌ Issues Found"
            print(f"\n    ❌ Sheet 'Opportunity_product' not found.")

        # ----------------------- Check for Duplicate Oppty Legacy Id -----------------------

        print(f"\n🔍 Step 6: Check if there are any duplicate Oppty Legacy id")
        if "Opportunity" in xls.sheet_names:
            try:    
                df_opportunity = pd.read_excel(xls, sheet_name='Opportunity')
                df_opportunity = df_opportunity.drop_duplicates()
                df_opportunity = df_opportunity.dropna(how='all')
                if df_opportunity['opportunity_legacy_id_c'].duplicated().any():
                    file_status = "❌ Issues Found"
                    duplicated_values = df_opportunity[df_opportunity['opportunity_legacy_id_c'].duplicated(keep=False)]
                    print("\n   ❗️ Duplicate values found in 'opportunity_legacy_id_c'")
                else:
                    print("\n   ✅ No duplicate values found in 'opportunity_legacy_id_c'.")
            except Exception as e:  
                file_status = "❌ Issues Found"
                print(f"\n    ❗️ Error: {e}")                     
        else:
            file_status = "❌ Issues Found"
            print(f"\n    ❌ Sheet 'Opportunity' not found.")

        # Append file result to summary
        summary_list.append({"File Name": file, "Status": file_status})

title = "✅  File Validation Done ✅"
show_title(title)

# ✅ Final Summary Report
title = "📊 Final Summary 📊"
show_title(title)

print(tabulate(summary_list, headers="keys", tablefmt="fancy_grid"))
