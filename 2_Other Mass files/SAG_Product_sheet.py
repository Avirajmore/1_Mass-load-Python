# Importing all the necessary Libraries

import os
import sys
import pandas as pd
import tkinter as tk
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import Tk, Label, Button, Listbox, Scrollbar, MULTIPLE, END, Toplevel, StringVar, filedialog
from tkinter import Tk, Label, Button, Checkbutton, IntVar, Toplevel, Scrollbar, Frame, Canvas

print("=" * 100)
print(" " * 33 + "📝 PRODUCT SHEET EXECUTION 📝")
print("=" * 100)


# ======================================================================
# Select the file to process
# ======================================================================

print("\n\n🔍 Step 1: Select the file to Process")

# Set the directory to search for Excel files
directory = "/Users/avirajmore/Downloads"

# Create a hidden root window (used for file dialog)
root = tk.Tk()
root.withdraw()

# Ask the user to select an Excel file
file_path = filedialog.askopenfilename(
    initialdir=directory,
    title="📄 Select an Excel file",
    filetypes=[("Excel files", "*.xlsx")]
)

# Print the selected file path
if file_path:
    filename = str(file_path.split('/')[-1])
    print(f"\n    ✅ File selected: '{filename}'.")
else:
    print("\n    ❌ No file selected. Exiting the program. ❌")
    sys.exit()

# ======================================================================
# Step 2: Rename the sheet to Opportunity_product
# ======================================================================

import openpyxl

print("\n\n🔍 Step 2: Rename the sheet to Opportunity_product")

# Load the workbook
try:
    # file_path = "your_excel_file.xlsx"  # Replace with your actual file path
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames

    # Check if the sheet already exists with the correct name
    if "Opportunity_product" in sheets:
        print("\n    ✅ The sheet 'Opportunity_product' already exists with the correct name. Skipping renaming.")
    else:
        # Display available sheets
        print("\n    📂 Available sheets:")
        for idx, sheet in enumerate(sheets, start=1):
            print(f"\n      {idx}. {sheet}")

        # Ask user to select a sheet to rename
        while True:
            try:
                choice = int(input("\n    👉 Enter the number of the sheet to rename: "))
                if 1 <= choice <= len(sheets):
                    old_name = sheets[choice - 1]
                    break
                else:
                    print("\n    ❗️ Invalid choice, please select a valid number.")
            except ValueError:
                print("\n    ❗️ Please enter a valid number.")

        # Rename the selected sheet
        wb[old_name].title = "Opportunity_product"

        # Save the workbook
        wb.save(file_path)
        print(f"\n    ✅ Sheet '{old_name}' has been renamed to 'Opportunity_product' successfully!")

except FileNotFoundError:
    print("\n    ❌ Error: File not found. Please provide a valid file path.")
except Exception as e:
    print(f"\n    ❌ An error occurred: {e}")

# ======================================================================
# Step 3:Renaming the columns correctly
# ======================================================================


print("\n\n🔍 Step 3: Changing the product columns names ")

# Load the Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    # Read the specific sheet
    df = pd.read_excel(file_path, sheet_name="Opportunity_product")

    # Drop the "Product" column if it exists
    if 'Product' in df.columns:
        df.drop(columns=['Product'], inplace=True)

    # Rename "product_code_family" to "Product" if it exists
    if 'product_code_family' in df.columns:
        df.rename(columns={'product_code_family': 'Product'}, inplace=True)

    # Write back to the same sheet
    df.to_excel(writer, sheet_name="Opportunity_product", index=False)

print("\n    ✅ Column updated and file saved successfully.")


# ======================================================================
# Step 4: Converting headers to lowercase
# ======================================================================

print("\n\n🔍 Step 4: Converting headers to lowercase...")

# Read the Excel file with all sheets, initially treating all data as strings
xls = pd.ExcelFile(file_path)

# Dictionary to hold modified dataframes
sheets_dict = {}

# Columns that should remain numeric
numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

# Iterate through each sheet
for sheet_name in xls.sheet_names:
    # Read each sheet into a dataframe with all columns as strings
    df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
    
    # Convert headers to lowercase
    df.columns = [col.lower() for col in df.columns]
    
    # Convert specific columns back to numeric types
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')  # Convert to numeric, setting errors to NaN
    
    # Save modified dataframe to dictionary
    sheets_dict[sheet_name] = df

# Write the modified dataframes back to the Excel file
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for sheet_name, df in sheets_dict.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n    ✅ Headers of all sheets in the file have been converted to lowercase and data types preserved as specified.")


# ======================================================================
# Step 5: Count the rows and columns in the beginning of the process
# ======================================================================

print("\n\n🔍 Step 5: Counting the rows and columns...")

# Name of the sheet to target
product_sheet_name = 'Opportunity_product'

# Read the Excel file into a DataFrame
df = pd.read_excel(file_path, sheet_name = 'Opportunity_product')

# Get the number of rows and columns
product_initial_num_rows = df.shape[0]     # Number of rows in the DataFrame
product_num_columns = df.shape[1]          # Number of columns in the DataFrame

# Print the number of rows and columns
print(f"\n    ✅ Initial row count: {product_initial_num_rows}")
# print(f"\n    ✅ Initial column count: {product_num_columns}")


# ======================================================================
# Step 6:- Extract legacy id to fetch currency
# ======================================================================

print("\n\n🔍 Step 6: Extract legacy id to fetch currency...")


# file_path = "your_file.xlsx"  # Change this to your actual file
output_file = "/Users/avirajmore/Downloads/Currency_Extract.xlsx"
column_name = "legacy opportunityid"  # Column to process

# Read the specific sheet and column
df = pd.read_excel(file_path, sheet_name="Opportunity_product", usecols=[column_name], dtype=str)

# Process column: Trim, format with inverted commas, remove duplicates
df[column_name] = df[column_name].str.strip().dropna()  # Trim spaces & remove NaNs
df[column_name] = "'" + df[column_name] + "',"  # Add inverted commas and comma
df = df.drop_duplicates()  # Remove duplicates

# Save to a new Excel file
df.to_excel(output_file, index=False, engine="openpyxl")

print(f"\n    ✅ Processed file saved as {output_file}")

# ======================================================================
# Step 7:- Copy the data from Currency data to csv
# ======================================================================

print("\n\n🔍 Step 7: Copy Currency data to main...")

# Specify the CSV file path and the destination Excel file
csv_file_path = "/Users/avirajmore/Downloads/currency.csv"
# file_path = "output.xlsx"  # Change this to your desired file path

# Check if the CSV file exists, and prompt to retry if not
while not os.path.exists(csv_file_path):
    print(f"\n    ❌ Error: The currency file at path '{csv_file_path}' does not exist.")
    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
    while try_again not in ['yes', 'no']:
        print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
    if try_again == 'no':
        print("\n          🚫 Exiting the program.")
        sys.exit()

try:
    # Read the CSV file
    df = pd.read_csv(csv_file_path)

    # Write to the given file_path in a new sheet called "Currency"
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="Currency", index=False)

    print(f"\n    ✅ Data successfully copied from {csv_file_path} to 'Currency' sheet.")

except Exception as e:
    print(f"\n    ❌ Error: {e}")


# ======================================================================
# Step 8: Create new Currency column in the sheet...
# ======================================================================


print("\n\n🔍 Step 8: Create new Currency column in the sheet...")


# file_path= "/Users/avirajmore/Downloads/ISC Price and type correction file 02042025 copy.xlsx"

# Define the sheet names
file_pathopportunity_product_sheet_name = "Opportunity_product"
opportunity_currency_sheet = "Currency"

try:
    # Load data from the specified sheets
    opportunity_product_df = pd.read_excel(file_path, sheet_name=file_pathopportunity_product_sheet_name)
    opportunity_currency_df = pd.read_excel(file_path, sheet_name=opportunity_currency_sheet)

    # Clean and normalize the 'email' columns for consistency
    opportunity_product_df["legacy opportunityid"] = opportunity_product_df["legacy opportunityid"].str.strip().str.lower()
    opportunity_currency_df["Legacy_Opportunity_Split_Id__c"] = opportunity_currency_df["Legacy_Opportunity_Split_Id__c"].str.strip().str.lower()

    # Perform a left join to match emails and retrieve IDs
    result_df = pd.merge(
        opportunity_product_df,
        opportunity_currency_df[["Legacy_Opportunity_Split_Id__c", "CurrencyIsoCode"]],
        left_on="legacy opportunityid",
        right_on="Legacy_Opportunity_Split_Id__c",
        how="left"
    )


    # Rename the 'Id' column to 'OwnerId'
    result_df.rename(columns={"CurrencyIsoCode": "opportunity currency"}, inplace=True)

    # Save the updated data back to the 'Opportunity_team' sheet
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        result_df.to_excel(writer, sheet_name=file_pathopportunity_product_sheet_name, index=False)
    
    print("\n    ✅ Currency has been updated")


except FileNotFoundError:
    print(f"\n    ❌ File not found at path: {file_path}.")

except KeyError as e:
    print(f"\n    ❌ KeyError: Column '{e}' not found. Please check the column names in your Excel sheets.")

except Exception as e:
    print(f"\n    ❌ An unexpected error occurred: {e}")

# ======================================================================
# Step 9: Formatting the date column
# ======================================================================

print("\n\n🔍 Step 9: Formatting the date column in the 'Opportunity_product' sheet...")

product_sheet_name = 'Opportunity_product'  # Replace with the actual sheet name
date_column = 'expiration date'  # Replace with the actual column name containing the dates

try:
    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=product_sheet_name)

    # Check if the specified column exists in the DataFrame
    if date_column not in df.columns:
        print(f"\n    ❌ Error: The column '{date_column}' is missing from the sheet '{product_sheet_name}'. ❌")
        sys.exit(1)  # Exit the script if the column is missing

    # Ensure the date column is in datetime format and then format it as YYYY-MM-DD
    # print(f"\n    🔄 Formatting the '{date_column}' column...")
    df[date_column] = pd.to_datetime(df[date_column]).dt.strftime('%Y-%m-%d')

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    # Success message
    print(f"\n    ✅ Date column formatted successfully in the '{product_sheet_name}' sheet. ✅")

except FileNotFoundError:
    # Handle file not found
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 10:- Adding the 'Quantity' column
# ======================================================================


print("\n\n🔍 Step 10: Adding the 'Quantity' column in the 'Opportunity_product' sheet...")

# =============== Code to delete quantity column first ===============
try:
    # Load workbook and target sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[product_sheet_name]

    # Check if 'quantity' column exists
    header_row = [cell.value for cell in sheet[1]]
    if 'quantity' in header_row:
        col_index = header_row.index('quantity') + 1

        # Shift columns left to delete the 'quantity' column
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=sheet.max_column):
            for cell_index, cell in enumerate(row[:-1]):
                cell.value = row[cell_index + 1].value
            row[-1].value = None  # Clear the last column cell

    # Save changes
    wb.save(file_path)
except Exception:
    pass

# =============== Main code ===============

product_sheet_name = 'Opportunity_product'  # Replace with the actual sheet name
new_column_name = 'Quantity'  # Column name to be added
default_value = 1  # Default value for the new column

try:
    # Load the workbook and target the specified sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[product_sheet_name]

    # Add new column header at the end of existing headers
    # print(f"\n    🔄 Adding the '{new_column_name}' column header...")
    sheet.cell(row=1, column=sheet.max_column + 1, value=new_column_name)

    # Iterate over rows and set default value for the new column
    # print(f"\n    🔄 Setting the default value '{default_value}' for the new column...")
    for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming headers in row 1)
        sheet.cell(row=row, column=sheet.max_column, value=default_value)

    # Save the workbook
    wb.save(file_path)

    # Success message
    print(f"\n    ✅ A new column '{new_column_name}' has been added to the '{product_sheet_name}' sheet with default value '{default_value}'. ✅")

except FileNotFoundError:
    # Handle file not found
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 11: To delete unwanted columns from the sheet
# ======================================================================

print("\n\n🔍 Step 11: Deleting unwanted columns from the 'Opportunity_product' sheet...")

# Sheet name and columns to delete
product_sheet_name = "Opportunity_product"
columns_to_delete = [
    "created_by",
    "current quarter revenue",
    "modified_by",
    "created_date",
    "modified_date",
    "product_code_family",
    "pricebookentryid"
]

try:
    # Load the workbook and target the specified sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[product_sheet_name]

    columns_deleted = []

    # Check and delete each specified column
    for col_name in columns_to_delete:
        found = False
        for col in sheet.iter_cols():
            if col[0].value == col_name:
                sheet.delete_cols(col[0].column)
                columns_deleted.append(col_name)
                found = True
                # print(f"\n        🔸 '{col_name}' deleted. ✅")
                break
        if not found:
            print(f"\n        🔸 '{col_name}' not found. ❌")

    # Save the workbook
    wb.save(file_path)

    # Success or no deletion message
    if columns_deleted:
        print(f"\n    ✅ Successfully deleted mentioned columns from the '{product_sheet_name}' sheet:")
    else:
        print(f"\n    ❌ No columns from the specified list were found in the '{product_sheet_name}' sheet. ❌")

except FileNotFoundError:
    # Handle file not found error
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()

# ======================================================================
# Step 12: To get the "Product_Code_Family" column
# ======================================================================

print("\n\n🔍 Step 12: Creating 'Product_Code_Family' column in the 'Opportunity_product' sheet...")

product_sheet_name = "Opportunity_product"

try:
    # Load the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=product_sheet_name)

    # Check if required columns exist
    if "product" not in df.columns:
        print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()
    elif "product_type" not in df.columns:
        print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()

    # Concatenate values from 'product' and 'product_type' columns with a hyphen
    # print(f"    🔄 Creating 'Product_Code_Family' column by concatenating 'product' and 'product_type'...")
    df["Product_Code_Family"] = df["product"] + "-" + df["product_type"]

    # Save the updated DataFrame back to the same sheet
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    # Success message
    print(f"\n    ✅ The 'Product_Code_Family' column has been created and saved in the '{product_sheet_name}' sheet. ✅")

except FileNotFoundError:
    # Handle file not found error
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 13: To get the "Practise_Multiple country" column
# ======================================================================

print("\n\n🔍 Step 13: Creating 'Practise_Multiple country' column in the 'Opportunity_product' sheet...")

product_sheet_name = "Opportunity_product"

try:
    # Load the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=product_sheet_name)

    # Check if required columns exist
    if "product" not in df.columns:
        print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()
    elif "product_type" not in df.columns:
        print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()
    elif "opportunity currency" not in df.columns:
        print(f"\n    ❌ Error: Column 'opportunity currency' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()

    # Concatenate values from 'product', 'product_type', and 'opportunity currency' columns with a hyphen
    # print(f"    🔄 Creating 'Practise_Multiple country' column by concatenating 'product', 'product_type', and 'opportunity currency'...")
    df["Practise_Multiple country"] = df["product"] + "-" + df["product_type"] + "-" + df["opportunity currency"]

    # Save the updated DataFrame back to the same sheet
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    # Success message
    print(f"\n    ✅ The 'Practise_Multiple country' column has been created and saved in the '{product_sheet_name}' sheet. ✅")

except FileNotFoundError:
    # Handle file not found error
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 14: To Concatenate the Currency and Product Family
# ======================================================================

print("\n\n🔍 Step 14: Concatenating 'Currency' and 'Product Family' columns...")

product_sheet_name = "Opportunity_product"

try:
    # Load the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=product_sheet_name)

    # Check if required columns exist
    if "Product_Code_Family" not in df.columns:
        print(f"\n    ❌ Error: Column 'Product_Code_Family' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()
    elif "opportunity currency" not in df.columns:
        print(f"\n    ❌ Error: Column 'opportunity currency' not found in '{product_sheet_name}' sheet. ❌")
        sys.exit()

    # Task 1: Concatenate values from 'Product_Code_Family' column with inverted commas and commas
    # print(f"    🔄 Concatenating 'Product_Code_Family' column with inverted commas and commas...")
    df["Concatenated Product Family"] = "'" + df["Product_Code_Family"] + "', "

    # Task 2: Concatenate values from 'opportunity currency' column with inverted commas and commas
    # print(f"    🔄 Concatenating 'opportunity currency' column with inverted commas and commas...")
    df["Concatenated Currency"] = "'" + df["opportunity currency"] + "', "

    # Save the updated DataFrame back to the same sheet
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    # Success message
    print(f"\n    ✅ 'Concatenated Product Family' and 'Concatenated Currency' columns have been added successfully. ✅")

except FileNotFoundError:
    # Handle file not found error
    print(f"\n    ❌ Error: File not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 15: To keep the decimal values as 2
# ======================================================================

print("\n\n🔍 Step 15: Formatting decimal values to two decimal places...")

product_sheet_name = 'Opportunity_product'
headers_to_format = ['unitprice', 'expiring amount']

try:
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[product_sheet_name]

    # Function to format numbers to two decimal places and convert to float
    def format_to_float(value):
        if isinstance(value, (int, float)):
            return float(f'{value:.2f}')
        return value

    # Find column indices based on headers
    column_indices = {}
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header in headers_to_format:
            column_indices[header] = col

    # Iterate through each header and format numbers
    for header in headers_to_format:
        col = column_indices.get(header)
        if col:
            # print(f"    🔄 Formatting '{header}' column...")
            for row in range(2, sheet.max_row + 1):  # Start from row 2 assuming headers in row 1
                cell = sheet.cell(row=row, column=col)
                formatted_value = format_to_float(cell.value)
                sheet.cell(row=row, column=col, value=formatted_value)

    # Save the workbook
    wb.save(file_path)

    # Success message
    print(f"\n    ✅ Numbers in columns {', '.join(headers_to_format)} have been formatted to two decimal places. ✅")

except FileNotFoundError:
    # Handle file not found error
    print(f"\n    ❌ Error: File '{file_path}' not found. ❌")
    sys.exit()
except KeyError:
    # Handle missing sheet error
    print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ❌")
    sys.exit()
except Exception as e:
    # Handle any unexpected errors
    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
    sys.exit()


# ======================================================================
# Step 16: To extract the concatenated values
# ======================================================================

print("\n\n🔍 Step 16: Extracting concatenated values...")

def process_excel_file(file_path, sheet_name, required_columns, output_file):
    # Check if the input file exists
    if not os.path.exists(file_path):
        print(f"\n    ❌ Error: The input file '{file_path}' does not exist. ❌")
        return

    try:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"\n    ❌ Error: Failed to read the Excel file. Details: {e} ❌")
        return

    # Initialize a dictionary to hold cleaned data for each column
    cleaned_data_dict = {}

    # Process each required column
    for column in required_columns:
        if column in df.columns:
            # Remove blank and duplicate values
            cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
            cleaned_data_dict[column.replace("Concatenated", "").strip()] = cleaned_data
        else:
            print(f"\n    ❌ Error: Column '{column}' is missing. ❌")

    # Create an empty DataFrame for the output
    output_df = pd.DataFrame()

    # Add each cleaned column as a separate DataFrame and concatenate them
    for key, cleaned_data in cleaned_data_dict.items():
        output_df = pd.concat([output_df, pd.DataFrame({key: cleaned_data})], axis=1, ignore_index=False)

    # Write the processed data to a new Excel file if there's any data to write
    if not output_df.empty:
        try:
            output_df.to_excel(output_file, index=False)
            print(f"\n    ✅ Data written to '{output_file}'. ✅")
        except Exception as e:
            print(f"\n    ❌ Error: Failed to write the Excel file. Details: {e} ❌")
    else:
        print("\n    ❌ Error: No data to process. ❌")

# Specify the input file path, sheet name, required columns, and output file path
sheet_name = "Opportunity_product"  # Specify the sheet name
required_columns = ["Concatenated Product Family", "Concatenated Currency"]
output_file = "/Users/avirajmore/Downloads/ProductFamily_and_Currency_extract.xlsx"  # Specify the output file path

# Process the Excel file
process_excel_file(file_path, sheet_name, required_columns, output_file)


# ======================================================================
# Step 17:- To copy the data from CSV file
# ======================================================================

print("\n\n🔍 Step 17: Copying data from CSV file to Excel...")

# Define the CSV file path
csv_file_path = "/Users/avirajmore/Downloads/productfamily.csv"

# Check if the CSV file exists, and prompt to retry if not
while not os.path.exists(csv_file_path):
    print(f"\n    ❌ Error: The CSV file at path '{csv_file_path}' does not exist.")
    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
    while try_again not in ['yes', 'no']:
        print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
    if try_again == 'no':
        print("\n          🚫 Exiting the program.")
        sys.exit()

# Read data from the CSV file
df = pd.read_csv(csv_file_path)

# Specify the Excel file path and sheet name
sheet_name = "Opportunity_product_Copy"

# Write data to the specified sheet in the Excel file
with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"\n    ✅ Data from the CSV file has been successfully copied to the '{sheet_name}' sheet.")


# ======================================================================
# Step 18:- Create 'Practise_Multiple country' column in Product Copy sheet
# ======================================================================
print("\n\n🔍 Step 18: Create 'Practise_Multiple country' column in Product Copy sheet" )

# Read the Excel file
df = pd.read_excel(file_path, sheet_name="Opportunity_product_Copy")

# Concatenate values from "product" and "product_type" columns with a hyphen
df["Practise_Multiple country"] = df["Product2.Product_Code_Family__c"] + "-" + df["CurrencyIsoCode"]

# Save the updated DataFrame to the same sheet
with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name="Opportunity_product_Copy", index=False)

print(f"\n    ✅ The values from have been successfully concatenated and saved in the 'Practise_Multiple country' column.")


# ======================================================================
# Step 19:- Getting the PricebookEntry id
# ======================================================================

print("\n\n🔍 Step 19: Getting the PricebookEntry id ...")

# New code to check if the pricebook entry is active or not
product_sheet_name = 'Opportunity_product'
opportunity_copy_sheet_name = 'Opportunity_product_Copy'

# Function to standardize column names
def standardize_columns(df):
    df.columns = df.columns.str.strip().str.lower()
    return df

# Function to standardize column values
def standardize_column_values(df, column_name):
    df[column_name] = df[column_name].str.strip().str.lower()
    return df

# Read the data from both sheets
opportunity_df = pd.read_excel(file_path, sheet_name=product_sheet_name)
opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

# Standardize column names
opportunity_df = standardize_columns(opportunity_df)
opportunity_copy_df = standardize_columns(opportunity_copy_df)

# Standardize column values for the merge key
opportunity_df = standardize_column_values(opportunity_df, 'practise_multiple country')
opportunity_copy_df = standardize_column_values(opportunity_copy_df, 'practise_multiple country')

# Check if 'pricebookentryid' already exists in opportunity_df
if 'pricebookentryid' in opportunity_df.columns:
    raise KeyError("❌ Error: Column 'pricebookentryid' already exists in 'Opportunity_product'. Please check your data processing steps.")

# Filter the opportunity_copy_df based on 'IsActive' column
opportunity_copy_df['pricebookentryid'] = opportunity_copy_df.apply(
    lambda row: row['id'] if row['isactive'] else 'Not Active', axis=1
)

# Perform a left join to get PriceBookEntryid for each Practise_Multiple country
merged_df = pd.merge(opportunity_df, 
                     opportunity_copy_df[['practise_multiple country', 'pricebookentryid']], 
                     left_on='practise_multiple country', 
                     right_on='practise_multiple country',
                     how='left')

# Fill missing values with 'No Pricebookid found'
merged_df['pricebookentryid'] = merged_df['pricebookentryid'].fillna('No Pricebookid found')

# Count occurrences of 'No Pricebookid found' and 'Not Active'
count_no_pricebookid_found = (merged_df['pricebookentryid'] == 'No Pricebookid found').sum()
count_not_active = (merged_df['pricebookentryid'] == 'Not Active').sum()

# Save the updated DataFrame back to the same Excel file
with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
    merged_df.to_excel(writer, sheet_name=product_sheet_name, index=False)

print(f"\n    ✅ The 'Opportunity_product' sheet has been successfully updated with the 'PriceBookEntryid' column.")
print(f"\n        ❗️ Count of 'No Pricebookid found': {count_no_pricebookid_found}")
print(f"\n        ❗️ Count of 'Not Active': {count_not_active}")

# ======================================================================
# Step 20: Rearranging the Columns in Sequence
# ======================================================================

print("\n\n🔍 Step 20: Rearranging Columns in the 'Opportunity_product' Sheet...")

# Name of the sheet to target
sheet_name = 'Opportunity_product'

# Specify the desired order of columns
desired_column_order = [
    'legacy opportunityid',
    'opportunityid',
    'quantity',
    'product',
    'product_type',
    'product_code_family',
    'opportunity currency',
    'practise_multiple country',
    'pricebookentryid'
]

try:
    # Read the Excel file
    excel_data = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Check if the sheet exists
    if isinstance(excel_data, pd.DataFrame):
        # Check if all specified columns exist
        missing_columns = [col for col in desired_column_order if col not in excel_data.columns]
        extra_columns = [col for col in excel_data.columns if col not in desired_column_order]

        # Rearrange columns
        rearranged_columns = [col for col in desired_column_order if col in excel_data.columns]

        # Add extra columns to the end
        rearranged_columns += extra_columns

        # Write the modified DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            excel_data[rearranged_columns].to_excel(writer, sheet_name=sheet_name, index=False)

        # Notify the user about the changes
        if missing_columns:
            print("\n    ❌ The following columns were missing and were skipped:")
            for col in missing_columns:
                print(f"\n      🔸 {col}")
        
        # if extra_columns:
        #     print("\n    🔄 The following extra columns were moved to the end of the sheet:")
        #     for col in extra_columns:
        #         print(f"\n      🔸 {col}")

        print(f"\n    ✅ Columns successfully rearranged in the '{sheet_name}'")
    else:
        print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the Excel file.")
except FileNotFoundError:
    print(f"\n    ❌ Error: File '{file_path}' not found.")


# ======================================================================
# Step 21: Changing the values in Renewal status according to API name
# ======================================================================

print("\n\n🔍 Step 21: Changing the values in Renewal status according to API name...")
import pandas as pd
import re

# Load the Excel file
# file_path = "your_file.xlsx"  # Change this to your actual file path
sheet_name = "Opportunity_product"

# Read the Excel file
df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

# Rename columns while handling case sensitivity
df.columns = [col.strip().lower() for col in df.columns]  # Normalize column names
rename_mapping = {
    "renewal status/attrition indicator": "renewal status",
    "expiring amount/expiring tcv": "expiring amount"
}
df.rename(columns={k.lower(): v.lower() for k, v in rename_mapping.items()}, inplace=True)

# Define mapping for replacement
renewal_mapping = {
    "propose": "RENEW_PROPOSE",
    "negotiate": "RENEW_NEGOTIATE",
    "design": "RENEW_DESIGN",
    "closing": "RENEW_CLOSING",
    "lost": "RENEW_LOST",
    "won": "RENEW_WON"
}

# Define the valid renewal statuses
valid_renewal_statuses = [
    "RENEW_PROPOSE",
    "RENEW_NEGOTIATE",
    "RENEW_DESIGN",
    "RENEW_CLOSING",
    "RENEW_LOST",
    "RENEW_WON"
]

# Function to clean and replace renewal status values
def format_renewal_status(value):
    if isinstance(value, str) and "renew" in value.lower():
        for key, replacement in renewal_mapping.items():
            if re.search(rf"\b{key}\b", value, re.IGNORECASE):
                return replacement
    return value  # Keep unchanged if no match

# Apply transformation to "renewal status" column
df["renewal status"] = df["renewal status"].apply(format_renewal_status)

# Check for invalid renewal status values (not in valid ones, and not blank)
invalid_count = df["renewal status"].apply(lambda x: x not in valid_renewal_statuses and pd.notna(x)).sum()

# If there are invalid values, ask for confirmation to continue or terminate
if invalid_count > 0:
    while True:
        response = input(f"\n    ❗️ There are {invalid_count} invalid values in the 'renewal status' column. Do you want to continue (y/n)? ").strip().lower()
        if response == 'y':
            print("\n       ✅ Continuing with the process.")
            break
        elif response == 'n':
            print("\n       ❌ Terminating the program.")
            exit()  # Terminates the program
        else:
            print("\n       ❗️ Invalid response. Please type 'y' to continue or 'n' to terminate.")

# Save the modified data back to the same file
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n    ✅ Excel file updated successfully.")

# ======================================================================
# Step 22: Rename the Columns
# ======================================================================

print("\n\n🔍 Step 22: Renaming Columns in the 'Opportunity_product' Sheet...")

# Name of the sheet to target
sheet_name = 'Opportunity_product'

# Dictionary mapping old column names to new column names
column_rename_mapping = {
    'legacy opportunityid': 'Legacy_Opportunity_Split_Id__c',
    'quantity': 'Quantity',
    'product_code_family': 'Product_Family__c',
    'pricebookentryid': 'PricebookEntryId',
    'unitprice': 'UnitPrice',
    'term': 'Term__c',
    'classification type': 'Classification__c',
    'type': 'Type__c',
    'renewal type': 'Renewal_Type__c',
    'renewal status': 'Renewal_Status__c',
    'expiration date': 'Expiration_Date__c',
    'expiring term': 'Expiring_Term__c',
    'expiring amount': 'Expiring_Amount__c',
    'external id': 'External_IDs__c',
}

# Read the Excel file
try:
    excel_data = pd.read_excel(file_path, sheet_name=None)
    
    # Check if the specified sheet exists
    if sheet_name in excel_data:
        # Access the specified sheet
        df = excel_data[sheet_name]
        
        # Check if all specified columns exist
        missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]
        
        # If any specified column is missing, notify the user and ask if they want to proceed
        if missing_columns:
            print("\n    ❌ The following columns are missing and cannot be renamed:")
            for col in missing_columns:
                print(f"\n      🔸 {col}")
            
            # Loop until a valid response is entered
            while True:
                proceed = input("\n    🔹 Do you want to proceed with the execution? (yes/no): ").lower()
                if proceed == 'yes':
                    break
                elif proceed == 'no':
                    print("\n      🚫 Operation aborted by the user.")
                    exit()
                else:
                    print("\n      ❗️ Invalid input. Please enter 'yes' or 'no'.")
        
        # Rename specified columns
        df.rename(columns=column_rename_mapping, inplace=True)
        
        # Write the modified DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\n    ✅ Columns renamed successfully in the '{sheet_name}'")
    
    else:
        print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the Excel file.")
except FileNotFoundError:
    print(f"\n    ❌ Error: File '{file_path.split('/')[-1]}' not found.")

# ======================================================================
# Step 23: Final Row and Column Count
# ======================================================================

print("\n\n🔍 Step 23: Final Row and Column Count...")

# Name of the sheet to target
product_sheet_name = 'Opportunity_product'

# Read the Excel file into a DataFrame
df = pd.read_excel(file_path, sheet_name=product_sheet_name)

# Get the number of rows and columns in the DataFrame
product_final_num_rows = df.shape[0]
product_final_num_columns = df.shape[1]

# Display the final row count
print(f"\n    ✅ Final row count: {product_final_num_rows}")
# print(f"\n    ✅ Final column count: {product_final_num_columns}")

# Check if the number of rows has changed
if product_initial_num_rows != product_final_num_rows:
    print(f"\n    ❗️ Row count mismatch detected!")
    print(f"\n       📊 Initial: {product_initial_num_rows}")
    print(f"\n       📊 Final: {product_final_num_rows}")

    while True:
        # Ask the user whether to continue or stop
        user_input = input(
            f"\n    🔹 Do you want to continue? Type 'continue' to proceed or 'no' to stop: "
        ).strip().lower()

        if user_input == "continue":
            print("\n      🔄 Continuing the program...")
            break  # Exit the loop and continue execution
        elif user_input == "no":
            print("\n      🚫 Terminating the program...")
            exit()  # Terminate the program
        else:
            print("\n      ❗️ Invalid input. Please type 'continue' to proceed or 'no' to stop.")



# ======================================================================
print('\n')
print("=" * 100)
print(" " * 33 + "📝 PRODUCT SHEET COMPLETED 📝")
print("=" * 100)

# ======================================================================


