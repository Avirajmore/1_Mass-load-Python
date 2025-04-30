# Importing all the necessary Libraries
import time
import os
import re
import sys
import shutil
import openpyxl
import pyperclip
import numpy as np
import pandas as pd
from tkinter import *
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import SheetTitleException

# =========================================================================================================================================
                                                 # FOLDER CREATION & FILE MOVEMENT
# =========================================================================================================================================

# =========================================================
# Define the base paths for storing mass load files
# =========================================================

# Path of the folder Where you want to save the Mass Load Files. 
# ❗️ Change this path if you want to store files in a different location
BASE_DIR = os.path.expanduser("~/Documents/Office Docs/Massload Files/2025") 

# for Example :-  os.path.expanduser("~/Downloads") == "/Users/avirajmore/Downloads" 

# Path of the folder Where you have saved a the template of the Summary file
# ❗️ Change this path if your summary file is stored elsewhere
REF_SUMMARY_FILE_PATH = os.path.expanduser("~/Documents/Office Docs/Massload Files/Reference File/Reference_Summary_file.xlsx") #Change it to where you want to store the summary file

# Path for the Downloads Folder
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")

# =========================================================
# Folder Creation starts
# =========================================================

# Function to display a title with a decorative line
def show_title(title):

    line_width = 100
    line = "=" * line_width
    print(f"\n{line}")
    print(title.center(line_width))
    print(f"{line}\n")

# Display the title for the folder creation and file movement process
title = "📂 FOLDER CREATION & FILE MOVEMENT 📂"
show_title(title)

# =========================================================
# Function to validate folder names 
# =========================================================

# Function to validate folder names
def is_valid_folder_name(name):
    invalid_chars = set('\\/:*?\"<>|')
    return name and not any(char in invalid_chars for char in name) # return True if name is not empty and Does not have Invalid Characters

# =========================================================
print("\n\n🔍 Step 1: Creating Main Folder")

# Prompt the user for a valid folder name, Keep prompting the user until a valid folder name is provided 
while True:

    Current_Iteration_Folder_Name  = input("\n    📂 Enter the name of the Folder: ").strip()
    
    # Check if the folder name is valid
    if is_valid_folder_name(Current_Iteration_Folder_Name ): 
        break # If valid, exit the loop and proceed
    
    else:
        # If invalid, display an error message and prompt again
        print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

# Create the main folder and subfolders

# Construct a Full path for the user-specified Folder inside BASE_DIR
Current_iteration_Folder = os.path.join(BASE_DIR, Current_Iteration_Folder_Name )

# Create the main subfolder (if it doesn't already exist)
os.makedirs(Current_iteration_Folder, exist_ok=True)

# Define paths for additional subfolders "Copy Files and Final Iteration File"
copy_file_path = os.path.join(Current_iteration_Folder, "Copy files")
final_iteration_file_path = os.path.join(Current_iteration_Folder, "Final iteration files")

# Create the additional subfolders "Copy Files and Final Iteration File"
os.makedirs(copy_file_path, exist_ok=True)
os.makedirs(final_iteration_file_path, exist_ok=True)

print(f"\n        ✅ Subfolders 'Copy files' and 'Final iteration files' created successfully")

# =========================================================
# Step 2: Move Mass load files from Downloads to Current Iteration Folder
# =========================================================

# Define supported Excel file extensions
excel_extensions = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm')

# List to keep track of successfully moved files
files_moved = []

print(f"\n\n🔍 Step 2: Moving Excel files")

# Check if any Excel files exist in Downloads folder before moving them
if os.path.exists(DOWNLOADS_DIR):
    
    # Loop through all files in the Downloads folder
    for file_name in os.listdir(DOWNLOADS_DIR):
    
        # Check if the file has an Excel file extension (case-insensitive)    
        if file_name.lower().endswith(excel_extensions):
            
            # Construct full source and target file paths
            source_path = os.path.join(DOWNLOADS_DIR, file_name)
            target_path = os.path.join(Current_iteration_Folder, file_name)
            
            # Move the file from Downloads to the target folder
            shutil.move(source_path, target_path)

            # Add the file name to the list of moved files
            files_moved.append(file_name)


# Populate the excel_files list with all Excel files in the Current Iteration Folder

excel_files_Curr_Iter = [] # Initialize an empty List to store All the File

for f in os.listdir(Current_iteration_Folder):
    if f.lower().endswith(excel_extensions): # Check if the file has a valid Excel extension
        excel_files_Curr_Iter.append(f)

# Display the files that were moved from the Downloads folder (if any)

if not files_moved:
    print("\n    📥 Moved Files:")
    print("\n        1) ❗️ No files were moved from Downloads. ")
else:
    print("\n    📥 Moved Files:")
    for index, file_name in enumerate(files_moved, start=1):
        print(f"\n        {index}) {file_name} ✅") # Show each moved file with an index

# =========================================================
# Step 4: Make a Copy of Original Files and save them to "Copy Folder" 
# =========================================================

print("\n\n🔍 Step 3: Copying files")

files_copied = [] # Initialize an empty List

# Iterate through all Excel files in the current iteration folder

for file_name in excel_files_Curr_Iter: 
    
    # Rename Oriiginal File By adding "_Copy"
    copy_file_name = f"{os.path.splitext(file_name)[0]}_Copy{os.path.splitext(file_name)[1]}" 
    
    # Construct full source and target file paths
    source_path = os.path.join(Current_iteration_Folder, file_name)
    target_path = os.path.join(copy_file_path, copy_file_name)
    

    if os.path.exists(target_path):
        files_copied.append((file_name, "skipped"))
    else:
        shutil.copy(source_path, target_path)
        files_copied.append((file_name, "copied"))

# Display the results of the file copying operation

print("\n    📤 Copied Files:")
for index, (file_name, status) in enumerate(files_copied, start=1):
    if status == "copied":
        print(f"\n        {index}) {file_name} ✅") # File successfully copied
    elif status == "skipped":
        print(f"\n        {index}) {file_name} - Skipped 🛑 ") # File was skipped because it already existed


# =========================================================
# Step 5: Create Folders in "Final iteration files" to store Final processed Files and other Data
# =========================================================

print("\n\n🔍 Step 4: Creating folders in 'Final iteration files'")
Final_File_Folders_created = []

# Iterate through all Excel files in Currenct Iteration to create "Final Files" folders with that File's Name
for file_name in excel_files_Curr_Iter:

    # Extract file name without extension (.xlxs) by spliting the file name base on "." (dot) to Rename the Folders in Final Files Foldrt
    Final_File_Folder_Name = os.path.splitext(file_name)[0] 

    # Define the Folder path of Each File
    Final_File_Folder_Path = os.path.join(final_iteration_file_path, Final_File_Folder_Name) 

    # Ensure that the final folder, named after the Excel file, exists (create it if necessary).
    if not os.path.exists(Final_File_Folder_Path):
        os.makedirs(Final_File_Folder_Path, exist_ok=True)
        folder_status = "created"
    else:
        folder_status = "exists"

    # Create Given subfolders inside the main folder (always ensure they exist)
    subfolders = ["Removed Rows", "Success and error files" , "CSV Files"]
    
    for subfolder in subfolders:
        SubFolder_Path = os.path.join(Final_File_Folder_Path, subfolder)
        if not os.path.exists(SubFolder_Path):
            os.makedirs(SubFolder_Path, exist_ok=True)

    Final_File_Folders_created.append((Final_File_Folder_Name, folder_status))

# Display folder creation results
print("\n    🗂️ Folder Created:")
for index, (Final_File_Folder_Name, status) in enumerate(Final_File_Folders_created, start=1):
    if status == "created":
        print(f"\n        {index}) {Final_File_Folder_Name} ✅ (Main folder created)")
    elif status == "exists":
        print(f"\n        {index}) {Final_File_Folder_Name} - Main folder already exists 🛑 (Subfolders ensured)")

# Check if the "Copy files" folder exists before proceeding
if not os.path.exists(copy_file_path):
    print("\n     ❗️ 'Copy files' folder does not exist. ")
    sys.exit()

# Retrieve the list of files present in the "Copy files" folder

files_in_copy_folder = [] # Initialize an empty List to Save File in Copy Folder

# Iterate over all items in the directory
for f in os.listdir(copy_file_path):
    file_path = os.path.join(copy_file_path, f)
    
    # Check if the item is a file
    if os.path.isfile(file_path):
        files_in_copy_folder.append(f)

if not files_in_copy_folder:
    print("\n     🚫 No files found in 'Copy files' folder. ")
    sys.exit()
# =========================================================================================================================================
#                                                OPPORTUNITY SHEET EXECUTION
# =========================================================================================================================================


# ======================================================================
# Determine Which File to Process and Set the File Path Accordingly
# ======================================================================

while True:
    
    # Display available files for selection

    print("\n====================================================================================================")
    print("\n📂 Please select a file to process:")
    print("\n    🔸 List of Files in Copy Folder: ")

    for idx, file_name in enumerate(files_in_copy_folder, start=1):
        print(f"\n        📄 {idx}. {file_name}")

    # Prompt the user to select a file from the list
    while True:
        
        user_input = input("\n    👉 Enter the number of the file to process (or type 'exit' to quit): ").strip()
        
        # Allow the user to exit the selection process
        if user_input.lower() == 'exit':
            print("\n        ❌ File selection has been canceled. Exiting process. \n")
            sys.exit()

        try:
            # Convert user input to an index and validate selection
            selected_index = int(user_input) - 1
            
            if 0 <= selected_index < len(files_in_copy_folder):
                file_path = os.path.join(copy_file_path, files_in_copy_folder[selected_index])
                print(f"\n        ✅ You selected the file: {files_in_copy_folder[selected_index]} \n")
                break # Exit the loop if a valid file is selected

            else:
                print(f"\n        ❗ Invalid selection. Please select a number between 1 and {len(files_in_copy_folder)}.")
        
        except ValueError:
            print("\n        ❗ Invalid input. Please enter a valid number or type 'exit' to cancel.")

    # ================================================================================
    # Code to Construct Path to Folder of current Running file in "Final Iteration file" Folder
    # ================================================================================
    start_time = time.time()   # Record start time
    # Extract the selected file name from the full file path
    selected_file_name = os.path.basename(file_path.split("/")[-1])

    # Remove '_Copy' from the file name (if present) and remove the file extension
    folder_file_name = os.path.splitext(re.sub(r'_Copy', '', selected_file_name))[0]

    # Define the output folder path where processed data will be stored
    output = os.path.join(final_iteration_file_path, folder_file_name)

    # Define paths for subdirectories where different types of processed data will be saved
    csv_file_dir = os.path.join(output, "CSV Files") # Folder for storing CSV Files
    removed_rows_dir = os.path.join(output, "Removed Rows") # Folder for storing Removed Rows Files

    # ================================================================================
    # Check for missing required sheets and rename if necessary
    # ================================================================================

    symbol = "="
    print(symbol*100)

    print("\n\n🔍 Check if all the Required Sheets are present or not")
 
    # Load the Excel workbook 
    '''
        Info:
        We use openpyxl here instead of pandas because:
        - openpyxl gives access to sheet names, cell formatting, and workbook structure.
        - pandas is mainly for working with data tables (DataFrames), not the workbook structure.
    '''
    wb = openpyxl.load_workbook(file_path)
    # 📌 [NEW] Auto-Rename Known Variants to Correct Names

    # Define the list of required sheet names
    # 'Tags' is considered optional and will not be treated as missing if absent
    required_sheets = ['Opportunity', 'Opportunity_product','Opportunity_Team ', 'Reporting_codes', 'Tags']
    
    variant_mapping = {
        'Opportunity_products': 'Opportunity_product',
        'Opportunity_Team': 'Opportunity_Team '  # note the trailing space
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in variant_mapping:
            ws = wb[sheet_name]
            correct_name = variant_mapping[sheet_name]
            ws.title = correct_name
            print(f"\n    ✅  Renamed '{sheet_name}' to '{correct_name}' automatically.")

    # Get the list of sheet names present in the current workbook
    sheets_in_file = wb.sheetnames

    # Identify missing required sheets (excluding 'Tags' which is optional)
    missing_sheets = [] # Initialize Missing Sheet List

    for sheet in required_sheets:
        # Check if the sheet is not 'Tags' and is not in the list of sheets in the file
        if sheet != 'Tags' and sheet not in sheets_in_file:
            missing_sheets.append(sheet)

    # Check if all required sheets are present
    if not missing_sheets:
        # All required sheets are present — no further action needed
        print("\n    ✅ All required sheets are already present! 🎉")

    else:
        print("\n    ❌ The following required sheets are missing: ")

        # Print the missing sheets
        for i, sheet in enumerate(missing_sheets, 1):
            print(f"\n        {i}. {sheet}")

        # Identify extra sheets in the workbook that are NOT in the required list
        available_sheets = [] # Initialize Available Sheet List

        for s in sheets_in_file:
            # Check if the sheet is not in the list of required sheets
            if s not in required_sheets:
                available_sheets.append(s)

        # If there are available sheets, prompt user to rename them
        if available_sheets:
            print("\n    📋 Here are the available sheets to rename: ")
            # Display the available sheets as a numbered list
            for i, s in enumerate(available_sheets, 1):
                print(f"\n        {i}. {s}")
        
        used_indices = []  # keep track of already used sheet indices

        # Loop through each missing sheet and ask the user if they want to rename one of the available sheets
        for sheet in missing_sheets:
            if len(used_indices) == len(available_sheets):
                print(f"\n    ⏭️  No sheets available to rename. Automatically skipping '{sheet}'!")
                continue

            while True:
                choice = input(f"\n    🔸 Enter the index of the sheet to rename to '{sheet}' or type 'skip': ")

                if choice.lower() == 'skip':
                    print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
                    break

                try:
                    choice = int(choice)
                    if 1 <= choice <= len(available_sheets):
                        if choice in used_indices:
                            print("\n        ❗ That sheet has already been used. Choose a different one.")
                            continue

                        rename_sheet = available_sheets[choice - 1]
                        ws = wb[rename_sheet]
                        ws.title = sheet

                        print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")

                        used_indices.append(choice)
                        break

                    else:
                        print("\n        ❗ Invalid number selected. Please choose a valid option.")
                except ValueError:
                    print("\n        ❗ Invalid input, please enter a valid number or 'skip'.")
        # Save the modified workbook (if any renaming was done)
        wb.save(file_path)
        print("\n    💾 Workbook saved with changes!")
    
    
    # ==========================================
    # To Handle Sheets with Similar Names but Different Casing
    # ==========================================

    # Create a mapping of sheet names that may have different cases to their standard names
    sheet_name_mapping = {
        'Opportunity1': 'Opportunity',
        'Opportunity_product1': 'Opportunity_product',
        'Opportunity_team1': 'Opportunity_team',
        'Reporting_codes1': 'Reporting_codes'
    }

    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet in wb.sheetnames:

        # Check if the sheet's name is present in the mapping dictionary
        if sheet in sheet_name_mapping:

            # If a match is found, get the corresponding new name
            new_name = sheet_name_mapping[sheet]
            
            # Access the worksheet with the current name
            ws = wb[sheet]

            # Rename the sheet to the mapped new name
            ws.title = new_name

    # Save the workbook with the renamed sheets (the content will remain unchanged)
    wb.save(file_path)
    
    # ==========================================
    # Convert all column headers to lowercase
    # ==========================================
    
    print("\n\n🔍 Converting all columns headers to lowercase...")

    # Read the Excel file with all sheets, initially treating all data as strings
    xls = pd.ExcelFile(file_path)

    # Create a dictionary to store updated DataFrames for each sheet
    sheets_dict = {}

    # List of column names that should be kept as numeric (e.g., for calculations)
    numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

    # Iterate through each sheet
    for sheet_name in xls.sheet_names:
        # Read each sheet into a dataframe with all columns as strings
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        
        # Convert all column headers to lowercase for consistency
        df.columns = [col.lower() for col in df.columns]
        
        # Convert specified columns back to numeric (e.g., prices, amounts, terms)
        for col in numeric_columns:
            if col in df.columns:
                # Store original values before conversion
                original_values = df[col].copy()
                
                # Convert the column to numeric, replacing errors with NaN
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Identify and display values that failed to convert
                invalid_mask = df[col].isna() & original_values.notna() & (original_values != "")
                if invalid_mask.any():
                    print(f"\n    ❗️ Warning: The following values in column '{col}' could not be converted to numeric and were set to NaN:")
                    for i, val in original_values[invalid_mask].items():
                        print(f"\n       🔸 Row {i + 2}: '{val}'")  # +2 to adjust for header and 0-based index
                
        # Save modified dataframe to dictionary
        sheets_dict[sheet_name] = df

    # Write the modified dataframes back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\n    ✅ Headers of all sheets in the file have been converted to lowercase and data types preserved as specified.")

    # ==========================================
    # Creating Blank sheets for rough work
    # ==========================================

    print("\n\n🔍 Creating Blank sheets for rough work...")

    # Names of the sheets to add
    sheet_names = [
        "Opportunity_Copy",
        "Opportunity_product_Copy",
        "Opportunity_team_Copy",
        "Reporting_codes_Copy",
        "Tags_Copy"
    ]

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Add new sheets
        for name in sheet_names:
            wb.create_sheet(title=name)

        # Save the workbook
        wb.save(file_path)
        print("\n    ✅ Blank Copy Sheets added successfully.")

    except Exception as e:
        print(f"\n    📝 An error occurred: {e}")

    # ======================================================================
    # Print Opportunity Script Execution 📝                               
    # ======================================================================
    
    # Display the title for the Opportunity Sheet Execution
    title = "📝 OPPORTUNITY SHEET EXECUTION 📝"
    show_title(title)
    
    # ======================================================================
    # Step 1: File Existence Check
    # ======================================================================

    print("\n\n🔍 Step 1: Checking if the file exists...")
    
    def check_file_exists(file_path):
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            print(f"\n    ✅ File '{filename}' exists at the specified path.")
            return (filename)
        else:
            print("\n    ❌ Error: File does not exist or the path is invalid.\n")
            sys.exit()  # Exit the program if file is not found

    # Example usage:
    filename = check_file_exists(file_path)
    
    # ======================================================================
    #  Extract the Mpp_Number__c column and save it to a new Excel file
    # ======================================================================

    opportunity_df = pd.read_excel(file_path, sheet_name="Opportunity", dtype={"Mpp_Number__c": str})

    if 'Mpp_Number__c' in opportunity_df.columns:
        # Extract the two required columns
        columns_to_extract = ["opportunity_legacy_id_c", "Mpp_Number__c"]
        mpp_df = opportunity_df[columns_to_extract]

        # Save to a new Excel file, ensuring Mpp_Number__c remains text
        mpp_output_file = os.path.join(csv_file_dir,"MPP_Column.xlsx")
        with pd.ExcelWriter(mpp_output_file, engine='openpyxl') as writer:
            mpp_df.to_excel(writer, index=False)

    # ======================================================================
    # Step 2: Removing Duplicates and Blank Rows
    # ======================================================================

    print("\n\n🔍 Step 2: Removing duplicate rows and blank rows...")
    
    # Define Opportunity Sheet Name
    opportunity_sheet_name = "Opportunity"

    def clean_sheet(df, sheet_name, remove_duplicates=False):
        try:
            # Remove duplicate rows if requested
            if remove_duplicates:
                df = df.drop_duplicates()

            # Remove rows where all cells are NaN (blank rows)
            df = df.dropna(how='all')
            if sheet_name == "Opportunity":
                df = df.dropna(subset=['opportunity_legacy_id_c'], how='all')

            if remove_duplicates:
                print(f"\n    ✅ Removed duplicate and blank rows from '{sheet_name}' sheet. ")
            else:
                print(f"\n    ✅ Removed blank rows from '{sheet_name}' sheet. ")

        except Exception as e:
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
            sys.exit()
        return df

    opportunity_df = clean_sheet(opportunity_df, opportunity_sheet_name, remove_duplicates=True)


    # ======================================================================
    # Function to check for duplicates in the 'opportunity_legacy_id_c' column of the 'Opportunity' sheet
    # ======================================================================

    # Check if 'opportunity_legacy_id_c' column exists
    if 'opportunity_legacy_id_c' not in opportunity_df.columns:
        print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found in the sheet.")
        sys.exit(1)

    # Check for duplicate values
    if opportunity_df['opportunity_legacy_id_c'].duplicated().any():
        print("\n    ❌ Error: Duplicate values still Present in 'opportunity_legacy_id_c' column.")
        sys.exit(1)

    # ======================================================================
    # Step 3: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 3: Counting the rows and columns...")

    def count_row(df):
        # Get the number of rows and columns
        num_rows = df.shape[0]     # Number of rows in the DataFrame

        return num_rows

    oppty_initial_num_rows = count_row(opportunity_df)
    
    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {oppty_initial_num_rows}")

    # =============================
    # Step 4:- Check for Already Existing Opportunities in ISC
    # =============================
    
    print("\n\n🔍 Step 4: Looking for Already existing Oppties")

    # Set paths
    legacy_csv = os.path.expanduser("~/Downloads/legacyid.csv") 

    while not os.path.exists(legacy_csv):

        print(f"\n    ❌ File 'legacyid.csv' does not exist. Did you query the Legacy Id?")

        legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

        while legacy_choice not in ['yes', 'exit']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'exit'.")
            legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

        if legacy_choice != 'yes':
            print ("\n           🚫 Skipping this Step")
            break
    
    if os.path.exists(legacy_csv):
        # Read CSV file
        csv_df = pd.read_csv(legacy_csv)
        csv_ids = set(csv_df['Opportunity_Legacy_Id__c'].dropna().astype(str))

        try:
            
            if 'opportunity_legacy_id_c' in opportunity_df.columns:
                # Ensure the column is string for safe comparison
                opportunity_df['opportunity_legacy_id_c'] = opportunity_df['opportunity_legacy_id_c'].astype(str)
                
                # Add Found/Not Found column
                opportunity_df['Already Exist'] = opportunity_df['opportunity_legacy_id_c'].apply(
                    lambda x: 'Already Exist in ISC' if x in csv_ids else 'Does not Exist in ISC'
                )

                count_not_exist = (opportunity_df['Already Exist'] == 'Already Exist in ISC').sum()

                print(f"\n    ✅ Added 'Already Exist' column ")
                if count_not_exist > 0:
                    print(f"\n    ❗️ Count of Already Exsisting Opportunities in ISC' : {count_not_exist}")
                else:
                    print(f"\n    ✅ All Opptys are new")
                
            else:
                print(f"\n    ❌ 'opportunity_legacy_id_c' not found ")
        except Exception as e:
            print(f"\n    ❌ Error processing : {e}")
    
    
    # ======================================================================
    # Step 5: Convert the email ids to lowercase and fill missing values with a default value
    # ======================================================================

    print("\n\n🔍 Step 5: Converting email ids to lowercase and filling missing values...")

    # Columns to convert to lowercase and fill blanks
    columns_to_process = ['ownerid', 'created_by']
    
    # Default value to fill in case of missing email IDs
    data_migration_email = "iscdmig2@in.ibm.com"

    
    # Initialize a dictionary to track how many blanks are filled per column
    filled_counts = {}
    for col in columns_to_process:
        filled_counts[col] = 0
    
    # Fill blank cells with specified value and count filled blanks
    for column in columns_to_process:
        if column in opportunity_df.columns:
            # Count missing values before filling
            blank_count = opportunity_df[column].isnull().sum()
            filled_counts[column] = blank_count
            
            # Fill missing values with default and convert all to lowercase
            opportunity_df[column] = opportunity_df[column].fillna(data_migration_email)
            opportunity_df[column] = opportunity_df[column].apply(lambda x: x.lower() if isinstance(x, str) else x)

        else:
            # If the column doesn't exist, print an error and exit
            print(f"\n    ❌ Error: Column '{column}' not found in the '{opportunity_sheet_name}' sheet. Terminating the Program.")
            sys.exit()
    

    # Display the count of blank columns filled for each column
    for col, count in filled_counts.items():
        if count > 0:
            print(f"\n    ❗️ Blank Values filled with Data migration Id in {col} column: {count}")
        else:
            print(f"\n    ✅ All Valid Email ids in {col} column")

    # ======================================================================
    # Step 6: Add Pricebook and RecordType id column in the sheet
    # ======================================================================

    print("\n\n🔍 Step 6: Adding Pricebook and RecordType id columns...")

    # Add the two new columns with the specified values
    opportunity_df['Pricebook2Id'] = '01s3h000003KXvoAAG'
    opportunity_df['RecordTypeId'] = '0123h000000kppcAAA'

    print("\n    ✅ 'Price_Book' and 'Record_Type' Columns added successfully.")

    # ======================================================================
    # Step 7: Change the format of the Date Column
    # ======================================================================

    print("\n\n🔍 Step 7: Formatting the Date column...")

    date_column = 'expected_close_date'

    try:
        # Check if the specified column exists
        if date_column not in opportunity_df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing in the sheet '{opportunity_sheet_name}'.")
            sys.exit(1)  # Exit the script with an error code

        # Ensure the date column is in datetime format, allowing for errors to be coerced to NaT
        opportunity_df[date_column] = pd.to_datetime(opportunity_df[date_column], errors='coerce')

        # Check for blank values (NaT) after processing
        if opportunity_df[date_column].isnull().any():
            print(f"\n    ❌ Error: The column '{date_column}' contains blank or invalid values after processing.")
            print("\n    ❗️ Please review the data, correct the issues, and try again.")
            sys.exit(1)  # Exit the script with an error code

        # Format the valid dates to YYYY-MM-DD
        opportunity_df[date_column] = opportunity_df[date_column].dt.strftime('%Y-%m-%d')

        print("\n    ✅ Date column formatted to YYYY-MM-DD successfully.")
    
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")


    # ======================================================================
    # Step 8: Create new "legacy_opportunity_split_id_c" column if it does not exist
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'legacy_opportunity_split_id_c' column...")

    # Check if the column "legacy_opportunity_split_id_c" already exists (case-insensitive)
    existing_columns = [col.lower() for col in opportunity_df.columns]
    if "legacy_opportunity_split_id_c" in existing_columns:
        print("\n    ✅ 'legacy_opportunity_split_id_c' column already exists in the sheet.")
        
        # Check for blank (NaN) values in the 'legacy_opportunity_split_id_c' column
        if opportunity_df['legacy_opportunity_split_id_c'].isnull().any():
            print("\n    ❌ Error: 'legacy_opportunity_split_id_c' column contains blank (NaN) values. Please review. Exiting process.")
            sys.exit()  # Exit the code if blank values are found
    else:
        # Check if "opportunity_legacy_id_c" column exists
        if "opportunity_legacy_id_c" not in existing_columns:
            print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found. Exiting process.")
            sys.exit()  # Exit if "opportunity_legacy_id_c" is not found
        
        # Create the new column 'legacy_opportunity_split_id_c' and populate it with 'opportunity_legacy_id_c' values
        opportunity_df['legacy_opportunity_split_id_c'] = opportunity_df['opportunity_legacy_id_c']

        print('\n    ✅ New legacy_opportunity_split_id_c column added to sheet. Process completed successfully.')

    # ======================================================================
    # Step 9: Create new column with Trimmed Account_id and Email_id column
    # ======================================================================

    print('\n\n🔍 Step 9: Creating new column with Trimmed Account_id and Email_id...\n')

    columns_to_trim = ['accountid', 'ownerid']  

    # Check if specified columns exist in the DataFrame
    missing_columns = [col for col in columns_to_trim if col not in opportunity_df.columns]

    if missing_columns:
        print(f"    ❌ Error: The following columns were not found in the sheet '{opportunity_sheet_name}': {', '.join(missing_columns)}")
        sys.exit(1)

    # Trim the values for whitespaces and create new columns for the trimmed values
    for column in columns_to_trim:
        # If the column is 'accountid', remove internal spaces in addition to trimming
        if column == 'accountid':
            opportunity_df[column] = opportunity_df[column].str.replace(r'\s+', '', regex=True).str.strip()
        else:
            opportunity_df[column] = opportunity_df[column].str.strip()

    print("    ✅ Account and Email Columns trimmed successfully, and internal spaces in 'accountid' removed.")


    # ======================================================================
    # Step 10: Remove the country code from DC Accounts
    #   • For columns with both DC and DB accounts, remove country codes from DC accounts, as they are invalid.
    #   • And keep The DB values as it is
    # ======================================================================

    print("\n\n🔍 Step 10: Processing Accounts with correct format...\n")

    accountid_column = 'accountid'
    new_column_name = 'AccountNumber'  

    # Define a function to process the values
    def process_value(value):
        if isinstance(value, str) and value.startswith('DC'):
            return value.split('-')[0]
        return value

    # Apply the function to the accountid column and store results in the new column
    opportunity_df[new_column_name] = opportunity_df[accountid_column].apply(process_value)

    print("    ✅ New column with formatted DC values created and added to the sheet successfully.")


    # ==========================================================================================
    # Step 11: Copy Extracted Data to the Main Excel File
    # ------------------------------------------------------------------------------------------
    # • This step transfers the processed account and user data into a separate sheet in the main Excel file to support lookups and validations (e.g., VLOOKUPs).
    # • If the required CSV files are not found in the designated directory, the script attempts to automatically rename and move downloaded bulk query files.
    # • If renaming fails, it prompts the user to retry the query manually, offering the exact query in clipboard for convenience.
    # ==========================================================================================


    print("\n\n🔍 Step 11: Checking for Extract Files...")

    # Define expected CSV file paths
    accounts_csv = DOWNLOADS_DIR+"/accounts.csv"  
    userid_csv = DOWNLOADS_DIR+"/userid.csv" 
    
    # Check if the CSV files exist, and prompt to retry if not

    while not os.path.exists(accounts_csv):

        print(f"\n    ❌ File 'accounts.csv' does not exist. Did you query the accounts?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            break
    if os.path.exists(accounts_csv):
        print("\n    ✅ Account CSV files is present.")

    # Check if the CSV files exist, and prompt to retry if not
    while not os.path.exists(userid_csv):
    
        print(f"\n    ❌ File 'userid.csv' does not exist. Did you query the Userid?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n         ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()
    if os.path.exists(userid_csv):
        print("\n    ✅ Userid CSV files is present.")

    # ======================================================================
    # Step 12: Check how many Accounts are present in ISC
    # ======================================================================

    print("\n\n🔍 Step 12: Checking how many Accounts are present in ISC...")


    try:
        
        # Load data from the account Extract sheets
        accounts_df = pd.read_csv(accounts_csv, usecols=[0, 1])  # Read first two columns

        # Remove rows from accounts_df where 'Id' is missing
        account_df_no_nan = accounts_df.dropna(subset=['Id'])
        
        # Identify AccountNumbers that appear more than once with different Id values
        duplicate_accounts = account_df_no_nan[
            account_df_no_nan.duplicated(subset=['AccountNumber'], keep=False)
        ]
        
        if not duplicate_accounts.empty:
            print("\n    ❗️ Duplicate AccountNumbers found with multiple Id values:")
            
            # Loop through each group of duplicates by AccountNumber
            for account_number, group in duplicate_accounts.groupby('AccountNumber'):
                print(f"\n        🔹 AccountNumber: {account_number}")
                
                # Display Ids and their corresponding row numbers in Excel
                for idx, row in group.iterrows():
                    excel_row_number = idx + 2  # Adjust for Excel row numbering
                    print(f"\n           🔸 Id: {row['Id']} (Excel Row {excel_row_number})")
                
                # Ask user to select the correct Id to retain for this AccountNumber
                valid_ids = group['Id'].tolist()
                while True:
                    chosen_id = input(f"\n        🔹 Select id for AccountNumber {account_number} from above Ids: ").strip()
                    if chosen_id in valid_ids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id from {valid_ids}. ")
                
                # Keep only the row with the chosen Id for the current AccountNumber
                accounts_df = accounts_df[
                    ~((accounts_df['AccountNumber'] == account_number) & 
                    (accounts_df['Id'] != chosen_id))
                ]
        
        # Merge the original opportunity_df with the filtered Ids from accounts_df
        merged_df = pd.merge(opportunity_df, accounts_df[['AccountNumber', 'Id']],
                            on='AccountNumber', how='left')
        
        # Count how many AccountNumbers are missing (i.e., not found in ISC)
        not_in_isc_count = merged_df["Id"].isna().sum()

        # Replace missing Ids with a placeholder text
        merged_df['Id'] = merged_df['Id'].fillna('Not in ISC')
        
        # Optionally, you could replace with AccountNumber instead of 'Not in ISC' using combine_first
        # merged_df['Id'] = merged_df['Id'].combine_first(opportunity_df['AccountNumber'])
        
        # Rename the 'Id' column to indicate ISC status
        merged_df.rename(columns={'Id': 'In ISC or Not'}, inplace=True)
        opportunity_df = merged_df
        
        # Display summary of missing accounts
        if not_in_isc_count > 0:
            print(f"\n    ❗️ Count of accounts Not in ISC: {not_in_isc_count}")
        else:
            print(f"\n    ✅ All Accounts are Present")
    
    except FileNotFoundError:
        print("\n    ❌ Error: The specified file was not found. Please check the file path.")
    except ValueError as e:
        print(f"\n    ❌ Error: {e}")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
    
    # ======================================================================
    # Step 13: Get the IDs of the Opportunity Owner
    # ======================================================================

    print("\n\n🔍 Step 13: Fetching IDs of Opportunity Owners...")
    DEFAULT_USERID = '0053h000000sdCVAAY'

    try:

        userid_df = pd.read_csv(userid_csv)  # Read first two columns
        
        # Clean and normalize the 'ownerid' column in opportunity_df
        if 'ownerid' in opportunity_df.columns:
            opportunity_df['ownerid'] = opportunity_df['ownerid'].str.strip().str.lower()
            # print("\n    ✅ 'ownerid' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'ownerid' not found in the Opportunity sheet.")
            sys.exit()

        # Clean and normalize the 'Email' column in userid_df
        if 'Email' in userid_df.columns:
            userid_df['Email'] = userid_df['Email'].str.strip().str.lower()
            # print("\n    ✅ 'Email' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Email' not found in the Opportunity_Copy sheet.")
            sys.exit()

        # Remove rows with missing 'userid' and detect duplicate Email entries with different 'userid' values
        ownerid_copy_df = userid_df.dropna(subset=['Id'])
        duplicate_emails = ownerid_copy_df[
            ownerid_copy_df.duplicated(subset=['Email'], keep=False)
        ]

        # If duplicate emails are found, prompt the user to resolve them
        if not duplicate_emails.empty:
            print("\n    ❗️ Duplicate Email IDs with multiple UserIDs found:")
            for email, group in duplicate_emails.groupby(['Email']):
                print(f"\n        📧 Email: {email}")

                # Show all UserIDs associated with the duplicated email
                for idx, row in group.iterrows():
                    excel_row = idx + 2  # Adjust for Excel row indexing (1-based + header)
                    print(f"\n           🔸 UserID: {row['Id']} (Row {excel_row})")
                
                # Collect valid UserIDs for this Email
                valid_userids = group['Id'].tolist()
                
                # Ask the user to select the correct UserID for the current email
                while True:
                    chosen_userid = input(f"\n        🔹 Select Id for UserId '{email}' from above Ids: ").strip()
                    if chosen_userid in valid_userids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id . ")
                
                # Keep only the row with the selected UserID for the current email
                userid_df = userid_df[
                    ~((userid_df['Email'] == email) & (userid_df['Id'] != chosen_userid))
                ]
            print("\n    ✅ Duplicate emails handled successfully.")

        # Perform a left join to map the 'ownerid' from opportunity_df to 'userid' in userid_df
        result_df = pd.merge(
            opportunity_df,
            userid_df[['Email', 'Id']],
            left_on='ownerid',
            right_on='Email',
            how='left'
        )
        
        # Count how many 'userid' entries are missing (NaN) before filling them
        nan_before = result_df['Id'].isna().sum()
        
        # Fill missing userids with a default fallback value
        result_df['Id'] = result_df['Id'].fillna(DEFAULT_USERID)

        # Remove redundant 'Email' column and rename 'userid' to 'OwnerId'
        result_df.drop(columns=['Email'], inplace=True)
        result_df.rename(columns={'Id': 'OwnerId'}, inplace=True)
        opportunity_df = result_df

        print("\n    ✅ Success: IDs for Opportunity Owners updated successfully.")
        
        # Notify user if any invalid userids were replaced
        if nan_before > 0:
            print(f"\n    ❗️ Number of invalid 'ownerid' values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")

    # ======================================================================
    # Step 14: To get IDs of the Created By
    # ======================================================================

    print("\n\n🔍 Step 14: Fetching IDs of 'Created By'...")

    try:

        # Check if 'created_by' column exists and contains any non-empty values
        if 'created_by' not in opportunity_df.columns or opportunity_df['created_by'].dropna().empty:
            if 'created_by' not in opportunity_df.columns:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column does not exist in 'Opportunity' sheet.")
            elif opportunity_df['created_by'].dropna().empty:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column is empty in 'Opportunity' sheet.")
        else:

            # Perform a left join (VLOOKUP-like) to map 'created_by' email to 'Id'
            merged_df = pd.merge(
                opportunity_df,
                userid_df[['Email', 'Id']],
                left_on='created_by',
                right_on='Email',
                how='left'
            )

            # Rename 'Id' to 'createdbyid' for clarity
            merged_df.rename(columns={'Id': 'createdbyid'}, inplace=True)

            # Count how many 'createdbyid' values are missing before filling them
            nan_before = merged_df['createdbyid'].isna().sum()

            # Replace missing IDs with the default fallback user ID
            merged_df['createdbyid'] = merged_df['createdbyid'].fillna(DEFAULT_USERID)

            opportunity_df = merged_df
            print("\n    ✅ Success: IDs for  'Created By' updated successfully.")
            
            # Save the updated dataframe back to the original Opportunity sheet
            if nan_before > 0:
                print(f"\n    ❗️ Number of invalid 'createdbyid' values replaced with Data Migration Id: {nan_before}")
        

    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")

    # ======================================================================
    # Step 15: Renaming Columns
    # ======================================================================


    print("\n\n🔍 Step 15: Renaming Columns...")



    def safe_rename_columns(df, column_rename_mapping):
        try:
            # Find missing columns
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]

            if missing_columns:
                print(f"\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n        🔸 {col}")
                
                while True:
                    proceed = input("\n    🔹 Do you want to proceed with renaming the available columns? (yes/no): ").strip().lower()
                    if proceed == 'yes':
                        break
                    elif proceed == 'no':
                        print("\n        ❌ Operation aborted.")
                        sys.exit(1)
                    else:
                        print("\n        ❗️ Invalid choice. Please enter 'yes' or 'no'.")

            # Rename columns
            df.rename(columns=column_rename_mapping, inplace=True)
            print(f"\n    ✅ Columns renamed successfully.")

        except Exception as e:
            print(f"\n    ❌ An unexpected error occurred while renaming columns in {e}")
            sys.exit(1)
        return df

    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunity_legacy_id_c': 'opportunity_legacy_id__c',
        'legacy_opportunity_split_id_c': 'Legacy_Opportunity_Split_Id__c',
        'In ISC or Not': 'AccountId',
        'sales_stage': 'StageName',
        'won reason': 'Won_Reason__c',
        'lost category': 'Lost_Category__c',
        'lost reason': 'Lost_Reason__c',
        'currency_code': 'CurrencyIsoCode',
        'next_step': 'NextStep',
        'oi_source': 'OI_Group__c',
        'expected_close_date': 'CloseDate',
        'ownerid': 'Email',
    }

    # Call Function to Rename the columns
    opportunity_df = safe_rename_columns(opportunity_df, column_rename_mapping)

    # ======================================================================
    # Step 16: Rearrange the Columns in the Opportunity Copy
    #   • Rearrange columns to prioritize important fields, grouping related ones (e.g., account number and account ID) and moving less important ones to the end.
    # ======================================================================


    print("\n\n🔍 Step 16: Rearranging Columns...")

    def rearrange_and_save_columns(df, desired_column_order, sheet_name):

        try:
            # Check for missing and extra columns
            missing_columns = [col for col in desired_column_order if col not in df.columns]
            extra_columns = [col for col in df.columns if col not in desired_column_order]

            # Rearrange columns
            rearranged_columns = [col for col in desired_column_order if col in df.columns]
            rearranged_columns += extra_columns  # Add extra columns to the end

            if missing_columns:
                print("\n    ❌ The following columns are missing and were skipped:")
                for col in missing_columns:
                    print(f"\n        🔸  {col}")

            if sheet_name.lower() == 'opportunity':
                if extra_columns:
                    print("\n    🔷 The following extra columns were moved to the end:")
                    for col in extra_columns:
                        print(f"\n        🔸  {col}")

            df = df[rearranged_columns]

            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"\n    ✅ Columns successfully rearranged in the '{sheet_name}' sheet of the file: {file_path.split('/')[-1]}")

        except FileNotFoundError:
            print(f"\n    ❌ File '{file_path}' not found.")
            sys.exit(1)

        return df  # Return in case you want to use it afterward


    desired_column_order = [
        'opportunity_legacy_id__c','Legacy_Opportunity_Split_Id__c','name','AccountNumber','AccountId',
    'StageName','Won_Reason__c','Lost_Category__c','Lost_Reason__c','CloseDate','CurrencyIsoCode','Email','OwnerId',
    'NextStep','OI_Group__c','created_by','createdbyid','Pricebook2Id','RecordTypeId','modified_by','created_date',
    'modified_date','pricebook','recordtypeid','accountid'
    ]
    
    # Call Function to Rearrange Columns in a sheet
    opportunity_df =rearrange_and_save_columns(opportunity_df, desired_column_order, opportunity_sheet_name)

    # ======================================================================
    # Step 17: Final Row and Column Count
    # ======================================================================


    print("\n\n🔍 Step 17: Final Row and Column Count...")

    oppty_final_num_rows = count_row(opportunity_df)

    print(f"\n    ✅ Final row count: {oppty_final_num_rows}")

    # Code to Check if the row count has changed
    def check_row_count(initial_num_rows, final_num_rows,label = ""):
        """
        Check if the row count of a DataFrame has changed.
        """
        if initial_num_rows != final_num_rows:
            print(f"\n    ❗️ Row count mismatch detected{f' in {label}' if label else ''}!")
            print(f"\n       📊 Initial: {initial_num_rows}")
            print(f"\n       📊 Final: {final_num_rows}")

            while True:
                user_input = input("\n    🔹 Do you want to continue? Type 'continue' to proceed or 'no' to abort: ").strip().lower()
                if user_input == "continue":
                    print("\n       ✅ Continuing the program...")
                    break
                elif user_input == "no":
                    print("\n       ❌ Terminating the program...")
                    sys.exit(1)
                else:
                    print("\n       ❌ Invalid input. Please type 'continue' to proceed or 'no' to stop.")

    check_row_count(oppty_initial_num_rows, oppty_final_num_rows, label = "Opportunity")
    # =========================================================================================================================================
    #                                                PRODUCT SHEET EXECUTION
    # =========================================================================================================================================


    title = "📝 PRODUCT SHEET EXECUTION 📝"
    show_title(title)

    product_sheet_name = 'Opportunity_product'

    product_df = pd.read_excel(file_path, sheet_name=product_sheet_name)
    # ======================================================================
    # Step 1: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 1: Counting the rows and columns...")

    # Function to count rows in a specific sheet
    product_initial_num_rows = count_row(product_df)

    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {product_initial_num_rows}")
    
    # ======================================================================
    # Step 2:- Removing duplicate rows and blank rows...
    # ======================================================================

    print("\n\n🔍 Step 2: Removing blank rows...")

    product_df = clean_sheet(product_df, product_sheet_name, remove_duplicates=False)


    # ======================================================================
    # Step 3 :- Add Exsising column, To check if the given Opportunities are present in the Opportunity Sheet 
    # ======================================================================

    print("\n\n🔍 Step 3: Verifying opportunities in the 'Opportunity' sheet...")
    
    def verify_opportunity(df,sheet_name):
        try:
            # opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

            # Validate the required columns
            if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet. ")
                sys.exit()
            elif 'opportunityid' not in df.columns:
                print(f"\n    ❌ Column 'opportunityid' not found in the '{sheet_name}' sheet. ")
                sys.exit()

            # Perform the comparison
            df['existing'] = df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

            # Calculate the number of false values
            false_count = (~df['existing']).sum()

            # Success message with false count
            print(f"\n    ✅ Verification completed. 'existing' column has been added to the '{sheet_name}' sheet. ")
        
            if false_count > 0:
                print(f"\n    ❗️ Number of False values in 'existing' column: {false_count}")
            else:
                print(f"\n    ✅ All Opportunities Exist In Opportunity Sheet")
        
        except Exception as e:
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e}")
            sys.exit()
        return df

    # Call the function to verify opportunities
    product_df = verify_opportunity(product_df,product_sheet_name) 

    

    # ======================================================================
    # Step 4: Formatting the date column
    # ======================================================================

    print("\n\n🔍 Step 4: Formatting the date column in the 'Opportunity_product' sheet...")

    date_column = 'expiration date'  # The column containing the dates to be formatted

    try:
        # Check if the specified column exists in the DataFrame
        if date_column not in product_df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing from the sheet '{product_sheet_name}'. ")
            sys.exit(1)   # Exit the script if the required column is not found

        # Convert the values in the date column to datetime format and standardize to 'YYYY-MM-DD'
        product_df[date_column] = pd.to_datetime(product_df[date_column]).dt.strftime('%Y-%m-%d')

        # Success message
        print(f"\n    ✅ Date column formatted successfully in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()

    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 5: Adding Quantity Columns
    # ======================================================================

    print("\n\n🔍 Step 5: Adding the 'Quantity' column in the 'Opportunity_product' sheet...")
    
    # Attempt to delete existing 'quantity' column if it already exists
    if 'quantity' in product_df.columns:
        product_df = product_df.drop(columns=['quantity'])

    # Now add a fresh 'Quantity' column to the sheet
    new_column_name = 'Quantity'  # Column name to be added
    default_value = 1  # Default value for the new column

    # Add the new column with the default value
    product_df[new_column_name] = default_value
    
    # Success message
    print(f"\n    ✅ A new column '{new_column_name}' has been added to the '{product_sheet_name}' sheet with default value '{default_value}'.")



    # ======================================================================
    # Step 6: To create a new Currency column
    # ======================================================================
    
    print("\n\n🔍 Step 6: Creating or Overwriting the 'opportunity currency' column in the 'Opportunity_product' sheet...")

    try:
        # Read both relevant sheets from the Excel file into pandas DataFrames
        # opportunity_df = pd.read_excel(file_path, sheet_name = opportunity_sheet_name)

        # Create a mapping from 'opportunity_legacy_id__c' to 'CurrencyIsoCode'
        currency_mapping = opportunity_df.set_index("opportunity_legacy_id__c")["CurrencyIsoCode"]

        # Use this mapping to populate the 'opportunity currency' column in the product sheet
        product_df["opportunity currency"] = product_df["opportunityid"].map(currency_mapping).fillna("Not Found")

        # Success message
        print("\n    ✅ Process completed. The 'opportunity currency' column has been successfully created in the 'Opportunity_product' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found.")
        sys.exit()

    except KeyError as e:
        # Handle missing column error
        print(f"\n    ❌ Error: The required column '{e.args[0]}' is missing.")
        sys.exit()

    except Exception as e:
        # Handle any other unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e}")
        sys.exit()


    # ======================================================================
    # Step 7: To delete unwanted columns from the sheet
    # ======================================================================

    print("\n\n🔍 Step 7: Deleting unwanted columns from the 'Opportunity_product' sheet...")

    # Define columns to be removed
    columns_to_delete = [
        "created_by","current quarter revenue","modified_by",
        "created_date","modified_date","product_code_family","pricebookentryid"
    ]

    try:
        unnamed_columns = []
        for col in product_df.columns:
            if pd.isna(col) or str(col).strip() == "" or str(col).startswith("Unnamed"):
                unnamed_columns.append(col)

        # Combine all columns to remove
        all_columns_to_delete = unnamed_columns + columns_to_delete
        
        # Track deleted columns
        deleted_columns = []

        # Drop only the columns that exist
        for col in all_columns_to_delete:
            if col in product_df.columns:
                product_df = product_df.drop(columns=[col])
                deleted_columns.append(col)
                # print(f"\n    🔸 '{col}'")

        if deleted_columns:
            print(f"\n    ✅ Successfully deleted {len(deleted_columns)} columns from the '{product_sheet_name}' sheet.")
        else:
            print(f"\n    ✅ No columns were deleted from the '{product_sheet_name}' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()



    # ======================================================================
    # Step 8: To get the "Product_Code_Family" column
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'Product_Code_Family' column in the 'Opportunity_product' sheet...")

    try:
        # Validate required columns
        if "product" not in product_df.columns:
            print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "product_type" not in product_df.columns:
            print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Create new column by concatenating product and product_type
        product_df["Product_Code_Family"] = product_df["product"] + "-" + product_df["product_type"]
        
        # Success message
        print(f"\n    ✅ The 'Product_Code_Family' column has been created and saved. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()



    # ======================================================================
    # Step 9: To get the "Practise_Multiple country" column
    # ======================================================================

    print("\n\n🔍 Step 9: Creating 'Practise_Multiple country' column in the 'Opportunity_product' sheet...")


    try:

        # Validate required columns
        required_columns = ["product", "product_type", "opportunity currency"]
        
        for col in required_columns:
            if col not in product_df.columns:
                print(f"\n    ❌ Error: Column '{col}' not found in '{product_sheet_name}' sheet.")
                sys.exit()

        # Create new column by concatenating product, product_type, and opportunity currency
        product_df["Practise_Multiple country"] = product_df["product"] + "-" + product_df["product_type"] + "-" + product_df["opportunity currency"]

        # Success message
        print(f"\n    ✅ The 'Practise_Multiple country' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 11: To keep the decimal values as 2
    # ======================================================================

    print("\n\n🔍 Step 11: Formatting decimal values to two decimal places...")

    headers_to_format = ['unitprice', 'expiring amount']

    try:
        # Format each specified column to 2 decimal places
        for col in headers_to_format:
            if col in product_df.columns:
                product_df[col] = pd.to_numeric(product_df[col], errors='coerce').round(2)

        print(f"\n    ✅ Numbers in columns {', '.join(headers_to_format)} have been formatted to two decimal places. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 12:- To copy the data from CSV file
    # ======================================================================

    print("\n\n🔍 Step 12: Check If the Product Extract File is present...")

    # Define the path to the 'productfamily.csv' file
    product_family_csv = DOWNLOADS_DIR+"/productfamily.csv"

    # Continuously check if the file exists
    # If not, prompt the user to retry or exit
    while not os.path.exists(product_family_csv):

        print(f"\n    ❌ File 'productFamily.csv' does not exist. Did you query the ProductFamily?")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        if try_again == 'no':
            print("\n          🚫 Exiting the program.")
            sys.exit()
    
    if os.path.exists(accounts_csv):
        print(f"\n    ✅ Product CSV file is present.")




    # ======================================================================
    # Step 13:- Create 'Practise_Multiple country' column in Product Copy sheet
    # ======================================================================

    print("\n\n🔍 Step 13: Create 'Practise_Multiple country' column in Product Copy sheet" )


    product_copy_df = pd.read_csv(product_family_csv)  # Read first two columns
    # Create a new column by combining Product Code Family and Currency with a hyphen
    product_copy_df["Practise_Multiple country"] = product_copy_df["Product2.Product_Code_Family__c"] + "-" + product_copy_df["CurrencyIsoCode"]

    print(f"\n    ✅ The values from have been successfully concatenated and saved in the 'Practise_Multiple country' column.")


    # ======================================================================
    # Step 14:- Getting the PricebookEntry id
    # ======================================================================

    print("\n\n🔍 Step 14: Getting the PricebookEntry id ...")

    # Function to standardize column names
    def standardize_columns(df):
        df.columns = df.columns.str.strip().str.lower()
        return df

    # Function to standardize column values
    def standardize_column_values(df, column_name):
        df[column_name] = df[column_name].str.strip().str.lower()
        return df

    # Standardize column names and merge key values in both DataFrames
    product_df = standardize_columns(product_df)
    product_copy_df = standardize_columns(product_copy_df)

    # Standardize column values for the merge key
    product_df = standardize_column_values(product_df, 'practise_multiple country')
    product_copy_df = standardize_column_values(product_copy_df, 'practise_multiple country')

    # Prevent accidental overwrite if 'pricebookentryid' already exists
    if 'pricebookentryid' in product_df.columns:
        raise KeyError("❌ Error: Column 'pricebookentryid' already exists in 'Opportunity_product'. Please check your data processing steps.")

    # 1️⃣ Keep only active entries and get their ids
    active_product_copy_df = product_copy_df[product_copy_df['isactive'] == True].copy()
    active_product_copy_df['pricebookentryid'] = active_product_copy_df['id']
    active_product_copy_df = active_product_copy_df[['practise_multiple country', 'pricebookentryid']].drop_duplicates(subset='practise_multiple country', keep='first')

    # 2️⃣ Merge active pricebook ids into product_df
    merged_df = pd.merge(product_df, 
                        active_product_copy_df, 
                        on='practise_multiple country', 
                        how='left')

    # 3️⃣ Find rows with no active pricebook id found (null values)
    missing_pricebook_mask = merged_df['pricebookentryid'].isna()

    # 4️⃣ Now check if any inactive entries exist for those countries
    inactive_product_copy_df = product_copy_df[product_copy_df['isactive'] == False].copy()
    inactive_countries = inactive_product_copy_df['practise_multiple country'].unique()

    # 5️⃣ Assign 'Not Active' to countries with inactive records
    merged_df.loc[merged_df['practise_multiple country'].isin(inactive_countries) & missing_pricebook_mask, 'pricebookentryid'] = 'Not Active'

    # 6️⃣ Fill remaining missing values with 'No Pricebookid found'
    merged_df['pricebookentryid'] = merged_df['pricebookentryid'].fillna('No Pricebookid found')

    # 7️⃣ Count stats
    count_no_pricebookid_found = (merged_df['pricebookentryid'] == 'No Pricebookid found').sum()
    count_not_active = (merged_df['pricebookentryid'] == 'Not Active').sum()

    product_df = merged_df

    # 9️⃣ Print completion message and stats
    print(f"\n    ✅ The 'Opportunity_product' sheet has been successfully updated with the 'PriceBookEntryid' column.")

    if count_no_pricebookid_found > 0 or count_not_active > 0:
        print(f"\n        ❗️ Count of 'No Pricebookid found': {count_no_pricebookid_found}")
        print(f"\n        ❗️ Count of 'Not Active': {count_not_active}")
    else:
        print(f"\n    ✅ All Products are Valid")





    # ======================================================================
    # Step 15: Rename the Columns
    # ======================================================================

    print("\n\n🔍 Step 15: Renaming Columns in the 'Opportunity_product' Sheet...")


    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunityid': 'Legacy_Opportunity_Split_Id__c',
        'quantity': 'Quantity',
        'product_code_family': 'Product_Family__c',
        'pricebookentryid': 'PricebookEntryId',
        'unitprice': 'UnitPrice',
        'term': 'Term__c',
        'classification type': 'Classification__c',
        'type': 'Type__c',
        'renewal type': 'Renewal_Type__c',
        'renewal status': 'Renewal_Status__c',
        'expiration date': 'Expiration_Date__c',
        'expiring term': 'Expiring_Term__c',
        'expiring amount': 'Expiring_Amount__c',
        'external id': 'External_IDs__c',
    }

    product_df = safe_rename_columns(product_df, column_rename_mapping)
    # ======================================================================
    # Step 16: Rearranging the Columns in Sequence
    # ======================================================================

    print("\n\n🔍 Step 16: Rearranging Columns in the 'Opportunity_product' Sheet...")

    # Specify the desired order of columns
    desired_column_order = [
        'Legacy_Opportunity_Split_Id__c','existing','Quantity','product','product_type','PricebookEntryId',
        'Product_Family__c','opportunity currency','practise_multiple country'
    ]

    # Call Function to Rearrange Columns in a sheet
    product_df =rearrange_and_save_columns(product_df, desired_column_order, product_sheet_name)


    # ======================================================================
    # Step 18: Final Row and Column Count
    # ======================================================================

    print("\n\n🔍 Step 18: Final Row and Column Count...")

    product_final_num_rows = count_row(product_df)

    # Display the final row count
    print(f"\n    ✅ Final row count: {product_final_num_rows}")

    # Check if the number of rows has changed
    check_row_count(product_initial_num_rows, product_initial_num_rows, label = "Opportunity_product")

    # ======================================================================

    title = "📝 PRODUCT SHEET COMPLETED 📝"
    show_title(title)
    # ======================================================================




    end_time = time.time()   # Record start time
    elapsed_time = end_time - start_time
    print(f"\n    ✅ Total time taken: {elapsed_time:.2f} seconds")
