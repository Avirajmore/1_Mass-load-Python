# Importing all the necessary Libraries
import os
import re
import sys
import shutil
import openpyxl
import pyperclip
import numpy as np
import pandas as pd
from tkinter import *
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import SheetTitleException

# =========================================================================================================================================
                                                 # FOLDER CREATION & FILE MOVEMENT
# =========================================================================================================================================


# =========================================================
# Define the base paths for storing mass load files
# =========================================================

# Path of the folder Where you want to save the Mass Load Files. 
# ❗️ Change this path if you want to store files in a different location
BASE_DIR = os.path.expanduser("~/Documents/Office Docs/Massload Files/2025") 

# for Example :-  os.path.expanduser("~/Downloads") == "/Users/avirajmore/Downloads" 

# Path of the folder Where you have saved a the template of the Summary file
# ❗️ Change this path if your summary file is stored elsewhere
REF_SUMMARY_FILE_PATH = os.path.expanduser("~/Documents/Office Docs/Massload Files/Reference File/Reference_Summary_file.xlsx") #Change it to where you want to store the summary file

# Path for the Downloads Folder
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")

# =========================================================
# Folder Creation starts
# =========================================================

# Function to display a title with a decorative line
def show_title(title):

    line_width = 100
    line = "=" * line_width
    print(f"\n{line}")
    print(title.center(line_width))
    print(f"{line}\n")

# Display the title for the folder creation and file movement process
title = "📂 FOLDER CREATION & FILE MOVEMENT 📂"
show_title(title)

# =========================================================
# Function to validate folder names 
# =========================================================

# Function to validate folder names
def is_valid_folder_name(name):
    invalid_chars = set('\\/:*?\"<>|')
    return name and not any(char in invalid_chars for char in name) # return True if name is not empty and Does not have Invalid Characters

# =========================================================
# Step 1: Create the main Sprint folder
# =========================================================

# print("\n🔍 Step 1: Creating Sprint folder")
# while True:
#     Sprint_Number = input("\n    📂 Enter the Sprint number: ").strip()
#     if is_valid_folder_name(Sprint_Number):
#         break
#     else:
#         print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

# main_folder_path = os.path.join(BASE_DIR, Sprint_Number)
# os.makedirs(main_folder_path, exist_ok=True)
# print(f"\n        ✅ Folder '{Sprint_Number}' created successfully")

# =========================================================
# Step 2: Create Current Iteration Folder and "Copy File" and "Final iteration file" folders
# =========================================================

print("\n\n🔍 Step 1: Creating Main Folder")

# Prompt the user for a valid folder name, Keep prompting the user until a valid folder name is provided 
while True:

    Current_Iteration_Folder_Name  = input("\n    📂 Enter the name of the Folder: ").strip()
    
    # Check if the folder name is valid
    if is_valid_folder_name(Current_Iteration_Folder_Name ): 
        break # If valid, exit the loop and proceed
    
    else:
        # If invalid, display an error message and prompt again
        print("\n        ❗️ Error: Invalid folder name. Please avoid using invalid characters like \\ / : * ? \" < > |.")

# Create the main folder and subfolders

# Construct a Full path for the user-specified Folder inside BASE_DIR
Current_iteration_Folder = os.path.join(BASE_DIR, Current_Iteration_Folder_Name )

# Create the main subfolder (if it doesn't already exist)
os.makedirs(Current_iteration_Folder, exist_ok=True)

# Define paths for additional subfolders "Copy Files and Final Iteration File"
copy_file_path = os.path.join(Current_iteration_Folder, "Copy files")
final_iteration_file_path = os.path.join(Current_iteration_Folder, "Final iteration files")

# Create the additional subfolders "Copy Files and Final Iteration File"
os.makedirs(copy_file_path, exist_ok=True)
os.makedirs(final_iteration_file_path, exist_ok=True)

print(f"\n        ✅ Subfolders 'Copy files' and 'Final iteration files' created successfully")

# =========================================================
# Step 3: Move Mass load files from Downloads to Current Iteration Folder
# =========================================================

# Define supported Excel file extensions
excel_extensions = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm')

# List to keep track of successfully moved files
files_moved = []

print(f"\n\n🔍 Step 2: Moving Excel files")

# Check if any Excel files exist in Downloads folder before moving them
if os.path.exists(DOWNLOADS_DIR):
    
    # Loop through all files in the Downloads folder
    for file_name in os.listdir(DOWNLOADS_DIR):
    
        # Check if the file has an Excel file extension (case-insensitive)    
        if file_name.lower().endswith(excel_extensions):
            
            # Construct full source and target file paths
            source_path = os.path.join(DOWNLOADS_DIR, file_name)
            target_path = os.path.join(Current_iteration_Folder, file_name)
            
            # Move the file from Downloads to the target folder
            shutil.move(source_path, target_path)

            # Add the file name to the list of moved files
            files_moved.append(file_name)


# Populate the excel_files list with all Excel files in the Current Iteration Folder

excel_files_Curr_Iter = [] # Initialize an empty List to store All the File

for f in os.listdir(Current_iteration_Folder):
    if f.lower().endswith(excel_extensions): # Check if the file has a valid Excel extension
        excel_files_Curr_Iter.append(f)

# Display the files that were moved from the Downloads folder (if any)

if not files_moved:
    print("\n    📥 Moved Files:")
    print("\n        1) ❗️ No files were moved from Downloads. ")
else:
    print("\n    📥 Moved Files:")
    for index, file_name in enumerate(files_moved, start=1):
        print(f"\n        {index}) {file_name} ✅") # Show each moved file with an index

# =========================================================
# Step 4: Make a Copy of Original Files and save them to "Copy Folder" 
# =========================================================

print("\n\n🔍 Step 3: Copying files")

files_copied = [] # Initialize an empty List

# Iterate through all Excel files in the current iteration folder

for file_name in excel_files_Curr_Iter: 
    
    # Rename Oriiginal File By adding "_Copy"
    copy_file_name = f"{os.path.splitext(file_name)[0]}_Copy{os.path.splitext(file_name)[1]}" 
    
    # Construct full source and target file paths
    source_path = os.path.join(Current_iteration_Folder, file_name)
    target_path = os.path.join(copy_file_path, copy_file_name)
    

    if os.path.exists(target_path):
        files_copied.append((file_name, "skipped"))
    else:
        shutil.copy(source_path, target_path)
        files_copied.append((file_name, "copied"))

# Display the results of the file copying operation

print("\n    📤 Copied Files:")
for index, (file_name, status) in enumerate(files_copied, start=1):
    if status == "copied":
        print(f"\n        {index}) {file_name} ✅") # File successfully copied
    elif status == "skipped":
        print(f"\n        {index}) {file_name} - Skipped 🛑 ") # File was skipped because it already existed


# =========================================================
# Step 5: Create Folders in "Final iteration files" to store Final processed Files and other Data
# =========================================================

print("\n\n🔍 Step 4: Creating folders in 'Final iteration files'")
Final_File_Folders_created = []

# Iterate through all Excel files in Currenct Iteration to create "Final Files" folders with that File's Name
for file_name in excel_files_Curr_Iter:

    # Extract file name without extension (.xlxs) by spliting the file name base on "." (dot) to Rename the Folders in Final Files Foldrt
    Final_File_Folder_Name = os.path.splitext(file_name)[0] 

    # Define the Folder path of Each File
    Final_File_Folder_Path = os.path.join(final_iteration_file_path, Final_File_Folder_Name) 

    # Ensure that the final folder, named after the Excel file, exists (create it if necessary).
    if not os.path.exists(Final_File_Folder_Path):
        os.makedirs(Final_File_Folder_Path, exist_ok=True)
        folder_status = "created"
    else:
        folder_status = "exists"

    # Create Given subfolders inside the main folder (always ensure they exist)
    subfolders = ["Removed Rows", "Success and error files" , "CSV Files"]
    
    for subfolder in subfolders:
        SubFolder_Path = os.path.join(Final_File_Folder_Path, subfolder)
        if not os.path.exists(SubFolder_Path):
            os.makedirs(SubFolder_Path, exist_ok=True)

    Final_File_Folders_created.append((Final_File_Folder_Name, folder_status))

# Display folder creation results
print("\n    🗂️ Folder Created:")
for index, (Final_File_Folder_Name, status) in enumerate(Final_File_Folders_created, start=1):
    if status == "created":
        print(f"\n        {index}) {Final_File_Folder_Name} ✅ (Main folder created)")
    elif status == "exists":
        print(f"\n        {index}) {Final_File_Folder_Name} - Main folder already exists 🛑 (Subfolders ensured)")

# Check if the "Copy files" folder exists before proceeding
if not os.path.exists(copy_file_path):
    print("\n     ❗️ 'Copy files' folder does not exist. ")
    sys.exit()

# Retrieve the list of files present in the "Copy files" folder

files_in_copy_folder = [] # Initialize an empty List to Save File in Copy Folder

# Iterate over all items in the directory
for f in os.listdir(copy_file_path):
    file_path = os.path.join(copy_file_path, f)
    
    # Check if the item is a file
    if os.path.isfile(file_path):
        files_in_copy_folder.append(f)

if not files_in_copy_folder:
    print("\n     🚫 No files found in 'Copy files' folder. ")
    sys.exit()
# =========================================================================================================================================
#                                                OPPORTUNITY SHEET EXECUTION
# =========================================================================================================================================


# ======================================================================
# Determine Which File to Process and Set the File Path Accordingly
# ======================================================================

while True:
    
    # Display available files for selection

    print("\n====================================================================================================")
    print("\n📂 Please select a file to process:")
    print("\n    🔸 List of Files in Copy Folder: ")

    for idx, file_name in enumerate(files_in_copy_folder, start=1):
        print(f"\n        📄 {idx}. {file_name}")

    # Prompt the user to select a file from the list
    while True:
        
        user_input = input("\n    👉 Enter the number of the file to process (or type 'exit' to quit): ").strip()
        
        # Allow the user to exit the selection process
        if user_input.lower() == 'exit':
            print("\n        ❌ File selection has been canceled. Exiting process. \n")
            sys.exit()

        try:
            # Convert user input to an index and validate selection
            selected_index = int(user_input) - 1
            
            if 0 <= selected_index < len(files_in_copy_folder):
                file_path = os.path.join(copy_file_path, files_in_copy_folder[selected_index])
                print(f"\n        ✅ You selected the file: {files_in_copy_folder[selected_index]} \n")
                break # Exit the loop if a valid file is selected

            else:
                print(f"\n        ❗ Invalid selection. Please select a number between 1 and {len(files_in_copy_folder)}.")
        
        except ValueError:
            print("\n        ❗ Invalid input. Please enter a valid number or type 'exit' to cancel.")

    # ================================================================================
    # Code to Construct Path to Folder of current Running file in "Final Iteration file" Folder
    # ================================================================================

    # Extract the selected file name from the full file path
    selected_file_name = os.path.basename(file_path.split("/")[-1])

    # Remove '_Copy' from the file name (if present) and remove the file extension
    folder_file_name = os.path.splitext(re.sub(r'_Copy', '', selected_file_name))[0]

    # Define the output folder path where processed data will be stored
    output = os.path.join(final_iteration_file_path, folder_file_name)

    # Define paths for subdirectories where different types of processed data will be saved
    csv_file_dir = os.path.join(output, "CSV Files") # Folder for storing CSV Files
    removed_rows_dir = os.path.join(output, "Removed Rows") # Folder for storing Removed Rows Files

    # ================================================================================
    # Check for missing required sheets and rename if necessary
    # ================================================================================

    symbol = "="
    print(symbol*100)

    print("\n\n🔍 Check if all the Required Sheets are present or not")
 
    # Load the Excel workbook 
    '''
        Info:
        We use openpyxl here instead of pandas because:
        - openpyxl gives access to sheet names, cell formatting, and workbook structure.
        - pandas is mainly for working with data tables (DataFrames), not the workbook structure.
    '''
    wb = openpyxl.load_workbook(file_path)
    # 📌 [NEW] Auto-Rename Known Variants to Correct Names

    # Define the list of required sheet names
    # 'Tags' is considered optional and will not be treated as missing if absent
    required_sheets = ['Opportunity', 'Opportunity_product','Opportunity_Team ', 'Reporting_codes', 'Tags']
    
    variant_mapping = {
        'Opportunity_products': 'Opportunity_product',
        'Opportunity_Team': 'Opportunity_Team '  # note the trailing space
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in variant_mapping:
            ws = wb[sheet_name]
            correct_name = variant_mapping[sheet_name]
            ws.title = correct_name
            print(f"\n    ✅  Renamed '{sheet_name}' to '{correct_name}' automatically.")

    # Get the list of sheet names present in the current workbook
    sheets_in_file = wb.sheetnames

    # Identify missing required sheets (excluding 'Tags' which is optional)
    missing_sheets = [] # Initialize Missing Sheet List

    for sheet in required_sheets:
        # Check if the sheet is not 'Tags' and is not in the list of sheets in the file
        if sheet != 'Tags' and sheet not in sheets_in_file:
            missing_sheets.append(sheet)

    # Check if all required sheets are present
    if not missing_sheets:
        # All required sheets are present — no further action needed
        print("\n    ✅ All required sheets are already present! 🎉")

    else:
        print("\n    ❌ The following required sheets are missing: ")

        # Print the missing sheets
        for i, sheet in enumerate(missing_sheets, 1):
            print(f"\n        {i}. {sheet}")

        # Identify extra sheets in the workbook that are NOT in the required list
        available_sheets = [] # Initialize Available Sheet List

        for s in sheets_in_file:
            # Check if the sheet is not in the list of required sheets
            if s not in required_sheets:
                available_sheets.append(s)

        # If there are available sheets, prompt user to rename them
        if available_sheets:
            print("\n    📋 Here are the available sheets to rename: ")
            # Display the available sheets as a numbered list
            for i, s in enumerate(available_sheets, 1):
                print(f"\n        {i}. {s}")
        
        used_indices = []  # keep track of already used sheet indices

        # Loop through each missing sheet and ask the user if they want to rename one of the available sheets
        for sheet in missing_sheets:
            if len(used_indices) == len(available_sheets):
                print(f"\n    ⏭️  No sheets available to rename. Automatically skipping '{sheet}'!")
                continue

            while True:
                choice = input(f"\n    🔸 Enter the index of the sheet to rename to '{sheet}' or type 'skip': ")

                if choice.lower() == 'skip':
                    print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
                    break

                try:
                    choice = int(choice)
                    if 1 <= choice <= len(available_sheets):
                        if choice in used_indices:
                            print("\n        ❗ That sheet has already been used. Choose a different one.")
                            continue

                        rename_sheet = available_sheets[choice - 1]
                        ws = wb[rename_sheet]
                        ws.title = sheet

                        print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")

                        used_indices.append(choice)
                        break

                    else:
                        print("\n        ❗ Invalid number selected. Please choose a valid option.")
                except ValueError:
                    print("\n        ❗ Invalid input, please enter a valid number or 'skip'.")
        # Save the modified workbook (if any renaming was done)
        wb.save(file_path)
        print("\n    💾 Workbook saved with changes!")
    
    
    # ==========================================
    # To Handle Sheets with Similar Names but Different Casing
    # ==========================================

    # Create a mapping of sheet names that may have different cases to their standard names
    sheet_name_mapping = {
        'Opportunity1': 'Opportunity',
        'Opportunity_product1': 'Opportunity_product',
        'Opportunity_team1': 'Opportunity_team',
        'Reporting_codes1': 'Reporting_codes'
    }

    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets in the workbook
    for sheet in wb.sheetnames:

        # Check if the sheet's name is present in the mapping dictionary
        if sheet in sheet_name_mapping:

            # If a match is found, get the corresponding new name
            new_name = sheet_name_mapping[sheet]
            
            # Access the worksheet with the current name
            ws = wb[sheet]

            # Rename the sheet to the mapped new name
            ws.title = new_name

    # Save the workbook with the renamed sheets (the content will remain unchanged)
    wb.save(file_path)

    # ======================================================================
    # Print Opportunity Script Execution 📝                               
    # ======================================================================
    
    # Display the title for the Opportunity Sheet Execution
    title = "📝 OPPORTUNITY SHEET EXECUTION 📝"
    show_title(title)
    
    # ======================================================================
    # Step 1: File Existence Check
    # ======================================================================
    #  i am here
    print("\n\n🔍 Step 1: Checking if the file exists...")
    def check_file_exists(file_path):
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            print(f"\n    ✅ File '{filename}' exists at the specified path.")
            return (filename)
        else:
            print("\n    ❌ Error: File does not exist or the path is invalid.\n")
            sys.exit()  # Exit the program if file is not found

    # Example usage:
    filename = check_file_exists(file_path)
    
    # ======================================================================
    #  Extract the Mpp_Number__c column and save it to a new Excel file
    # ======================================================================

    df = pd.read_excel(file_path, sheet_name="Opportunity", dtype={"Mpp_Number__c": str})

    if 'Mpp_Number__c' in df.columns:
        # Extract the two required columns
        columns_to_extract = ["opportunity_legacy_id_c", "Mpp_Number__c"]
        mpp_df = df[columns_to_extract]

        # Save to a new Excel file, ensuring Mpp_Number__c remains text
        mpp_output_file = os.path.join(csv_file_dir,"MPP_Column.xlsx")
        with pd.ExcelWriter(mpp_output_file, engine='openpyxl') as writer:
            mpp_df.to_excel(writer, index=False)

    # ======================================================================
    # Step 2: Removing Duplicates and Blank Rows
    # ======================================================================


    print("\n\n🔍 Step 2: Removing duplicate rows and blank rows...")
    
    # Define Opportunity Sheet Name
    opportunity_sheet_name = "Opportunity"

    def clean_sheet(file_path, sheet_name, remove_duplicates=False):
        try:
            # Read the spreadsheet with the given sheet name
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # Remove duplicate rows if requested
            if remove_duplicates:
                df = df.drop_duplicates()

            # Remove rows where all cells are NaN (blank rows)
            df = df.dropna(how='all')

            # Save the cleaned data back to the same file
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            if remove_duplicates:
                print(f"\n    ✅ Removed duplicate and blank rows from '{sheet_name}' sheet. ")
            else:
                print(f"\n    ✅ Removed blank rows from '{sheet_name}' sheet. ")

        except Exception as e:
            print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
            sys.exit()

    clean_sheet(file_path, opportunity_sheet_name, remove_duplicates=True)
    
    # =========================================================

    # Function to check for duplicates in the 'opportunity_legacy_id_c' column of the 'Opportunity' sheet

    def check_duplicates_in_excel(file_path):
        try:
            # Load the Excel file with openpyxl explicitly for .xlsx
            df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name, engine='openpyxl')

            # Check if 'opportunity_legacy_id_c' column exists
            if 'opportunity_legacy_id_c' not in df.columns:
                print("Error: 'opportunity_legacy_id_c' column not found in the sheet.")
                sys.exit(1)

            # Check for duplicate values
            if df['opportunity_legacy_id_c'].duplicated().any():
                print("\n    ❌ Error: Duplicate values still Present in 'opportunity_legacy_id_c' column.")
                sys.exit(1)
            
            # If no duplicates found, indicate successful validation
            print("\n    ✅ No duplicates found. Program executed successfully.")

        except FileNotFoundError:
            print("\n    ❌ Error: The specified Excel file was not found.")
            sys.exit(1)
        except ValueError:
            print("\n    ❌ Error: Failed to read the Excel file. It may be corrupted or not a valid Excel format.")
        except Exception as e:
            print(f"\n    ❌ An unexpected error occurred: {e}")
            sys.exit(1)

    # Call the function to validate uniqueness in the cleaned data
    check_duplicates_in_excel(file_path)


    # ======================================================================
    # Step 3: Check for required columns and blank values
    #   • Ensure required columns are present; prompt to continue or stop if any are missing.
    #   • Extra columns are listed but do not halt the process.
    #   • Identify blank values in critical columns, report their count, and prompt to proceed or stop.

    # ======================================================================


    print("\n\n🔍 Step 3: Checking required columns and blank values...")

    # List of required columns that must exist in the sheet (case insensitive)
    required_columns = [
        'opportunity_legacy_id_c','name','accountid','sales_stage','won reason','lost category','lost reason',
        'expected_close_date','currency_code','ownerid','next_step','oi_source','created_by','modified_by','created_date',
        'modified_date','pricebook','recordtypeid'
    ]

    # Define a subset of columns that must not contain blank (empty or null) values
    required_non_blank_columns = [
        'name',
        'accountid',
        'sales_stage',
        'expected_close_date',
        'currency_code',
        'oi_source'
    ]

    try:
        # Read the Opportunity sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    
    except Exception as e:
        print(f"\n    ❌ Error reading Excel file: {str(e)} ")
        sys.exit()

    # Normalize column names to lowercase to simplify case-insensitive comparisons
    df.columns = df.columns.str.lower()

    # Identify required columns that are missing in the DataFrame
    missing_columns = []
    for col in required_columns:
        if col.lower() not in df.columns:
            missing_columns.append(col)

    # Identify columns in the sheet that are not part of the expected list
    extra_columns = []
    for col in df.columns:
        if col.lower() not in required_columns:
            extra_columns.append(col)
    
    # Display missing and extra columns (only once)
    if missing_columns:
        print("\n    ❗ The following required columns are missing in the Opportunity sheet:")
        for col in missing_columns:
            print(f"\n        🔸 {col}")
    
    # Display any extra columns that are not needed (for awareness)
    if extra_columns:
        print("\n    ❗️ The following extra columns are present in the Opportunity sheet:")
        for col in extra_columns:
            print(f"\n        🔸 {col}")

    # Check for blank (empty or null) values in key fields
    blank_values = {}
    
    for col in required_non_blank_columns:
        if col.lower() in df.columns:
            blank_rows = df[col.lower()].isnull() | (df[col.lower()] == "")
            blank_values[col] = blank_rows.sum()

    # Check if there are any blank values in key columns or if any required columns are missing
    if any(blank_values.values()) or missing_columns:
        
        # Show blank value summary if applicable
        if any(blank_values.values()):
            print("\n    ❗️ The following columns have blank values:")
            for col, count in blank_values.items():
                if count > 0:
                    print(f"\n        🔸 {col}: {count} blank values")

        # Ask user whether to continue if missing columns or blank values are found
        while True:
            choice = input("\n    👉 Do you want to continue with the operation? (yes/no): ").strip().lower()
            if choice == 'yes':
                print("\n        ✅ Continuing the operation... ")
                break
            elif choice == 'no':
                print("\n        ❌ Operation terminated as requested. \n")
                sys.exit()
            else:
                print("\n      ❗ Invalid choice. Please enter 'yes' or 'no'.")
    else:
        print("\n    ✅ All required columns are present in the Opportunity sheet. ")


    # ======================================================================
    # Step 4: Count the rows and columns in the beginning of the process
    #   • Count rows after cleaning to ensure no discrepancies.
    #   • Compare row counts before and after processing; prompt to continue or exit if mismatched
    # ======================================================================

    print("\n\n🔍 Step 4: Counting the rows and columns...")

    def count_row(file_path, opportunity_sheet_name):

        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

        # Get the number of rows and columns
        num_rows = df.shape[0]     # Number of rows in the DataFrame

        return num_rows

    oppty_initial_num_rows = count_row(file_path, opportunity_sheet_name)
    
    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {oppty_initial_num_rows}")


    # ======================================================================

    # Step 5: Convert the headers of all the sheets to lowercase
    #   • Convert all column headers to lowercase to avoid case-sensitivity issues.
    #   • Ensure numeric columns retain their data types.

    # ======================================================================

    print("\n\n🔍 Step 5: Converting headers to lowercase...")

    # Read the Excel file with all sheets, initially treating all data as strings
    xls = pd.ExcelFile(file_path)

    # Create a dictionary to store updated DataFrames for each sheet
    sheets_dict = {}

    # List of column names that should be kept as numeric (e.g., for calculations)
    numeric_columns = ['unitprice', 'expiring amount', 'term', 'expiring term']

    # Iterate through each sheet
    for sheet_name in xls.sheet_names:
        # Read each sheet into a dataframe with all columns as strings
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        
        # Convert all column headers to lowercase for consistency
        df.columns = [col.lower() for col in df.columns]
        
        # Convert specified columns back to numeric (e.g., prices, amounts, terms)
        for col in numeric_columns:
            if col in df.columns:
                # Store original values before conversion
                original_values = df[col].copy()
                
                # Convert the column to numeric, replacing errors with NaN
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
                # Identify and display values that failed to convert
                invalid_mask = df[col].isna() & original_values.notna() & (original_values != "")
                if invalid_mask.any():
                    print(f"\n    ❗️ Warning: The following values in column '{col}' could not be converted to numeric and were set to NaN:")
                    for i, val in original_values[invalid_mask].items():
                        print(f"\n       🔸 Row {i + 2}: '{val}'")  # +2 to adjust for header and 0-based index
                
        # Save modified dataframe to dictionary
        sheets_dict[sheet_name] = df

    # Write the modified dataframes back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("\n    ✅ Headers of all sheets in the file have been converted to lowercase and data types preserved as specified.")

    # =============================
    # Defining Function to Rename bulk Query csv
    # =============================

    def rename_and_move_bulkquery_file(new_name, csv_file_dir):
        """
        Searches the downloads folder for a file with 'bulkQuery' in the name and 
        renames/moves it to the designated CSV directory using the provided new name.
        """
        
        for filename in os.listdir(DOWNLOADS_DIR):
            if "bulkQuery" in filename and filename.endswith(".csv"):
                old_path = os.path.join(DOWNLOADS_DIR, filename)
                new_path = os.path.join(DOWNLOADS_DIR, new_name)
                shutil.move(old_path, new_path)
                return True  # Successful rename and move
        return False  # No matching file found
    
    # =============================
    # Vlookup for legacy id
    # =============================
    
    print("\n\n🔍 Step 6: Looking for Already existing Oppties")

    # Set paths
    legacy_csv = os.path.expanduser("~/Downloads/legacyid.csv") 

    # Ensure the output directory exists
    os.makedirs('Delete', exist_ok=True)

    while not os.path.exists(legacy_csv):
        if rename_and_move_bulkquery_file('legacyid.csv',csv_file_dir):
            continue  # If renaming was successful, check again if the file exists

        print(f"\n    ❌ File 'legacyid.csv' does not exist. Did you query the Legacy Id?")

        # Read the "Opportunity" sheet
        df = pd.read_excel(file_path, sheet_name='Opportunity')

        # Extract unique, non-null values from 'opportunity_legacy_id_c'
        unique_values = df['opportunity_legacy_id_c'].dropna().unique()

        # Convert to string with inverted commas and comma separation
        formatted_values = ",".join(f"'{val}'" for val in unique_values)

        # Prepare the final query
        query = f"""SELECT Opportunity_Legacy_Id__c, Id,Name,Owned_By_Name__c,OwnerId 
        FROM Opportunity 
        WHERE Opportunity_Legacy_Id__c IN ({formatted_values})"""
    

    
        # Write the query to a text file
        with open('Delete/0_Legacyids.txt', 'w') as file:
            file.write(query)
        
        pyperclip.copy(query)
        
        legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

        while legacy_choice not in ['yes', 'exit']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'exit'.")
            legacy_choice = input("\n        🔸 Do you want to try again? (yes/exit): ").strip().lower()

        if legacy_choice != 'yes':
            print ("\n           🚫 Skipping this Step")
            break
        
    
    if os.path.exists(legacy_csv):
        # Read CSV file
        csv_df = pd.read_csv(legacy_csv)
        csv_ids = set(csv_df['Opportunity_Legacy_Id__c'].dropna().astype(str))

        try:
            # Read 'Opportunity' sheet
            df = pd.read_excel(file_path, sheet_name='Opportunity')
            
            if 'opportunity_legacy_id_c' in df.columns:
                # Ensure the column is string for safe comparison
                df['opportunity_legacy_id_c'] = df['opportunity_legacy_id_c'].astype(str)
                
                # Add Found/Not Found column
                df['Already Exist'] = df['opportunity_legacy_id_c'].apply(
                    lambda x: 'Already Exist in ISC' if x in csv_ids else 'Does not Exist in ISC'
                )
                
                # Save back to the same Excel file, updating the 'Opportunity' sheet
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Opportunity', index=False)
                count_not_exist = (df['Already Exist'] == 'Already Exist in ISC').sum()

                print(f"\n    ✅ Added 'Already Exist' column ")
                if count_not_exist > 0:
                    print(f"\n    ❗️ Count of Already Exsisting Opportunities in ISC' : {count_not_exist}")
                else:
                    print(f"\n    ✅ All Opptys are new")
                
            else:
                print(f"\n    ❌ 'opportunity_legacy_id_c' not found ")
        except Exception as e:
            print(f"\n    ❌ Error processing : {e}")

    # ======================================================================

    # Step 7: Convert the email ids to lowercase and fill missing values with a default value
    #   • Convert email IDs of “owner id” and “created by” column to lowercase for consistent matching (e.g., during lookups like owner ID by email).
    #   • If any of the values are blank, it should automatically fill it with Data migration id
    #   • The code will show how many blank values were filled with Data migration id for reference

    # ======================================================================


    print("\n\n🔍 Step 7: Converting email ids to lowercase and filling missing values...")

    # Columns to convert to lowercase and fill blanks
    columns_to_process = ['ownerid', 'created_by']
    
    # Default value to fill in case of missing email IDs
    data_migration_email = "iscdmig2@in.ibm.com"

    # Load the entire Excel file into memory (all sheets)
    excel_data = pd.read_excel(file_path, sheet_name=None)

    # Check if the specified sheet exists
    if opportunity_sheet_name in excel_data:
        # Access the specified sheet
        df = excel_data[opportunity_sheet_name]
        
        # Initialize a dictionary to track how many blanks are filled per column
        filled_counts = {}
        for col in columns_to_process:
            filled_counts[col] = 0
        
        # Fill blank cells with specified value and count filled blanks
        for column in columns_to_process:
            if column in df.columns:
                # Count missing values before filling
                blank_count = df[column].isnull().sum()
                filled_counts[column] = blank_count
                
                # Fill missing values with default and convert all to lowercase
                df[column] = df[column].fillna(data_migration_email)
                df[column] = df[column].apply(lambda x: x.lower() if isinstance(x, str) else x)

            else:
                # If the column doesn't exist, print an error and exit
                print(f"\n    ❌ Error: Column '{column}' not found in the '{opportunity_sheet_name}' sheet. Terminating the Program.")
                sys.exit()
        
        # Replace the existing data in the sheet with the modified values
        excel_data[opportunity_sheet_name] = df

        # Write the modified Excel data back to the file
        with pd.ExcelWriter(file_path) as writer:
            for sheet, data in excel_data.items():
                data.to_excel(writer, sheet_name=sheet, index=False)

        # Display the count of blank columns filled for each column
        for col, count in filled_counts.items():
            if count > 0:
                print(f"\n    ❗️ Blank Values filled with Data migration Id in {col} column: {count}")
            else:
                print(f"\n    ✅ All Valid Email ids in {col} column")

        
    else:
        print(f"\n    ❌ Error: Sheet '{opportunity_sheet_name}' not found in the Excel file.")


    # ======================================================================
    # Step 8: Create Blank sheets in the excel for rough work
    #   • Add blank sheets for rough work, where queried data for vlookups can be pasted.
    # ======================================================================


    print("\n\n🔍 Step 8: Creating Blank sheets for rough work...")

    # Names of the sheets to add
    sheet_names = [
        "Opportunity_Copy",
        "Opportunity_product_Copy",
        "Opportunity_team_Copy",
        "Reporting_codes_Copy",
        "Tags_Copy"
    ]

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Add new sheets
        for name in sheet_names:
            wb.create_sheet(title=name)

        # Save the workbook
        wb.save(file_path)
        print("\n    ✅ Blank Copy Sheets added successfully.")

    except Exception as e:
        print(f"\n    📝 An error occurred: {e}")


    # ======================================================================

    # Step 9: Add Pricebook and RecordType id column in the sheet
    #   • Add two new columns, "Pricebook" and "RecordType id," with predefined values for all rows.

    # ======================================================================

    print("\n\n🔍 Step 9: Adding Pricebook and RecordType id columns...")

    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Add the two new columns with the specified values
    df['Pricebook2Id'] = '01s3h000003KXvoAAG'
    df['RecordTypeId'] = '0123h000000kppcAAA'

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

    print("\n    ✅ 'Price_Book' and 'Record_Type' Columns added successfully.")


    # ======================================================================
    # Step 10: Change the format of the Date Column
    #   • Format the date column to YYYY-MM-DD.
    #   • If invalid dates are found, return an error and exit, as closeDate is critical
    # ======================================================================

    print("\n\n🔍 Step 10: Formatting the Date column...")

    date_column = 'expected_close_date'

    try:
        # Load the specific sheet into a DataFrame
        try:
            df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        except ValueError:
            print(f"\n    ❌ Error: The sheet '{opportunity_sheet_name}' is missing in the file.")
            sys.exit(1)  # Exit the script with an error code

        # Check if the specified column exists
        if date_column not in df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing in the sheet '{opportunity_sheet_name}'.")
            sys.exit(1)  # Exit the script with an error code

        # Ensure the date column is in datetime format, allowing for errors to be coerced to NaT
        df[date_column] = pd.to_datetime(df[date_column], errors='coerce')

        # Check for blank values (NaT) after processing
        if df[date_column].isnull().any():
            print(f"\n    ❌ Error: The column '{date_column}' contains blank or invalid values after processing.")
            print("\n    ❗️ Please review the data, correct the issues, and try again.")
            sys.exit(1)  # Exit the script with an error code

        # Format the valid dates to YYYY-MM-DD
        df[date_column] = df[date_column].dt.strftime('%Y-%m-%d')

        # Save the updated DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

        print("\n    ✅ Date column formatted to YYYY-MM-DD successfully.")

    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the script with an error code


    # ======================================================================
    # Step 11: Create new "legacy_opportunity_split_id_c" column if it does not exist
    #   • Skip if the column already exists.
    #   • Otherwise, create it and copy values from the "opportunity_legacy_id_c" column.

    # ======================================================================


    print("\n\n🔍 Step 11: Creating 'legacy_opportunity_split_id_c' column...")

    # Read the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Check if the column "legacy_opportunity_split_id_c" already exists (case-insensitive)
    existing_columns = [col.lower() for col in df.columns]
    if "legacy_opportunity_split_id_c" in existing_columns:
        print("\n    ✅ 'legacy_opportunity_split_id_c' column already exists in the sheet.")
        
        # Check for blank (NaN) values in the 'legacy_opportunity_split_id_c' column
        if df['legacy_opportunity_split_id_c'].isnull().any():
            print("\n    ❌ Error: 'legacy_opportunity_split_id_c' column contains blank (NaN) values. Please review. Exiting process.")
            sys.exit()  # Exit the code if blank values are found
    else:
        # Check if "opportunity_legacy_id_c" column exists
        if "opportunity_legacy_id_c" not in existing_columns:
            print("\n    ❌ Error: 'opportunity_legacy_id_c' column not found. Exiting process.")
            sys.exit()  # Exit if "opportunity_legacy_id_c" is not found
        
        # Create the new column 'legacy_opportunity_split_id_c' and populate it with 'opportunity_legacy_id_c' values
        df['legacy_opportunity_split_id_c'] = df['opportunity_legacy_id_c']

        # Write the modified DataFrame back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=opportunity_sheet_name)

        print('\n    ✅ New legacy_opportunity_split_id_c column added to sheet. Process completed successfully.')



    # ======================================================================
    # Step 12: Create new column with Trimmed Account_id and Email_id column
    #   • Remove extra spaces (including within values) from 'accountid' and 'ownerid' columns.
    #   • Create a new column with trimmed values.
    #   • Throw an error and stop if these columns are missing.
    # ======================================================================


    print('\n\n🔍 Step 12: Creating new column with Trimmed Account_id and Email_id...\n')

    columns_to_trim = ['accountid', 'ownerid']  

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    except ValueError as e:
        print(f"    ❌ Error: The sheet '{opportunity_sheet_name}' was not found in the file.")
        sys.exit(1)

    # Check if specified columns exist in the DataFrame
    missing_columns = [col for col in columns_to_trim if col not in df.columns]

    if missing_columns:
        print(f"    ❌ Error: The following columns were not found in the sheet '{opportunity_sheet_name}': {', '.join(missing_columns)}")
        sys.exit(1)

    # Trim the values for whitespaces and create new columns for the trimmed values
    for column in columns_to_trim:
        new_column_name = f'Trimmed_{column}'
        # If the column is 'accountid', remove internal spaces in addition to trimming
        if column == 'accountid':
            df[new_column_name] = df[column].str.replace(r'\s+', '', regex=True).str.strip()
        else:
            df[new_column_name] = df[column].str.strip()

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

    print("    ✅ Account and Email Columns trimmed successfully, and internal spaces in 'accountid' removed.")


    # ======================================================================
    # Step 13: Remove the country code from DC Accounts
    #   • For columns with both DC and DB accounts, remove country codes from DC accounts, as they are invalid.
    #   • And keep The DB values as it is
    # ======================================================================


    print("\n\n🔍 Step 13: Processing Accounts with correct format...\n")

    accountid_column = 'Trimmed_accountid'
    new_column_name = 'AccountNumber'  

    # Load the specific sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Define a function to process the values
    def process_value(value):
        if isinstance(value, str) and value.startswith('DC'):
            return value.split('-')[0]
        return value

    # Apply the function to the accountid column and store results in the new column
    df[new_column_name] = df[accountid_column].apply(process_value)

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

    print("    ✅ New column with formatted DC values created and added to the sheet successfully.")


    # ======================================================================
    # Step 14: Concatenate the Values
    #   • Add apostrophes and commas to account IDs and emails to format them for Salesforce query use.
    # ======================================================================

    print("\n\n🔍 Step 14: Concatenating the Values...\n")

    # Columns to process and their corresponding new column names
    columns_to_concatenate = ['AccountNumber', 'Trimmed_ownerid', 'created_by'] 
    new_column_names = ['Concatenatedaccountid', 'Concatenatedownerid', 'concatenatedcreatedby']  # Names for the new columns with concatenated values

    # Load the sheet containing opportunity data
    df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Check for missing columns before processing
    missing_columns = []
    for col in columns_to_concatenate:
        if col not in df.columns:
            missing_columns.append(col)

    # If any required columns are missing, notify the user and prompt for confirmation
    if missing_columns:
        print(f"    ❗️ The following columns are missing: {', '.join(missing_columns)}")
        user_input = input("\n    📝 Do you want to continue? (yes/no): ").lower()
        if user_input != 'yes':
            print("    ❌ Operation aborted.")
            exit()

    # Format each specified column by wrapping values in apostrophes and adding a comma
    # Example: 12345 → '12345',
    for i, column in enumerate(columns_to_concatenate):
        if column in df.columns:
            # Convert the column to string, handle NaNs by filling with empty strings
            df[column] = df[column].astype(str).fillna('')
            df[new_column_names[i]] = "'" + df[column] + "',"

    # Save the updated DataFrame back to the Excel file, replacing the existing sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

    print("    ✅ Columns concatenated and new columns with concatenated values added successfully.")


    # ======================================================================
    # Step 15: Extract Concatenated values
    #   • Save formatted values to a separate file for easier copying and pasting into the Workbench query without opening the main file.
    # ======================================================================


    print("\n\n🔍 Step 15: Extracting Concatenated values...\n")
    
    # Create the output folder if it doesn't already exist
    os.makedirs('Extracts', exist_ok=True)

    # Columns to extract (created in the previous step)
    required_columns = ["Concatenatedaccountid", "Concatenatedownerid", "concatenatedcreatedby"]

    # Check if the input file exists before proceeding
    if not os.path.exists(file_path):
        print(f"    ❌ The input file '{file_path}' does not exist.")
        
        # Create an empty DataFrame with expected columns in case the file is missing
        df = pd.DataFrame(columns=required_columns)
    else:
        # Load the Excel file and read the opportunity sheet
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

    # Create an empty DataFrame to store the final extracted data
    output_df = pd.DataFrame()

    # Process each column: clean data and add it to the output DataFrame
    for column in required_columns:
        if column in df.columns:
            
            # Remove empty and duplicate entries, then reset the index
            cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
            
            # Add cleaned data to the output with simplified column names (remove "Concatenated")
            output_df[column.replace("Concatenated", "")] = cleaned_data
        else:
            print(f"    ❌ Column '{column}' is missing in the input file.")
    
    # Save the cleaned and structured data to a new Excel file (if any data exists)
    if not output_df.empty:
        output_file = "Extracts/Account_and_Ownerid_extract.xlsx"
        output_df.to_excel(output_file, index=False)
        print(f"    ✅ Processed data has been written to Account_and_Ownerid_extract.xlsx")
    else:
        print("    ❌ No columns were processed due to missing columns.")

    # ====================================================
    # To extract account id to text file 
    # ====================================================

    # Path to the extracted Excel file (created in the previous step)
    extract_file_path = "Extracts/Account_and_Ownerid_extract.xlsx"  # Change this to your actual file path

    # Load the Excel file into a DataFrame
    df = pd.read_excel(extract_file_path)

    # Extract the "accountid" column values

    # Check if the 'accountid' column exists before processing
    if "accountid" in df.columns:
        
        # Drop empty values and convert to string format
        account_ids = df["accountid"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save the cleaned account IDs to a text file (one per line)
        with open("Delete/1_account_ids.txt", "w") as f:
            f.write("\n".join(account_ids))

    else:
        print("Column 'accountid' not found in the sheet.")
    
    # ====================================================
    # To extract user id to text file 
    # ====================================================

    # Path to the extracted Excel file
    extract_file_path = "Extracts/Account_and_Ownerid_extract.xlsx"  # Change this to your actual file path

    # Load the Excel data into a DataFrame
    df = pd.read_excel(extract_file_path)

    # Extract non-empty string values from the relevant columns
    ownerid_values = df["ownerid"].dropna().astype(str).tolist() if "ownerid" in df.columns else []
    createdby_values = df["concatenatedcreatedby"].dropna().astype(str).tolist() if "concatenatedcreatedby" in df.columns else []

    # Combine both sets of user IDs
    all_values = ownerid_values + createdby_values  # Concatenating both lists

    # Save combined user IDs to a text file (one ID per line)
    with open("Delete/2_userid.txt", "w") as f:
        f.write("\n".join(all_values))

    # ====================================================================================
    # Remove Last Character from Last Line of a File
    # ====================================================================================
    
    def remove_last_char_from_last_line(extract_file):
        try:
            # Read all lines from the file
            with open(extract_file, 'r') as file:
                lines = file.readlines()

            # Check if the file is not empty
            if lines:
                # Strip the last character from the final line
                lines[-1] = lines[-1][:-1]

                # Write the cleaned lines back to the file
                with open(extract_file, 'w') as file:
                    file.writelines(lines)
        
        except Exception as e:
            print(f"Error: {e}")
    

    # Calling Function
    remove_last_char_from_last_line('Delete/1_account_ids.txt')    
    remove_last_char_from_last_line('Delete/2_userid.txt')
        
        
    # ==========================================================================================
    # Step 16: Copy Extracted Data to the Main Excel File
    # ------------------------------------------------------------------------------------------
    # • This step transfers the processed account and user data into a separate sheet in the main Excel file to support lookups and validations (e.g., VLOOKUPs).
    # • If the required CSV files are not found in the designated directory, the script attempts to automatically rename and move downloaded bulk query files.
    # • If renaming fails, it prompts the user to retry the query manually, offering the exact query in clipboard for convenience.
    # ==========================================================================================


    print("\n\n🔍 Step 16: Copying extracted data to main file...")

    # Define expected CSV file paths
    accounts_csv = DOWNLOADS_DIR+"/accounts.csv"  
    userid_csv = DOWNLOADS_DIR+"/userid.csv" 
    
    # Check if the CSV files exist, and prompt to retry if not
    while not os.path.exists(accounts_csv):
        # Try renaming a bulkQuery file first
        if rename_and_move_bulkquery_file('accounts.csv',csv_file_dir):
            continue  # If renaming was successful, check again if the file exists

        # Read account IDs to generate the SOQL query
        with open("Delete/1_account_ids.txt", "r", encoding="utf-8") as file:
            cliptext = file.read()

        # Copy SQL query to clipboard
        account_query = f'Select AccountNumber,id from Account where AccountNumber in ({cliptext})'
        pyperclip.copy(account_query)

        print(f"\n    ❌ File 'accounts.csv' does not exist. Did you query the accounts?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            break

    # Check if the CSV files exist, and prompt to retry if not
    while not os.path.exists(userid_csv):

        if rename_and_move_bulkquery_file('userid.csv',csv_file_dir):
            continue  # If renaming was successful, check again if the file exists
        
        # Read user identifiers to generate the SOQL query
        with open("Delete/2_userid.txt", "r", encoding="utf-8") as file:
            cliptext = file.read()  # Read all lines as a single string

        # Copy the query to clipboard
        user_query = f"select email,id,Profile.Name,isactive from user where email in ({cliptext}) and Profile.Name != 'IBM Partner Community Login User' and IsActive = true "
        pyperclip.copy(user_query)
        
        print(f"\n    ❌ File 'userid.csv' does not exist. Did you query the Userid?")
        try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n         ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔸 Do you want to try again? (yes/no): ").strip().lower()
        if try_again != 'yes':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    # Read the CSV files
    accounts_df = pd.read_csv(accounts_csv, usecols=[0, 1])  # Read first two columns
    userid_df = pd.read_csv(userid_csv, usecols=[0, 1, 2, 3])  # Read first four columns

    # Load the Excel file or create a new one if it doesn't exist
    opportunity_copy_sheet_name = "Opportunity_Copy"

    if os.path.exists(file_path):
        book = openpyxl.load_workbook(file_path)
        if opportunity_copy_sheet_name not in book.sheetnames:
            sheet = book.create_sheet(title = opportunity_copy_sheet_name)
        else:
            sheet = book[opportunity_copy_sheet_name]
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = opportunity_copy_sheet_name

    # Write the headers to the "Opportunity_Copy" sheet
    for col_index, header in enumerate(accounts_df.columns, start=1):
        sheet.cell(row=1, column=col_index, value=header)
    for col_index, header in enumerate(userid_df.columns, start=4):  # Change start index to 4
        sheet.cell(row=1, column=col_index, value=header)

    # Write the data to the "Opportunity_Copy" sheet
    max_length = max(len(accounts_df), len(userid_df))

    for row_index in range(max_length):
        if row_index < len(accounts_df):
            sheet.cell(row=row_index + 2, column=1, value=accounts_df.iloc[row_index, 0])
            sheet.cell(row=row_index + 2, column=2, value=accounts_df.iloc[row_index, 1])
        if row_index < len(userid_df):
            sheet.cell(row=row_index + 2, column=4, value=userid_df.iloc[row_index, 0])  # Change column to 4
            sheet.cell(row=row_index + 2, column=5, value=userid_df.iloc[row_index, 1])  # Change column to 5
            sheet.cell(row=row_index + 2, column=6, value=userid_df.iloc[row_index, 2])  # Change column to 6
            sheet.cell(row=row_index + 2, column=7, value=userid_df.iloc[row_index, 3])  # Change column to 7

    # Save the changes
    book.save(file_path)
    print(f"\n    ✅ 'Accounts' and 'Userid' Data has been successfully copied '{file_path.split('/')[-1]}'.")


    # ======================================================================
    # Step 17: Check how many Accounts are present in ISC
    #   • Perform vlookup on the 'Accountid' column using the rough sheet to fetch Salesforce IDs.
    #   • Handle duplicate Salesforce IDs by prompting you to select one.
    #   • Populate unmatched accounts with "Not present in ISC" and display the count of such accounts.
    # ======================================================================


    print("\n\n🔍 Step 17: Checking how many Accounts are present in ISC...")


    try:
        
        # Load data from the specified Excel sheets
        opportunity_df = pd.read_excel(file_path, sheet_name = opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name = opportunity_copy_sheet_name)
        
        # Remove rows from opportunity_copy_df where 'Id' is missing
        opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['Id'])
        
        # Identify AccountNumbers that appear more than once with different Id values
        duplicate_accounts = opportunity_copy_df_no_nan[
            opportunity_copy_df_no_nan.duplicated(subset=['AccountNumber'], keep=False)
        ]
        
        if not duplicate_accounts.empty:
            print("\n    ❗️ Duplicate AccountNumbers found with multiple Id values:")
            
            # Loop through each group of duplicates by AccountNumber
            for account_number, group in duplicate_accounts.groupby('AccountNumber'):
                print(f"\n        🔹 AccountNumber: {account_number}")
                
                # Display Ids and their corresponding row numbers in Excel
                for idx, row in group.iterrows():
                    excel_row_number = idx + 2  # Adjust for Excel row numbering
                    print(f"\n           🔸 Id: {row['Id']} (Excel Row {excel_row_number})")
                
                # Ask user to select the correct Id to retain for this AccountNumber
                valid_ids = group['Id'].tolist()
                while True:
                    chosen_id = input(f"\n        🔹 Select id for AccountNumber {account_number} from above Ids: ").strip()
                    if chosen_id in valid_ids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id from {valid_ids}. ")
                
                # Keep only the row with the chosen Id for the current AccountNumber
                opportunity_copy_df = opportunity_copy_df[
                    ~((opportunity_copy_df['AccountNumber'] == account_number) & 
                    (opportunity_copy_df['Id'] != chosen_id))
                ]
        
        # Merge the original opportunity_df with the filtered Ids from opportunity_copy_df
        merged_df = pd.merge(opportunity_df, opportunity_copy_df[['AccountNumber', 'Id']],
                            on='AccountNumber', how='left')
        
        # Count how many AccountNumbers are missing (i.e., not found in ISC)
        not_in_isc_count = merged_df["Id"].isna().sum()

        # Replace missing Ids with a placeholder text
        merged_df['Id'] = merged_df['Id'].fillna('Not in ISC')
        
        # Optionally, you could replace with AccountNumber instead of 'Not in ISC' using combine_first
        # merged_df['Id'] = merged_df['Id'].combine_first(opportunity_df['AccountNumber'])
        
        # Rename the 'Id' column to indicate ISC status
        merged_df.rename(columns={'Id': 'In ISC or Not'}, inplace=True)
        
        # Save the updated merged data back to the original Excel file, replacing the existing sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            merged_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)
        
        # Display summary of missing accounts
        if not_in_isc_count > 0:
            print(f"\n    ❗️ Count of accounts Not in ISC: {not_in_isc_count}")
        else:
            print(f"\n    ✅ All Accounts are Present")

    except FileNotFoundError:
        print("\n    ❌ Error: The specified file was not found. Please check the file path.")
    except ValueError as e:
        print(f"\n    ❌ Error: {e}")
    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")


    # ======================================================================
    # Step 18: Rename 'Id' to 'userid' in Opportunity_Copy sheet
    #   • Rename the duplicate 'Id' column (from the Userid file) to 'Userid' for clarity after Step 15 merges CSV data into rough sheets.
    # ======================================================================


    # Define constants

    DEFAULT_USERID = '0053h000000sdCVAAY'

    print("\n\n🔍 Step 18: Renaming 'Id' to 'userid' in Opportunity_Copy sheet")

    try:
        # Load the Excel workbook
        wb = load_workbook(file_path)
        sheet_name = 'Opportunity_Copy'

        # Check if sheet exists
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet['E1'] = 'userid'
            wb.save(file_path)
            print("\n    ✅ Success: Renamed 'Id' to 'userid' in Opportunity_Copy sheet.")
        else:
            print(f"\n    ❗️Warning: Sheet '{sheet_name}' not found in the workbook.")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 19: Get the IDs of the Opportunity Owner
    #   • Perform vlookup on 'Ownerid' to retrieve Salesforce IDs.
    #   • Handle duplicate IDs by prompting selection.
    #   • Populate unmatched emails with the "Datamigration" Salesforce ID and display their count.
    # ======================================================================


    print("\n\n🔍 Step 19: Fetching IDs of Opportunity Owners...")

    try:

        # Load data from Excel sheets: Opportunity and Opportunity_Copy
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

        # Clean and normalize the 'Trimmed_ownerid' column in opportunity_df
        if 'Trimmed_ownerid' in opportunity_df.columns:
            opportunity_df['Trimmed_ownerid'] = opportunity_df['Trimmed_ownerid'].str.strip().str.lower()
            # print("\n    ✅ 'Trimmed_ownerid' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Trimmed_ownerid' not found in the Opportunity sheet.")
            sys.exit()

        # Clean and normalize the 'Email' column in opportunity_copy_df
        if 'Email' in opportunity_copy_df.columns:
            opportunity_copy_df['Email'] = opportunity_copy_df['Email'].str.strip().str.lower()
            # print("\n    ✅ 'Email' column cleaned.")
        else:
            print("\n    ❌ Error: Column 'Email' not found in the Opportunity_Copy sheet.")
            sys.exit()

        # Remove rows with missing 'userid' and detect duplicate Email entries with different 'userid' values
        opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['userid'])
        duplicate_emails = opportunity_copy_df_no_nan[
            opportunity_copy_df_no_nan.duplicated(subset=['Email'], keep=False)
        ]

        # If duplicate emails are found, prompt the user to resolve them
        if not duplicate_emails.empty:
            print("\n    ❗️ Duplicate Email IDs with multiple UserIDs found:")
            for email, group in duplicate_emails.groupby('Email'):
                print(f"\n        📧 Email: {email}")

                # Show all UserIDs associated with the duplicated email
                for idx, row in group.iterrows():
                    excel_row = idx + 2  # Adjust for Excel row indexing (1-based + header)
                    print(f"\n           🔸 UserID: {row['userid']} (Row {excel_row})")
                
                # Collect valid UserIDs for this Email
                valid_userids = group['userid'].tolist()
                
                # Ask the user to select the correct UserID for the current email
                while True:
                    chosen_userid = input(f"\n        🔹 Select id for UserId '{email}' from above Ids: ").strip()
                    if chosen_userid in valid_userids:
                        break
                    else:
                        print(f"\n           ❌ Invalid input. Please choose a valid Id . ")
                
                # Keep only the row with the selected UserID for the current email
                opportunity_copy_df = opportunity_copy_df[
                    ~((opportunity_copy_df['Email'] == email) & (opportunity_copy_df['userid'] != chosen_userid))
                ]
            print("\n    ✅ Duplicate emails handled successfully.")

        # Perform a left join to map the 'Trimmed_ownerid' from opportunity_df to 'userid' in opportunity_copy_df
        result_df = pd.merge(
            opportunity_df,
            opportunity_copy_df[['Email', 'userid']],
            left_on='Trimmed_ownerid',
            right_on='Email',
            how='left'
        )
        
        # Count how many 'userid' entries are missing (NaN) before filling them
        nan_before = result_df['userid'].isna().sum()
        
        # Fill missing userids with a default fallback value
        result_df['userid'] = result_df['userid'].fillna(DEFAULT_USERID)

        # Remove redundant 'Email' column and rename 'userid' to 'OwnerId'
        result_df.drop(columns=['Email'], inplace=True)
        result_df.rename(columns={'userid': 'OwnerId'}, inplace=True)

        # Write the updated result_df back to the same Excel sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            result_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

        print("\n    ✅ Success: IDs for Opportunity Owners updated successfully.")
        
        # Notify user if any invalid userids were replaced
        if nan_before > 0:
            print(f"\n    ❗️ Number of invalid 'userid' values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File not found at path: {file_path}. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 20: To get IDs of the Created By
    #   • Same as Step 18, but applied to the 'Created By' column.
    # ======================================================================


    print("\n\n🔍 Step 20: Fetching IDs of 'Created By'...")

    try:
        # Load data from the Excel sheets: 'Opportunity' and 'Opportunity_Copy'
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        opportunity_copy_df = pd.read_excel(file_path, sheet_name=opportunity_copy_sheet_name)

        # Check if 'created_by' column exists and contains any non-empty values
        if 'created_by' not in opportunity_df.columns or opportunity_df['created_by'].dropna().empty:
            if 'created_by' not in opportunity_df.columns:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column does not exist in 'Opportunity' sheet.")
            elif opportunity_df['created_by'].dropna().empty:
                print("    ❌ Skipping VLOOKUP-like operation. Reason: 'created_by' column is empty in 'Opportunity' sheet.")
        else:
            
            # Remove rows where either 'Email' or 'userid' is missing
            opportunity_copy_df_no_nan = opportunity_copy_df.dropna(subset=['Email', 'userid'])

            # Identify duplicate 'Email' entries with multiple 'userid' values
            duplicate_emails = opportunity_copy_df_no_nan[opportunity_copy_df_no_nan.duplicated(subset=['Email'], keep=False)]

            # If duplicates exist, prompt the user to resolve them manually
            if not duplicate_emails.empty:
                print("\n    ❗️ Duplicate Email IDs found with multiple UserIDs:")
                for email, group in duplicate_emails.groupby('Email'):
                    print(f"\n        📧 Email: {email}")
                    
                    # Show all 'userid' values for this email along with their Excel row numbers
                    for idx, row in group.iterrows():
                        excel_row_number = idx + 2  # Adjust for 0-based index and Excel rows starting from 2
                        print(f"\n           🔸 UserID: {row['userid']} (Excel Row {excel_row_number})")

                    # Ask the user to choose the correct UserID to retain
                    valid_userids = group['userid'].tolist()
                    while True:
                        chosen_userid = input(f"\n        🔹 Enter the UserID to keep for Email '{email}' from the above options: ").strip()
                        if chosen_userid in valid_userids:
                            break
                        else:
                            print("\n           ❌ Invalid input. Please choose a valid UserID from the options above.")

                    # Retain only the selected UserID for the current email in the dataframe
                    opportunity_copy_df = opportunity_copy_df[
                        ~((opportunity_copy_df['Email'] == email) & (opportunity_copy_df['userid'] != chosen_userid))
                    ]
                print("\n    ✅ Duplicate emails handled successfully.")

            # Perform a left join (VLOOKUP-like) to map 'created_by' email to 'userid'
            merged_df = pd.merge(
                opportunity_df,
                opportunity_copy_df[['Email', 'userid']],
                left_on='created_by',
                right_on='Email',
                how='left'
            )

            # Rename 'userid' to 'createdbyid' for clarity
            merged_df.rename(columns={'userid': 'createdbyid'}, inplace=True)

            # Count how many 'createdbyid' values are missing before filling them
            nan_before = merged_df['createdbyid'].isna().sum()

            # Replace missing IDs with the default fallback user ID
            merged_df['createdbyid'] = merged_df['createdbyid'].fillna(DEFAULT_USERID)

            # Save the updated dataframe back to the original Opportunity sheet
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                merged_df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

            print("\n    ✅ Successfully fetched and updated 'Created By' IDs.")
            
            # Save the updated dataframe back to the original Opportunity sheet
            if nan_before > 0:
                print(f"\n    ❗️ Number of invalid 'createdbyid' values replaced with Data Migration Id: {nan_before}")

    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

    except KeyError as e:
        print(f"\n    ❌ Error: Column '{e}' not found. Please check the column names in your sheets.")

    except Exception as e:
        print(f"\n    ❌ Error: An unexpected error occurred - {e}")


    # ======================================================================
    # Step 21: Renaming Columns
    #   • Rename all columns in the opportunity sheet to match API names for seamless mass loading.
    #   • Prompt to continue or abort if required columns are missing.

    # ======================================================================


    print("\n\n🔍 Step 21: Renaming Columns...")


    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunity_legacy_id_c': 'opportunity_legacy_id__c',
        'legacy_opportunity_split_id_c': 'Legacy_Opportunity_Split_Id__c',
        'In ISC or Not': 'AccountId',
        'sales_stage': 'StageName',
        'won reason': 'Won_Reason__c',
        'lost category': 'Lost_Category__c',
        'lost reason': 'Lost_Reason__c',
        'currency_code': 'CurrencyIsoCode',
        'next_step': 'NextStep',
        'oi_source': 'OI_Group__c',
        'expected_close_date': 'CloseDate',
        'ownerid': 'Email',
    }

    try:
        # Read the Excel file
        excel_data = pd.read_excel(file_path, sheet_name=None)

        if opportunity_sheet_name in excel_data:
            df = excel_data[opportunity_sheet_name]

            # Check for missing columns
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]

            if missing_columns:
                print("\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n        - {col}")
                
                while True:  # Loop until a valid input is provided
                    proceed = input("\n    ❓ Do you want to proceed with renaming the available columns? (yes/no): ").strip().lower()
                    if proceed == 'yes':
                        break  # Exit the loop and proceed with the operation
                    elif proceed == 'no':
                        print("\n        ❌ Operation aborted.")
                        sys.exit(1)  # Exit the program
                    else:
                        print("\n        ❗️ Invalid choice. Please enter 'yes' or 'no'.")  # Prompt for valid input

            # Rename columns
            df.rename(columns=column_rename_mapping, inplace=True)

            # Save the changes back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

            print("\n    ✅ Columns renamed successfully.")
        else:
            print(f"\n    ❌ Sheet '{opportunity_sheet_name}' not found in the Excel file.")

    except Exception as e:
        print(f"\n    ❌ An unexpected error occurred: {e}")
        sys.exit(1)


    # ======================================================================
    # Step 22: Rearrange the Columns in the Opportunity Copy
    #   • Rearrange columns to prioritize important fields, grouping related ones (e.g., account number and account ID) and moving less important ones to the end.
    # ======================================================================


    print("\n\n🔍 Step 22: Rearranging Columns...")

    desired_column_order = [
        'opportunity_legacy_id__c',
        'Legacy_Opportunity_Split_Id__c',
        'name',
        'AccountNumber',
        'AccountId',
        'StageName',
        'Won_Reason__c',
        'Lost_Category__c',
        'Lost_Reason__c',
        'CloseDate',
        'CurrencyIsoCode',
        'Email',
        'OwnerId',
        'NextStep',
        'OI_Group__c',
        'created_by',
        'createdbyid',
        'Pricebook2Id',
        'RecordTypeId',
        'modified_by',
        'created_date',
        'modified_date',
        'pricebook',
        'recordtypeid',
        'Trimmed_accountid',
        'Trimmed_ownerid',
        'accountid',
        'Concatenatedaccountid',
        'Concatenatedownerid',
        'concatenatedcreatedby'
    ]

    try:
        df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)

        # Check for missing and extra columns
        missing_columns = [col for col in desired_column_order if col not in df.columns]
        extra_columns = [col for col in df.columns if col not in desired_column_order]

        # Rearrange columns
        rearranged_columns = [col for col in desired_column_order if col in df.columns]
        rearranged_columns += extra_columns  # Add extra columns to the end

        # Save the changes back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df[rearranged_columns].to_excel(writer, sheet_name=opportunity_sheet_name, index=False)

        if missing_columns:
            print("\n    ❌ The following columns are missing and were skipped:")
            for col in missing_columns:
                print(f"\n        🔸  {col}")

        if extra_columns:
            print("\n    🔷 The following extra columns were moved to the end:")
            for col in extra_columns:
                print(f"\n        🔸  {col}")

        print("\n    ✅ Columns rearranged successfully.")
    except FileNotFoundError:
        print(f"\n    ❌ File '{file_path}' not found.")
        sys.exit(1)


    # ======================================================================
    # Step 23: Final Row and Column Count
    #   • Recount rows after processing to ensure no extra rows were added mistakenly.
    #   • If there's a mismatch, prompt to either continue or stop.
    # ======================================================================


    print("\n\n🔍 Step 23: Final Row and Column Count...")

    oppty_final_num_rows = count_row(file_path, opportunity_sheet_name)
    
    print(f"\n    ✅ Final row count: {oppty_final_num_rows}")

    # Code to Check if the row count has changed
    def check_row_count(initial_num_rows, final_num_rows,label = ""):
        """
        Check if the row count of a DataFrame has changed.
        """
        if initial_num_rows != final_num_rows:
            print(f"\n    ❗️ Row count mismatch detected{f' in {label}' if label else ''}!")
            print(f"\n       📊 Initial: {initial_num_rows}")
            print(f"\n       📊 Final: {final_num_rows}")

            while True:
                user_input = input("\n    🔹 Do you want to continue? Type 'continue' to proceed or 'no' to abort: ").strip().lower()
                if user_input == "continue":
                    print("\n       ✅ Continuing the program...")
                    break
                elif user_input == "no":
                    print("\n       ❌ Terminating the program...")
                    sys.exit(1)
                else:
                    print("\n       ❌ Invalid input. Please type 'continue' to proceed or 'no' to stop.")

    check_row_count(oppty_initial_num_rows, oppty_final_num_rows, label = "Opportunity")

    # =========================================================================================================================================
    #                                                PRODCUT SHEET EXECUTION
    # =========================================================================================================================================


    title = "📝 PRODUCT SHEET EXECUTION 📝"
    show_title(title)

    product_sheet_name = 'Opportunity_product'

    # ======================================================================
    # Step 1: Count the rows and columns in the beginning of the process
    # ======================================================================

    print("\n\n🔍 Step 1: Counting the rows and columns...")

    # Function to count rows in a specific sheet
    product_initial_num_rows = count_row(file_path, product_sheet_name)

    # Print the number of rows and columns
    print(f"\n    ✅ Initial row count: {product_initial_num_rows}")

    # ======================================================================
    # Step 2:- Removing duplicate rows and blank rows...
    # ======================================================================

    print("\n\n🔍 Step 2: Removing blank rows...")

    clean_sheet(file_path, product_sheet_name, remove_duplicates=False)

    # ======================================================================
    # Step 3 :- Add Exsising column, To check if the given Opportunities are present in the Opportunity Sheet 
    # ======================================================================

    print("\n\n🔍 Step 3: Verifying opportunities in the 'Opportunity' sheet...")


    try:
        # Load the sheets into DataFrames
        opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
        product_df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Validate the required columns
        if 'opportunity_legacy_id__c' not in opportunity_df.columns:
            print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet. ")
            sys.exit()
        elif 'opportunityid' not in product_df.columns:
            print(f"\n    ❌ Column 'opportunityid' not found in the '{product_sheet_name}' sheet. ")
            sys.exit()

        # Perform the comparison
        product_df['existing'] = product_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

        # Calculate the number of false values
        false_count = (~product_df['existing']).sum()

        # Save the updated data back to the Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            product_df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message with false count
        print(f"\n    ✅ Verification completed. 'existing' column has been added to the '{product_sheet_name}' sheet. ")
        if false_count > 0:
            print(f"\n    ❗️ Number of False values in 'existing' column: {false_count}")
        else:
            print(f"\n    ✅ All Opportunities Exist In Opportunity Sheet")


    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File not found. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 4: Formatting the date column
    # ======================================================================

    print("\n\n🔍 Step 4: Formatting the date column in the 'Opportunity_product' sheet...")

    date_column = 'expiration date'  # The column containing the dates to be formatted

    try:
        # Load the specific sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Check if the specified column exists in the DataFrame
        if date_column not in df.columns:
            print(f"\n    ❌ Error: The column '{date_column}' is missing from the sheet '{product_sheet_name}'. ")
            sys.exit(1)   # Exit the script if the required column is not found

        # Convert the values in the date column to datetime format and standardize to 'YYYY-MM-DD'
        df[date_column] = pd.to_datetime(df[date_column]).dt.strftime('%Y-%m-%d')

        # Write the updated DataFrame back to the Excel file, replacing the existing sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ Date column formatted successfully in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 5: Adding Quantity Columns
    # ======================================================================

    # Attempt to delete existing 'quantity' column if it already exists
    try:
        # Load workbook and target sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

            # Get the first row (header) to locate the 'quantity' column
        header_row = [cell.value for cell in sheet[1]]
        if 'quantity' in header_row:
            col_index = header_row.index('quantity') + 1

            # Shift each cell value to the left for all rows to delete the column
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=sheet.max_column):
                for cell_index, cell in enumerate(row[:-1]):
                    cell.value = row[cell_index + 1].value
                row[-1].value = None  # Clear the last column cell

        # Save changes
        wb.save(file_path)
    except Exception:
        # Silently pass if there's an error (e.g. column not found)
        pass

    # Now add a fresh 'Quantity' column to the sheet
    print("\n\n🔍 Step 5: Adding the 'Quantity' column in the 'Opportunity_product' sheet...")

    new_column_name = 'Quantity'  # Column name to be added
    default_value = 1  # Default value for the new column

    try:
        # Reload the workbook to work with updated sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        # Add the new column header at the end of the first row
        sheet.cell(row=1, column=sheet.max_column + 1, value=new_column_name)

        # Populate the new column with the default value for all data rows
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming headers in row 1)
            sheet.cell(row=row, column=sheet.max_column, value=default_value)

        # Save the workbook
        wb.save(file_path)

        # Success message
        print(f"\n    ✅ A new column '{new_column_name}' has been added to the '{product_sheet_name}' sheet with default value '{default_value}'.")

    except FileNotFoundError:
        # Handle file not found
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 6: To create a new Currency column
    # ======================================================================
    print("\n\n🔍 Step 6: Creating or Overwriting the 'opportunity currency' column in the 'Opportunity_product' sheet...")

    try:
        # Read both relevant sheets from the Excel file into pandas DataFrames
        opportunity_product_df = pd.read_excel(file_path, sheet_name = product_sheet_name)
        opportunity_df = pd.read_excel(file_path, sheet_name = opportunity_sheet_name)

        # Create a mapping from 'opportunity_legacy_id__c' to 'CurrencyIsoCode'
        currency_mapping = opportunity_df.set_index("opportunity_legacy_id__c")["CurrencyIsoCode"]

        # Use this mapping to populate the 'opportunity currency' column in the product sheet
        opportunity_product_df["opportunity currency"] = opportunity_product_df["opportunityid"].map(currency_mapping).fillna("Not Found")

        # Save the updated DataFrame back to the Excel file, replacing the sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            opportunity_product_df.to_excel(writer, sheet_name= product_sheet_name, index=False)

        # Success message
        print("\n    ✅ Process completed. The 'opportunity currency' column has been successfully created in the 'Opportunity_product' sheet.")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found.")
        sys.exit()

    except KeyError as e:
        # Handle missing column error
        print(f"\n    ❌ Error: The required column '{e.args[0]}' is missing.")
        sys.exit()

    except Exception as e:
        # Handle any other unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e}")
        sys.exit()

    # ======================================================================
    # Step 7: To delete unwanted columns from the sheet
    # ======================================================================

    print("\n\n🔍 Step 7: Deleting unwanted columns from the 'Opportunity_product' sheet...")

    # Define columns to be removed
    columns_to_delete = [
        "created_by",
        "current quarter revenue",
        "modified_by",
        "created_date",
        "modified_date",
        "product_code_family",
        "pricebookentryid"
    ]

    try:
        # Load the workbook and target the specified sheet
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

        columns_deleted = []

        # Extract the header row (first row in sheet)
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Delete unnamed columns starting from the right (to prevent shifting issues)
        unnamed_column_indexes = [
            idx + 1 for idx, col in enumerate(header)
            if col is None or str(col).strip() == "" or str(col).startswith("Unnamed:")
        ]

        # Delete unnamed columns from right to left (to preserve indexes)
        if unnamed_column_indexes:
            print("\n    ❗️ Removing columns with blank or unnamed headers:")
            for col_idx in sorted(unnamed_column_indexes, reverse=True):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"\n        🔸 Column '{header[col_idx - 1]}'")
                sheet.delete_cols(col_idx)
        
        # Delete each user-specified column by matching column name
        for col_name in columns_to_delete:
            found = False
            for col in sheet.iter_cols():
                if col[0].value == col_name:
                    sheet.delete_cols(col[0].column)
                    columns_deleted.append(col_name)
                    found = True
                    # print(f"\n        🔸 '{col_name}' deleted. ")
                    break
            if not found:
                print(f"\n        🔸 '{col_name}' not found. ")

        # Save the workbook
        wb.save(file_path)

        # Success or no deletion message
        if columns_deleted:
            print(f"\n    ✅ Successfully deleted mentioned columns from the '{product_sheet_name}' sheet:")
        else:
            print(f"\n    ❌ No columns from the specified list were found in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()

    # ======================================================================
    # Step 8: To get the "Product_Code_Family" column
    # ======================================================================

    print("\n\n🔍 Step 8: Creating 'Product_Code_Family' column in the 'Opportunity_product' sheet...")

    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Validate required columns
        if "product" not in df.columns:
            print(f"\n    ❌ Error: Column 'product' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "product_type" not in df.columns:
            print(f"\n    ❌ Error: Column 'product_type' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Create new column by concatenating product and product_type
        df["Product_Code_Family"] = df["product"] + "-" + df["product_type"]

        # Save the updated DataFrame back to the sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ The 'Product_Code_Family' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 9: To get the "Practise_Multiple country" column
    # ======================================================================

    print("\n\n🔍 Step 9: Creating 'Practise_Multiple country' column in the 'Opportunity_product' sheet...")


    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Validate required columns
        required_columns = ["product", "product_type", "opportunity currency"]
        for col in required_columns:
            if col not in df.columns:
                print(f"\n    ❌ Error: Column '{col}' not found in '{product_sheet_name}' sheet.")
                sys.exit()

        # Create new column by concatenating product, product_type, and opportunity currency
        df["Practise_Multiple country"] = df["product"] + "-" + df["product_type"] + "-" + df["opportunity currency"]

        # Save the updated DataFrame back to the sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ The 'Practise_Multiple country' column has been created and saved in the '{product_sheet_name}' sheet. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 10: To Concatenate the Currency and Product Family
    # ======================================================================

    print("\n\n🔍 Step 10: Concatenating 'Currency' and 'Product Family' columns...")

    try:
        # Load the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=product_sheet_name)

        # Validate that required columns exist in the sheet
        if "Product_Code_Family" not in df.columns:
            print(f"\n    ❌ Error: Column 'Product_Code_Family' not found in '{product_sheet_name}' sheet. ")
            sys.exit()
        elif "opportunity currency" not in df.columns:
            print(f"\n    ❌ Error: Column 'opportunity currency' not found in '{product_sheet_name}' sheet. ")
            sys.exit()

        # Create a new column by wrapping values from 'Product_Code_Family' in single quotes followed by a comma
        df["Concatenated Product Family"] = "'" + df["Product_Code_Family"] + "',"

        # Do the same for the 'opportunity currency' column
        df["Concatenated Currency"] = "'" + df["opportunity currency"] + "',"

        # Save the updated DataFrame back to the same sheet
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=product_sheet_name, index=False)

        # Success message
        print(f"\n    ✅ 'Concatenated Product Family' and 'Concatenated Currency' columns have been added successfully. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 11: To keep the decimal values as 2
    # ======================================================================

    print("\n\n🔍 Step 11: Formatting decimal values to two decimal places...")

    headers_to_format = ['unitprice', 'expiring amount']

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[product_sheet_name]

    # Helper function to format values as float with 2 decimal places
        def format_to_float(value):
            if isinstance(value, (int, float)):
                return float(f'{value:.2f}')
            return value

        # Identify the column indices of the target headers
        column_indices = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header in headers_to_format:
                column_indices[header] = col

        # Format each cell under the specified columns
        for header in headers_to_format:
            col = column_indices.get(header)
            if col:
                # print(f"    🔄 Formatting '{header}' column...")
                for row in range(2, sheet.max_row + 1):  # Start from row 2 assuming headers in row 1
                    cell = sheet.cell(row=row, column=col)
                    formatted_value = format_to_float(cell.value)
                    sheet.cell(row=row, column=col, value=formatted_value)

        # Save the workbook
        wb.save(file_path)

        # Success message
        print(f"\n    ✅ Numbers in columns {', '.join(headers_to_format)} have been formatted to two decimal places. ")

    except FileNotFoundError:
        # Handle file not found error
        print(f"\n    ❌ Error: File '{file_path}' not found. ")
        sys.exit()
    except KeyError:
        # Handle missing sheet error
        print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the file. ")
        sys.exit()
    except Exception as e:
        # Handle any unexpected errors
        print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ")
        sys.exit()


    # ======================================================================
    # Step 12: To extract the concatenated values
    # ======================================================================

    print("\n\n🔍 Step 12: Extracting concatenated values...")

    def process_excel_file(file_path, sheet_name, required_columns, output_file):
        # Check if the input file exists
        if not os.path.exists(file_path):
            print(f"\n    ❌ Error: The input file '{file_path}' does not exist. ")
            return

        try:
            # Read the Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            print(f"\n    ❌ Error: Failed to read the Excel file. Details: {e} ")
            return

        # Dictionary to store cleaned (unique & non-blank) data for each column
        cleaned_data_dict = {}

        # Process each required column
        for column in required_columns:
            if column in df.columns:

                # Remove empty (NaN) and duplicate values
                cleaned_data = df[column].dropna().drop_duplicates().reset_index(drop=True)
                cleaned_data_dict[column.replace("Concatenated", "").strip()] = cleaned_data
            else:
                print(f"\n    ❌ Error: Column '{column}' is missing. ")

        # Initialize an output DataFrame and add each cleaned series
        output_df = pd.DataFrame()

        # Add each cleaned column as a separate DataFrame and concatenate them
        for key, cleaned_data in cleaned_data_dict.items():
            output_df = pd.concat([output_df, pd.DataFrame({key: cleaned_data})], axis=1, ignore_index=False)

        # Write the processed data to a new Excel file if there's any data to write
        if not output_df.empty:
            try:
                output_df.to_excel(output_file, index=False)
                print(f"\n    ✅ Data written to '{output_file}'. ")
            except Exception as e:
                print(f"\n    ❌ Error: Failed to write the Excel file. Details: {e} ")
        else:
            print("\n    ❌ Error: No data to process. ")

    # Define sheet, columns, and destination file for extraction
    sheet_name = product_sheet_name  # Specify the sheet name
    required_columns = ["Concatenated Product Family", "Concatenated Currency"]
    output_file = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Specify the output file path

    # Process the Excel file
    process_excel_file(file_path, sheet_name, required_columns, output_file)


    # ======================================================================
    # To extact product code to text file
    # ======================================================================

    import pandas as pd
    # Load the Excel file
    extract_file_path = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Change this to your actual file path
    df = pd.read_excel(extract_file_path)

    # Extract the "accountid" column values
    if "Product Family" in df.columns:
        Product_Family = df["Product Family"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/3_product_code.txt", "w") as f:
            f.write("\n".join(Product_Family))

    else:
        print(f"Column not found in the sheet.")

    # ======================================================================
    # To extact currency code to text file
    # ======================================================================

    # Specify the path of the Excel file containing the extracted data
    extract_file_path = "Extracts/ProductFamily_and_Currency_extract.xlsx"  # Change this to your actual file path

    # Load the Excel file into a DataFrame
    df = pd.read_excel(extract_file_path)

    # Check if the "Currency" column exists
    if "Currency" in df.columns:
        currency_values = df["Currency"].dropna().astype(str)  # Drop NaN values and convert to string

        # Save to a text file
        with open("Delete/4_currency.txt", "w") as f:
            f.write("\n".join(currency_values))

    else:
        print(f"Column not found in the sheet.")
    
    # ========================================================================  
    
    # Remove the last character (e.g., trailing comma) from the last line of both files
    remove_last_char_from_last_line('Delete/3_product_code.txt')    
    remove_last_char_from_last_line('Delete/4_currency.txt')
    
    # ========================================================================  
    # Step: Read the cleaned text files and prepare SOQL query
    # ========================================================================  
    
    # Read the entire contents of the product code text file as a string
    with open("Delete/3_product_code.txt", "r", encoding="utf-8") as file:
        product_code_txt = file.read()  # Read all lines as a single string

    # Read the entire contents of the currency text file as a string
    with open("Delete/4_currency.txt", "r", encoding="utf-8") as file:
        currency_txt = file.read()  # Read all lines as a single string

    # Construct the SOQL query using the extracted product codes and currency codes
    pricebook_query = f'select  Product2.Product_Code_Family__c,CurrencyIsoCode,id,isactive from PricebookEntry where Product2.Product_Code_Family__c in ({product_code_txt}) and CurrencyIsoCode in ({currency_txt})'
    
    # Copy the final query to the clipboard for easy use
    pyperclip.copy(pricebook_query)

    # ======================================================================
    # Step 13:- To copy the data from CSV file
    # ======================================================================

    print("\n\n🔍 Step 13: Copying data from CSV file to Excel...")

    # Define the path to the 'productfamily.csv' file
    product_family_csv = DOWNLOADS_DIR+"/productfamily.csv"

    # Continuously check if the file exists
    # If not, prompt the user to retry or exit
    while not os.path.exists(product_family_csv):

        if rename_and_move_bulkquery_file('productfamily.csv',csv_file_dir):
            continue  # If renaming was successful, check again if the file exists

        print(f"\n    ❌ File 'productFamily.csv' does not exist. Did you query the ProductFamily?")
        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        while try_again not in ['yes', 'no']:
            print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
            try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
        if try_again == 'no':
            print("\n          🚫 Exiting the program.")
            sys.exit()

    # Read data from the CSV file
    df = pd.read_csv(product_family_csv)

    # Specify the Excel file path and sheet name
    product_copy_sheet_name = "Opportunity_product_Copy"

    # Append the data to the Excel file, replacing the existing sheet if it exists  
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=product_copy_sheet_name, index=False)

    print(f"\n    ✅ Data from the CSV file has been successfully copied to the '{product_copy_sheet_name}' sheet.")


    # ======================================================================
    # Step 14:- Create 'Practise_Multiple country' column in Product Copy sheet
    # ======================================================================

    print("\n\n🔍 Step 14: Create 'Practise_Multiple country' column in Product Copy sheet" )

    # Read the copied product data from the Excel sheet
    df = pd.read_excel(file_path, sheet_name= product_copy_sheet_name)

    # Create a new column by combining Product Code Family and Currency with a hyphen
    df["Practise_Multiple country"] = df["Product2.Product_Code_Family__c"] + "-" + df["CurrencyIsoCode"]

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name = product_copy_sheet_name, index=False)

    print(f"\n    ✅ The values from have been successfully concatenated and saved in the 'Practise_Multiple country' column.")

    # ======================================================================
    # Step 15:- Getting the PricebookEntry id
    # ======================================================================

    print("\n\n🔍 Step 15: Getting the PricebookEntry id ...")

    # Function to standardize column names
    def standardize_columns(df):
        df.columns = df.columns.str.strip().str.lower()
        return df

    # Function to standardize column values
    def standardize_column_values(df, column_name):
        df[column_name] = df[column_name].str.strip().str.lower()
        return df

    # Read both the original and copied product data from Excel
    product_df = pd.read_excel(file_path, sheet_name=product_sheet_name)
    product_copy_df = pd.read_excel(file_path, sheet_name=product_copy_sheet_name)

    # Standardize column names and merge key values in both DataFrames
    product_df = standardize_columns(product_df)
    product_copy_df = standardize_columns(product_copy_df)

    # Standardize column values for the merge key
    product_df = standardize_column_values(product_df, 'practise_multiple country')
    product_copy_df = standardize_column_values(product_copy_df, 'practise_multiple country')

    # Prevent accidental overwrite if 'pricebookentryid' already exists
    if 'pricebookentryid' in product_df.columns:
        raise KeyError("❌ Error: Column 'pricebookentryid' already exists in 'Opportunity_product'. Please check your data processing steps.")

    # 1️⃣ Keep only active entries and get their ids
    active_product_copy_df = product_copy_df[product_copy_df['isactive'] == True].copy()
    active_product_copy_df['pricebookentryid'] = active_product_copy_df['id']
    active_product_copy_df = active_product_copy_df[['practise_multiple country', 'pricebookentryid']].drop_duplicates(subset='practise_multiple country', keep='first')

    # 2️⃣ Merge active pricebook ids into product_df
    merged_df = pd.merge(product_df, 
                        active_product_copy_df, 
                        on='practise_multiple country', 
                        how='left')

    # 3️⃣ Find rows with no active pricebook id found (null values)
    missing_pricebook_mask = merged_df['pricebookentryid'].isna()

    # 4️⃣ Now check if any inactive entries exist for those countries
    inactive_product_copy_df = product_copy_df[product_copy_df['isactive'] == False].copy()
    inactive_countries = inactive_product_copy_df['practise_multiple country'].unique()

    # 5️⃣ Assign 'Not Active' to countries with inactive records
    merged_df.loc[merged_df['practise_multiple country'].isin(inactive_countries) & missing_pricebook_mask, 'pricebookentryid'] = 'Not Active'

    # 6️⃣ Fill remaining missing values with 'No Pricebookid found'
    merged_df['pricebookentryid'] = merged_df['pricebookentryid'].fillna('No Pricebookid found')

    # 7️⃣ Count stats
    count_no_pricebookid_found = (merged_df['pricebookentryid'] == 'No Pricebookid found').sum()
    count_not_active = (merged_df['pricebookentryid'] == 'Not Active').sum()

    # 8️⃣ Save the updated DataFrame back to the same Excel file
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        merged_df.to_excel(writer, sheet_name=product_sheet_name, index=False)

    # 9️⃣ Print completion message and stats
    print(f"\n    ✅ The 'Opportunity_product' sheet has been successfully updated with the 'PriceBookEntryid' column.")

    if count_no_pricebookid_found > 0 or count_not_active > 0:
        print(f"\n        ❗️ Count of 'No Pricebookid found': {count_no_pricebookid_found}")
        print(f"\n        ❗️ Count of 'Not Active': {count_not_active}")
    else:
        print(f"\n    ✅ All Products are Valid")


    # ======================================================================
    # Step 16: Rearranging the Columns in Sequence
    # ======================================================================

    print("\n\n🔍 Step 16: Rearranging Columns in the 'Opportunity_product' Sheet...")


    # Specify the desired order of columns
    desired_column_order = [
        'opportunityid',
        'existing',
        'quantity',
        'product',
        'product_type',
        'product_code_family',
        'opportunity currency',
        'practise_multiple country',
        'pricebookentryid'
    ]

    try:
        # Read the Excel file
        excel_data = pd.read_excel(file_path, sheet_name=product_sheet_name)
        
        # Check if the sheet exists
        if isinstance(excel_data, pd.DataFrame):
            # Check if all specified columns exist
            missing_columns = [col for col in desired_column_order if col not in excel_data.columns]
            extra_columns = [col for col in excel_data.columns if col not in desired_column_order]

            # Rearrange columns
            rearranged_columns = [col for col in desired_column_order if col in excel_data.columns]

            # Add extra columns to the end
            rearranged_columns += extra_columns

            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                excel_data[rearranged_columns].to_excel(writer, sheet_name=product_sheet_name, index=False)

            # Notify the user about the changes
            if missing_columns:
                print("\n    ❌ The following columns were missing and were skipped:")
                for col in missing_columns:
                    print(f"\n      🔸 {col}")
            
            # if extra_columns:
            #     print("\n    🔄 The following extra columns were moved to the end of the sheet:")
            #     for col in extra_columns:
            #         print(f"\n      🔸 {col}")

            print(f"\n    ✅ Columns successfully rearranged in the '{product_sheet_name}' sheet of the file: {file_path.split('/')[-1]}")
        else:
            print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path}' not found.")


    # ======================================================================
    # Step 17: Rename the Columns
    # ======================================================================

    print("\n\n🔍 Step 17: Renaming Columns in the 'Opportunity_product' Sheet...")


    # Dictionary mapping old column names to new column names
    column_rename_mapping = {
        'opportunityid': 'Legacy_Opportunity_Split_Id__c',
        'quantity': 'Quantity',
        'product_code_family': 'Product_Family__c',
        'pricebookentryid': 'PricebookEntryId',
        'unitprice': 'UnitPrice',
        'term': 'Term__c',
        'classification type': 'Classification__c',
        'type': 'Type__c',
        'renewal type': 'Renewal_Type__c',
        'renewal status': 'Renewal_Status__c',
        'expiration date': 'Expiration_Date__c',
        'expiring term': 'Expiring_Term__c',
        'expiring amount': 'Expiring_Amount__c',
        'external id': 'External_IDs__c',
    }

    # Read the Excel file
    try:
        excel_data = pd.read_excel(file_path, sheet_name=None)
        
        # Check if the specified sheet exists
        if product_sheet_name in excel_data:
            # Access the specified sheet
            df = excel_data[product_sheet_name]
            
            # Check if all specified columns exist
            missing_columns = [col for col in column_rename_mapping.keys() if col not in df.columns]
            
            # If any specified column is missing, notify the user and ask if they want to proceed
            if missing_columns:
                print("\n    ❌ The following columns are missing and cannot be renamed:")
                for col in missing_columns:
                    print(f"\n      🔸 {col}")
                
                # Loop until a valid response is entered
                while True:
                    proceed = input("\n    🔹 Do you want to proceed with the execution? (yes/no): ").lower()
                    if proceed == 'yes':
                        break
                    elif proceed == 'no':
                        print("\n      🚫 Operation aborted by the user.")
                        exit()
                    else:
                        print("\n      ❗️ Invalid input. Please enter 'yes' or 'no'.")
            
            # Rename specified columns
            df.rename(columns=column_rename_mapping, inplace=True)
            
            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=product_sheet_name, index=False)
            
            print(f"\n    ✅ Columns renamed successfully in the '{product_sheet_name}' sheet of the file: {file_path.split('/')[-1]}")
        
        else:
            print(f"\n    ❌ Error: Sheet '{product_sheet_name}' not found in the Excel file.")
    except FileNotFoundError:
        print(f"\n    ❌ Error: File '{file_path.split('/')[-1]}' not found.")

    # ======================================================================
    # Step 18: Final Row and Column Count
    # ======================================================================

    print("\n\n🔍 Step 18: Final Row and Column Count...")

    product_final_num_rows = count_row(file_path, product_sheet_name)

    # Display the final row count
    print(f"\n    ✅ Final row count: {product_final_num_rows}")

    # Check if the number of rows has changed
    check_row_count(product_initial_num_rows, product_initial_num_rows, label = "Opportunity_product")

    # ======================================================================

    title = "📝 PRODUCT SHEET COMPLETED 📝"
    show_title(title)
    # ======================================================================


    # =========================================================================================================================================
    #                                                TEAM MEMBER SHEET EXECUTION
    # =========================================================================================================================================
    
    
    # Display the header once
    print("\n\n📄 Execute Next Sheet:")

    original_team_sheet = 'Opportunity_Team '
    
    def is_sheet_empty(file_path, original_team_sheet):
        """
        Check if a given Excel sheet is empty or contains only headers.

            Returns:
                - (True, None): If the sheet is empty or has only headers.
                - (False, DataFrame): If the sheet contains data (returns first 4 rows).
                - (None, None): If an error occurs while reading the sheet.
        """
        try:
            df = pd.read_excel(file_path, sheet_name=original_team_sheet)
            
            # Check if the sheet is empty or only contains headers
            if df.empty or df.dropna(how='all').shape[0] == 0:
                return True, None  # Sheet is empty or has only headers
            
            return False, df.head(4)  # Sheet contains data, return first 4 rows
        except Exception as e:
            print(f"\n❗️ Error reading sheet '{original_team_sheet}': {e}\n")
            return None, None
        
    # Check if Opportunity_Team sheet has data
    is_empty, preview = is_sheet_empty(file_path, original_team_sheet)

    # Display results based on sheet contents
    if is_empty:
        print(f"\n📂 The sheet '{original_team_sheet}' is empty or contains only headers.")
        choice = 'no'
    elif is_empty is None:
        print("\n❗️ Could not process the sheet due to an error.\n")
        choice = 'no'
    else:
        choice = 'yes'
        print(f"\n✅ The sheet '{original_team_sheet}' contains data. Here are the first 4 rows:\n")
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    # Ask user to proceed with execution of Team Member sheet
    while True:
        print(f"\n    🔹 Do you want to execute the Team member Sheet ? (yes/no): {choice}")
        if choice == "yes":
            team_member_choice = 'yes'
            print(f"\n        ⏳ Executing the Sheet: Teammember sheet ")
            
            # ======================================================================

            title = "📝 TEAM MEMBER SHEET EXECUTION 📝"
            show_title(title)

            # ======================================================================

            # ======================================================================
            # Step 1:-  To rename the sheet to Opportunity Team
            # ======================================================================

            print("\n\n🔍 Step 1: Renaming Team Sheet...")

            # Target sheet name to find and rename
            Opportunity_team_sheet_name = 'Opportunity_team'  # Modified to lowercase 'team'

            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)

            # Function to normalize and clean sheet names for comparison
            def normalize_sheet_name(name):
                return name.strip().replace(' ', '').replace('-', '').lower().replace('_', '')

            # Flag to track if the sheet is found
            sheet_found = False
            sheet_to_rename = None

            # Iterate through sheet names and rename if found
            for sheet_name in workbook.sheetnames:
                if normalize_sheet_name(sheet_name) == normalize_sheet_name(Opportunity_team_sheet_name):
                    workbook[sheet_name].title = 'Opportunity_team_2'  # Renaming to lowercase 'team'
                    sheet_found = True
                    sheet_to_rename = sheet_name
                    print(f"\n    ✅ Sheet '{sheet_name}' has been successfully renamed to 'Opportunity_team_2'.")
                    break

            # If sheet not found, raise an error with the name of the sheet
            if not sheet_found:
                sheet_names = ", ".join(workbook.sheetnames)
                print(f"\n    ❌ ERROR: Sheet '{Opportunity_team_sheet_name}' not found.")
                sys.exit()

            # Save the workbook
            workbook.save(file_path)

            # ======================================================================
            # Step 2:- Check count of rows and columns
            # ======================================================================

            print("\n\n🔍 Step 2: Checking the Number of Rows and Columns...")

            # Name of the sheet to target
            team_sheet2 = 'Opportunity_team_2'

            # Function to count rows in a specific sheet
            team__initial_num_rows = count_row(file_path, team_sheet2)
            
            # Print the number of rows and columns
            print(f"\n    ✅ Initial rows count '{team_sheet2}' sheet: {team__initial_num_rows}\n")

            # ======================================================================
            # Step 3: Creating New Records for Multiple Emails in a Cell
            # ======================================================================

            print("\n\n🔍 Step 3: Creating New Records for Multiple Emails in a Cell...")

            # Load the existing workbook
            wb = load_workbook(file_path)

            # Load the DataFrame from the 'Opportunity_team_2' sheet
            df_opportunity_team_2 = pd.read_excel(file_path, sheet_name = team_sheet2)

            # Initialize an empty list to store rows for the new DataFrame
            new_rows = []
            duplicate_count = 0  # Counter for duplicate rows
            skipped_blank_count = 0  # Counter for rows skipped due to blank values

            # Set to track duplicate rows
            seen_rows = set()

            # Iterate through each row in the original DataFrame
            for index, row in df_opportunity_team_2.iterrows():
                # Check if the row is already seen (duplicate)
                row_tuple = tuple(row.items())
                if row_tuple in seen_rows:
                    duplicate_count += 1
                    continue  # Skip duplicate rows
                seen_rows.add(row_tuple)

                # Handle blank or NaN values in 'opportunityid' and 'email'
                opportunity_id = row.get('opportunityid', None)
                emails = row.get('email', None)

                if pd.isna(opportunity_id) or pd.isna(emails):
                    skipped_blank_count += 1  # Increment skipped rows count
                    continue  # Skip rows with missing 'opportunityid' or 'email'

                # Split emails if multiple are present
                emails = str(emails).split(',')

                if len(emails) > 1:
                    for email in emails:
                        email = email.strip()  # Remove any whitespace
                        if email:  # Skip blank emails
                            new_row = row.copy()
                            new_row['email'] = email  # Assign a single email
                            new_rows.append(new_row)
                else:
                    # Handle rows with a single email
                    if emails[0].strip():  # Skip rows with blank single email
                        new_rows.append(row)

            # Create a new DataFrame with the processed rows
            df_Opportunity_team = pd.DataFrame(new_rows, columns=df_opportunity_team_2.columns)

            # Total row counts before and after removing duplicates
            total_rows_before = len(df_opportunity_team_2)
            total_rows_after = len(df_Opportunity_team)

            # Drop duplicate rows (if necessary)
            df_Opportunity_team = df_Opportunity_team.drop_duplicates()

            # Create a new sheet in the workbook
            ws = wb.create_sheet(title=Opportunity_team_sheet_name)

            # Convert DataFrame to rows and append to the new sheet
            for r_idx, row in enumerate(dataframe_to_rows(df_Opportunity_team, index=False, header=True), 1):
                ws.append(row)

            # Save the workbook
            wb.save(file_path)

            # Print results
            print(f"\n    ✅ New sheet '{Opportunity_team_sheet_name}' has been successfully created in the Excel file.\n")
            print(f"\n        🔸 Total rows before processing: {total_rows_before}")
            print(f"\n        🔸 Total rows after processing: {total_rows_after}")
            rows_difference =  total_rows_before - total_rows_after
            if rows_difference > 0 :
                print(f"\n        🔸 Rows removed: {rows_difference}")
            else:
                print(f"\n        🔸 Rows Added: {abs(rows_difference)}")



            # ======================================================================
            # Step 4: Concatenating Email Values
            # ======================================================================

            print("\n\n🔍 Step 4: Concatenating Email Values...")

            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Access the 'Opportunity_team' sheet
            sheet = wb[Opportunity_team_sheet_name]

            # Locate the 'email' column index in the header row (row 1)
            email_column_index = None
            for col in sheet.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    if cell.value == 'email':
                        email_column_index = cell.column
                        break
                if email_column_index is not None:
                    break

            # If 'email' column is not found, raise an error and stop execution
            if email_column_index is None:
                print("\n    ❌ ERROR: Column 'email' not found in the 'Opportunity_team' sheet.")
                raise ValueError("Column 'email' not found.")

            # Define the column header for the new column
            Concat_T_M_column_header = 'Concat_T_M'

            # Calculate the max row in the email column
            max_row = sheet.max_row

            # Process each row starting from the second row (assuming the first row is the header)
            rows_processed = 0  # Counter for processed rows
            
            # Iterate through each row starting from the second row (skip header)
            for row in range(2, max_row + 1):
                # Get the value from the email column
                email_value = sheet.cell(row=row, column=email_column_index).value

                # Check if the email_value is not None
                if email_value is not None:
                    # Concatenate with inverted commas and comma
                    concatenated_value = f"'{email_value}',"

                    # Write the concatenated value to the new column
                    Concat_T_M_cell = sheet.cell(row=row, column=email_column_index + 1)
                    Concat_T_M_cell.value = concatenated_value

                    rows_processed += 1

            # Add the header for the new column
            sheet.cell(row=1, column=email_column_index + 1, value=Concat_T_M_column_header)

            # Save the workbook
            wb.save(file_path)

            # Print completion message
            print(f"\n    ✅ Concatenated email values.")


            # ======================================================================
            # Step 5 :- Checking if Opportunities Exist in the 'Opportunity' Sheet
            # ======================================================================

            print("\n\n🔍 Step 5: Checking if Opportunities Exist in the 'Opportunity' Sheet...")

            try:
                # Load the sheets into DataFrames
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                Opportunity_team_df = pd.read_excel(file_path, sheet_name=Opportunity_team_sheet_name)

                # Check for required columns
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ ERROR: Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet.")
                elif 'opportunityid' not in Opportunity_team_df.columns:
                    print(f"\n    ❌ ERROR: Column 'opportunityid' not found in the '{Opportunity_team_sheet_name}' sheet.")
                else:
                    # Add a new 'existing' column
                    Opportunity_team_df['existing'] = Opportunity_team_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])
                    
                    # Count rows where 'existing' is False
                    false_count = len(Opportunity_team_df[~Opportunity_team_df['existing']])

                    # Save the updated DataFrame back to the sheet
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        Opportunity_team_df.to_excel(writer, sheet_name=Opportunity_team_sheet_name, index=False)
                    
                    # Display results
                    print(f"\n    ✅ The 'existing' column has been added to the '{Opportunity_team_sheet_name}' sheet.")
                    if false_count > 0:
                        print(f"\n    ❗️ Number of Opportnities Missing in Team sheet: {false_count}")
                    else:
                        print(f"\n    ✅ All Opportunities Exist In Opportunity Sheet")

            except FileNotFoundError:
                print(f"\n    ❌ ERROR: File '{file_path}' not found.")
            except KeyError as e:
                print(f"\n    ❌ ERROR: {str(e)}")
            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {str(e)}")

            # ======================================================================
            # Step 6 :- Extracting Concatenated Values
            # ======================================================================

            print("\n\n🔍 Step 6: Extracting Concatenated Values...")

            # Check if the input file exists
            if not os.path.exists(file_path):
                print(f"\n    ❌ ERROR: The input file '{file_path}' does not exist. Please check the file path and try again.")
                exit()

            try:
                # Read the Excel file
                df = pd.read_excel(file_path, sheet_name=Opportunity_team_sheet_name)

                # Specify the column to extract concatenated values from
                column_name = "Concat_T_M"

                # Check if the column exists in the dataframe
                if column_name not in df.columns:
                    print(f"\n    ❌ ERROR: Column '{column_name}' is missing in the sheet '{Opportunity_team_sheet_name}' of the input file.")
                    exit()

                # Remove blank values and drop duplicates
                cleaned_data = df[column_name].dropna().drop_duplicates().reset_index(drop=True)

                # Create a new DataFrame for the output
                output_df = pd.DataFrame({column_name: cleaned_data})

                # Define the output file path and name
                output_file_path = "Extracts/Team_Member_extract.xlsx"

                # Write the processed data to a new Excel file
                output_df.to_excel(output_file_path, index=False)

                # Success message
                print(f"\n    ✅ Created 'Team_Member_extract' file and saved in Downloads")

            except Exception as e:
                print(f"\n    ❌ ERROR: An unexpected error occurred: {str(e)}")

            # ==================================================================================================================
            
            # Load the Excel file
            extract_file_path = "Extracts/Team_Member_extract.xlsx"  # Change this to your actual file path
            
            # Load the Excel file into a DataFrame
            df = pd.read_excel(extract_file_path)

            # Extract the "accountid" column values
            if "Concat_T_M" in df.columns:
                account_ids = df["Concat_T_M"].dropna().astype(str)  # Drop NaN values and convert to string

                # Save to a text file
                with open("Delete/5_teammember.txt", "w") as f:
                    f.write("\n".join(account_ids))

            else:
                print("Column 'accountid' not found in the sheet.")


            #---------- Clean-up: Remove the last character from the final line of the file (likely an extra comma)------
            
            remove_last_char_from_last_line('Delete/5_teammember.txt')
            
            #---------- Read the cleaned-up text file and prepare a SOQL query to copy to clipboard ---------- 
            
            # Open the file and read its contents as a single string
            with open("Delete/5_teammember.txt", "r", encoding="utf-8") as file:
                cliptext = file.read()  # Read all lines as a single string

            # Construct a Salesforce SOQL query using the list of emails
            team_query = f"select email,id,Profile.Name,isactive from user where email in ({cliptext}) and Profile.Name != 'IBM Partner Community Login User' and IsActive = true"
            
            # Copy the query to the clipboard for easy pasting
            pyperclip.copy(team_query)

            # ======================================================================
            # 🔍 Step 7: Copying Data from CSV File
            # ======================================================================

            print("\n\n🔍 Step 7: Copying Data from CSV File...")

            # Define the file path for the CSV file
            team_csv = DOWNLOADS_DIR+"/teammember.csv" #As email id are store in Userid csv
            
            # Flag to determine whether to proceed with processing the team member CSV
            run_team_code = False

            # Check if the file already exists
            if os.path.exists(team_csv):
                run_team_code = True  # File is available, proceed with processing
            else:
                # Loop until the file is found or the user decides to skip
                while not os.path.exists(team_csv):

                    # Attempt to locate and rename the downloaded bulk query file if available
                    if rename_and_move_bulkquery_file('teammember.csv',csv_file_dir):
                        if os.path.exists(team_csv):
                            run_team_code = True # If the file now exists after renaming, proceed
                            break
                        continue  # Continue checking if the file exists after renaming attempt

                    # If the file still doesn't exist, prompt the user
                    print(f"\n    ❌ File 'teammember.csv' does not exist. Did you Query the Team member?")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
                    
                    # Validate input until user provides 'yes' or 'no'
                    while try_again not in ['yes', 'no']:
                        print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()
                    
                    if try_again == 'yes':
                        # Check if the file now exists after user agreed to try again
                        if os.path.exists(team_csv):
                            # If the file exists, set the flag to run the team code
                            run_team_code = True
                            break
                        else:
                            # If the file still doesn't exist, continue the loop and prompt again
                            continue

                    # User chose not to retry — skip this step
                    if try_again == 'no':
                        print("\n          🚫 Skipping Team Member Sheet.")
                        team_member_choice = 'no'
                        break
                
            if run_team_code:
                # If the file is found, proceed to copy its data into the Excel sheet
                try:
                    # Read data from the CSV file
                    df = pd.read_csv(team_csv)

                    # Specify the Excel file path and sheet name
                    Opportunity_team_copy_sheet = "Opportunity_team_Copy"

                    # Write data to the specified sheet in the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=Opportunity_team_copy_sheet, index=False)

                    # Success message
                    print(f"\n    ✅ Data from the CSV file has been successfully copied to the '{Opportunity_team_copy_sheet}' sheet in the Excel file:")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: Excel file '{file_path}' not found.")
                except Exception as e:
                    print(f"\n    ❌ Error: An unexpected error occurred: {e}")

                # ======================================================================
                # 🔍 Step 8: Fetching User IDs of Team Members...
                # ======================================================================

                print ('\n\n🔍 Step 8: Fetching User IDs of Team Members...')

                try:
                    # Load data from the relevant Excel sheets:
                    # - 'Opportunity_team' contains the list of team members needing IDs.
                    # - 'Opportunity_team_Copy' contains user records with Email and Id.
                    opportunity_team_df = pd.read_excel(file_path, sheet_name=Opportunity_team_sheet_name)
                    opportunity_team_copy_df = pd.read_excel(file_path, sheet_name=Opportunity_team_copy_sheet)

                    # Clean and standardize email formats to ensure accurate matching:
                    # - Remove leading/trailing whitespaces.
                    # - Convert emails to lowercase.
                    opportunity_team_df["email"] = opportunity_team_df["email"].str.strip().str.lower()
                    opportunity_team_copy_df["Email"] = opportunity_team_copy_df["Email"].str.strip().str.lower()

                    # Perform a left join:
                    # - Match emails in the 'Opportunity_team' sheet with 'Email' in the copy sheet.
                    # - Bring the 'Id' field from the copy sheet into the result.
                    result_df = pd.merge(
                        opportunity_team_df,
                        opportunity_team_copy_df[["Email", "Id"]],
                        left_on="email",
                        right_on="Email",
                        how="left"
                    )

                    # Count how many user IDs are missing (NaN) before replacement
                    nan_before = result_df["Id"].isna().sum()

                    # Fill missing IDs with "Inactive" to flag unmatched users
                    result_df["Id"] = result_df["Id"].fillna("Inactive")

                    # Count missing IDs after replacement (should be zero now)
                    nan_after = result_df["Id"].isna().sum()

                    # Calculate how many IDs were marked as "Inactive"
                    nan_replaced = nan_before - nan_after

                    # Remove the now redundant 'Email' column (came from the right dataframe)
                    result_df.drop(columns=["Email"], inplace=True)

                    # Rename 'Id' column to 'OwnerId' for clarity or standardization
                    result_df.rename(columns={"Id": "OwnerId"}, inplace=True)

                    # Save the updated data back to the 'Opportunity_team' sheet
                    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        result_df.to_excel(writer, sheet_name=Opportunity_team_sheet_name, index=False)

                    # Output success message
                    if nan_replaced > 0:
                        print(f"\n    ❗️ Number of 'Inactive' values : {nan_replaced}")
                    else:
                        print(f"\n    ✅ All Team Member are valid")

                except FileNotFoundError:
                    print(f"\n    ❌ File not found at path: {file_path}.")

                except KeyError as e:
                    print(f"\n    ❌ KeyError: Column '{e}' not found. Please check the column names in your Excel sheets.")

                except Exception as e:
                    print(f"\n    ❌ An unexpected error occurred: {e}")


                # ======================================================================
                # 🔍 Step 9: Rearranging Columns in Sequence...
                # ======================================================================

                print("\n\n🔍 Step 9: Rearranging Columns in Sequence...")

                # Specify the preferred order for key columns in the sheet
                desired_column_order = [
                    "opportunityid",
                    "existing",
                    "opportunityaccesslevel",
                    "teammemberrole",
                    "email",
                    "OwnerId",
                    "Concat_T_M"
                ]

                try:
                    # Load the data from the specified sheet in the Excel file
                    excel_data = pd.read_excel(file_path, sheet_name=Opportunity_team_sheet_name)

                    # Check if the data was loaded correctly into a DataFrame
                    if isinstance(excel_data, pd.DataFrame):

                        # Identify any columns from the desired order that are missing in the actual sheet
                        missing_columns = []
                        for col in desired_column_order:
                            if col not in excel_data.columns:
                                missing_columns.append(col)
                        
                        # Identify any extra columns present in the actual sheet but not in the desired order
                        extra_columns = []
                        for col in excel_data.columns:
                            if col not in desired_column_order:
                                extra_columns.append(col)

                        # Build a new column sequence:
                        # - Start with the desired columns (if they exist in the actual data)
                        # - Append any extra columns to the end to preserve all data
                        rearranged_columns = []
                        for col in desired_column_order:
                            if col in excel_data.columns:
                                rearranged_columns.append(col)
                                
                        rearranged_columns += extra_columns

                        # Rearrange the columns in the DataFrame and save the updated sheet
                        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                            excel_data[rearranged_columns].to_excel(writer, sheet_name=Opportunity_team_sheet_name, index=False)

                        # Display success message with details
                        print(f"\n    ✅ Columns rearranged successfully in the team sheet")

                        # Notify about missing columns
                        if missing_columns:
                            print("\n    ❗️ The following columns were missing and skipped:")
                            for col in missing_columns:
                                print(f"        🔸 {col}")

                        # Notify about extra columns
                        if extra_columns:
                            print("\n    🔹 The following extra columns were moved to the end:")
                            for col in extra_columns:
                                print(f"\n        🔸 {col}")
                    else:
                        print(f"\n    ❌ Sheet '{Opportunity_team_sheet_name}' not found in the Excel file.")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

                except Exception as e:
                    print(f"\n    ❌ An unexpected error occurred: {e}")


                # ======================================================================
                # 🔍 Step 10: Renaming Columns...
                # ======================================================================

                print("\n\n🔍 Step 10: Renaming Columns...")

                # Dictionary mapping old column names to new column names
                column_rename_mapping = {
                    'opportunityid': 'OpportunityId',
                    'teammemberrole': 'TeamMemberRole',
                    'opportunityaccesslevel': 'OpportunityAccessLevel',
                    'OwnerId': 'UserId',
                }

                # file_path = "your_excel_file.xlsx"  # Specify the path to your Excel file

                try:
                    # Read the Excel file
                    excel_data = pd.read_excel(file_path, sheet_name=None)

                    # Check if the specified sheet exists
                    if Opportunity_team_sheet_name in excel_data:
                        
                        # Access the specified sheet
                        df = excel_data[Opportunity_team_sheet_name]

                        # Check if all specified columns exist
                        missing_columns = []
                        for col in column_rename_mapping.keys():
                            if col not in df.columns:
                                missing_columns.append(col)

                        # If any specified column is missing, notify the user and ask if they want to proceed
                        if missing_columns:
                            
                            print("\n    ❗️The following columns are missing and cannot be renamed:")
                            
                            for col in missing_columns:
                                print(f"\n      🔸 {col}")

                            while True:  # Loop until a valid response is given
                                proceed = input("\n    🔹 Do you want to proceed with the execution? (yes/no): ").strip().lower()

                                if proceed == 'yes':
                                    break  # Exit the loop and proceed
                                elif proceed == 'no':
                                    print("\n      ⛔️ Operation aborted. Exiting...")
                                    exit()  # Exit the program
                                else:
                                    print("\n      ❗️ Invalid input. Please enter 'yes' or 'no'.")

                        # Rename specified columns
                        df.rename(columns=column_rename_mapping, inplace=True)

                        # Write the modified DataFrame back to the Excel file
                        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                            df.to_excel(writer, sheet_name=Opportunity_team_sheet_name, index=False)

                        # Success message
                        print(f"\n    ✅ Columns renamed successfully in the Team sheet.")

                    else:
                        print(f"\n    ❌ Sheet '{Opportunity_team_sheet_name}' not found in the Excel file.")

                except FileNotFoundError:
                    
                    print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")

                except Exception as e:
                    
                    print(f"\n    ❌ An unexpected error occurred: {e}")


                # ======================================================================
                # Step 11: Count the Number of Rows and Columns
                # ======================================================================

                print("\n\n🔍 Step 11: Counting the Number of Rows and Columns...")

                # check count of rows
                team_final_num_rows = count_row(file_path, Opportunity_team_sheet_name)
                
                # Print the number of rows and columns
                print(f"\n    ✅ Final rows count: {team_final_num_rows}")

    
            # ======================================================================
            # Team Member Sheet Completed  
            # ======================================================================

            title = "📝 TEAM MEMBER SHEET COMPLETED 📝"
            show_title(title)

            break

        elif choice == "no":
            team_member_choice = 'no'
            print("\n        🚫 Team Member sheet execution skipped!")
            print("\n")

            title = "📝 TEAM MEMBER SHEET SKIPPED 📝"
            show_title(title)

            break  # Just breaking without running Block 2

        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")


    # =========================================================================================================================================
    #                                                STRATEGY SHEET EXECUTION
    # =========================================================================================================================================

    # Display the header once
    print("\n\n📄 Execute Next Sheet:")
    original_codes_sheet_name = 'Reporting_codes'
    
    # Check if the 'Reporting_codes' sheet is empty or contains only headers
    is_empty, preview = is_sheet_empty(file_path, original_codes_sheet_name)

    if is_empty:
        # If the sheet is empty or contains only headers, skip execution
        print(f"\n📂 The sheet '{original_codes_sheet_name}' is empty or contains only headers.\n")
        choice = 'no'
    elif is_empty is None:
        # If an error occurred while processing the sheet
        print("\n❗️ Could not process the sheet due to an error.\n")
        choice = 'no'
    else:
        # Sheet contains data, proceed with processing
        choice = 'yes'
        
        # Read all sheets from the Excel file into a dictionary of DataFrames
        xls = pd.ExcelFile(file_path)
        dfs = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}

        # Extract the DataFrame for the 'Reporting_codes' sheet
        df = dfs[original_codes_sheet_name]

        tag_column = 'tags'

        # Ensure 'tags' column exists; if not, create it with empty values
        if tag_column not in df.columns:
            df['tags'] = None  # Create an empty 'tags' column

        # Check if 'tags' column exists (case-insensitive)
        tags_column = [col for col in df.columns if col.strip().lower() == 'tags']

        # Check if 'Reporting Codes' column exists (case-insensitive)
        reporting_codes_column = [col for col in df.columns if col.strip().lower() == 'reporting_codes']

        # Check if 'Opportunity_id' column exists (case-insensitive)
        opportunity_id_column = [col for col in df.columns if col.strip().lower() == 'opportunity_id']

        # If 'tags' column doesn't exist or is empty, add or update the 'tags' column
        if reporting_codes_column and opportunity_id_column:
            reporting_codes_column = reporting_codes_column[0]
            opportunity_id_column = opportunity_id_column[0]

            # If the 'tags' column doesn't exist or is completely empty
            if not tags_column or df[tags_column[0]].isna().all():
                # Convert 'tags' column to string type to avoid issues when assigning text
                df['tags'] = df['tags'].astype(str)  # Convert to string type to avoid dtype mismatch
                
                # Populate 'tags' column by copying 'Reporting Codes' values for each unique opportunity_id
                for opportunity_id in df[opportunity_id_column].unique():
                    
                    # Get all rows for the current opportunity_id
                    opportunity_rows = df[df[opportunity_id_column] == opportunity_id]
                    
                    # Assign corresponding 'Reporting Codes' values to the 'tags' column
                    df.loc[df[opportunity_id_column] == opportunity_id, 'tags'] = opportunity_rows[reporting_codes_column].values

        # Update the dictionary with the modified 'Reporting_codes' DataFrame
        dfs[original_codes_sheet_name] = df

        # Write all sheets back to the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet, data in dfs.items():
                data.to_excel(writer, sheet_name=sheet, index=False)


        print(f"\n✅ The sheet '{original_codes_sheet_name}' contains data. Here are the first 4 rows:\n")
        df = pd.read_excel(file_path, sheet_name=original_codes_sheet_name)
        preview = df.head(4)
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))

    while True:

        print(f"\n    🔹 Do you want to execute the Strategy Sheet ? (yes/no): {choice}")
        
        if choice == "yes": # Automatically taken

            strategy_choice = 'yes'
            print(f"\n        ⏳ Executing the Sheet: Strategy sheet ")

            print("\n")
            title = "📝 STRATEGY SHEET EXECUTION 📝"
            show_title(title)

            # =========================================
            # Step 1: Rename Reporting Codes Sheet
            # =========================================

            print("\n\n🔍 Step 1: Renaming Reporting Codes Sheet...")

            def rename_reporting_codes(wb):
                """
                Rename the sheet whose normalized name matches 'reportingcodes' 
                (ignoring case, spaces, and underscores) to 'Reporting_codes_2'.
                """
                
                for sheetname in wb.sheetnames:
                    # Normalize sheet name by removing spaces and underscores, and converting to lowercase
                    normalized_name = sheetname.strip().replace('_', '').replace(' ', '').lower()
                    if normalized_name == 'reportingcodes':
                        # Rename the matched sheet
                        wb[sheetname].title = 'Reporting_codes_2'
                        return True
                return False

            try:
                # Try to load the workbook
                wb = openpyxl.load_workbook(file_path)

                # Attempt to rename the sheet
                if rename_reporting_codes(wb):
                    # Save the changes to the file
                    wb.save(file_path)
                    print("\n    ✅ Sheet 'Reporting_codes' renamed to 'Reporting_codes_2' successfully.")
                else:
                    print("\n    ❌ Error: Sheet 'Reporting_codes' not found in the workbook.")
                    sys.exit()

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found. Please check the file path and try again.")
            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")

            # ========================================================================
            # Step 2: Rename the columns in "Reporting_codes_2" sheet
            # ========================================================================

            print("\n\n🔍 Step 2: Renaming Columns in 'Reporting_codes_2' Sheet...\n")


            # Load the workbook
            try:
                wb = load_workbook(file_path)

                # Define the target sheet name
                target_sheet_name = 'Reporting_codes_2'

                # Check if the target sheet exists in the workbook
                if target_sheet_name in wb.sheetnames:
                    # Load the target sheet
                    ws = wb[target_sheet_name]

                    # Define normalization mapping for column headers
                    column_name_mappings = {
                        'tags': 'tag',
                        'Opportunity Id': 'opportunityid',
                        'opportunity_id': 'opportunityid',
                        'Opportunityid': 'opportunityid',
                        'opportunityid': 'opportunityid'
                    }

                    # List to track renamed columns
                    renamed_columns = []

                    # Loop through the header row and rename applicable columns
                    for col_idx, cell in enumerate(ws[1], start=1):  # start=1 for 1-based index
                        if isinstance(cell.value, str):
                            normalized_name = cell.value.strip().lower()  # normalize to lowercase and strip whitespace
                            if normalized_name in column_name_mappings:
                                new_column_name = column_name_mappings[normalized_name]
                                ws.cell(row=1, column=col_idx, value=new_column_name)  # Rename the column header
                                renamed_columns.append(f"'{cell.value}' ➔ '{new_column_name}'")

                    # Save the modified workbook
                    wb.save(file_path)

                    # Provide output
                    if renamed_columns:
                        print(f"\n    ✅ Columns renamed successfully in '{target_sheet_name}' sheet:")

                    else:
                        print(f"\n    ❗️ No columns were renamed. All columns were already in the desired format.")
                    
                else:
                    print(f"\n    ❗️ '{target_sheet_name}' sheet not found in the Excel file. No action taken.")

            except FileNotFoundError:
                print(f"\n    ❌ Error: The file '{file_path}' was not found. Please check the file path and try again.")
                exit()

            except Exception as e:
                print(f"\n    ❌ An unexpected error occurred: {e}")
                exit()

            # ========================================================================
            # Step 3: Create separate sheet for "Tags"
            # ========================================================================

            print("\n\n🔍 Step 3: Creating 'Tags' Sheet...")

            # Load the Excel file
            try:
                wb = load_workbook(file_path)
            except FileNotFoundError:
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")
                exit()

            # Load the specific sheet into a DataFrame
            sheet_name = 'Reporting_codes_2'
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            except KeyError:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in '{file_path}'.")
                wb.close()
                exit()

            # Normalize column names to lowercase for consistency
            df.columns = df.columns.str.lower()

            # Check if 'opportunity_id' and 'tag' columns are present
            expected_columns = ['opportunityid', 'tag']
            missing_columns = [col for col in expected_columns if col not in df.columns]
            if missing_columns:
                print(f"\n    ❌ Error: Missing columns: {', '.join(missing_columns)}. ")
                wb.close()
                exit()

            # Filter rows where 'tag' column contains a value
            filtered_df = df[df['tag'].notnull()]

            # Check if 'Tags_2' sheet already exists
            if 'Tags_2' in wb.sheetnames:
                overwrite_input = input("\n    ❗️ Sheet 'Tags_2' already exists. Do you want to overwrite it? (yes/no): ").lower().strip()
                if overwrite_input == 'yes':
                    # Remove the existing 'Tags_2' sheet
                    existing_sheet = wb['Tags_2']
                    wb.remove(existing_sheet)
                    print(f"\n       ✅ 'Tags_2' sheet cleared for overwrite.")
                else:
                    print("\n       ❌ Not overwriting 'Tags_2' sheet. Aborting further processing.")
                    wb.close()
                    exit()

            # Create a new sheet named 'Tags_2' and write filtered data
            ws = wb.create_sheet('Tags_2')

            # Append header row
            ws.append(expected_columns)

            # Append filtered data rows
            copied_rows = 0
            for r_idx, row in enumerate(dataframe_to_rows(filtered_df[expected_columns], index=False, header=False)):
                ws.append(row)
                copied_rows += 1

            # Save the updated workbook
            try:
                wb.save(file_path)
                print(f"\n    ✅ Filtered data ({copied_rows} rows) has been successfully written to 'Tags_2' sheet.")
            except Exception as e:
                print(f"\n    ❌ Error occurred while saving: {e}")
            finally:
                wb.close()

            # ========================================================================
            # Step 4: To Delete the "Tag" Column from Reporting Codes Sheet
            # ========================================================================

            print("\n\n🔍 Step 4: Deleting 'Tag' Column from Reporting Codes Sheet...")

            # Load the Excel file
            try:
                wb = load_workbook(file_path)

            except FileNotFoundError:
                print(f"\n    ❌ Error: File '{file_path}' not found. Please check the file path and try again.")
                exit()

            # Load a specific sheet
            sheet_name = 'Reporting_codes_2'
            if sheet_name not in wb.sheetnames:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in '{file_path}'. Aborting further processing.")
                wb.close()
                exit()

            # Get the sheet
            ws = wb[sheet_name]

            # Find and delete 'tag' column if it exists (case insensitive)
            tag_column_found = False
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                if col[0].value and col[0].value.strip().lower() == 'tag':
                    ws.delete_cols(col_idx)
                    tag_column_found = True
                    print(f"\n    ✅ 'Tag' column found and deleted from '{sheet_name}' sheet.")
                    break

            if not tag_column_found:
                print(f"\n    ❗️ Column 'tag' not found in '{sheet_name}' sheet. No action taken.")

            # Save and close the workbook
            try:
                wb.save(file_path)
            except Exception as e:
                print(f"\n    ❌ Error occurred while saving: {e}")
            finally:
                wb.close()

            # ========================================================================
            # Step 5: To Remove Comma Separated Values from 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 5: Removing Comma Separated Values and Duplicates from 'Reporting_codes' Sheet...\n")


            def process_reporting_codes(file_path, sheet_name):
                try:
                    # Load the existing workbook
                    wb = load_workbook(file_path)
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to load workbook. {e}")
                    return
                
                try:
                    # Read the specific sheet into a DataFrame
                    df_reporting_codes_2 = pd.read_excel(file_path, sheet_name=sheet_name)
                except Exception as e:
                    print(f"\n    ❌ Error: Failed to read sheet '{sheet_name}'. {e}")
                    return
                
                # Initialize an empty list to store rows for the new DataFrame
                new_rows_reporting = []
                
                # Iterate through each row in the original DataFrame
                for index, row in df_reporting_codes_2.iterrows():
                    reporting_codes = []
                    if isinstance(row['reporting_codes'], str) and row['reporting_codes'].strip():
                        reporting_codes = row['reporting_codes'].split(',')
                    
                    # If there are multiple reporting codes in the cell, create new rows
                    if reporting_codes:
                        for code in reporting_codes:
                            new_row_reporting = row.copy()
                            new_row_reporting['reporting_codes'] = code.strip()
                            new_rows_reporting.append(new_row_reporting)
                    else:
                        new_rows_reporting.append(row)
                
                # Create a new DataFrame from the list of rows
                df_Reporting_codes = pd.DataFrame(new_rows_reporting, columns=df_reporting_codes_2.columns)
                
                # Drop duplicate rows
                df_Reporting_codes.drop_duplicates(inplace=True)
                
                # Create a new sheet in the existing workbook
                sheet_name_reporting = 'Reporting_codes'
                if sheet_name_reporting in wb.sheetnames:
                    print(f"    ❗️ Sheet '{sheet_name_reporting}' already exists. It will be replaced.")
                    ws_reporting = wb[sheet_name_reporting]
                    wb.remove(ws_reporting)
                    ws_reporting = wb.create_sheet(title=sheet_name_reporting)
                else:
                    ws_reporting = wb.create_sheet(title=sheet_name_reporting)
                
                # Convert DataFrame to rows and append to the sheet

                for r_idx, row in enumerate(dataframe_to_rows(df_Reporting_codes, index=False, header=True), 1):
                    ws_reporting.append(row)
                
                # Save the workbook with the new sheet
                try:
                    wb.save(file_path)
                    print(f"\n    ✅ New sheet '{sheet_name_reporting}' has been successfully created in the Excel file.")
                    print(f"\n        🔸 Total rows before processing: {len(df_reporting_codes_2)}")
                    print(f"\n        🔸 Total rows after processing: {len(df_Reporting_codes)}")

                    duplicate_rows_difference = len(df_reporting_codes_2) - len(df_Reporting_codes)
                    if duplicate_rows_difference >= 0:
                        print(f"\n        🔸 Rows removed: {duplicate_rows_difference}.")
                    else:
                        print(f"\n        🔸 Rows Added: {abs(duplicate_rows_difference)}.")

                except Exception as e:
                    print(f"\n    ❌ Error: Failed to save workbook. {e}")

            process_reporting_codes(file_path, 'Reporting_codes_2')

            # ========================================================================
            # Step 6: To Remove Comma Separated Values from 'Tags_2' Sheet
            # ========================================================================

            print("\n\n🔍 Step 6: Removing Comma Separated Values and Duplicates from 'Tags_2' Sheet...")

            # Load the existing workbook
            try:
                wb = load_workbook(file_path)
            except Exception as e:
                print(f"\n    ❌ Error: Failed to load workbook. {e}")
                exit()

            # Initialize variables for sheet and DataFrame
            sheet_name = 'Tags_2'
            df_tag_2 = None

            # Attempt to load the sheet using openpyxl, handle if it doesn't exist
            try:
                sheet = wb[sheet_name]
                df_tag_2 = pd.DataFrame(sheet.values)
                df_tag_2.columns = df_tag_2.iloc[0]  # Assuming the first row contains column headers
                df_tag_2 = df_tag_2[1:]  # Skip the first row (headers)
            except KeyError:
                print(f"\n    ❌ Error: Sheet '{sheet_name}' not found in the workbook.")
                exit()

            # If sheet exists and DataFrame is loaded, process the data
            if df_tag_2 is not None:
                new_rows_tag = []
                duplicate_count_tag = 0  # Initialize count for duplicate rows
                seen_rows_tag = set()

                # Iterate through each row in the original dataframe
                for index, row in df_tag_2.iterrows():
                    row_tuple_tag = tuple(row.items())
                    if row_tuple_tag in seen_rows_tag:
                        duplicate_count_tag += 1  # Increment duplicate count
                    else:
                        seen_rows_tag.add(row_tuple_tag)
                        # Check if 'tag' column exists and is a non-empty string
                        if isinstance(row['tag'], str) and row['tag'].strip():
                            tags = row['tag'].split(',')
                        else:
                            tags = []
                        
                        # If there are multiple tags in the cell, create new rows
                        if len(tags) > 0:
                            for tag in tags:
                                new_row_tag = row.copy()  # Create a copy of the original row
                                new_row_tag['tag'] = tag.strip()  # Assign a single tag
                                new_rows_tag.append(new_row_tag)
                        else:
                            new_rows_tag.append(row)

                # Create a new DataFrame from the list of rows
                df_Tag = pd.DataFrame(new_rows_tag, columns=df_tag_2.columns)

                # Count total rows before and after removing duplicates
                total_rows_before_tag = len(df_tag_2)
                total_rows_after_tag = len(df_Tag)

                # Drop duplicate rows based on all columns (if necessary)
                df_Tag = df_Tag.drop_duplicates()

                # Count how many duplicates were removed
                removed_duplicates_count_tag = total_rows_before_tag - total_rows_after_tag

                # Create a new sheet in the existing workbook
                sheet_name_tag = 'Tags'
                if sheet_name_tag in wb.sheetnames:
                    print(f"\n    ❗️ Sheet '{sheet_name_tag}' already exists. It will be replaced.")
                    ws_tag = wb[sheet_name_tag]
                    wb.remove(ws_tag)
                ws_tag = wb.create_sheet(title=sheet_name_tag)

                # Convert DataFrame to rows and append to the sheet
                for r_idx, row in enumerate(dataframe_to_rows(df_Tag, index=False, header=True), 1):
                    ws_tag.append(row)

                # Save the workbook with the new sheet
                try:
                    wb.save(file_path)
                    print(f"\n    ✅ New sheet '{sheet_name_tag}' has been successfully created in the Excel file.")
                    print(f"\n        🔸 Total rows before processing: {total_rows_before_tag}")
                    print(f"\n        🔸 Total rows after processing: {total_rows_after_tag}")
                    if removed_duplicates_count_tag >= 0:
                        print(f"\n        🔸 Rows removed: {removed_duplicates_count_tag}")
                    else:
                        print(f"\n        🔸 Rows Added: {abs(removed_duplicates_count_tag)}")

                except Exception as e:
                    print(f"\n    ❌ Error: Failed to save workbook. {e}")

            # ========================================================================
            # Step 7: Add existing Column to 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 7: Adding 'existing' Column to 'Reporting_codes' Sheet...")

            # Specify the file path of the Excel file
            # file_path = os.path.expanduser("~/Downloads/your_excel_file.xlsx")

            # Specify the sheet names
            opportunity_sheet_name = 'Opportunity'
            reporting_codes_sheet_name = 'Reporting_codes'

            try:
                # Read data from the specified sheets
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                reporting_codes_df = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)

                # Check if required columns exist
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in 'Reporting codes' sheet.")
                    sys.exit()
                elif 'opportunityid' not in reporting_codes_df.columns:
                    print(f"\n    ❌ Column 'opportunityid' not found in 'Reporting codes' sheet.")
                    sys.exit()
                else:
                    # Create a new column 'existing' in reporting_codes_df
                    reporting_codes_df['existing'] = reporting_codes_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

                    # Write back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        reporting_codes_df.to_excel(writer, sheet_name=reporting_codes_sheet_name, index=False)

                    # Notify the successful operation
                    print(f"\n    ✅ 'existing' column has been successfully added to the 'Reporting codes' sheet.")

            except FileNotFoundError:
                print(f"    ❌ File not found.")
                sys.exit()
            except Exception as e:
                print(f"    ❌ Error: {e}")
                sys.exit()


            # ========================================================================
            # Step 8: Add existing Column to 'Tags' Sheet
            # ========================================================================

            print("\n\n🔍 Step 8: Adding 'existing' Column to 'Tags' Sheet...")


            # Specify the sheet names
            opportunity_sheet_name = 'Opportunity'
            tags_sheet_name = 'Tags'

            try:
                # Read data from the specified sheets
                opportunity_df = pd.read_excel(file_path, sheet_name=opportunity_sheet_name)
                tags_df = pd.read_excel(file_path, sheet_name=tags_sheet_name)

                # Check if required columns exist
                if 'opportunity_legacy_id__c' not in opportunity_df.columns:
                    print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in '{opportunity_sheet_name}' sheet.")
                    sys.exit()
                elif 'opportunityid' not in tags_df.columns:
                    print(f"\n    ❌ Column 'opportunityid' not found in '{tags_sheet_name}' sheet.")
                    sys.exit()
                else:
                    # Create a new column 'existing' in tags_df
                    tags_df['existing'] = tags_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

                    # Write back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        tags_df.to_excel(writer, sheet_name=tags_sheet_name, index=False)

                    # Notify the successful operation
                    print(f"\n    ✅ 'existing' column has been successfully added to the '{tags_sheet_name}' sheet.")

            except FileNotFoundError:
                print(f"    ❌ File '{file_path}' not found.")
                sys.exit()
            except Exception as e:
                print(f"    ❌ Error: {e}")
                sys.exit()

            # ========================================================================
            # Step 9: To Concatenate Values in 'Reporting_codes' Sheet
            # ========================================================================

            print("\n\n🔍 Step 9: Concatenating Values in 'Reporting_codes' Sheet...")

            sheet_name = 'Reporting_codes'

            try:
                # Load the workbook and specify append mode to modify the sheet
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    
                    # Read the specified sheet into a DataFrame
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Check if 'reporting_codes' column exists (case-insensitive)
                    lowercase_columns = [col.lower() for col in df.columns]
                    
                    if 'reporting_codes' in lowercase_columns:
                        # Find the actual column name in the original case
                        actual_reporting_codes_column = df.columns[lowercase_columns.index('reporting_codes')]
                        
                        # Create a new column 'Concatcodes' with concatenated values
                        df['Concatcodes'] = df[actual_reporting_codes_column].apply(lambda x: f"'{x}'," if pd.notnull(x) else x)
                        
                        # Write the modified DataFrame back to the specified sheet
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Notify successful operation
                        print(f"\n    ✅ The 'Concatcodes' column has been successfully added to '{sheet_name}' sheet.")
                    else:
                        print(f"\n    ❗️ The '{sheet_name}' sheet does not have a column named 'reporting_codes'.")
            except FileNotFoundError:
                print(f"\n    ❌ File '{file_path}' not found.")
                sys.exit()
            except Exception as e:
                print(f"\n    ❌ Error: {e}")
                sys.exit()

            # ========================================================================
            # Step 10:- To concatenate values in "Tags" sheet
            # ========================================================================

            print("\n\n🔍 Step 10: Adding 'Concattags' Column to 'Tags' Sheet...")

            # file_path = 'your_file_path.xlsx'
            sheet_name = 'Tags'

            # Check if the sheet exists in the workbook
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                if sheet_name in sheet_names:
                    # Load the entire workbook
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # Read the specified sheet into a DataFrame
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # Check if 'tag' column exists (case-insensitive)
                        lowercase_columns = [col.lower() for col in df.columns]
                        if 'tag' in lowercase_columns:
                            # Find the actual column name in the original case
                            actual_tag_column = df.columns[lowercase_columns.index('tag')]

                            # Create a new column 'Concattags' with concatenated values
                            df['Concattags'] = df[actual_tag_column].apply(lambda x: f"'{x}'," if pd.notnull(x) else x)

                            # Write the modified DataFrame back to the specified sheet
                            df.to_excel(writer, sheet_name=sheet_name, index=False)

                            # Print a meaningful message
                            print("\n    ✅ 'Concattags' column has been successfully added to the 'Tags' sheet.")
                        else:
                            print(f"\n    ❗️ The '{sheet_name}' sheet does not contain a column named 'tag'. Please verify the column name.")
                else:
                    print(f"\n    ❗️ The '{sheet_name}' sheet does not exist in the workbook. Please check the sheet name.")
                    sys.exit()
            except FileNotFoundError:
                print(f"\n    ❌ The file '{file_path}' was not found. Please verify the file path.")
                sys.exit()
            except Exception as e:
                print(f"\n    ❌ An error occurred: {str(e)}")
                sys.exit()

            # ========================================================================
            # Step 11:- To extract concatenated values
            # ========================================================================


            # Step 11:- Extracting Data from 'Reporting_codes' and 'Tags' Sheets

            print("\n\n🔍 Step 11: Extracting Data from 'Reporting_codes' and 'Tags' Sheets...")

            # Define the input file path and sheet names
            reporting_codes_sheet_name = "Reporting_codes"
            tags_sheet_name = "Tags"

            # Define the columns to extract from each sheet
            reporting_codes_column = "Concatcodes"
            tags_column = "Concattags"

            # Check if the input file exists
            if not os.path.exists(file_path):
                print(f"\n    ❌ The input file '{file_path}' does not exist. Please verify the file path.")
                exit()

            # Initialize flags to check if sheets are found
            reporting_codes_found = False
            tags_found = False

            # Read data from the Reporting_codes sheet if it exists
            if reporting_codes_sheet_name in pd.ExcelFile(file_path).sheet_names:
                df_reporting_codes = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)
                reporting_codes_found = True

            else:
                print(f"\n    ❗️ Sheet '{reporting_codes_sheet_name}' not found.")
                sys.exit()
            
            # Read data from the Tags sheet if it exists
            if tags_sheet_name in pd.ExcelFile(file_path).sheet_names:
                df_tags = pd.read_excel(file_path, sheet_name=tags_sheet_name)
                tags_found = True
            else:
                print(f"    ❗️ Sheet '{tags_sheet_name}' not found.")
                sys.exit()
            
            # Extract the required columns if sheets are found
            concatcodes_values = []
            concattags_values = []

            # Extract non-null, unique values from 'Concatcodes' column in 'Reporting_codes'
            if reporting_codes_found:
                concatcodes_values = df_reporting_codes[reporting_codes_column].dropna().unique()
            else:
                print(f"\n    ❗️ Column '{reporting_codes_column}' not found in '{reporting_codes_sheet_name}' sheet.")

            # Extract non-null, unique values from 'Concattags' column in 'Tags'
            if tags_found:
                concattags_values = df_tags[tags_column].dropna().unique()
            else:
                print(f"\n    ❗️ Column '{tags_column}' not found in '{tags_sheet_name}' sheet.")

            # Ensure both lists are the same length by padding with None
            max_length = max(len(concatcodes_values), len(concattags_values))
            
            if len(concatcodes_values) < max_length:
                concatcodes_values = list(concatcodes_values) + [None] * (max_length - len(concatcodes_values))
            if len(concattags_values) < max_length:
                concattags_values = list(concattags_values) + [None] * (max_length - len(concattags_values))

            # Create a new DataFrame for the extracted values
            output_df = pd.DataFrame({
                reporting_codes_column: concatcodes_values,
                tags_column: concattags_values
            })

            # Define the output file path and name
            output_file_path = "Extracts/Reporting_codes_and_Tags_extract.xlsx"

            # Write the extracted data to a new Excel file
            output_df.to_excel(output_file_path, index=False)

            # Final success message
            print(f"\n    ✅ Extracted data has been successfully written to '{output_file_path}'.")

            # ==================================================================================================================
            # To extract tags and code to text file
            # ==================================================================================================================
            
            # Load the Extract file
            extract_file_path = "Extracts/Reporting_codes_and_Tags_extract.xlsx"  # Change this to your actual file path

            df = pd.read_excel(extract_file_path)

            # Extract values from 'ownerid' and 'concatenatedcreatedby' (even if their lengths differ)
            Concatcodes_values = df["Concatcodes"].dropna().astype(str).tolist() if "Concatcodes" in df.columns else []
            Concattags_values = df["Concattags"].dropna().astype(str).tolist() if "Concattags" in df.columns else []

            # Combine both lists while maintaining all values
            all_values = Concatcodes_values + Concattags_values  # Concatenating both lists

            # Save to a text file
            with open("Delete/6_strategy.txt", "w") as f:
                f.write("\n".join(all_values))
            
            # ========================================================================  
            
            # Remove the last character from the last line of the text file (e.g., trailing comma)
            remove_last_char_from_last_line('Delete/6_strategy.txt')

            # Open the cleaned text file and read its contents as a single string
            with open("Delete/6_strategy.txt", "r", encoding="utf-8") as file:
                cliptext = file.read()  # Read all lines as a single string

            # Construct the SOQL query using the cleaned list of strategy names
            tags_query = f'Select id,name,Record_Type_Name__c from Strategy__c where name in ({cliptext})'
            
            # Copy the constructed query to the clipboard for immediate use
            pyperclip.copy(tags_query)

            # ========================================================================
            # Step 12:- Processing CSV File and Adding Filtered Data to Excel
            # ========================================================================

            print("\n\n🔍 Step 12: Processing CSV File and Adding Filtered Data to Excel...")

            # Define the file path for the Exported csv file
            tags_csv = DOWNLOADS_DIR+ "/tags.csv"

            # Flag to determine whether to proceed with processing
            run_code_strategy = False
            
            # Check if the file already exists
            if os.path.exists(tags_csv):
                run_code_strategy = True  # Proceed if the file is already available
            else:         
                
                # Keep trying to find or rename the file until it's available or user chooses to skip
                while not os.path.exists(tags_csv):

                    # Try to rename and move the bulk query file if it's downloaded with the default name
                    if rename_and_move_bulkquery_file('tags.csv',csv_file_dir):
                        if os.path.exists(tags_csv):
                            run_code_strategy = True  # Proceed if the file becomes available after renaming
                            break
                        continue  # Retry the loop in case file still doesn't exist after renaming
                    
                    # Inform that the file is still missing and prompt for next action
                    print(f"\n    ❌ File 'tags.csv' does not exist. Did you query the tags?")
                    try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                    # Keep asking until valid input is received
                    while try_again not in ['yes', 'no']:
                        print("\n          ❗️ Invalid input. Please enter 'yes' or 'no'.")
                        try_again = input("\n        🔹 Do you want to try again? (yes/no): ").strip().lower()

                    if try_again == 'yes':
                        if os.path.exists(tags_csv):
                            run_code_strategy = True
                            break
                        else:
                            continue

                    if try_again == 'no':
                        run_code_strategy = False
                        strategy_choice = 'no'
                        print("\n          🚫 Skipping Strategy Sheet.")
                        break

            # If the csv file exists or becomes available, proceed with processing it
            if run_code_strategy:
                
                # If the file exists, process the file
                try:
                    # Read CSV file into a DataFrame
                    df = pd.read_csv(tags_csv)

                    # Specify the column name to filter
                    filter_column = "Record_Type_Name__c"
                    filter_value = "Reporting codes"

                    # Check if the filter column exists in the CSV file
                    if filter_column not in df.columns:
                        print(f"\n    ❌ Error: Column '{filter_column}' not found in the CSV file.")
                        sys.exit()  # Exit if the required column is not found

                    # Filter rows where the specified column equals the specified value
                    df_filtered = df[df[filter_column] == filter_value]

                    # Specify the sheet name in the Excel file
                    sheet_name = "Reporting_codes_Copy"

                    # Write filtered data to the specified sheet in the Excel file
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

                    print(f"\n    ✅ Filtered data has been successfully copied to the '{sheet_name}' sheet in the Excel file.")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: The Excel file '{file_path}' was not found.")
                except KeyError:
                    print(f"\n    ❌ Error: The column '{filter_column}' is not found in the CSV file.")
                except Exception as e:
                    print(f"\n    ❌ Error: {e}")

                # ========================================================================
                # Step 13:- To copy data from Tag csv file
                # ========================================================================


                # Step 13:- Processing CSV File and Adding Filtered Data to Excel
                print("\n\n🔍 Step 13: Processing CSV File and Adding Filtered Data to Excel...")

                # If the file exists, process the file
                try:
                    # Read CSV file into a DataFrame
                    df = pd.read_csv(tags_csv)

                    # Specify the column name to filter
                    filter_column = "Record_Type_Name__c"
                    filter_value = "Tags"

                    # Check if the filter column exists in the CSV file
                    if filter_column not in df.columns:
                        print(f"\n    ❌ Error: Column '{filter_column}' not found in the CSV file.")
                        sys.exit()  # Exit if the required column is not found

                    # Filter rows where the specified column equals the specified value
                    df_filtered = df[df[filter_column] == filter_value]

                    # Specify the sheet name in the Excel file
                    sheet_name = "Tags_Copy"

                    # Write filtered data to the specified sheet in the Excel file
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_filtered.to_excel(writer, sheet_name=sheet_name, index=False)

                    print(f"\n    ✅ Filtered data has been successfully copied to the '{sheet_name}' sheet in the Excel file.")

                except FileNotFoundError:
                    print(f"\n    ❌ Error: The Excel file '{file_path}' was not found.")
                except KeyError:
                    print(f"\n    ❌ Error: The column '{filter_column}' is not found in the CSV file.")
                except Exception as e:
                    print(f"\n    ❌ Error: {e}")

                # =======================================================================
                # Step 14: To get strategy Ids for Codes
                # ========================================================================

                print("\n\n🔍 Step 14: To get strategy Ids for Codes...")

                reporting_codes_sheet_name = 'Reporting_codes'
                reporting_codes_copy_sheet_name = 'Reporting_codes_Copy'

                def vlookup_operation(file_path, reporting_codes_sheet_name, reporting_codes_copy_sheet_name):
                    # Check if the file exists
                    if not os.path.exists(file_path):
                        print(f"\n    ❌ Error: File '{file_path}' not found.")
                        return

                    try:
                        # Read the data from both sheets
                        reporting_codes_df = pd.read_excel(file_path, sheet_name=reporting_codes_sheet_name)
                        reporting_codes_copy_df = pd.read_excel(file_path, sheet_name=reporting_codes_copy_sheet_name)

                        # Ensure column names are standardized and lowercase
                        reporting_codes_df.columns = reporting_codes_df.columns.str.strip().str.lower()
                        reporting_codes_copy_df.columns = reporting_codes_copy_df.columns.str.strip().str.lower()

                        # Convert relevant columns to lowercase for case-insensitive merge
                        reporting_codes_df['reporting_codes'] = reporting_codes_df['reporting_codes'].astype(str).str.strip().str.lower()
                        reporting_codes_copy_df['name'] = reporting_codes_copy_df['name'].str.lower()

                        # Check if the necessary columns exist (case-insensitive)
                        if 'reporting_codes' not in reporting_codes_df.columns:
                            print(f"\n    ❌ Column 'reporting_codes' not found in '{reporting_codes_sheet_name}' sheet.")
                            sys.exit()
                        if 'name' not in reporting_codes_copy_df.columns:
                            print(f"\n    ❌ Column 'name' not found in '{reporting_codes_copy_sheet_name}' sheet.")
                            sys.exit()
                        if 'id' not in reporting_codes_copy_df.columns:
                            print(f"\n    ❌ Column 'id' not found in '{reporting_codes_copy_sheet_name}' sheet.")
                            sys.exit()

                        # Perform merge using standardized column names
                        merged_df = pd.merge(reporting_codes_df,
                                            reporting_codes_copy_df[['name', 'id']],
                                            left_on='reporting_codes', right_on='name',
                                            how='left')

                        # Count NaN values before replacing them
                        na_count = merged_df['id'].isna().sum()

                        # Create a new column 'StrategyId' and fill missing values with 'Not found'
                        merged_df['StrategyId'] = merged_df['id'].fillna('Not found')

                        # Drop unnecessary columns after merging
                        merged_df.drop(['name', 'id'], axis=1, inplace=True)

                        # Save the updated DataFrame back to the same Excel file
                        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                            merged_df.to_excel(writer, sheet_name=reporting_codes_sheet_name, index=False)
                        if na_count > 0:
                            print(f"\n    ❗️ Count of Codes  'Not found': {na_count}")
                        else:
                            print(f"\n    ✅ All Codes are Present")


                    except FileNotFoundError:
                        print(f"\n    ❌ Error: File '{file_path}' not found.")
                        sys.exit()
                    except KeyError as e:
                        print(f"\n    ❌ {str(e)}")
                        sys.exit()
                    except Exception as e:
                        print(f"\n    ❌ An error occurred: {str(e)}")
                        sys.exit()

                vlookup_operation(file_path, reporting_codes_sheet_name, reporting_codes_copy_sheet_name)

                # ========================================================================
                # Step 15:- To get strategy Ids for tags
                # ========================================================================

                print("\n\n🔍 Step 15: To get strategy Ids for Tags...")

                tags_sheet_name = 'Tags'
                tags_copy_sheet_name = 'Tags_Copy'

                def vlookup_tags(file_path, tags_sheet_name, tags_copy_sheet_name):
                    # Check if the file exists
                    if not os.path.exists(file_path):
                        print(f"\n    ❌ Error: File '{file_path}' not found.")
                        return

                    try:
                        # Read the data from both sheets
                        tags_df = pd.read_excel(file_path, sheet_name=tags_sheet_name)
                        tags_copy_df = pd.read_excel(file_path, sheet_name=tags_copy_sheet_name)

                        # Ensure column names are standardized and lowercase
                        tags_df.columns = tags_df.columns.str.strip().str.lower()
                        tags_copy_df.columns = tags_copy_df.columns.str.strip().str.lower()

                        # Convert relevant columns to lowercase for case-insensitive merge
                        tags_df['tag'] = tags_df['tag'].str.lower()
                        tags_copy_df['name'] = tags_copy_df['name'].str.lower()

                        # Check if the necessary columns exist (case-insensitive)
                        if 'tag' not in tags_df.columns:
                            print(f"\n    ❌ Column 'tag' not found in '{tags_sheet_name}' sheet.")
                        if 'name' not in tags_copy_df.columns:
                            print(f"\n    ❌ Column 'name' not found in '{tags_copy_sheet_name}' sheet.")
                        if 'id' not in tags_copy_df.columns:
                            print(f"\n    ❌ Column 'id' not found in '{tags_copy_sheet_name}' sheet.")

                        # Perform merge using standardized column names

                        merged_df = pd.merge(tags_df,
                                            tags_copy_df[['name', 'id']],
                                            left_on='tag', right_on='name',
                                            how='left')

                        # Count NaN values before replacing them
                        na_count = merged_df['id'].isna().sum()

                        # Create a new column 'StrategyId' and fill missing values with 'Not found'
                        merged_df['StrategyId'] = merged_df['id'].fillna('Not found')

                        # Drop unnecessary columns after merging
                        merged_df.drop(['name', 'id'], axis=1, inplace=True)

                        # Save the updated DataFrame back to the same Excel file
                        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                            merged_df.to_excel(writer, sheet_name=tags_sheet_name, index=False)

                        if na_count > 0:
                            print(f"\n    ❗️ Count of tags 'Not found': {na_count}")
                        else:
                            print(f"\n    ✅ All tags are Present")

                    except FileNotFoundError:
                        print(f"\n    ❌ Error: File '{file_path}' not found.")
                        sys.exit()
                    except KeyError as e:
                        print(f"\n    ❌ {str(e)}")
                        sys.exit()
                    except Exception as e:
                        print(f"\n    ❌ An error occurred: {str(e)}")
                        sys.exit()

                # Example usage:
                # file_path = 'your_file_path.xlsx'
                vlookup_tags(file_path, tags_sheet_name, tags_copy_sheet_name)
                
                
                # ===================================================
                
                # Code to create List of tags 
                output_for_tags = csv_file_dir + "/Tags_to_be_inserted.csv" # Change to your desired output file path


                def process_excel(file_path, output_for_tags):
                    # Read the Excel file
                    xls = pd.ExcelFile(file_path)
                    
                    # Iterate over all sheets to find the relevant one
                    sheet_name = "Tags" 
                    df = pd.read_excel(xls, sheet_name)
                    
                    # Check if required columns exist
                    required_columns = {'opportunityid', 'tag', 'existing', 'concattags', 'StrategyId'}
                    if required_columns.issubset(df.columns):
                        
                        # Filter rows where StrategyId has 'Not found'
                        filtered_df = df[df['StrategyId'] == 'Not found']

                        # Drop duplicate tags
                        filtered_df = filtered_df.drop_duplicates(subset='tag')
                        
                        # Create the output DataFrame
                        output_df = pd.DataFrame({
                            'Name': filtered_df['tag'],
                            'Strategy_Full_Name__c': '',
                            'RecordTypeId': '0123h000000kqchAAA',
                            'Record_Type_Name__c': 'Tags',
                            'IsDeleted': False,
                            'Active__c': True
                        })
                        
                        # Save to new Excel file
                        if not output_df.empty:
                            output_df.to_csv(output_for_tags, index=False)
                        # print(f"Output file saved as {output_for_tags}")
                        return
                    
                    # print("No valid sheet found with the required columns.")

                # Example usage
                process_excel(file_path, output_for_tags)
                # ========================================================================
                print("\n")
                title = "📝 STRATEGY SHEET COMPLETED 📝"
                show_title(title)
                # ========================================================================

            break
        elif choice == "no":
            strategy_choice = 'no'
            print("\n        🚫 Strategy Sheet execution skipped!")

            print("\n")
            title = "📝 STRATEGY SHEET SKIPPED 📝"
            show_title(title)

            break  # Just breaking without running Block 2
        
        else:
            print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.")

    # ==========================================================================
                        # Contact Role Sheet Execution
    # ==========================================================================

    # Display the header once
    print("\n\n📄 Execute Next Sheet:")
    required_columns = {'opportunityid', 'contactid'}
    original_contact_sheet = 'Contact Roles'

    wb = load_workbook(file_path)
    if original_contact_sheet in wb.sheetnames:
        # Read the sheet
        df = pd.read_excel(file_path, sheet_name=original_contact_sheet)

        # Check if 'role' column exists
        if 'role' in df.columns:
            # Replace blank or NaN values with 'Other'
            df['role'] = df['role'].replace(r'^\s*$', np.nan, regex=True)  # Treat empty strings as NaN
            df['role'] = df['role'].fillna('Other')

            # Write back to the same sheet
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=original_contact_sheet, index=False)
            # print("Updated 'role' column successfully.")

        is_empty, preview = is_sheet_empty(file_path, original_contact_sheet)
        
        print(tabulate(preview, headers='keys', tablefmt='fancy_grid', showindex=False))


    if original_contact_sheet not in wb.sheetnames:
        print(f"\n📂 The sheet '{original_contact_sheet}' is missing.\n")
        print("\n    🚫 Contact Role sheet execution skipped!")
        contact_choice = 'no'

    elif is_empty:
        print(f"\n📂 The sheet '{original_contact_sheet}' is empty or contains only headers.\n")
        print("\n    🚫 Contact Role sheet execution skipped!")
        contact_choice = 'no'

    elif not required_columns.issubset(preview.columns.str.lower()):
        print(f"\n    🚫 Required columns {required_columns} not found in the sheet. Skipping execution!\n")
        contact_choice = 'no'

    elif preview[['opportunityid', 'contactid']].isnull().all().any():
        # Check if either column is completely NaN or blank
        print(f"\n    🚫 One of the required columns has no data. Skipping execution!\n")
        contact_choice = 'no'

    elif is_empty is None:
        print("\n    ❗️ Could not process the sheet due to an error.\n")
        contact_choice = 'no'

    else:
        while True:
            user_choice = input(f"\n🔹 Do you want to execute the Contact Roles Sheet ? (yes/no): ").strip().lower()
            
            if user_choice == 'yes':
                
                contact_choice = 'yes'

                print(f"\n        ⏳ Executing the Sheet: Contact Roles")
                
                # ======================================================================
                print("\n")
                title = "📝 CONTACT ROLES SHEET EXECUTION 📝"
                show_title(title)

                # ======================================================================

                print("\n🔍 Step 1: Checking if the file exists...")

                check_file_exists(file_path)

                # ===================================================================================
                # ===================================================================================

                print("\n🔍 Step 2: Verifying opportunities in the 'Opportunity' sheet...")

                opportunity_sheet_name = 'Opportunity'
                contact_sheet_name = 'Contact Roles'

                try:
                    # Load the sheets into DataFrames
                    all_sheets = pd.read_excel(file_path, sheet_name=None)  # Load all sheets into a dictionary
                    sheet_names = [sheet.lower() for sheet in all_sheets.keys()]  # Convert sheet names to lowercase

                    # Check if the required sheets exist (case-insensitive)
                    if opportunity_sheet_name.lower() not in sheet_names:
                        print(f"\n    ❌ Sheet '{opportunity_sheet_name}' not found. ❌")
                        sys.exit()
                    if contact_sheet_name.lower() not in sheet_names:
                        print(f"\n    ❌ Sheet '{contact_sheet_name}' not found. ❌")
                        sys.exit()

                    # Load the relevant sheets into DataFrames (case-insensitive)
                    opportunity_df = all_sheets[list(all_sheets.keys())[sheet_names.index(opportunity_sheet_name.lower())]]
                    contact_df = all_sheets[list(all_sheets.keys())[sheet_names.index(contact_sheet_name.lower())]]

                    # Validate the required columns (case-insensitive)
                    opportunity_columns = [col.lower() for col in opportunity_df.columns]
                    product_columns = [col.lower() for col in contact_df.columns]

                    if 'opportunity_legacy_id__c'.lower() not in opportunity_columns:
                        print(f"\n    ❌ Column 'opportunity_legacy_id__c' not found in the '{opportunity_sheet_name}' sheet. ❌")
                        sys.exit()
                    elif 'opportunityid'.lower() not in product_columns:
                        print(f"\n    ❌ Column 'opportunityid' not found in the '{contact_sheet_name}' sheet. ❌")
                        sys.exit()

                    # Perform the comparison
                    contact_df['existing'] = contact_df['opportunityid'].isin(opportunity_df['opportunity_legacy_id__c'])

                    # Calculate the number of false values
                    false_count = (~contact_df['existing']).sum()

                    # Save the updated data back to the Excel file
                    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                        contact_df.to_excel(writer, sheet_name=contact_sheet_name, index=False)

                    # Success message with false count
                    print(f"\n    ✅ Verification completed. 'existing' column has been added to the '{contact_sheet_name}' sheet. ✅")
                    print(f"\n    ❗️ Number of False values in 'existing' column: {false_count}")

                except FileNotFoundError:
                    # Handle file not found
                    print(f"\n    ❌ Error: File not found. ❌")
                    sys.exit()
                except Exception as e:
                    # Handle any unexpected errors
                    print(f"\n    ❌ Error: An unexpected error occurred. Details: {e} ❌")
                    sys.exit()

                # =================================

                # Read the "Contact Roles" sheet from the Excel file
                df = pd.read_excel(file_path, sheet_name='Contact Roles')

                # Print out the column names to confirm the correct name
                # print(df.columns)

                # Assuming the correct column name is found, apply the transformation
                # Adjust 'contactid' to match the correct column name from the printout
                if 'contactid' in df.columns:
                    df['contactid'] = df['contactid'].apply(lambda x: str(int(x)))
                else:
                    print("Column 'contactid' not found!")

                # Use ExcelWriter to write the changes back to the same file, replacing the existing "Contact Roles" sheet
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Contact Roles', index=False)

                print("\n")
                title = "📝 CONTACT ROLES SHEET COMPLETED 📝"
                show_title(title)

                break

            elif user_choice == 'no':
                contact_choice = 'no'
                print("\n        🚫 Contact Role sheet execution skipped!\n")
                
                print("\n")
                title = "📝 CONTACT ROLE SHEET SKIPPED 📝"
                show_title(title)
                break
            else:
                print("\n        ❗️ Invalid input. Please enter 'yes' or 'no'.\n")


    # ========================================================================
    # Rearranging the sheets!
    # ========================================================================
    print("\n")
    title = "📝 Rearranging the sheets 📝"
    show_title(title)

    #  Rearranging Sheets in Workbook
    print("\n\n📄 Rearranging Sheets in Workbook...")

    # Define the file paths
    # file_path = os.path.expanduser("~/Downloads/avi.xlsx")
    new_file_path = csv_file_dir + "/Rearranged_file.xlsx"

    # Check if the input file exists
    if not os.path.exists(file_path):
        print(f"\n    ❌ Error: The file '{file_path}' does not exist.")
    else:
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)

            # Desired order of sheets
            desired_order = [
                "opportunity",
                "opportunity_Copy",
                "opportunity_product",
                "opportunity_product_Copy",
                "Opportunity_team",
                "Opportunity_team_Copy",
                "Reporting_codes",
                "Reporting_codes_Copy",
                "Tags",
                "Tags_Copy"
            ]

            # Normalize sheet names to lower case for comparison
            sheet_names = {sheet.title.lower(): sheet for sheet in wb.worksheets}

            # List to store ordered sheets
            ordered_sheets = []

            # Track missing sheets
            missing_sheets = []

            for sheet_name in desired_order:
                # Check if the normalized name exists in the workbook
                normalized_name = sheet_name.lower()
                if normalized_name in sheet_names:
                    ordered_sheets.append(sheet_names[normalized_name])
                else:
                    missing_sheets.append(sheet_name)

            # If any sheets are missing, show the missing sheets and exit
            if missing_sheets:
                print(f"\n    ❌ The following sheets were missing and will be skipped:\n")
                for missing_sheet in missing_sheets:
                    print(f"\n        🔸 {missing_sheet}")
            else:
                # Append any remaining sheets that were not in the desired order
                remaining_sheets = [sheet for sheet in wb.worksheets if sheet not in ordered_sheets]
                ordered_sheets.extend(remaining_sheets)

                # Create a new workbook to hold the sheets in the desired order
                new_wb = openpyxl.Workbook()
                new_wb.remove(new_wb.active)  # Remove the default sheet created by openpyxl

                for sheet in ordered_sheets:
                    new_sheet = new_wb.create_sheet(title=sheet.title)
                    for row in sheet.iter_rows(values_only=True):
                        new_sheet.append(row)

                # Save the new workbook
                new_wb.save(new_file_path)

                print(f"\n    ✅ Rearranged workbook saved as '{os.path.basename(new_file_path)}'.")

        except Exception as e:
            print(f"\n    ❌ An error occurred: {str(e)}")

    # =========================================================================================================================================
    #                                                FINAL FILE EXECUTION
    # =========================================================================================================================================


    # ========================================================================
    print("\n")
    title = "📝 FINAL SHEET EXECUTION 📝"
    show_title(title)
    # ========================================================================

    # ========================================================================
    # Step 1:-Initialize file name
    # ========================================================================


    print("\n\n🔍 Initializing file name and determining output path...")

    # Check if the output directory exists and is valid
    if os.path.exists(output) and os.path.isdir(output):
        print(f"\n    ✅ Output folder selected automatically:")
        
        # Show a relative path by trimming the base directory for cleaner display
        print(f"\n        📂 {output.split(BASE_DIR, 1)[-1]}")

    else:
        # If the path doesn't exist or is invalid, show an error and print full path
        print("\n    ❌ Error: The corresponding output folder does not exist.")
        print(f"\n       ❗️ Please check the path: {output}")


    # Extract the folder name from the output path to use in naming CSV files
    file_name = output.split('/')[-1]
    
    # Dynamically generate standardized CSV file names for different sheets
    opportunity = "1" + '_Opportunity load.csv'

    opportunity_product = "2" + '_Opportunity Product load.csv'

    opportunity_team = "3" + '_Opportunity Team member Load.csv'

    reporting_codes = "4" + '_Reporting_codes.csv'

    tags = "5" + '_Tags.csv'

    Contact_role = "6" + '_Contact_Roles.csv'

    # Print the generated file names for confirmation
    print("\n    📄 CSV File Names Generated:")
    print(f"\n        1️⃣ {opportunity}")
    print(f"\n        2️⃣ {opportunity_product}")
    print(f"\n        3️⃣ {opportunity_team}")
    print(f"\n        4️⃣ {reporting_codes}")
    print(f"\n        5️⃣ {tags}")
    print(f"\n        6️⃣ {Contact_role}")


    # ======================================================================
    # Step 1: Creating Opportunity Sheet
    # ======================================================================
    print()
    print("=" * 100)
    print("\n\n🔍 CREATING OPPORTUNITY FILE")
    

    # --------------------------- Define Path for Processed data and Removed Rows Data  ----------------------

    def get_file_paths(sheet_name, output_dir, removed_rows_dir, output_filename, removed_row_filename):
        # Construct the full paths based on provided filenames
        output_file = os.path.join(output_dir, output_filename)  # Processed CSV output path
        removed_rows_file = os.path.join(removed_rows_dir, removed_row_filename)  # Removed rows CSV
        
        return output_file, removed_rows_file

    sheet_name = 'Opportunity'
    opppty_processed_file, removed_rows_oppty = get_file_paths(sheet_name, output, removed_rows_dir,opportunity,'Removed_Rows - Oppty.csv')


    # -------------------------------------------- Define Columns --------------------------------------------

    predefined_columns_oppty = [
        'Already Exist','AccountNumber', 'Email', 'created_by', 'modified_by', 'created_date',
        'modified_date', 'Trimmed_accountid', 'Trimmed_ownerid', 'Type Of Opportunity',"mpp_number__c",
        'Concatenatedaccountid', 'Concatenatedownerid', 'concatenatedcreatedby','accountid','type of opportunity','Email.1'
    ]

    # List of columns to exclude from user selection in the GUI
    excluded_columns_oppty = [
        'opportunity_legacy_id__c', 'Legacy_Opportunity_Split_Id__c', 'name', 'StageName','description',
        'Won_Reason__c', 'Lost_Category__c', 'Lost_Reason__c', 'CloseDate', 'CurrencyIsoCode',
        'OwnerId', 'NextStep', 'OI_Group__c','AccountId','createdbyid','Pricebook2Id','RecordTypeId'
    ]

    # ---------------------- Read Opportunity Sheet ----------------------
    
    opportunity_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # ---------------------- Clean DataFrame ----------------------

    def clean_dataframe(df: pd.DataFrame, drop_duplicates: bool = True) -> pd.DataFrame:
        """
        Cleans the given DataFrame by:
        - Dropping columns that are completely empty
        - Dropping rows that are completely empty
        - Optionally dropping duplicate rows

        Parameters:
            df (pd.DataFrame): The DataFrame to clean.
            drop_duplicates (bool): Whether to drop duplicate rows. Defaults to True.

        Returns:
            pd.DataFrame: A cleaned DataFrame.
        """
        df.dropna(axis=1, how='all', inplace=True)  # Drop blank columns
        df.dropna(axis=0, how='all', inplace=True)  # Drop blank rows
        if drop_duplicates:
            df.drop_duplicates(inplace=True)        # Drop duplicate rows
        return df


    # call the function to clean the DataFrame
    opportunity_df = clean_dataframe(opportunity_df)

    # Initialize DataFrame to store removed rows and reasons
    removed_rows_oppty_df = pd.DataFrame(columns=opportunity_df.columns.tolist() + ['Reason'])

    # Track all columns that are dropped (predefined + user-selected)
    dropped_columns_oppty = []

    # ---------------------- Remove Rows: Already Exist is 'Already Exist in ISC' ----------------------

    def remove_rows_based_on_condition(df, removed_rows_df , column, value,reason):
    
        # Initialize the count to 0 for tracking removed rows
        row_removed_count = 0

        if column in df.columns:
            # Identify rows to remove based on the condition
            rows_to_remove = df[df[column] == value].copy()
            
            # Count the rows that are being removed
            row_removed_count = len(rows_to_remove)  # Count the rows to be removed
            
            if not rows_to_remove.empty:
                # Add a reason for removal
                rows_to_remove['Reason'] = reason
                # Add the removed rows to the removed_rows_df
                removed_rows_df = pd.concat([removed_rows_df, rows_to_remove], ignore_index=True)
            
            # Update the original DataFrame to keep only rows that don't match the condition
            df = df[df[column] != value]    # Keep only unmatched rows

        return df, removed_rows_df, row_removed_count

    opportunity_df, removed_rows_oppty_df, count_duplicate_oppty = remove_rows_based_on_condition(opportunity_df, removed_rows_oppty_df , 'Already Exist' , "Already Exist in ISC" ,  "Duplicate Opportunity")

    # ---------------------- Delete Predefined Columns ----------------------

    def delete_predefined_columns(df, predefined_columns,dropped_columns):
            

        columns_to_delete_predefined = []
        for col in predefined_columns:
            if col in df.columns:
                columns_to_delete_predefined.append(col)

        if columns_to_delete_predefined:
            df.drop(columns=columns_to_delete_predefined, inplace=True)
            dropped_columns.extend(columns_to_delete_predefined)
        
        return df, dropped_columns
    
    # Call the function to delete predefined columns
    opportunity_df,dropped_columns_oppty = delete_predefined_columns(opportunity_df, predefined_columns_oppty,dropped_columns_oppty)

    # ---------------------- GUI for User-Selectable Column Deletion ----------------------

    def show_column_deletion_gui(df, excluded_columns, dropped_columns):
        """
        Displays a GUI for selecting columns to delete from a DataFrame.

        Parameters:
        - opportunity_df (pd.DataFrame): The DataFrame from which columns may be deleted.
        - excluded_columns (list): List of columns that should not be shown in the GUI.
        - dropped_columns (list): A list to extend with any columns deleted by the user.

        Returns:
        - pd.DataFrame: The modified DataFrame with selected columns removed.
        """
        # Filter columns to be shown in the GUI (excluding important ones)

        columns_for_ui = []
        for col in df.columns:
            if col not in excluded_columns:
                columns_for_ui.append(col)
        
        checkboxes = {}
        if columns_for_ui:
            root = Tk()
            root.title("Select Columns to Delete")
            root.geometry("500x600")
            root.resizable(False, False)

            # Scrollable canvas setup
            canvas = Canvas(root)
            scrollbar = Scrollbar(root, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas to hold checkboxes
            frame = Frame(canvas)
            canvas.create_window((0, 0), window=frame, anchor="nw")

            for column in columns_for_ui:
                var = IntVar()
                checkboxes[column] = var
                checkbutton = Checkbutton(frame, text=column, variable=var, font=('Helvetica', 12), anchor="w", padx=10)
                checkbutton.pack(anchor="w", pady=5)

            # Submit button to close GUI
            button_frame = Frame(root)
            submit_button = Button(button_frame, text="Submit", command=root.quit,
                                font=('Helvetica', 12, 'bold'), relief='flat', padx=20, pady=10)
            submit_button.pack(side="right")
            button_frame.pack(anchor="ne", padx=20, pady=10)
            frame.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))

            root.mainloop()
            root.destroy()
        else:
            print("\n    ✅ No user-selectable columns available for deletion. Skipping GUI.")

        # Delete selected columns
        columns_to_delete_from_user = []
        for col, var in checkboxes.items():
            if var.get() == 1:
                columns_to_delete_from_user.append(col)

        if columns_to_delete_from_user:
            df.drop(columns=columns_to_delete_from_user, inplace=True)
            dropped_columns.extend(columns_to_delete_from_user)
            print("\n    ✅ Additional columns deleted:")
            for col in columns_to_delete_from_user:
                print(f"\n        🔸 {col}")
        else:
            print("\n    ✅ No additional columns selected for deletion.")

        return df
    # Call the GUI function to show the column deletion options
    
    opportunity_df = show_column_deletion_gui(opportunity_df, excluded_columns_oppty, dropped_columns_oppty)

# ---------------------- Remove Rows: AccountId is 'Not in ISC' ----------------------
    
    opportunity_df, removed_rows_oppty_df, count_not_in_isc = remove_rows_based_on_condition(opportunity_df, removed_rows_oppty_df , 'AccountId' , "Not in ISC" ,  "AccountId is 'Not in ISC'")

# ---------------------- Remove Rows with Invalid PricebookEntryId ----------------------

    count_invalid_pricebook = 0

    try:
        opportunity_product_df = pd.read_excel(file_path, sheet_name='Opportunity_product')
        
        # Identify invalid rows based on PricebookEntryId
        invalid_pricebook_ids = opportunity_product_df[
            opportunity_product_df['PricebookEntryId'].isin(['Not Active', 'No Pricebookid found'])
        ]['Legacy_Opportunity_Split_Id__c'].unique()

        rows_to_remove_invalid_pricebook = opportunity_df[
            opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
        ].copy()
        count_invalid_pricebook = len(rows_to_remove_invalid_pricebook)  # Count the rows to be removed
        
        if count_invalid_pricebook > 0:
            rows_to_remove_invalid_pricebook['Reason'] = "Invalid PricebookEntryId"
            opportunity_df = opportunity_df[
                ~opportunity_df['opportunity_legacy_id__c'].isin(invalid_pricebook_ids)
            ]
            removed_rows_oppty_df = pd.concat([removed_rows_oppty_df, rows_to_remove_invalid_pricebook], ignore_index=True)

    except Exception as e:
        print(f"\n    ❌ Error processing invalid PricebookEntryId rows: {e}")

# ---------------------- Row Removal Summary ----------------------

    def print_removal_summary(reason_counts: dict, label: str = "rows"):
        total_removed = sum(reason_counts.values())

        if total_removed > 0:
            print(f"\n    ❗️ Total {label} removed: {total_removed}")
            for reason, count in reason_counts.items():
                print(f"\n        🔸 {reason}: {count}")
        else:
            print(f"\n    ✅ No {label} removed")
    
    print_removal_summary({
        "Removed due to invalid PricebookEntryId": count_invalid_pricebook,
        "Remove due Account Not in ISC": count_not_in_isc,
        "Removed due to Duplicate Opportunity": count_duplicate_oppty
    })
        
# ---------------------- Clean Removed Rows DF ----------------------
   
    def drop_dropped_columns_from_removed_rows(dropped_columns,removed_rows_df):
        columns_to_drop_from_removed = []
        for col in dropped_columns:
            if col in removed_rows_df.columns:
                columns_to_drop_from_removed.append(col)

        if columns_to_drop_from_removed:
            removed_rows_df.drop(columns=columns_to_drop_from_removed, inplace=True)

        return removed_rows_df

    removed_rows_oppty_df = drop_dropped_columns_from_removed_rows(dropped_columns_oppty,removed_rows_oppty_df)

    # ---------------------- Save Cleaned Opportunity Data ----------------------

    def save_dataframe(df, file_path, label):
        try:
            df.to_csv(file_path, index=False)
            shortened = "/".join(file_path.split("/")[-4:])
            print(f"\n    ✅ {label} data saved to:\n\n        📂 {shortened}")
        except Exception as e:
            print(f"\n    ❌ Error saving the {label.lower()} file: {e}")
    
    save_dataframe(opportunity_df, opppty_processed_file, "Opportunity")


# ---------------------- Save Removed Rows Data ----------------------

    def save_removed_rows(df_removed, removed_path, label):
        if not df_removed.empty:
            try:
                df_removed.to_csv(removed_path, index=False)
                shortened = "/".join(removed_path.split("/")[-5:])
                print(f"\n    ✅ Removed rows saved to:\n\n        📂 {shortened}")
            except Exception as e:
                print(f"\n    ❌ Error saving the removed rows file for {label.lower()}: {e}")

    save_removed_rows(removed_rows_oppty_df, removed_rows_oppty, "Removed Rows_Opportunity")

    mpp_output_file = os.path.join(csv_file_dir,"MPP_Column.xlsx")
    
    if os.path.exists(mpp_output_file):
        csv_df = pd.read_csv(opppty_processed_file, dtype={'opportunity_legacy_id__c': str})
        excel_df = pd.read_excel(mpp_output_file, dtype={'opportunity_legacy_id_c': str, 'Mpp_Number__c': str})
        # Read the files, preserve leading zeros
        csv_df = pd.read_csv(opppty_processed_file, dtype={'opportunity_legacy_id__c': str})
        excel_df = pd.read_excel(mpp_output_file, dtype={'opportunity_legacy_id_c': str, 'Mpp_Number__c': str})

        # Convert matching columns to lowercase for case-insensitive comparison
        csv_df['opportunity_legacy_id__c_lower'] = csv_df['opportunity_legacy_id__c'].str.lower()
        excel_df['opportunity_legacy_id_c_lower'] = excel_df['opportunity_legacy_id_c'].str.lower()

        # Merge the Excel Mpp_Number__c into the CSV based on lowercased IDs
        csv_df = csv_df.merge(
            excel_df[['opportunity_legacy_id_c_lower', 'Mpp_Number__c']],
            how='left',
            left_on='opportunity_legacy_id__c_lower',
            right_on='opportunity_legacy_id_c_lower'
        )

        # Drop the helper lowercase columns
        csv_df.drop(columns=['opportunity_legacy_id__c_lower', 'opportunity_legacy_id_c_lower'], inplace=True)

        # Save back to the **same CSV file** (overwrite)
        csv_df.to_csv(opppty_processed_file, index=False)

        print(f"CSV file '{opppty_processed_file}' updated successfully with 'Mpp_Number__c' values.")

    # =======================================================
    # Step 2:- Creating the Opportunity Product File
    # =======================================================
    print()
    print("=" * 100)
    print("\n\n🔍 CREATING PRODUCT FILE")

    # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------

    sheet_name = 'Opportunity_product'
    product_processed_file, removed_rows_product = get_file_paths(sheet_name, output, removed_rows_dir,opportunity_product,'Removed_Rows - Product.csv')
    
    # -------------------------------------------- Define Columns --------------------------------------------

    predefined_columns_product = [
        'existing', 'product', 'product_type', 'Product_Family__c', 
        'opportunity currency', 'practise_multiple country', 
        'quantity.1', 'concatenated product family', 'concatenated currency'
    ]

    # Define columns that should never be shown to the user for deletion
    excluded_columns_product = [
    'Type__c','Renewal_Type__c','Renewal_Status__c','Expiration_Date__c','Expiring_Term__c','Expiring_Amount__c',
    'External_IDs__c','month 1 revenue','month 2 revenue','month 3 revenue','next quarter revenue','first 12 months revenue',
    'pre-contract planned revenue','pre-contract start date','pre-contract end date','loss reason/attition reason',
    'Legacy_Opportunity_Split_Id__c','PricebookEntryId','UnitPrice','Term__c','Classification__c','Quantity'
    ]

    try:
        # ---------------------- Load Product Sheet ----------------------

        product_df = pd.read_excel(file_path, sheet_name=sheet_name)

        # ---------------------- Clean DataFrame ----------------------
        
        product_df = clean_dataframe(product_df, drop_duplicates=False)  # Clean the opportunity_df without dropping duplicates
    
        # Initialize DataFrame to track removed rows with reasons
        removed_rows_product_df = pd.DataFrame(columns=product_df.columns.tolist() + ['Reason'])
        
        # Track all columns that are dropped (predefined + user-selected)
        dropped_columns_product = []
        
        # ---------------------- Remove Rows Based on existing ----------------------
        
        # Call the function to remove non-existing rows
        product_df, removed_rows_product_df, rows_dropped_existing_count = remove_rows_based_on_condition(product_df, removed_rows_product_df,'existing',False,"Opportunity Missing From Main sheet")

        # ---------------------- Remove Rows Based on Removed Oppties ----------------------

        def load_removed_rows_oppty(removed_rows_oppty,df,removed_rows_df,column_match):
            
            rows_dropped_legacy_count = 0
            if os.path.exists(removed_rows_oppty):
                
                oppty_removed_rows_df = pd.read_csv(removed_rows_oppty, usecols=["opportunity_legacy_id__c"])
                opportunity_ids_set = set(oppty_removed_rows_df["opportunity_legacy_id__c"].dropna().astype(str))

                # Filter rows where "Legacy_Opportunity_Split_Id__c" exists in opportunity_ids_set
                removed_rows_legacy = df[df[column_match].astype(str).isin(opportunity_ids_set)].copy()
                rows_dropped_legacy_count = len(removed_rows_legacy)

                if not removed_rows_legacy.empty:
                    removed_rows_legacy["Reason"] = "Opportunity not loaded"
                    removed_rows_df = pd.concat([removed_rows_df, removed_rows_legacy], ignore_index=True)
                    df = df[~df[column_match].astype(str).isin(opportunity_ids_set)]  # Keep only unmatched rows
            return df, removed_rows_df, rows_dropped_legacy_count

        product_df,removed_rows_product_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, product_df, removed_rows_product_df, "Legacy_Opportunity_Split_Id__c")
        
        # ---------------------- Remove Predefined Columns ----------------------

        product_df, dropped_columns_product = delete_predefined_columns(product_df, predefined_columns_product, dropped_columns_product)

        # ---------------------- User-Guided Column Deletion (GUI) ----------------------
        
        product_df = show_column_deletion_gui(product_df, excluded_columns_product, dropped_columns_product)
        
        # ---------------------- Row Removal Summary ----------------------

        print_removal_summary({
            "Due to 'existing' == False": rows_dropped_existing_count,
            "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
        })
        
        # ---------------------- Clean Removed Rows DF ----------------------
        
        # Drop all dropped columns from removed_rows_df (retain only 'Reason' and important columns)
        
        removed_rows_product_df = drop_dropped_columns_from_removed_rows(dropped_columns_product,removed_rows_product_df)
        
        # ---------------------- Save Processed Product File ----------------------
        
        save_dataframe(product_df, product_processed_file, "Product")

        # ---------------------- Save Removed Rows Data ----------------------
        
        save_removed_rows(removed_rows_product_df, removed_rows_product, "Removed Rows_Product")

    except Exception as e:
        print(f"\n    ❌ An error occurred: {e}")

    # =======================================================
    # Step 3: Processing Opportunity Team Data
    # =======================================================

    # Prompt: Check if user wants to run the Opportunity Team Member Sheet processing

    while True:
        print("\n================================================================================")
        print(f'\n📄 Do you want to run the team member Sheet? (yes/no): {team_member_choice}')
        
        # Use value from previous logic (assumed auto-input)
        user_input = team_member_choice # Automatically taking the user input from the above question from file processing
        
        if user_input == "yes":
            print("\n    ⏳ Running Opportunity Team Member Sheet...")
            print("\n================================================================================")

            # =======================================================
            # MAIN CODE
            # =======================================================

            print("\n\n🔍 CREATING TEAM MEMBER FILE")

            # ------------------------- Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Opportunity_team'
            team_procesed_file, removed_rows_team = get_file_paths(sheet_name, output, removed_rows_dir,opportunity_team,'Removed_Rows - Team.csv')
            
            # ---------------------- Define Columns ----------------------

            predefined_columns_Team = ['existing', 'email', 'Concat_T_M']
            
            # Columns that should not appear in the deletion GUI
            excluded_columns_Team=['OpportunityId','OpportunityAccessLevel','TeamMemberRole','UserId']

            try:
                
                # ---------------------- Load Sheet ----------------------

                team_df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # ---------------------- Clean DataFrame ----------------------

                team_df = clean_dataframe(team_df)

                #---------------------- Initialize Removed Rows df ----------------------
                
                removed_rows_team_df = pd.DataFrame(columns=team_df.columns.tolist() + ['Reason'])
                
                # ---------------------- columns that are dropped ----------------------

                dropped_columns_team = [] 
                
                # ---------------------- Remove rows where 'existing' is False ----------------------

                team_df,removed_rows_team_df,rows_dropped_existing_count =  remove_rows_based_on_condition(team_df, removed_rows_team_df,'existing',False,"Opportunity Missing From Main sheet")
                        
                # ---------------------- Remove Rows Based on Removed Oppties ----------------------
                
                team_df,removed_rows_team_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, team_df, removed_rows_team_df, "OpportunityId")

                # ---------------------- Remove Predefined Columns ----------------------
                
                team_df, dropped_columns_team = delete_predefined_columns(team_df, predefined_columns_Team, dropped_columns_team)
                
                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------

                # Call the GUI function to show the column deletion options
                team_df = show_column_deletion_gui(team_df, excluded_columns_Team, dropped_columns_team)
                
                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })                

                # ---------------------- Clean Removed Rows DF ----------------------
                
                removed_rows_team_df = drop_dropped_columns_from_removed_rows(dropped_columns_team,removed_rows_team_df)

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(team_df, team_procesed_file, "Team Member")
                
                # ---------------------- Save Removed Rows Data ----------------------

                save_removed_rows(removed_rows_team_df, removed_rows_team, "Removed Rows_Team")
                

            except Exception as e:
            
                print(f"\n    ❌ An error occurred: {e}")
            
            break # Exit loop if "yes" block executed successfully
        
        elif user_input == "no":
            print("\n    🛑 Skipping team member Sheet...")
            print("\n================================================================================")
            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 4: Processing Reporting code
    # =======================================================

    while True:
        print("\n================================================================================")
        print(f"\n📄 Do you want to run the Reporting codes Sheet? (yes/no)?: {strategy_choice}")
        user_input = strategy_choice
        if user_input == "yes":
            print("\n    ⏳ Running Reporting codes Sheet...")
            print("\n================================================================================")

            print("\n\n🔍 CREATING REPORTING CODES FILE")
            
            # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Reporting_codes'
            code_processed_file, removed_rows_codes = get_file_paths(sheet_name, output, removed_rows_dir,reporting_codes,'Removed_Rows - ReportingCodes.csv')
            
            # -------------------------------- Define Columns --------------------------------------
                      
            predefined_columns_Reportingcode = [ 'existing', 'concatcodes',"reporting_codes"]

            # Columns excluded from user deletion selection
            excluded_columns_codes = ['opportunityid','StrategyId']


            try:
                # ---------------------- Load Sheet ----------------------
                codes_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------

                codes_df = clean_dataframe(codes_df)
                
                #---------------------- Initialize Removed Rows df ----------------------
                
                removed_rows_code_df = pd.DataFrame(columns=codes_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------

                dropped_columns_codes = [] 

                
                # ---------------------- Remove all Rows If all Codes are not found ----------------------

                # Check if all StrategyId values are 'Not Found'
                def remove_if_strategy_not_found(df, removed_rows_df, removed_rows_path,reason,predefined_columns):                    
                    if 'StrategyId' in df.columns:
                        strategy_values = df['StrategyId'].astype(str).str.lower().dropna()

                        if not strategy_values.empty and strategy_values.nunique() == 1 and strategy_values.iloc[0] == 'not found':
                            df['Reason'] = reason
                            removed_rows_df = pd.concat([removed_rows_df, df], ignore_index=True)
                            removed_rows_df.drop(columns=[col for col in predefined_columns if col in removed_rows_df.columns], inplace=True)
                            removed_rows_df.to_csv(removed_rows_path, index=False)

                            print(f"\n    ❗️ All StrategyId values are 'Not Found'. Sheet skipped and data moved to removed rows.")
                            return df, removed_rows_df, True  # Signal to break or skip
                    return df, removed_rows_df, False
                                
                codes_df,removed_rows_code_df,should_break =  remove_if_strategy_not_found(codes_df, removed_rows_code_df, removed_rows_codes,'Code not found',predefined_columns_Reportingcode)

                if should_break:
                    break
                
                # ---------------------- Remove rows where 'existing' is False ----------------------
                
                codes_df, removed_rows_code_df, rows_dropped_existing_count = remove_rows_based_on_condition(codes_df, removed_rows_code_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------


                codes_df,removed_rows_code_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, codes_df, removed_rows_code_df, "opportunityid")
                
                # ---------------------- Remove Predefined Columns ----------------------
                
                codes_df, dropped_columns_codes = delete_predefined_columns(codes_df, predefined_columns_Reportingcode, dropped_columns_codes)
       
                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------

                codes_df = show_column_deletion_gui(codes_df, excluded_columns_codes, dropped_columns_codes)

                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })      
            
                # ---------------------- Clean Removed Rows DF ----------------------                

                removed_rows_code_df = drop_dropped_columns_from_removed_rows(dropped_columns_codes,removed_rows_code_df)

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(codes_df, code_processed_file, "Reporting Codes")
                
                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_code_df, removed_rows_codes, "Removed Rows_Codes")

            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break
        elif user_input == "no":
            print("\n    🛑 Skipping Reporting Code sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # =======================================================
    # Step 5: Processing Tags sheet
    # =======================================================

    while True:
        print("\n================================================================================")
        
        print(f"\n📄 Do you want to run the Tags Sheet? (yes/no)?: {strategy_choice}")
        
        user_input = strategy_choice # Automatically take user's earlier choice
        
        if user_input == "yes":
            print("\n    ⏳ Running Tags Sheet...")
            print("\n================================================================================")

            # ---------------------- Define Path for Processed data and Removed Rows Data  ----------------------
            
            sheet_name = 'Tags'
            tag_processed_file, removed_rows_tags = get_file_paths(sheet_name, output, removed_rows_dir,tags,'Removed_Rows - Tags.csv')
            
            # -------------------------------- Define Columns --------------------------------------
            
            predefined_columns_tags = ['existing', 'concattags','tag']
            
            excluded_columns_tags = ['opportunityid','StrategyId']


            try:
                print("\n\n🔍 CREATING TAGS FILE")

                # ---------------------- Load Sheet ----------------------

                tags_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------
                
                tags_df = clean_dataframe(tags_df)
                
                #---------------------- Initialize Removed Rows df ----------------------
                removed_rows_tag_df = pd.DataFrame(columns=tags_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------

                dropped_columns_tags = [] 

                # ---------------------- Remove all Rows If all Tags are not found ----------------------
                
                tags_df,removed_rows_tag_df,should_break =  remove_if_strategy_not_found(tags_df, removed_rows_tag_df, removed_rows_tags,'Tag not found',predefined_columns_tags)

                if should_break:
                    break
                
                # ---------------------- Remove rows where 'existing' is False ----------------------
                
                tags_df, removed_rows_tag_df, rows_dropped_existing_count = remove_rows_based_on_condition(tags_df, removed_rows_tag_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------
                
                tags_df,removed_rows_tag_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, tags_df, removed_rows_tag_df, "opportunityid")
                
                # ---------------------- Remove Predefined Columns ----------------------
                
                tags_df, dropped_columns_tags = delete_predefined_columns(tags_df, predefined_columns_tags, dropped_columns_tags)

                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------
                
                # Call the GUI function to show the column deletion options
                tags_df = show_column_deletion_gui(tags_df, excluded_columns_tags, dropped_columns_tags)

                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })      
                
                # ---------------------- Clean Removed Rows DF ----------------------
                
                removed_rows_tag_df = drop_dropped_columns_from_removed_rows(dropped_columns_tags,removed_rows_tag_df)    

                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(tags_df, tag_processed_file, "Tags")
                
                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_tag_df, removed_rows_tags, "Removed Rows_Tags")


            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
            break

        elif user_input == "no":
            print("\n    🛑 Skipping Tag Sheet...")
            print("\n================================================================================")

            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")


    # =======================================================
    # Step 6: Processing Contact Role
    # =======================================================
    while True:
        print("\n================================================================================")
        print(f'\n📄 Do you want to run the Contact Role Sheet? (yes/no): {team_member_choice}')
        
        # Automatically using the user input captured earlier from file processing
        user_input = contact_choice 
        
        if user_input == "yes":
            print("\n    ⏳ Running Contact Role Member Sheet...")
            print("\n================================================================================")

            print("\n🔍 CREATING CONTACTS FILE")
            
            # ------------------------ Define Path for Processed data and Removed Rows Data  ----------------------

            sheet_name = 'Contact Roles'
            contact_processed_file, removed_rows_contact = get_file_paths(sheet_name, output, removed_rows_dir,Contact_role,'Removed_Rows - contact.csv')

            # Define hardcoded columns to delete from contact sheet

            predefined_columns_contact = ['existing']
            
            excluded_columns_contact = ["opportunityid","contactid","role"]

            
            try:
                # ---------------------- Load Product Sheet ----------------------

                contact_df = pd.read_excel(file_path, sheet_name=sheet_name)

                # ---------------------- Clean DataFrame ----------------------
                
                contact_df = clean_dataframe(contact_df)

                #---------------------- Initialize Removed Rows df ----------------------
                removed_rows_contact_df = pd.DataFrame(columns=contact_df.columns.tolist() + ['Reason'])

                # ---------------------- columns that are dropped ----------------------
                
                dropped_columns_contacts = []
                
                # ---------------------- Remove rows where 'existing' is False ----------------------

                contact_df, removed_rows_contact_df, rows_dropped_existing_count = remove_rows_based_on_condition(contact_df, removed_rows_contact_df,'existing',False,"Opportunity Missing From Main sheet")

                # ---------------------- Remove Rows Based on Removed Oppties ----------------------

                contact_df,removed_rows_contact_df,rows_dropped_legacy_count = load_removed_rows_oppty(removed_rows_oppty, contact_df, removed_rows_contact_df, "opportunityid")

                # ---------------------- Remove Predefined Columns ----------------------
                
                contact_df, dropped_columns_contacts = delete_predefined_columns(contact_df, predefined_columns_contact, dropped_columns_contacts)

                # ---------------------- GUI: Let User Select Additional Columns to Delete ----------------------
                
                contact_df = show_column_deletion_gui(contact_df, excluded_columns_contact, dropped_columns_contacts)
                
                # ---------------------- Row Removal Summary ----------------------

                print_removal_summary({
                    "Due to 'existing' == False": rows_dropped_existing_count,
                    "Due to 'Opportunity not loaded'": rows_dropped_legacy_count
                })   
                
                # ---------------------- Clean Removed Rows DF ----------------------   

                removed_rows_contact_df = drop_dropped_columns_from_removed_rows(dropped_columns_contacts,removed_rows_contact_df)    
            
                # ---------------------- Save Processed Output ----------------------
                
                save_dataframe(contact_df, contact_processed_file, "Contact")

                # ---------------------- Save Removed Rows Data ----------------------
                
                save_removed_rows(removed_rows_contact_df, removed_rows_contact, "Removed Rows_Contact")

        
            except ValueError as ve:
                print(f"\n    ❌ ValueError: {ve}")
        
            except Exception as e:
                print(f"\n    ❌ An error occurred: {e}")
        
            break  # Exit loop if "yes" block executed successfully
        
        elif user_input == "no":
            print("\n    🛑 Skipping Contact Role Sheet...")
            print("\n================================================================================")
            break
        else:
            print("\n    ❗️ Invalid response. Please enter 'yes' or 'no'.")

    # ========================================================================
    # Last Step: Copy the Summary File to the Folder
    # ========================================================================

    print("\n\n🔍 Copying the Summary File to the Selected Folder...")

    # Check if the reference summary file exists at the specified location
    if not os.path.exists(REF_SUMMARY_FILE_PATH):
        print("\n    ❌ Error: Reference file does not exist at the specified path.")
        print(f"\n       📂 Path: {REF_SUMMARY_FILE_PATH}\n")
    else:
        try:
            # Extract the name of the selected folder (from the output path)
            selected_folder_name = os.path.basename(output)

            # Create the destination file path and rename the copied file accordingly
            destination_file_path = os.path.join(output, f"{selected_folder_name}_summary file.xlsx")

            # Copy the reference file to the destination location
            shutil.copy(REF_SUMMARY_FILE_PATH, destination_file_path)
            print(f"\n    ✅ Reference file copied successfully to the folder: {selected_folder_name}")

            # Load the copied file using openpyxl
            wb = load_workbook(destination_file_path)

            # Check if the 'Summary' sheet exists in the copied workbook
            if "Summary" in wb.sheetnames:
                ws = wb["Summary"]  # Access the "Summary" sheet

                # Write the selected folder name into cell D4
                ws['D4'] = selected_folder_name

                # Save the updated file (ensuring no other changes are made)
                wb.save(destination_file_path)
                print(f"\n    ✅ Folder name '{selected_folder_name}' written to cell D4 in the 'Summary' sheet.")
            else:
                print("\n    ❌ Error: 'Summary' sheet not found in the copied file.")
                print(f"\n       ❗️ Please check the file structure.")

        except Exception as e:
            print("\n    ❌ An error occurred while copying or modifying the file:")
            print(f"\n    ❗️ Details: {e}")


    # ========================================================================
    print("\n")
    title = "📝 FINAL SHEET COMPLETED 📝"
    show_title(title)
    # ========================================================================

    # =====================================================
    # Delete CSV Files
    # =====================================================
    # Hardcoded directory


    print("\n\n🔍 Delete the extract files")

    response = 'yes'
    # response = 'no'
    if response == 'yes':
        # Get list of files in the DOWNLOADS_DIR
        def delete_folder(folder_path):
            if os.path.exists(folder_path):
                shutil.rmtree(folder_path)
                print(f"\n        🗑️ Folder '{folder_path.split('/')[-1]}' and its contents have been deleted.")
            else:
                print(f"\n        ❗️ Folder '{folder_path}' does not exist.")


        delete_folder('Extracts')
        delete_folder('Delete')
    else:
        print("\n        🛑 No files were deleted.")

    print("\n")
    title = "✅ File Prepared: {filename} ✅"
    show_title(title)
    

    files_in_copy_folder.remove(files_in_copy_folder[selected_index])

    if files_in_copy_folder:
        while True:  # Inner loop
            continue_processing = input("\n 👉 Do you want to process another file? (yes/no): ").strip().lower()
            
            if continue_processing == 'yes':
                # Add logic for processing another file here
                print("\n    ⏳ Processing the file...")
                break  # Exit the inner loop and continue to the next iteration of the outer loop
            
            elif continue_processing == 'no':
                print(f"\n     🔚 End of Script\n")
                print("=" * 100)
                print("\n")
                break  # Exit the inner loop
            else:
                print("\n    ❗️ Invalid input. Please select 'yes' or 'no'.")
        
        if continue_processing == 'no':
            break  # Exit the outer loop if the user selects 'no'
    else:
        print(f"\n     🔚 End of Script\n")
        print("=" * 100)
        print("\n")
        break  # Exit the inner loop
