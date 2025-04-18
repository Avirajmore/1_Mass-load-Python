
import openpyxl
import os
import pandas as pd
import tkinter  as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import SheetTitleException
import sys

print(f"\n{"="*100}")

# =================================================================
# Step 1: Select a File to Process
# =================================================================

print("\n\n🔍 Step 1: Select a File to Process")

# Set the directory to search for Excel files
directory = os.path.expanduser("~/Downloads")

# Create a hidden root window (used for file dialog)
root = tk.Tk()
root.withdraw()

# Ask the user to select an Excel file
file_path = filedialog.askopenfilename(
    initialdir=directory,
    title="📄 Select an Excel file",
    filetypes=[("Excel files", "*.xlsx")]
)

# Print the selected file path
if file_path:
    filename = str(file_path.split('/')[-1])
    print(f"\n    ✅ File selected: '{filename}'.")
else:
    print("\n    ❌ No file selected. Exiting the program. ❌")
    sys.exit()

# =================================================================
# Step 2:- Renames the sheet to correct name if not present
# =================================================================

print("\n\n🔍 Step 2:- Check if all the Required Sheets are present or not")

# Path to the Excel file
# file_path = 'your_excel_file.xlsx'

# Load the Excel workbook
wb = openpyxl.load_workbook(file_path)

# List of required sheets (Tags is optional)
required_sheets = ['Opportunity', 'Opportunity_product', 'Opportunity_team', 'Reporting_codes', 'Tags']

# Get the list of sheets in the workbook
sheets_in_file = wb.sheetnames

# Check for missing required sheets (except Tags)
missing_sheets = [sheet for sheet in required_sheets if sheet != 'Tags' and sheet not in sheets_in_file]


if not missing_sheets:
    print("\n    ✅ All required sheets are already present! 🎉")
else:
    print("\n    ❌ The following required sheets are missing: ")
    for i, sheet in enumerate(missing_sheets, 1):
        print(f"\n        {i}. {sheet}")
    # Get the list of existing sheets that are not already required sheets
    available_sheets = [s for s in sheets_in_file if s not in required_sheets]
    
    if available_sheets:
        print("\n    📋 Here are the available sheets to rename: ")
        # Display the available sheets as a numbered list
        for i, s in enumerate(available_sheets, 1):
            print(f"\n        {i}. {s}")
    
    # If any required sheet is missing, ask for renaming
    for sheet in missing_sheets:
        while True:
            choice = input(f"\n    🔸 Enter the index of the sheet to rename to ' {sheet} ' or type 'skip': ")

            if choice.lower() == 'skip':
                print(f"\n        ⏭️  Skipped renaming '{sheet}'!")
                break  # Skip renaming this sheet
            try:
                # Convert the choice to an integer if it's a number
                choice = int(choice)
                if 1 <= choice <= len(available_sheets):
                    rename_sheet = available_sheets[choice - 1]
                    # Rename the selected sheet
                    ws = wb[rename_sheet]
                    ws.title = sheet
                    print(f"\n        ✅ Sheet '{rename_sheet}' renamed to '{sheet}' successfully! 🎉")
                    break  # Exit the loop after successful renaming
                else:
                    print("\n        ❗ Invalid number selected. Please choose a valid option.")
            except ValueError:
                print("\n        ❗ Invalid input, please enter a valid number or 'skip'. 😕")
    
    # Save the modified workbook (if any renaming was done)
    wb.save(file_path)
    print("\n    💾 Workbook saved with changes!")

# =================================================================
# Just in case if the files have similar name but different casing
# =================================================================

# Mapping of old sheet names to new sheet names
sheet_name_mapping = {
    'Opportunity1': 'Opportunity',
    'Opportunity_product1': 'Opportunity_product',
    'Opportunity_team1': 'Opportunity_team',
    'Reporting_codes1': 'Reporting_codes'
}

# Load the Excel file
# file_path = 'your_excel_file.xlsx'  # Replace with the actual file path
wb = openpyxl.load_workbook(file_path)

# Iterate through all sheets in the workbook
for sheet in wb.sheetnames:
    if sheet in sheet_name_mapping:
        # Rename sheet if it matches the mapping
        new_name = sheet_name_mapping[sheet]
        ws = wb[sheet]
        ws.title = new_name

# Save the workbook with the renamed sheets (the content will remain unchanged)
wb.save(file_path)

# ======================================================================
# Step 3:Renaming the columns correctly
# ======================================================================

print("\n\n🔍 Step 3: Changing the product columns names ")

# Load the Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    # Read the specific sheet
    df = pd.read_excel(file_path, sheet_name="Opportunity_product")

    # Drop the "Product" column if it exists
    if 'Product' in df.columns:
        df.drop(columns=['Product'], inplace=True)

    # Rename "product_code_family" to "Product" if it exists
    if 'product_code_family' in df.columns:
        df.rename(columns={'product_code_family': 'Product'}, inplace=True)

    # Write back to the same sheet
    df.to_excel(writer, sheet_name="Opportunity_product", index=False)

print("\n    ✅ Column updated and file saved successfully.")


# ======================================================================
# Step 4: Changing the values in Renewal status according to API name
# ======================================================================

print("\n\n🔍 Step 4: Changing the values in Renewal status according to API name...")
import pandas as pd
import re

# Load the Excel file
# file_path = "your_file.xlsx"  # Change this to your actual file path
sheet_name = "Opportunity_product"

# Read the Excel file
df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

# Rename columns while handling case sensitivity
df.columns = [col.strip().lower() for col in df.columns]  # Normalize column names
rename_mapping = {
    "renewal status/attrition indicator": "renewal status",
    "expiring amount/expiring tcv": "expiring amount"
}
df.rename(columns={k.lower(): v.lower() for k, v in rename_mapping.items()}, inplace=True)

# Define mapping for replacement
renewal_mapping = {
    "propose": "RENEW_PROPOSE",
    "negotiate": "RENEW_NEGOTIATE",
    "design": "RENEW_DESIGN",
    "closing": "RENEW_CLOSING",
    "lost": "RENEW_LOST",
    "won": "RENEW_WON"
}

# Define the valid renewal statuses
valid_renewal_statuses = [
    "RENEW_PROPOSE",
    "RENEW_NEGOTIATE",
    "RENEW_DESIGN",
    "RENEW_CLOSING",
    "RENEW_LOST",
    "RENEW_WON"
]

# Function to clean and replace renewal status values
def format_renewal_status(value):
    if isinstance(value, str) and "renew" in value.lower():
        for key, replacement in renewal_mapping.items():
            if re.search(rf"\b{key}\b", value, re.IGNORECASE):
                return replacement
    return value  # Keep unchanged if no match

# Apply transformation to "renewal status" column
df["renewal status"] = df["renewal status"].apply(format_renewal_status)

# Check for invalid renewal status values (not in valid ones, and not blank)
invalid_count = df["renewal status"].apply(lambda x: x not in valid_renewal_statuses and pd.notna(x)).sum()

# If there are invalid values, ask for confirmation to continue or terminate
if invalid_count > 0:
    while True:
        response = input(f"\n    ❗️ There are {invalid_count} invalid values in the 'renewal status' column. Do you want to continue (y/n)? ").strip().lower()
        if response == 'y':
            print("\n       ✅ Continuing with the process.")
            break
        elif response == 'n':
            print("\n       ❌ Terminating the program.")
            exit()  # Terminates the program
        else:
            print("\n       ❗️ Invalid response. Please type 'y' to continue or 'n' to terminate.")

# Save the modified data back to the same file
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n    ✅ Excel file updated successfully.")

