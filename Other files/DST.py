from openpyxl import load_workbook
import pandas as pd
import csv

# ===========================================
# Define the File path
# ===========================================
input_file = "Other files/input.txt"
output_file = "Other files/output.txt"

# ===========================================
# To read the content of input file
# ===========================================
with open(input_file , 'r') as file:
     a = file.readlines()

# ===========================================
# To erase all the content in the output file
# ===========================================
with open(output_file, "w") as file:
    pass  # This leaves the file empty

# ===========================================
# Main code
# ===========================================
lst= []

for i in a:
    i = i.strip()  # Remove leading/trailing whitespace including blank lines
    if not i:      # Skip empty lines
        continue
    
    if 'https:' in i:                               # To identify if it is URL or just an id
        id = i.split('/')[-2]
        print (id)
        lst.append(id)        
        with open (output_file, 'a') as file:        # Save the ids in ouput file
            file.write(f"{id}\n")

    else:
        id = i.split("\n")[0]        
        print(id)        
        lst.append(id)        
        with open (output_file, 'a') as file:          # Save the ids in ouput file
            file.write(f"{id}\n")

print("\n\n")
concatenated_values = str(tuple(lst))
Stripped_values  = concatenated_values.strip("()")
print (f'\n\n{Stripped_values}\n\n')

# ====================================================================================================
import csv
import pandas as pd
import os
from datetime import datetime

# Lost Reason Mapping

LOST_CATEGORY_MAP = {
    "Customer did not pursue": "LCDNP",
    "IBM did not pursue": "LIBMDNP",
    "Lost to competition": "LLTC"
}

# Lost Category Mapping
LOST_REASON_MAP = {
    "Bid / consulting delivery resources not available": "IBMRESC",
    "BP deal registration expired": "IBMREXP",
    "Business priority changed": "CUSPRIOR",
    "Client took in-house or status quo": "CUSSTATU",
    "Duplicated by another team": "IBMDUP",
    "Entered in error": "IBMERR",
    "Executive sponsorship changed": "CUSSPONS",
    "Insufficient customer budget": "IBMBUDG",
    "Lack of supply": "IBMSUPP",
    "Low budget": "CUSTBUDG",
    "Low odds of winning": "IBMLOW",
    "Migration/technical complexity/risk": "CUSTRISK",
    "No compelling reason to act": "CUSTACT",
    "Undesirable business": "IBMUNDES",
    "Unresponsive to our efforts": "IBMUNRSP"
}


# Function to process ID changes (for Opportunity and Deployment ownership)
def process_ids(txt_file, csv_file, excel_file, changes_made,notes):
    """Processes IDs from the text file and updates both CSV & Excel based on user inputs."""
    try:
        # Read IDs from text file
        with open(txt_file, "r") as f:
            ids = [line.strip() for line in f.readlines() if line.strip()]
        
        if not ids:
            print("No IDs found in the file.")
            return

        # Get user inputs
        owner_id = input("Enter Owner ID: ").strip()
        ticket_number = input("Enter Ticket Number: ").strip()

        if not owner_id or not ticket_number:
            print("Owner ID and Ticket Number cannot be empty.")
            return

        # Write to CSV file
        file_exists = os.path.exists(csv_file)

        with open(csv_file, "a", newline="") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["ID", "OwnerID"])  # Header
            
            for id_ in ids:
                writer.writerow([id_, owner_id])

        # Update Excel file with ticket number and changes made
        update_excel(excel_file, ticket_number, ids, changes_made,notes)

        print(f"Processed and updated {csv_file} and {excel_file}.")

    except Exception as e:
        print(f"Error: {e}")


def merge_adjacent_cells(filename, sheet_name, column_letter):
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    column_index = ws[column_letter + "1"].column  # Get column index
    start_row = None  # Track the starting row for merging
    prev_value = None  # Track previous cell value
    
    for row in range(2, ws.max_row + 1):  # Assuming the first row is a header
        cell = ws.cell(row=row, column=column_index)
        
        if cell.value == prev_value:
            if start_row is None:
                start_row = row - 1  # Mark the start of merging
        else:
            if start_row is not None:
                ws.merge_cells(start_row=start_row, start_column=column_index,
                               end_row=row - 1, end_column=column_index)
                start_row = None  # Reset start row
        
        prev_value = cell.value  # Update previous value
    
    # Merge the last group if necessary
    if start_row is not None:
        ws.merge_cells(start_row=start_row, start_column=column_index,
                       end_row=ws.max_row, end_column=column_index)
    
    wb.save(filename)


# Function to append data to the Excel file
def update_excel(excel_file, ticket_number, ids, changes_made,notes):
    """Appends data to the common Excel file with dynamically changing 'Changes Made'."""
    today_date = datetime.today().strftime("%Y-%m-%d")  # Get today's date

    try:
        if os.path.exists(excel_file):
            existing_df = pd.read_excel(excel_file, sheet_name="Sheet1", engine="openpyxl")
        else:
            existing_df = pd.DataFrame(columns=["Date", "Ticket Number", "ID", "Changes Made"])
    except ValueError:
        existing_df = pd.DataFrame(columns=["Date", "Ticket Number", "ID", "Changes Made"])
    
    new_df = pd.DataFrame({
        "Date": [today_date] * len(ids),
        "Ticket Number": [ticket_number] * len(ids),
        "ID": ids,
        "Changes Made": [changes_made] * len(ids),
        "Notes":notes
    })
    
    final_df = pd.concat([existing_df, new_df], ignore_index=True)
    
    with pd.ExcelWriter(excel_file, mode="w", engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Sheet1", index=False)



# Function to process "Opportunity Closed as Lost"
def close_opportunity(txt_file, csv_file, excel_file,notes,lost_reason_api,lost_category_api):
    """Processes 'Opportunity Closed as Lost' by selecting Lost Reason and Lost Category and updating both CSV & Excel."""
    try:
        # Read IDs from text file
        with open(txt_file, "r") as f:
            ids = [line.strip() for line in f.readlines() if line.strip()]
        
        if not ids:
            print("No IDs found in the file.")
            return

        # Get Ticket Number
        ticket_number = input("Enter Ticket Number: ").strip()
        if not ticket_number:
            print("Ticket Number cannot be empty.")
            return

        # Write to CSV file
        file_exists = os.path.exists(csv_file)
        
        with open(csv_file, "a", newline="") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["ID", "StageName", "Lost_Reason__c", "Lost_Category__c"])  # Header
            
            for id_ in ids:
                writer.writerow([id_, "Lost", lost_reason_api, lost_category_api])  # "Lost" instead of "Closed - Lost"

        # Update Excel file with ticket number and changes made
        update_excel(excel_file, ticket_number, ids, "Opportunity Closed as Lost",notes)

        print(f"Processed and updated {csv_file} and {excel_file}.")

    except Exception as e:
        print(f"Error: {e}")

def main():
    """Main function to handle user process selection dynamically."""
    txt_file = "Other files/output.txt"
    excel_file = "/Users/avirajmore/Downloads/Ticket.xlsx"  # Fixed Excel file

    while True:
        print("\nWhich process do you want to follow?")
        print("1. Change Oppty Ownership")
        print("2. Change Deployment Ownership")
        print("3. Close Oppty as Lost")
        print("4. Skip (Exit)")

        choice = input("Enter the option number: ").strip()
        
        if choice == "1":
            approval = input("Who gave the approval?")
            notes = f'DST Update:-\n\nThe given Opportunities ownership has been changed  based on the request and approval provided.\nCurrent OO is inactive\n\nOpportunities:- \n\nApproved By:-{approval} \n	\nThank you!'
            csv_file = "/Users/avirajmore/Downloads/Oppty_owner_change.csv"
            changes_made = "Opportunity Ownership Changed"
            process_ids(txt_file, csv_file, excel_file, changes_made,notes)
            break

        elif choice == "2":

            notes = f'DST Update:-\n\nThe given Deployments ownership has been changed based on the request and approval provided.\nCurrent OO is inactive\n\nDeployments:-\n\nThank you!'
            csv_file = "/Users/avirajmore/Downloads/Deployment_owner_change.csv"
            changes_made = "Deployment Ownership Changed"
            process_ids(txt_file, csv_file, excel_file, changes_made,notes)
            break

        elif choice == "3":

            # Choose Lost Category
            print("\nSelect Lost Category:")
            for i, category in enumerate(LOST_CATEGORY_MAP.keys(), start=1):
                print(f"{i}. {category}")

            category_choice = input("Enter the option number: ").strip()
            lost_category = list(LOST_CATEGORY_MAP.keys())[int(category_choice) - 1]
            lost_category_api = LOST_CATEGORY_MAP[lost_category]

            # Choose Lost Reason
            print("\nSelect Lost Reason:")
            for i, reason in enumerate(LOST_REASON_MAP.keys(), start=1):
                print(f"{i}. {reason}")

            reason_choice = input("\nEnter the option number: ").strip()
            lost_reason = list(LOST_REASON_MAP.keys())[int(reason_choice) - 1]
            lost_reason_api = LOST_REASON_MAP[lost_reason]

            approval = input("\nWho gave the approval?")
            notes = f'DST Update:-\nThe given Opportunities has been closed as Lost  based on the request and approval provided.\nCurrent OO is inactive\n\nOpportunities:- \n\nLost Category:-{lost_category} \nLost Reason:-{lost_reason}\nApproved By:-{approval} \n\nThank you!'
            csv_file = "/Users/avirajmore/Downloads/Oppty_closed_lost.csv"
            close_opportunity(txt_file, csv_file, excel_file,notes,lost_reason_api,lost_category_api)
            break

        elif choice == "4":
            print("Exiting program.")
            break

        else:
            print("Invalid choice. Please enter a valid option.")

if __name__ == "__main__":
    main()
